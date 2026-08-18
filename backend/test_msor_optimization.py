from __future__ import annotations

from array import array
from statistics import median
from types import SimpleNamespace
from typing import Optional
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from . import msor_converter as converter


def _reference_detect_break_from_series_km(
    values: list[float],
    curve_max_km: Optional[float],
    radius: int = 8,
) -> Optional[float]:
    """Original detector kept here as an output-equivalence oracle."""
    if curve_max_km is None or not values or len(values) < 80:
        return None

    smoothed = converter._moving_average([float(v) for v in values], radius=radius)
    tail = smoothed[int(len(smoothed) * 0.88):]
    pre = smoothed[:max(30, int(len(smoothed) * 0.60))]
    if not tail or not pre:
        return None

    floor = median(tail)
    plateau = converter._percentile(pre, 90)
    dynamic = plateau - floor
    if dynamic <= 0:
        plateau = max(pre)
        dynamic = plateau - floor
    if dynamic <= 0:
        return None

    threshold = floor + 0.18 * dynamic
    stable_count = max(16, int(len(smoothed) * 0.01))
    start_index = max(10, int(len(smoothed) * 0.05))

    target_index = None
    for idx in range(start_index, len(smoothed) - stable_count):
        window = smoothed[idx:idx + stable_count]
        if sum(value <= threshold for value in window) >= max(stable_count - 2, 1):
            target_index = idx
            break

    if target_index is None:
        diffs = [smoothed[i] - smoothed[i - 1] for i in range(1, len(smoothed))]
        search_from = max(1, int(len(diffs) * 0.20))
        target_index = min(range(search_from, len(diffs)), key=lambda i: diffs[i])

    break_km = (target_index / max(len(smoothed) - 1, 1)) * curve_max_km
    return round(break_km + 1e-12, 3)


class BreakDetectorOptimizationTests(unittest.TestCase):
    def test_rolling_detector_matches_original_detector(self) -> None:
        series_cases = [
            [5.0] * 120,
            [80.0] * 90 + [0.0] * 110,
            [80.0] * 90 + [0.0, 0.0, 40.0, 0.0] + [0.0] * 106,
            [100.0 - (index * 0.2) for index in range(240)],
            [
                55.0 if index < 130 else (2.0 if index % 11 else 24.0)
                for index in range(300)
            ],
        ]

        for values in series_cases:
            with self.subTest(points=len(values), tail=values[-5:]):
                self.assertEqual(
                    converter._detect_break_from_series_km(values, 42.5),
                    _reference_detect_break_from_series_km(values, 42.5),
                )


class CacheSafetyTests(unittest.TestCase):
    def test_bounded_cache_evicts_least_recently_used_item(self) -> None:
        cache = converter._BoundedCache(max_items=2)
        cache["a"] = 1
        cache["b"] = 2
        self.assertEqual(cache["a"], 1)

        cache["c"] = 3

        self.assertEqual(list(cache.keys()), ["a", "c"])
        self.assertNotIn("b", cache)

    def test_parse_cache_key_keeps_file_name_identity(self) -> None:
        raw = b"same content"
        self.assertNotEqual(
            converter._fr_cache_key("trace-a.sor", raw),
            converter._fr_cache_key("trace-b.sor", raw),
        )

    def test_excluded_output_sheets_are_removed_from_workbook(self) -> None:
        workbook = Workbook()
        for title in converter.EXCLUDED_OUTPUT_SHEETS:
            workbook.create_sheet(title)

        converter._fr_remove_excluded_output_sheets(workbook)

        self.assertTrue(
            converter.EXCLUDED_OUTPUT_SHEETS.isdisjoint(workbook.sheetnames)
        )


class StvSectionsWorkbookTests(unittest.TestCase):
    def test_stv_output_reuses_fastreporter_sections_sheet(self) -> None:
        summary = converter.FileSummary(
            file_name='fixture.sor',
            fiber='Fiber 1',
            wavelength_display='1550 nm',
            total_loss_db=0.2,
            length_km=1.0,
            attenuation_dbkm=0.2,
            splice_points=[],
            end_distance_km=1.0,
            graph_end_km=1.0,
            graph_curve_max_km=1.0,
            source_format='SOR',
            parsed_total_loss_db=0.2,
            route_corrected_total_loss_db=None,
            loss_source_used='fixture',
        )
        contexts = {
            'fixture.sor': {
                'events': [],
                'metadata': {'fiber_id': 'Fiber 1'},
            }
        }

        output = converter._stv_build_workbook(
            [summary],
            contexts,
            [],
            deviation_m=5.0,
            threshold_db=0.5,
            expected_route_km=None,
            jumper_excluded_m=0.0,
            graph_reach_tolerance_km=0.3,
            event_shortfall_tolerance_km=0.3,
            skipped=[],
            sections=[{'start_km': 0.0, 'end_km': 1.0, 'length_km': 1.0}],
            section_match_tolerance_m=100.0,
            section_measurement_mode='fit',
        )

        workbook = load_workbook(output, data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            [
                'Bảng sự kiện',
                'Sections',
                'Kiểm tra đồ thị',
                'Tệp bỏ qua',
                'Strict Validation',
            ],
        )
        sheet = workbook['Sections']
        self.assertEqual(sheet['C1'].value, 'Section 1')
        self.assertEqual(sheet['C2'].value, 1)
        self.assertEqual(sheet['D2'].value, 'ID S+E 1.000')
        self.assertEqual(sheet['C4'].value, 0.2)
        self.assertEqual(sheet['D4'].value, 0.2)
        self.assertEqual(sheet.freeze_panes, 'C4')
        self.assertEqual(contexts['fixture.sor']['section_pairs'], [(0.2, 0.2)])


class TotalLossSelectionTests(unittest.TestCase):
    @staticmethod
    def _summary(
        *,
        total_loss_db: float,
        route_candidate_db: float,
        source: str,
        reason: str,
    ) -> converter.FileSummary:
        return converter.FileSummary(
            file_name='s37.sor',
            fiber='Fiber 37',
            wavelength_display='1550 nm',
            total_loss_db=total_loss_db,
            length_km=42.242,
            attenuation_dbkm=round(total_loss_db / 42.242, 3),
            splice_points=[],
            end_distance_km=42.242,
            graph_end_km=42.242,
            graph_curve_max_km=42.242,
            source_format='SOR',
            parsed_total_loss_db=17.743,
            route_corrected_total_loss_db=route_candidate_db,
            loss_source_used=source,
            total_loss_selection_reason=reason,
        )

    def test_selected_total_does_not_use_rejected_route_candidate(self) -> None:
        summary = self._summary(
            total_loss_db=17.743,
            route_candidate_db=11.272,
            source='Tóm tắt SOR / điểm cuối sợi',
            reason='Giữ giá trị đã đọc vì hợp lệ.',
        )

        self.assertEqual(converter._selected_total_loss_db(summary), 17.743)

    def test_graph_check_sheet_writes_selection_reason_column(self) -> None:
        reason = 'Giữ giá trị đã đọc vì hợp lệ.'
        summary = self._summary(
            total_loss_db=17.743,
            route_candidate_db=11.272,
            source='Tóm tắt SOR / điểm cuối sợi',
            reason=reason,
        )
        workbook = Workbook()
        sheet = workbook.active

        converter._stv_fill_graph_check_sheet(
            sheet,
            [summary],
            {'s37.sor': {}},
        )

        self.assertEqual(sheet.cell(1, 19).value, 'Lý do chọn suy hao tổng')
        self.assertEqual(sheet.cell(2, 19).value, reason)

    def test_section_fallback_uses_selected_total(self) -> None:
        summary = self._summary(
            total_loss_db=17.743,
            route_candidate_db=11.272,
            source='Tóm tắt SOR / điểm cuối sợi',
            reason='Giữ giá trị đã đọc vì hợp lệ.',
        )
        summary.attenuation_dbkm = None

        loss, attenuation, method, _note = converter._section_span_attenuation_estimate(
            summary,
            0.0,
            10.0,
        )

        self.assertEqual(loss, 4.2003)
        self.assertEqual(attenuation, 0.42)
        self.assertEqual(method, 'span_attenuation_estimate')


class StandardSorTotalLossCorrectionTests(unittest.TestCase):
    @staticmethod
    def _event(
        event_no: int,
        event_type: str,
        distance_km: float,
        loss_db: Optional[float],
        slope_dbkm: Optional[float],
        label: str,
        *,
        reflectance_db: Optional[float] = 0.0,
        total_loss_db: Optional[float] = None,
    ) -> converter.EventRow:
        return converter.EventRow(
            file_name='fixture.sor',
            event_no=str(event_no),
            event_type=event_type,
            distance_m=distance_km * 1000.0,
            distance_km=distance_km,
            wavelength_nm='1550',
            loss_db=loss_db,
            reflectance_db=reflectance_db,
            slope_dbkm=slope_dbkm,
            total_loss_db=total_loss_db,
            note_original='',
            label=label,
        )

    def test_shared_filter_rejects_vendor_sentinel_in_stv_and_route(self) -> None:
        rows = [
            self._event(1, 'First Connector', 0.020, 0.0, 0.0, '1F9999LS'),
            self._event(2, 'Connector', 0.654, 0.0, 0.133, '1F9999LS', reflectance_db=-67.0),
            self._event(3, 'Splice', 6.269, 30.0, 32.767, '0F9999LS'),
            self._event(4, 'Fiber End', 6.289, 0.0, 0.0, '1E9999LS', reflectance_db=-87.0),
        ]

        route = converter._estimate_route_total_loss_details_from_standard_sor(rows, 6.289)
        stv = converter._estimate_stv_like_total_loss_from_standard_sor(rows, 6.289)
        anomaly = converter._find_preterminal_total_loss_anomaly(rows, 6.289)

        self.assertEqual(route.total_loss_db, 0.084)
        self.assertEqual(stv, 0.084)
        self.assertEqual(route.rejected_artifact_count, 1)
        self.assertEqual(route.confidence, 'low')
        self.assertLess(route.slope_coverage_ratio, 0.11)
        self.assertEqual(anomaly['reason_code'], 'sentinel_slope_high_loss')

    def test_low_coverage_fallback_is_not_auto_selected(self) -> None:
        rows = [
            self._event(1, 'First Connector', 0.020, 0.0, 0.0, '1F9999LS'),
            self._event(2, 'Connector', 0.654, 0.0, 0.133, '1F9999LS', reflectance_db=-67.0),
            self._event(3, 'Splice', 6.269, 30.0, 32.767, '0F9999LS'),
            self._event(4, 'Fiber End', 6.289, 0.0, 0.0, '1E9999LS', reflectance_db=-87.0, total_loss_db=0.084),
        ]
        meta = {
            'length_km': 6.289,
            'total_loss_db': 0.084,
            'total_loss_origin': 'stv_like_fallback',
            'tail_validation_status': 'rejected',
            'parse_mode': 'standard_sor_keyevents',
            'wavelength_display': '1550 nm',
        }

        summary = converter.summarize_file(
            'fixture.sor',
            b'',
            parsed_context=(rows, None, meta, 'standard_sor_keyevents', ''),
        )

        self.assertIsNone(summary.total_loss_db)
        self.assertEqual(summary.route_corrected_total_loss_db, 0.084)
        self.assertIn('10.1%', summary.total_loss_selection_reason)

    def test_non_reflective_terminal_sentinel_can_correct_validated_tail(self) -> None:
        rows = [
            self._event(1, 'First Connector', 0.020, 0.0, 0.0, '1F9999LS'),
            self._event(2, 'Splice', 45.000, 0.5, 0.2, '0F9999LS'),
            self._event(3, 'Splice', 49.980, 30.0, 32.767, '0F9999LS'),
            self._event(4, 'Fiber End', 50.000, 0.0, 0.0, '1E9999LS', reflectance_db=-87.0, total_loss_db=58.0),
        ]
        meta = {
            'length_km': 50.0,
            'total_loss_db': 58.0,
            'total_loss_origin': 'tail_validated',
            'tail_validation_status': 'accepted',
            'parse_mode': 'standard_sor_keyevents',
            'wavelength_display': '1550 nm',
        }

        summary = converter.summarize_file(
            'fixture.sor',
            b'',
            parsed_context=(rows, None, meta, 'standard_sor_keyevents', ''),
        )

        self.assertEqual(summary.route_corrected_total_loss_db, 9.496)
        self.assertEqual(summary.total_loss_db, 9.496)
        self.assertIn('sentinel_slope_high_loss', summary.total_loss_selection_reason)

    def test_fallback_provenance_does_not_receive_strong_tail_protection(self) -> None:
        rows = [
            self._event(1, 'First Connector', 0.020, 0.0, 0.0, '1F9999LS'),
            self._event(2, 'Splice', 49.980, 0.5, 0.2, '0F9999LS'),
            self._event(3, 'Fiber End', 50.000, 0.0, 0.0, '1E9999LS', reflectance_db=-87.0, total_loss_db=40.0),
        ]
        meta = {
            'length_km': 50.0,
            'total_loss_db': 40.0,
            'total_loss_origin': 'stv_like_fallback',
            'tail_validation_status': 'rejected',
            'parse_mode': 'standard_sor_keyevents',
            'wavelength_display': '1550 nm',
        }

        summary = converter.summarize_file(
            'fixture.sor',
            b'',
            parsed_context=(rows, None, meta, 'standard_sor_keyevents', ''),
        )

        self.assertEqual(summary.route_corrected_total_loss_db, 10.492)
        self.assertEqual(summary.total_loss_db, 10.492)
        self.assertEqual(summary.loss_source_used, 'Hiệu chỉnh theo tuyến')


class SorParseReuseTests(unittest.TestCase):
    def test_orl_extractors_use_preparsed_sor_metadata(self) -> None:
        meta = {"orl_db": 31.25}
        with patch.object(
            converter,
            "_parse_standard_sor_events_with_meta",
            side_effect=AssertionError("SOR must not be parsed again"),
        ):
            self.assertEqual(
                converter._fr_extract_orl_db(
                    "trace.sor",
                    b"not needed",
                    preparsed_sor_meta=meta,
                ),
                31.25,
            )
            candidate = converter._fr_extract_measured_orl_candidate(
                "trace.sor",
                b"not needed",
                preparsed_sor_meta=meta,
            )

        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["value_db"], 31.25)
        self.assertEqual(candidate["source_kind"], "measured_orl")

    def test_sor_trace_series_compacts_duplicate_sample_storage(self) -> None:
        values = [float(index) / 10.0 for index in range(100)]
        meta = {
            "trace_values_db": values,
            "trace_range_km": 1.0,
            "trace_calibrated_db": True,
            "trace_source": "SOR DataPts",
        }
        summary = SimpleNamespace(
            length_km=1.0,
            end_distance_km=1.0,
            graph_end_km=1.0,
        )

        series = converter._extract_raw_trace_series(
            "trace.sor",
            b"trace",
            summary,
            None,
            meta,
        )

        self.assertIsNotNone(series)
        self.assertIsInstance(series["x_km"], array)
        self.assertIsInstance(series["y_db"], array)
        self.assertEqual(list(series["y_db"]), values)
        self.assertEqual(meta["trace_values_db_count"], len(values))
        self.assertNotIn("trace_values_db", meta)


if __name__ == "__main__":
    unittest.main()
