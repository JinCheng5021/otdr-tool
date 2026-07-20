from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import struct
import time
import traceback
import zlib
from collections import Counter
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from statistics import median
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

PAT_EVENT = re.compile(r'<event\b([^>]*)>(.*?)</event>', re.S | re.I)
PAT_ATTR = re.compile(r'(\w+)="([^"]*)"')
PAT_CURVE_SCREEN = re.compile(r'CURVE:SCREEN\s*([\-0-9.]+),([\-0-9.]+),([\-0-9.]+),([\-0-9.]+)')
PAT_LASER = re.compile(r'<laser\b([^>]*)>(.*?)</laser>', re.S | re.I)
PAT_TEXT_FLOAT = re.compile(r'([+-]?\d+(?:\.\d+)?)')
RED_FILL = PatternFill(fill_type='solid', fgColor='FF0000')
GREEN_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill(fill_type='solid', fgColor='FFF2CC')
STV_DAT_FILL = PatternFill(fill_type='solid', fgColor='9DC3E6')
STV_DUT_FILL = PatternFill(fill_type='solid', fgColor='FF0000')
STV_SUY_HAO_FILL = PatternFill(fill_type='solid', fgColor='92D050')
ALLOWED_EXTENSIONS = {'.msor', '.sor', '.trc', '.crt'}
HEADER_FILL = PatternFill(fill_type='solid', fgColor='DCE6F2')
SUBHEADER_FILL = PatternFill(fill_type='solid', fgColor='EEF5FF')
LOG_INFO_FILL = PatternFill(fill_type='solid', fgColor='EAF2FF')
LOG_WARN_FILL = PatternFill(fill_type='solid', fgColor='FFF4CC')
LOG_ERROR_FILL = PatternFill(fill_type='solid', fgColor='FDE9E7')
_PARSE_CACHE: dict[str, dict] = {}
_MSOR_SECTIONS_CACHE: dict[str, dict[str, bytes]] = {}
_MINI_CURVE_CACHE: dict[str, Optional[list[int]]] = {}
_CURVE_MAX_CACHE: dict[str, Optional[float]] = {}
_TRACE_SERIES_CACHE: dict[str, Optional[dict]] = {}


_NOMINAL_WAVELENGTHS_NM = (1310, 1490, 1550, 1625, 1650)


def _snap_nominal_wavelength_nm(value: Optional[float], tolerance_nm: float = 8.0) -> Optional[int]:
    """Chuẩn hóa bước sóng danh định để tránh hiện 1554 thay vì 1550, 1308 thay vì 1310..."""
    if value is None:
        return None
    try:
        x = float(value)
    except Exception:
        return None
    best = None
    best_diff = None
    for nominal in _NOMINAL_WAVELENGTHS_NM:
        diff = abs(x - nominal)
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best = nominal
    if best is not None and best_diff is not None and best_diff <= float(tolerance_nm):
        return int(best)
    return int(round(x))



def _fr_cache_key(file_name: str, raw: bytes) -> str:
    return f"{Path(file_name).suffix.lower()}:{hashlib.sha1(raw).hexdigest()}"


def _fr_fast_raw_key(raw: bytes) -> str:
    """Cheap cache key for repeated preview helpers.

    Trace preview may call MiniCurve/section extraction several times for the
    same uploaded bytes.  A full SHA-1 over every file is accurate but costs
    noticeable time for large batches, so for these in-process helper caches we
    combine length plus head/tail fingerprints.
    """
    head = raw[:4096]
    tail = raw[-4096:] if len(raw) > 4096 else b''
    return f"{len(raw)}:{hashlib.sha1(head + tail).hexdigest()}"


def _fr_log(logs: list[dict], stage: str, level: str, message: str) -> None:
    logs.append({'time': datetime.now().strftime('%H:%M:%S'), 'stage': stage, 'level': level, 'message': message})


def _fr_init_context(summary, selected_rows, parse_mode, sor_meta, trc_trace, orl_db, metadata):
    return {
        'summary': summary,
        'events': selected_rows or [],
        'parse_mode': parse_mode or '',
        'sor_meta': sor_meta,
        'trc_trace': trc_trace,
        'orl_db': orl_db,
        'orl_display': None,
        'orl_status': 'Unknown',
        'orl_analysis': None,
        'orl_value_db': None,
        'orl_source_kind': '',
        'orl_source_detail': '',
        'orl_use_for_judgment': False,
        'orl_reason': '',
        'metadata': metadata or {},
        'parser_family': getattr(summary, 'parse_family', ''),
        'parser_family_confidence': getattr(summary, 'parse_family_confidence', ''),
        'parser_family_reason': getattr(summary, 'parse_family_reason', ''),
        'graph_assessment': None,
        'segment_assessment': None,
        'segment_event_rows': [],
    }

def _detect_parser_family(file_name: str, raw: bytes, ext: str, parse_mode: str, *, text: Optional[str] = None,
                          sor_meta: Optional[dict] = None, summary: Optional[FileSummary] = None,
                          events: Optional[list[EventRow]] = None) -> dict:
    """Nhận diện family file để ghi log và hỗ trợ tinh chỉnh theo họ file."""
    text = text or raw.decode('latin1', 'ignore')
    events = events or []
    reasons: list[str] = []
    family = 'unknown'
    confidence = 'Medium'

    if ext == '.msor':
        if any('keyevents loss enriched' in (e.note_original or '').lower() or 'smartlink + keyevents' in (e.note_original or '').lower() for e in events):
            family = 'msor_smart_link_keyevents_enriched'
            confidence = 'High'
            reasons.append('Có smart_link XML và đã ghép loss/slope từ KeyEvents nhị phân')
        elif any((e.note_original or '').lower().startswith('keyevents binary fallback') for e in events):
            family = 'msor_vendor_binary_keyevents'
            confidence = 'High'
            reasons.append('Event chỉ có trong KeyEvents nhị phân')
        elif '<smart_link>' in text.lower():
            family = 'msor_smart_link_xml'
            confidence = 'High'
            reasons.append('Có smart_link XML')
        elif _parse_msor_keyevents_summary(raw):
            family = 'msor_keyevents_summary_only'
            confidence = 'Medium'
            reasons.append('Có tóm tắt KeyEvents nhưng event không nằm trong XML')
        else:
            family = 'msor_generic'
            confidence = 'Low'
            reasons.append('MSOR chung, cần theo dõi thêm')
    elif ext == '.sor':
        map_info = _parse_standard_sor_map(raw)
        has_keyevents = bool(map_info and 'KeyEvents' in map_info.get('entries', {}))
        forced_route = bool(summary and summary.loss_source_used == 'Hiệu chỉnh theo tuyến')
        if has_keyevents and forced_route:
            family = 'sor_standard_keyevents_inflated_summary'
            confidence = 'High'
            reasons.append('SOR chuẩn có KeyEvents nhưng summary tail bị đội, ưu tiên route-corrected')
        elif has_keyevents:
            family = 'sor_standard_keyevents'
            confidence = 'High'
            reasons.append('SOR chuẩn đọc theo KeyEvents')
        elif parse_mode == 'sor_fallback':
            family = 'sor_fallback_text_or_xml'
            confidence = 'Medium'
            reasons.append('SOR dùng parser dự phòng XML/text')
        else:
            family = 'sor_generic'
            confidence = 'Low'
            reasons.append('SOR chung, cần theo dõi thêm')
    elif ext in {'.trc', '.crt'}:
        if parse_mode == 'trc_appregex_sections_points':
            family = 'trc_sections_points'
            confidence = 'Medium'
            reasons.append('TRC tách điểm theo sections/points')
        else:
            family = 'trc_standard'
            confidence = 'Medium'
            reasons.append('TRC chuẩn')
    else:
        family = f"{ext.lstrip('.')}_generic" if ext else 'unknown'
        confidence = 'Low'
        reasons.append('Family chưa có rule riêng')

    return {
        'family': family,
        'confidence': confidence,
        'reason': ' | '.join(reasons),
    }


def _to_vi_parser_family(family: str) -> str:
    mapping = {
        'msor_vendor_binary_keyevents': 'MSOR vendor - KeyEvents nhị phân',
        'msor_smart_link_keyevents_enriched': 'MSOR VIAVI/JDSU - smart_link + KeyEvents',
        'msor_smart_link_xml': 'MSOR - smart_link XML',
        'msor_keyevents_summary_only': 'MSOR - tóm tắt KeyEvents',
        'msor_generic': 'MSOR chung',
        'sor_standard_keyevents_inflated_summary': 'SOR chuẩn - KeyEvents, summary tail bị đội',
        'sor_standard_keyevents': 'SOR chuẩn - KeyEvents',
        'sor_fallback_text_or_xml': 'SOR - parser dự phòng XML/text',
        'sor_generic': 'SOR chung',
        'trc_sections_points': 'TRC - sections/points',
        'trc_standard': 'TRC chuẩn',
        'unknown': 'Chưa xác định',
    }
    return mapping.get(family, family)


def _fr_validate_template_or_raise(template_path: Path, wb: Workbook) -> list[str]:
    required = ['General Information', 'Sections', 'Link Results', 'Events']
    warnings = []
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise ValueError('Template thiếu sheet bắt buộc: ' + ', '.join(missing))
    if wb['Events'].max_column < 4:
        raise ValueError('Sheet Events của template không đủ cột mẫu.')
    if wb['Sections'].max_column < 4:
        raise ValueError('Sheet Sections của template không đủ cột mẫu.')
    if wb['Link Results'].max_column < 25:
        warnings.append('Sheet Link Results có ít cột hơn khuyến nghị, app sẽ cố gắng điền phần hiện có.')
    return warnings


def _fr_get_or_build_parse_bundle(file_name: str, raw: bytes, logs: list[dict]) -> dict:
    key = _fr_cache_key(file_name, raw)
    cached = _PARSE_CACHE.get(key)
    if cached is not None:
        _fr_log(logs, 'parse', 'INFO', f'Cache hit: {file_name}')
        return cached
    _fr_log(logs, 'parse', 'INFO', f'Cache miss: {file_name}')
    events, trc_trace, sor_meta, parse_mode, _text = _parse_events_with_context(file_name, raw)
    summary = summarize_file(file_name, raw, parsed_context=(events, trc_trace, sor_meta, parse_mode, _text))
    wavelength_nm = None
    m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
    if m:
        wavelength_nm = m.group(1)
    selected_rows = _fr_pick_rows_for_file(events, wavelength_nm)
    orl_db = _fr_extract_orl_db(file_name, raw)
    metadata = _extract_file_metadata(file_name, raw, summary, trc_trace, sor_meta)
    metadata['parser_family'] = summary.parse_family
    metadata['parser_family_confidence'] = summary.parse_family_confidence
    metadata['parser_family_reason'] = summary.parse_family_reason
    _fr_log(logs, 'parse', 'INFO', f"{file_name} | Family: {_to_vi_parser_family(summary.parse_family)} | Parse mode: {_to_vi_parse_mode(parse_mode)} | Confidence: {summary.parse_family_confidence or '-'}")
    if summary.parse_family_reason:
        _fr_log(logs, 'parse', 'INFO', f"{file_name} | Dấu hiệu nhận diện: {summary.parse_family_reason}")
    if summary.loss_source_used:
        _fr_log(logs, 'parse', 'INFO', f"{file_name} | Nguồn suy hao tổng sử dụng: {summary.loss_source_used}")
    bundle = {
        'summary': summary,
        'events': selected_rows,
        'parse_mode': parse_mode,
        'sor_meta': sor_meta,
        'trc_trace': trc_trace,
        'orl_db': orl_db,
        'metadata': metadata,
    }
    _PARSE_CACHE[key] = bundle
    return bundle


@dataclass
class EventRow:
    file_name: str
    event_no: str
    event_type: str
    distance_m: Optional[float]
    distance_km: Optional[float]
    wavelength_nm: Optional[str]
    loss_db: Optional[float]
    reflectance_db: Optional[float]
    slope_dbkm: Optional[float]
    total_loss_db: Optional[float]
    note_original: str
    label: str


@dataclass
class FileSummary:
    file_name: str
    fiber: str
    wavelength_display: str
    total_loss_db: Optional[float]
    length_km: Optional[float]
    attenuation_dbkm: Optional[float]
    splice_points: list[tuple[float, float]]
    end_distance_km: Optional[float]
    graph_end_km: Optional[float]
    graph_curve_max_km: Optional[float]
    source_format: str
    parsed_total_loss_db: Optional[float]
    route_corrected_total_loss_db: Optional[float]
    loss_source_used: str
    parse_family: str = ''
    parse_family_confidence: str = ''
    parse_family_reason: str = ''


@dataclass
class DistanceCluster:
    representative_km: float
    values_km: list[float]


@dataclass
class GraphAssessment:
    graph_end_km: Optional[float]
    jumper_excluded_km: float
    net_graph_length_km: Optional[float]
    expected_route_km: Optional[float]
    event_length_km: Optional[float]
    diff_km: Optional[float]
    graph_reach_tolerance_km: float
    event_shortfall_tolerance_km: float
    overlength_tolerance_km: float
    graph_reaches_expected: Optional[bool]
    verdict: str
    reason: str


@dataclass
class SegmentAssessment:
    start_km: float
    end_km: float
    span_km: float
    event_count: int
    segment_total_loss_db: Optional[float]
    segment_attenuation_dbkm: Optional[float]
    max_positive_event_loss_db: Optional[float]
    max_negative_event_loss_db: Optional[float]
    note: str
    recommendation: str
    method: str


@dataclass
class SectionFitResult:
    file_name: str
    section_index: int
    start_km: float
    end_km: float
    span_km: float
    source: str
    raw_points_total: int
    fit_points_used: int
    slope_dbkm: Optional[float]
    attenuation_dbkm: Optional[float]
    loss_db: Optional[float]
    intercept_db: Optional[float]
    r2: Optional[float]
    rms_residual_db: Optional[float]
    max_abs_residual_db: Optional[float]
    confidence: str
    used_for_section: bool
    fallback_method: str
    note: str
    # Phase 6.1: make section-fit provenance explicit.  These fields allow the
    # workbook to distinguish exact raw-fit, expanded-window fit, interpolation,
    # and event fallback instead of presenting all non-event results as one kind.
    fit_mode: str = ''
    fit_window_start_km: Optional[float] = None
    fit_window_end_km: Optional[float] = None
    r2_rms_scope: str = ''
    estimate_level: str = ''


@dataclass
class ORLAnalysis:
    file_name: str
    display: Optional[str]
    value_db: Optional[float]
    status: str
    advanced_status: str
    source_kind: str
    source_detail: str
    source_confidence: str
    pass_threshold_db: float
    use_for_judgment: bool
    lower_bound: bool
    reason: str
    recommendation: str
    physical_mode: str
    physical_attempted: bool
    physical_value_db: Optional[float]
    physical_status: str
    physical_reason: str
    strongest_reflectance_db: Optional[float]


@dataclass
class SectionEntry:
    name: str
    size: int


def _get_tag(body: str, tag: str) -> Optional[str]:
    m = re.search(rf'<{tag}(?:\s+[^>]*)?>(.*?)</{tag}>', body, re.S | re.I)
    if m:
        return m.group(1).strip()
    m = re.search(rf'<{tag}(?:\s+[^>]*)?\s*/>', body, re.S | re.I)
    if m:
        return ''
    return None


def _to_float(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    value = value.strip().replace(',', '.')
    if value == '':
        return None
    try:
        return float(value)
    except Exception:
        return None




def _normalize_sor_raw_wavelength_nm(raw_value: Optional[int]) -> Optional[int]:
    """Normalize vendor-specific SOR wavelength encodings to nominal OTDR values.

    Some SOR families store wavelength as:
      - raw_value * 0.1 nm (common Telcordia style)
      - raw_value in nm directly (e.g. 1554)
      - raw_value needing x10 expansion (e.g. 155 -> 1550)

    We keep the same parser flow but choose the most plausible nominal wavelength for
    reporting so files like the uploaded 1550-nm family do not appear as 155 nm.
    """
    if raw_value in (None, 0):
        return None
    candidates: list[float] = []
    try:
        raw_f = float(raw_value)
    except Exception:
        return None
    for cand in (raw_f * 0.1, raw_f, raw_f * 10.0):
        if 100.0 <= cand <= 2000.0:
            candidates.append(cand)
    if not candidates:
        return None
    best_nominal = None
    best_delta = None
    for cand in candidates:
        for nominal in _NOMINAL_WAVELENGTHS_NM:
            delta = abs(float(cand) - float(nominal))
            if delta <= 20.0 and (best_delta is None or delta < best_delta):
                best_delta = delta
                best_nominal = int(nominal)
    if best_nominal is not None:
        return int(best_nominal)
    # Fall back to the candidate that looks most realistic for telecom OTDR work.
    filtered = [cand for cand in candidates if cand >= 1000.0]
    chosen = filtered[0] if filtered else candidates[0]
    return int(round(chosen))


def _should_force_route_corrected_total_for_standard_sor(
    summary_total_loss_db: Optional[float],
    route_total_loss_db: Optional[float],
    length_km: Optional[float],
    fiber_end: Optional['EventRow'],
    wavelength_display: Optional[str],
) -> bool:
    """Detect the uploaded SOR family whose tail summary overstates route loss.

    These files expose a reasonable event table and length, but the terminal Fiber End
    row carries an anomalously large point loss (often > 10 dB with a bogus 32.7 dB/km
    slope) that inflates the summary total loss and average attenuation.  We do not
    change the route-loss algorithm; we only decide when to trust the already computed
    route-corrected total instead of the tail summary.
    """
    if summary_total_loss_db in (None, 0) or route_total_loss_db in (None, 0) or length_km in (None, 0):
        return False
    try:
        parsed_att = float(summary_total_loss_db) / float(length_km)
        route_att = float(route_total_loss_db) / float(length_km)
    except Exception:
        return False
    if fiber_end is None:
        return False
    terminal_loss = fiber_end.loss_db if isinstance(fiber_end.loss_db, (int, float)) else None
    terminal_slope = fiber_end.slope_dbkm if isinstance(fiber_end.slope_dbkm, (int, float)) else None
    malformed_wavelength = False
    if wavelength_display:
        try:
            m = re.search(r'(\d+(?:\.\d+)?)', str(wavelength_display))
            malformed_wavelength = bool(m and float(m.group(1)) < 300.0)
        except Exception:
            malformed_wavelength = False
    huge_terminal = terminal_loss is not None and float(terminal_loss) >= 5.0
    bogus_terminal_slope = terminal_slope is not None and float(terminal_slope) >= 10.0
    inflated_summary = parsed_att > max(route_att * 1.45, route_att + 0.10)
    return bool(inflated_summary and huge_terminal and (bogus_terminal_slope or malformed_wavelength))

def _safe_round3(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    return round(value + 1e-12, 3)

def _fiber_label_from_name(file_name: str) -> str:
    stem = Path(file_name).stem
    m = re.search(r'FiberCable\s*0*(\d+)', stem, re.I)
    if m:
        return f"Fiber {int(m.group(1)):03d}"
    m = re.search(r'(^|[^0-9])(\d{1,3})(?!\d)', stem)
    if m:
        return f"Fiber {int(m.group(2)):02d}"
    return stem or 'Fiber'


def _display_fiber_label(summary: FileSummary, meta: Optional[dict] = None) -> str:
    # Keep report labels aligned with the file numbering convention.
    # This avoids generic labels like 'Fiber' and matches operational usage.
    return _fiber_label_from_name(summary.file_name)


def _percentile(values: list[float], q: float) -> float:
    if not values:
        raise ValueError('empty values')
    if q <= 0:
        return min(values)
    if q >= 100:
        return max(values)
    ordered = sorted(values)
    pos = (len(ordered) - 1) * q / 100.0
    lo = int(pos)
    hi = min(lo + 1, len(ordered) - 1)
    frac = pos - lo
    return ordered[lo] * (1 - frac) + ordered[hi] * frac


def _moving_average(values: list[float], radius: int = 4) -> list[float]:
    result: list[float] = []
    for i in range(len(values)):
        a = max(0, i - radius)
        b = min(len(values), i + radius + 1)
        result.append(sum(values[a:b]) / (b - a))
    return result


def _parse_msor_sections(raw: bytes) -> dict[str, bytes]:
    key = _fr_fast_raw_key(raw)
    cached = _MSOR_SECTIONS_CACHE.get(key)
    if cached is not None:
        return cached
    root_end = raw.find(b'\x00', 0)
    if root_end < 0:
        _MSOR_SECTIONS_CACHE[key] = {}
        return {}
    try:
        map_size = struct.unpack('<I', raw[root_end + 3: root_end + 7])[0]
    except Exception:
        _MSOR_SECTIONS_CACHE[key] = {}
        return {}

    entries: list[SectionEntry] = []
    ptr = root_end + 9
    while ptr < map_size:
        end = raw.find(b'\x00', ptr)
        if end < 0 or end >= map_size:
            break
        try:
            name = raw[ptr:end].decode('latin1')
            size = struct.unpack('<I', raw[end + 3: end + 7])[0]
        except Exception:
            break
        entries.append(SectionEntry(name=name, size=size))
        ptr = end + 7

    sections: dict[str, bytes] = {}
    offset = map_size
    for entry in entries:
        sections[entry.name] = raw[offset: offset + entry.size]
        offset += entry.size
    _MSOR_SECTIONS_CACHE[key] = sections
    return sections




def _parse_msor_keyevents_summary(raw: bytes) -> Optional[dict]:
    """
    Parse VIAVI/JDSU MSOR KeyEvents tail summary.

    In many MSOR files the smart_link XML lists event positions only, while
    the real route summary lives in the binary KeyEvents tail. The last 44 bytes
    consistently contain:
      - bytes 22..25: total loss in milli-dB
      - bytes 30..33: loss-finish raw distance (vendor raw units)
      - bytes 34..35: ORL in milli-dB (optional)

    We primarily need the total route loss so attenuation can be computed even
    when XML event rows don't expose total_loss_dB.
    """
    sections = _parse_msor_sections(raw)
    payload = sections.get('KeyEvents')
    if not payload or len(payload) < 44:
        return None
    tail = payload[-44:]
    try:
        total_loss_raw = int.from_bytes(tail[22:26], 'little', signed=False)
        loss_start_raw = int.from_bytes(tail[26:30], 'little', signed=False)
        loss_finish_raw = int.from_bytes(tail[30:34], 'little', signed=False)
        orl_raw = int.from_bytes(tail[34:36], 'little', signed=False)
    except Exception:
        return None

    total_loss_db = None
    if 0 < total_loss_raw < 200000:
        total_loss_db = round(total_loss_raw / 1000.0, 3)

    orl_db = None
    if 0 < orl_raw < 100000:
        orl_db = round(orl_raw / 1000.0, 3)

    if total_loss_db is None and orl_db is None and loss_finish_raw == 0:
        return None

    return {
        'total_loss_db': total_loss_db,
        'loss_start_raw': loss_start_raw,
        'loss_finish_raw': loss_finish_raw,
        'orl_db': orl_db,
    }

def _extract_curve_max_km(raw: bytes) -> Optional[float]:
    key = _fr_fast_raw_key(raw)
    if key in _CURVE_MAX_CACHE:
        return _CURVE_MAX_CACHE[key]
    text = raw.decode('latin1', 'ignore')
    m = PAT_CURVE_SCREEN.search(text)
    if not m:
        _CURVE_MAX_CACHE[key] = None
        return None
    try:
        xmax_m = float(m.group(3))
        value = round(xmax_m / 1000.0, 6)
        _CURVE_MAX_CACHE[key] = value
        return value
    except Exception:
        _CURVE_MAX_CACHE[key] = None
        return None


def _extract_mini_curve_values(raw: bytes) -> Optional[list[int]]:
    key = _fr_fast_raw_key(raw)
    if key in _MINI_CURVE_CACHE:
        return _MINI_CURVE_CACHE[key]
    sections = _parse_msor_sections(raw)
    payload = sections.get('ActernaMiniCurve')
    if not payload:
        _MINI_CURVE_CACHE[key] = None
        return None

    candidates: list[tuple[float, int]] = []
    # Previous versions unpacked the full curve for every possible start offset.
    # For Trace Viewer this was the biggest cost on MSOR batches.  The score only
    # depends on the first ~160 samples, so score small samples first and unpack
    # the selected candidate once.
    max_start = min(60, len(payload) - 200)
    for start in range(44, max_start):
        count = (len(payload) - start) // 2
        if count < 200:
            continue
        sample_count = min(160, count)
        try:
            sample = list(struct.unpack('<' + 'h' * sample_count, payload[start: start + sample_count * 2]))
        except Exception:
            continue
        if not sample:
            continue
        proportion_small = sum(-2500 <= v <= 2500 for v in sample) / len(sample)
        q95 = abs(_percentile(sample, 95))
        score = proportion_small - (0.0001 * q95)
        candidates.append((score, start))

    if not candidates:
        _MINI_CURVE_CACHE[key] = None
        return None
    _, best_start = max(candidates, key=lambda item: item[0])
    count = (len(payload) - best_start) // 2
    try:
        best_values = list(struct.unpack('<' + 'h' * count, payload[best_start: best_start + count * 2]))
    except Exception:
        _MINI_CURVE_CACHE[key] = None
        return None
    _MINI_CURVE_CACHE[key] = best_values
    return best_values



def _detect_graph_end_km(raw: bytes) -> Optional[float]:
    curve_max_km = _extract_curve_max_km(raw)
    values = _extract_mini_curve_values(raw)
    if curve_max_km is None or not values or len(values) < 80:
        return None

    smoothed = _moving_average([float(v) for v in values], radius=4)
    tail = smoothed[int(len(smoothed) * 0.88):]
    pre = smoothed[: max(30, int(len(smoothed) * 0.60))]
    if not tail or not pre:
        return None

    floor = median(tail)
    plateau = _percentile(pre, 90)
    dynamic = plateau - floor
    if dynamic <= 0:
        plateau = max(pre)
        dynamic = plateau - floor
    if dynamic <= 0:
        return None

    threshold = floor + 0.18 * dynamic
    stable_count = max(8, int(len(smoothed) * 0.025))
    start_index = max(5, int(len(smoothed) * 0.12))

    target_index: Optional[int] = None
    for idx in range(start_index, len(smoothed) - stable_count):
        window = smoothed[idx: idx + stable_count]
        if sum(value <= threshold for value in window) >= max(stable_count - 1, 1):
            target_index = idx
            break

    if target_index is None:
        diffs = [smoothed[i] - smoothed[i - 1] for i in range(1, len(smoothed))]
        search_from = max(1, int(len(diffs) * 0.20))
        target_index = min(range(search_from, len(diffs)), key=lambda i: diffs[i])

    break_km = (target_index / max(len(smoothed) - 1, 1)) * curve_max_km
    return round(break_km + 1e-12, 3)


def _msor_terminal_event_length_km(rows: list[EventRow]) -> Optional[float]:
    if not rows:
        return None
    terminal_rows = [
        row for row in rows
        if row.distance_km is not None and row.event_type.lower() in {'fiber end', 'slope'}
    ]
    if terminal_rows:
        return max(float(row.distance_km or 0.0) for row in terminal_rows)
    usable = [float(row.distance_km or 0.0) for row in rows if row.distance_km is not None]
    return max(usable) if usable else None


def _msor_has_tail_after_terminal(rows: list[EventRow], terminal_km: Optional[float]) -> bool:
    if terminal_km in (None, 0):
        return False
    guard = max(0.8, 0.03 * float(terminal_km))
    for row in rows:
        if row.distance_km is None:
            continue
        if float(row.distance_km) > float(terminal_km) + guard:
            return True
    return False


def _score_msor_graph_candidate(candidate_km: Optional[float], *, source: str, terminal_km: Optional[float], max_event_km: Optional[float], curve_max_km: Optional[float], has_tail_after_terminal: bool) -> float:
    if candidate_km in (None, 0):
        return -1e9
    cand = float(candidate_km)
    score = 0.0

    if cand <= 0:
        return -1e9
    if curve_max_km not in (None, 0):
        curve_max = float(curve_max_km)
        if cand > curve_max * 1.03:
            score -= 80.0
        elif cand > curve_max:
            score -= 8.0

    if source == 'datapts':
        score += 5.0
    elif source == 'preview':
        score += 1.0

    ref_km = float(terminal_km) if terminal_km not in (None, 0) else None
    if ref_km is not None:
        diff = abs(cand - ref_km)
        rel = diff / max(ref_km, 1e-6)
        score -= min(diff, 25.0) * 0.40
        if rel <= 0.05:
            score += 10.0
        elif rel <= 0.10:
            score += 7.0
        elif rel <= 0.20:
            score += 4.0
        elif rel <= 0.35:
            score += 1.5
        else:
            score -= 4.0

        if ref_km > 10.0 and cand < 0.45 * ref_km:
            score -= 24.0
        if ref_km > 10.0 and cand < 0.65 * ref_km:
            score -= 8.0
        if (not has_tail_after_terminal) and ref_km > 10.0 and cand > 2.0 * ref_km:
            score -= 18.0

    if max_event_km not in (None, 0):
        max_event = float(max_event_km)
        if cand >= 0.85 * max_event:
            score += 2.0
        elif max_event > 10.0 and cand < 0.50 * max_event:
            score -= 10.0

    if has_tail_after_terminal:
        if source == 'event':
            score -= 8.0
        if max_event_km not in (None, 0) and cand >= 0.80 * float(max_event_km):
            score += 5.0

    return score


def _estimate_msor_graph_metrics(raw: bytes, rows: list[EventRow], event_length_km: Optional[float]) -> tuple[Optional[float], Optional[float]]:
    map_info = _parse_standard_sor_map(raw)
    fxd = _parse_standard_sor_fxdparams(raw, map_info) if map_info else None
    datapts_meta = _parse_standard_sor_datapts(raw, map_info, fxd) if map_info and fxd else None

    datapts_end = datapts_meta.get('graph_end_km') if datapts_meta else None
    datapts_max = datapts_meta.get('graph_curve_max_km') if datapts_meta else None
    preview_end = _detect_graph_end_km(raw)
    preview_max = _extract_curve_max_km(raw)

    curve_max_km = datapts_max or (round(float(fxd.get('range_km')), 3) if fxd and fxd.get('range_km') else None) or preview_max or event_length_km

    terminal_km = event_length_km if event_length_km not in (None, 0) else _msor_terminal_event_length_km(rows)
    max_event_km = max((float(row.distance_km or 0.0) for row in rows if row.distance_km is not None), default=None)
    has_tail_after_terminal = _msor_has_tail_after_terminal(rows, terminal_km)

    candidates = [
        ('datapts', datapts_end),
        ('preview', preview_end),
        ('event', terminal_km),
    ]

    chosen_source, chosen_value = max(
        candidates,
        key=lambda item: _score_msor_graph_candidate(
            item[1],
            source=item[0],
            terminal_km=terminal_km,
            max_event_km=max_event_km,
            curve_max_km=curve_max_km,
            has_tail_after_terminal=has_tail_after_terminal,
        ),
    )

    if chosen_value in (None, 0):
        chosen_value = terminal_km or datapts_end or preview_end
        chosen_source = 'event' if chosen_value == terminal_km else chosen_source

    if terminal_km not in (None, 0) and chosen_value not in (None, 0):
        guard = max(6.0, 0.35 * float(terminal_km))
        if (not has_tail_after_terminal) and abs(float(chosen_value) - float(terminal_km)) > guard:
            chosen_value = terminal_km
            chosen_source = 'event'

    if chosen_source == 'event' and curve_max_km in (None, 0):
        curve_max_km = terminal_km
    if curve_max_km not in (None, 0) and chosen_value not in (None, 0):
        curve_max_km = max(float(curve_max_km), float(chosen_value))

    return (_safe_round3(chosen_value), _safe_round3(curve_max_km))


def parse_embedded_xml_events(file_name: str, raw: bytes) -> list[EventRow]:
    text = raw.decode('latin1', 'ignore')
    default_wavelength = _preferred_msor_wavelength(file_name, raw, text) if Path(file_name).suffix.lower() == '.msor' else None
    rows: list[EventRow] = []
    for m in PAT_EVENT.finditer(text):
        attrs = dict(PAT_ATTR.findall(m.group(1)))
        body = m.group(2)
        event_type = attrs.get('eventType') or attrs.get('type') or ''
        if not event_type:
            continue

        wavelength_nm = default_wavelength
        loss_db = reflectance_db = slope_dbkm = total_loss_db = None
        laser_match = PAT_LASER.search(body)
        if laser_match:
            laser_attrs = dict(PAT_ATTR.findall(laser_match.group(1)))
            laser_body = laser_match.group(2)
            wavelength_nm = default_wavelength or laser_attrs.get('wavelength_nm') or laser_attrs.get('wavelength')
            loss_db = _to_float(_get_tag(laser_body, 'loss_dB') or _get_tag(laser_body, 'loss'))
            reflectance_db = _to_float(_get_tag(laser_body, 'reflectance_dB') or _get_tag(laser_body, 'reflectance'))
            slope_dbkm = _to_float(_get_tag(laser_body, 'slope_dBkm') or _get_tag(laser_body, 'slope'))
            total_loss_db = _to_float(_get_tag(laser_body, 'total_loss_dB') or _get_tag(laser_body, 'total_loss'))
        else:
            loss_db = _to_float(_get_tag(body, 'loss_dB') or _get_tag(body, 'loss'))
            reflectance_db = _to_float(_get_tag(body, 'reflectance_dB') or _get_tag(body, 'reflectance'))
            slope_dbkm = _to_float(_get_tag(body, 'slope_dBkm') or _get_tag(body, 'slope'))
            total_loss_db = _to_float(_get_tag(body, 'total_loss_dB') or _get_tag(body, 'total_loss'))

        distance_m = _to_float(_get_tag(body, 'distance_m') or _get_tag(body, 'distance'))
        if distance_m is None:
            distance_km = _to_float(_get_tag(body, 'distance_km'))
            if distance_km is not None:
                distance_m = distance_km * 1000.0

        rows.append(EventRow(
            file_name=file_name,
            event_no=attrs.get('no', attrs.get('id', '')),
            event_type=event_type,
            distance_m=distance_m,
            distance_km=round(distance_m / 1000.0, 6) if distance_m is not None else None,
            wavelength_nm=wavelength_nm,
            loss_db=loss_db,
            reflectance_db=reflectance_db,
            slope_dbkm=slope_dbkm,
            total_loss_db=total_loss_db,
            note_original=_get_tag(body, 'note') or '',
            label=_get_tag(body, 'label') or '',
        ))
    return rows


def _normalize_event_type(value: str) -> str:
    s = value.strip().lower()
    if not s:
        return 'Unknown'
    if 'first connector' in s or s in {'launch connector', 'start connector'}:
        return 'First Connector'
    if 'fiber end' in s or s in {'end', 'fiberend'}:
        return 'Fiber End'
    if 'splice' in s or 'non-reflective' in s:
        return 'Splice'
    if 'connector' in s or 'reflect' in s:
        return 'Connector'
    if 'gainer' in s:
        return 'Gainer'
    return value.strip()



def _extract_wavelength_from_text(text: str) -> Optional[str]:
    # Prefer contextual mentions that usually refer to the actual acquisition wavelength.
    patterns = [
        r'\b(1310|1490|1550|1625)\s*nm\b',
        r'\bwavelength\b[^\d]*(1310|1490|1550|1625)\b',
        r'(?<!\d)(1310|1490|1550|1625)(?!\d).{0,12}dBm',
        r'_(1310|1490|1550|1625)_(?:OE|CL|OL|OR|EV|SM|MM)\b',
        r'\b(?:FiberCable|Cable|Trace|File)[^\r\n]{0,40}_(1310|1490|1550|1625)_',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I | re.S)
        if m:
            return m.group(1)
    m = re.search(r'\bL(1310|1490|1550|1625)\b', text, re.I)
    if m:
        return m.group(1)
    return None


def _extract_wavelength_from_filename(file_name: str) -> Optional[str]:
    m = re.search(r'(?<!\d)(1310|1490|1550|1625)(?!\d)', Path(file_name).name, re.I)
    if m:
        return m.group(1)
    return None


def _extract_msor_binary_wavelength(raw: bytes) -> Optional[str]:
    sections = _parse_msor_sections(raw)
    payload = sections.get('FxdParams')
    if payload and len(payload) >= len('FxdParams') + 1 + 8:
        pos = len('FxdParams') + 1
        try:
            unit = payload[pos + 4: pos + 6].decode('latin1', 'ignore').lower()
        except Exception:
            unit = ''
        raw_nm = int.from_bytes(payload[pos + 6: pos + 8], 'little', signed=False)
        nm = raw_nm / 10.0 if raw_nm else 0.0
        if unit in {'km', 'mi', 'kf'} and 1000 <= nm <= 1700:
            return str(int(round(nm)))

    idx = raw.find(b'Fiber\x00')
    if idx >= 0:
        start = idx + len(b'Fiber\x00')
        for off in range(start, min(start + 16, len(raw) - 1)):
            nm = int.from_bytes(raw[off:off + 2], 'little', signed=False)
            if 1000 <= nm <= 1700:
                return str(int(nm))
    return None


def _preferred_msor_wavelength(file_name: str, raw: bytes, text: Optional[str] = None) -> Optional[str]:
    if text is None:
        try:
            text = raw.decode('latin1', 'ignore')
        except Exception:
            text = ''

    by_name = _extract_wavelength_from_filename(file_name)
    if by_name:
        return by_name

    m_embedded = re.search(
        r'\b(?:FiberCable|Cable|Trace|File)[^\r\n\x00]{0,60}_(1310|1490|1550|1625)_(?:OE|CL|OL|OR|EV|SM|MM)?',
        text,
        re.I,
    )
    if m_embedded:
        return m_embedded.group(1)

    sections = _parse_msor_sections(raw)
    pwm_text = sections.get('JDSUpwm_1', b'').decode('latin1', 'ignore') if sections else ''
    m_pwm = re.search(r'Module\s*:\s*\x00?(1310|1490|1550|1625)\x00', pwm_text) or re.search(
        r'Module\s*:\s*.*?\b(1310|1490|1550|1625)\b', pwm_text, re.I | re.S
    )
    if m_pwm:
        return m_pwm.group(1)

    binary = _extract_msor_binary_wavelength(raw)
    if binary:
        return binary

    return _extract_wavelength_from_text(text)


def _clean_visible_text(value: Optional[str]) -> str:
    if not value:
        return ''
    value = ''.join(ch for ch in str(value) if ch.isprintable())
    value = re.sub(r'\s+', ' ', value).strip().strip('|').strip()
    return value


def _split_pipe_section(payload: Optional[bytes]) -> list[str]:
    if not payload:
        return []
    txt = payload.decode('latin1', 'ignore').replace('\x00', '|')
    return [_clean_visible_text(part) for part in txt.split('|')]


def _extract_avg_seconds_from_wave_payload(payload: Optional[bytes]) -> Optional[float]:
    parts = _split_pipe_section(payload)
    if not parts:
        return None
    for i, part in enumerate(parts[:-1]):
        if part.upper() == 'AVG':
            value = _to_float(parts[i + 1])
            if value is not None and 0 < value <= 3600:
                return round(float(value), 3)
    txt = payload.decode('latin1', 'ignore').replace('\x00', '|') if payload else ''
    m = re.search(r'AVG\|\s*([0-9]+(?:\.[0-9]+)?)', txt, re.I)
    if m:
        value = _to_float(m.group(1))
        if value is not None and 0 < value <= 3600:
            return round(float(value), 3)
    return None


def _config_quoted(text: str, key: str) -> Optional[str]:
    if not text:
        return None
    m = re.search(rf'(?im)^\s*{re.escape(key)}\s*"([^"]*)"', text)
    if not m:
        m = re.search(rf'(?im)^\s*{re.escape(key)}\s*,"([^"]*)"', text)
    return _clean_visible_text(m.group(1)) if m else None


def _config_number(text: str, key: str) -> Optional[float]:
    if not text:
        return None
    m = re.search(rf'(?im)^\s*{re.escape(key)}\s*,?\s*([+-]?\d+(?:\.\d+)?)', text)
    return _to_float(m.group(1)) if m else None


def _parse_msor_fxdparams(raw: bytes) -> Optional[dict]:
    sections = _parse_msor_sections(raw)
    payload = sections.get('FxdParams')
    if not payload:
        return None
    pos = len('FxdParams') + 1
    if pos + 44 > len(payload):
        return None
    try:
        unit = payload[pos + 4: pos + 6].decode('latin1', 'ignore')
        wavelength_nm = _sor_get_uint(payload, pos + 6, 2) * 0.1
        pulse_us = _sor_get_uint(payload, pos + 18, 2) / 1000.0
        sample_spacing_usec = _sor_get_uint(payload, pos + 20, 4) * 1e-8
        num_points = _sor_get_uint(payload, pos + 24, 4)
        index = _sor_get_uint(payload, pos + 28, 4) * 1e-5
        range_km_raw = _sor_get_uint(payload, pos + 40, 4) * 2e-5
        # OTDR distance is one-way: c*t/(2*n). The previous metadata-only path omitted
        # the round-trip factor 2 and inflated the reported test Range. Keep analysis logic
        # elsewhere unchanged; only tune the metadata value used in the report template.
        dx_km = sample_spacing_usec * (299792.458 / 1.0e6) / (2.0 * index) if index else 0.0
        range_km = dx_km * num_points if dx_km and num_points else (range_km_raw / 2.0 if range_km_raw else None)
        return {
            'unit': unit,
            'wavelength_nm': round(wavelength_nm, 1) if wavelength_nm else None,
            'pulse_us': round(pulse_us, 3) if pulse_us else None,
            'sample_spacing_usec': sample_spacing_usec,
            'num_points': num_points,
            'index': index,
            'range_km': range_km,
        }
    except Exception:
        return None


def _parse_sor_supparams(raw: bytes, map_info: dict) -> dict:
    entry = map_info.get('entries', {}).get('SupParams') if map_info else None
    if not entry:
        return {}
    data = raw[entry['offset']: entry['offset'] + entry['size']]
    parts = _split_pipe_section(data)
    if len(parts) < 2:
        return {}
    version = parts[6] if len(parts) > 6 else ''
    serial_candidates = [p for p in parts[3:6] if p and p != version]
    serial_numeric = [p for p in serial_candidates if re.fullmatch(r'[A-Za-z0-9 -]{3,}', p)]
    serial = ''
    if serial_numeric:
        numeric_only = [p for p in serial_numeric if p.replace(' ', '').isdigit()]
        serial = numeric_only[-1] if numeric_only else serial_numeric[0]
    return {
        'company': parts[1] if len(parts) > 1 else '',
        'unit_model': parts[2] if len(parts) > 2 else '',
        'unit_serial': serial,
        'version': version,
    }


def _parse_msor_metadata(file_name: str, raw: bytes) -> dict:
    sections = _parse_msor_sections(raw)
    gen_parts = _split_pipe_section(sections.get('GenParams'))
    sup_parts = _split_pipe_section(sections.get('SupParams'))
    cfg_text = sections.get('ActernaConfig').decode('latin1', 'ignore') if sections.get('ActernaConfig') else ''
    wave_payload = sections.get('WaveMTSParams')
    fxd = _parse_msor_fxdparams(raw) or {}

    company = sup_parts[1] if len(sup_parts) > 1 else ''
    base_model = sup_parts[2] if len(sup_parts) > 2 else ''
    extra_model = sup_parts[4] if len(sup_parts) > 4 else ''
    if extra_model and extra_model != base_model and any(ch.isalpha() for ch in extra_model):
        unit_model = f'{base_model} / {extra_model}'.strip(' /')
    else:
        unit_model = base_model
    version = sup_parts[6] if len(sup_parts) > 6 else ''
    serial_candidates = [p for p in sup_parts[3:6] if p and p != version]
    numeric_only = [p for p in serial_candidates if p.replace(' ', '').isdigit()]
    unit_serial = numeric_only[-1] if numeric_only else (serial_candidates[0] if serial_candidates else '')

    cable_id = _config_quoted(cfg_text, 'FSETup:CABleid ORI,') or _config_quoted(cfg_text, 'FSETUP:CABleid ORI,')
    fiber_id = _config_quoted(cfg_text, 'FSETup:FIBerid ORI,') or _config_quoted(cfg_text, 'FSETUP:FIBerid ORI,')
    location_a = _config_quoted(cfg_text, 'FSETup:ORIGin') or _config_quoted(cfg_text, 'FSETUP:ORIGin')
    location_b = _config_quoted(cfg_text, 'FSETup:END') or _config_quoted(cfg_text, 'FSETUP:END')
    operator = _config_quoted(cfg_text, 'FSETup:OPERator') or _config_quoted(cfg_text, 'FSETUP:OPERator')
    job_id = _config_quoted(cfg_text, 'FSETup:JOBid') or _config_quoted(cfg_text, 'FSETUP:JOBid')
    comments = _config_quoted(cfg_text, 'FSETUP:SAVED:COMment')
    max_t = _config_number(cfg_text, 'OTDS:MAXT')

    if not cable_id and len(gen_parts) > 1:
        candidate = gen_parts[1]
        if candidate and candidate.upper() not in {'FR', 'EN'} and len(candidate) >= 3:
            cable_id = candidate
    if not fiber_id and len(gen_parts) > 2:
        fiber_id = gen_parts[2]
    if not location_a and len(gen_parts) > 3:
        location_a = re.sub(r'^[^A-Za-z0-9]+', '', gen_parts[3])
    if not location_b and len(gen_parts) > 4:
        location_b = gen_parts[4]

    duration_s = _extract_avg_seconds_from_wave_payload(wave_payload) or (round(max_t, 3) if max_t else None)

    return {
        'file_name': file_name,
        'company': company,
        'unit_model': unit_model,
        'unit_serial': unit_serial,
        'cable_id': cable_id or 'Cable',
        'fiber_id': fiber_id or 'Fiber',
        'location_a': location_a or '',
        'location_b': location_b or '',
        'operator_a': operator or '',
        'operator_b': '',
        'job_id': job_id or '',
        'comments': comments or '',
        'wavelength_nm': int(round(float(fxd.get('wavelength_nm')))) if fxd.get('wavelength_nm') else None,
        'range_km': round(float(fxd.get('range_km')), 4) if fxd.get('range_km') else None,
        'pulse_us': round(float(fxd.get('pulse_us')), 3) if fxd.get('pulse_us') else None,
        'duration_s': duration_s,
    }


def _parse_sor_metadata(file_name: str, raw: bytes) -> dict:
    map_info = _parse_standard_sor_map(raw)
    if not map_info:
        return {'file_name': file_name}
    fxd = _parse_standard_sor_fxdparams(raw, map_info) or {}
    sup = _parse_sor_supparams(raw, map_info)
    entry = map_info.get('entries', {}).get('WaveMTSParams')
    wave_payload = raw[entry['offset']: entry['offset'] + entry['size']] if entry else None
    fiber_label = _extract_standard_sor_fiber_label(raw, map_info) or 'Fiber'
    pulse_us = None
    fxd_entry = map_info.get('entries', {}).get('FxdParams')
    if fxd_entry:
        payload = raw[fxd_entry['offset']: fxd_entry['offset'] + fxd_entry['size']]
        pos = len('FxdParams') + 1
        if pos + 20 <= len(payload):
            p = _sor_get_uint(payload, pos + 18, 2) / 1000.0
            if 0 < p <= 100000:
                pulse_us = round(p, 3)
    return {
        'file_name': file_name,
        'company': sup.get('company') or '',
        'unit_model': sup.get('unit_model') or '',
        'unit_serial': sup.get('unit_serial') or '',
        'cable_id': 'Cable',
        'fiber_id': fiber_label or 'Fiber',
        'location_a': '',
        'location_b': '',
        'operator_a': '',
        'operator_b': '',
        'job_id': '',
        'comments': '',
        'wavelength_nm': int(round(float(fxd.get('wavelength_nm')))) if fxd.get('wavelength_nm') else None,
        'range_km': round(float(fxd.get('range_km')), 4) if fxd.get('range_km') else None,
        'pulse_us': pulse_us,
        'duration_s': _extract_avg_seconds_from_wave_payload(wave_payload),
        'reflection_threshold_db': round(float(fxd.get('reflection_threshold_db')), 3) if fxd.get('reflection_threshold_db') else None,
    }


def _parse_trc_metadata(file_name: str, raw: bytes) -> dict:
    events, _trc_trace, _sor_meta, _mode, _text = _parse_events_with_context(file_name, raw)
    wavelength_nm = None
    range_km = None
    pulse_us = None
    duration_s = None
    rows, meta = _parse_trc_events_appregex(file_name, raw)
    if meta:
        w = meta.get('Wavelength')
        if isinstance(w, (int, float)) and w:
            wavelength_nm = int(round(float(w)))
        dr = meta.get('DisplayRange') or meta.get('Range')
        if isinstance(dr, (int, float)) and dr:
            range_km = round(float(dr) / 1000.0, 4) if float(dr) > 1000 else round(float(dr), 4)
    combined = _decompress_trc_streams(raw)
    records = _collect_trc_appregex_records(combined) if combined else None
    if records:
        by_name = {}
        for rec in records:
            by_name.setdefault(rec['name'], []).append(rec)
        def last_scalar(name: str):
            arr = by_name.get(name) or []
            if not arr:
                return None
            try:
                return _parse_trc_value_appregex(combined, arr[-1])
            except Exception:
                return None
        pulse_val = last_scalar('Pulse')
        if isinstance(pulse_val, (int, float)) and 0 < float(pulse_val) < 1000000:
            pu = float(pulse_val) * 1e6 if float(pulse_val) < 0.01 else float(pulse_val)
            if 0 < pu <= 100000:
                pulse_us = round(pu, 3)
        avg_val = last_scalar('AverageTime') or last_scalar('Duration')
        if isinstance(avg_val, (int, float)) and 0 < float(avg_val) <= 3600:
            duration_s = round(float(avg_val), 3)
    return {
        'file_name': file_name,
        'company': '',
        'unit_model': '',
        'unit_serial': '',
        'cable_id': 'Cable',
        'fiber_id': Path(file_name).stem,
        'location_a': '',
        'location_b': '',
        'operator_a': '',
        'operator_b': '',
        'job_id': '',
        'comments': '',
        'wavelength_nm': wavelength_nm,
        'range_km': range_km,
        'pulse_us': pulse_us,
        'duration_s': duration_s,
    }


def _extract_file_metadata(file_name: str, raw: bytes, summary: FileSummary, trc_trace: Optional[dict], sor_meta: Optional[dict]) -> dict:
    ext = Path(file_name).suffix.lower()
    if ext == '.msor':
        meta = _parse_msor_metadata(file_name, raw)
    elif ext == '.sor':
        meta = _parse_sor_metadata(file_name, raw)
    elif ext in {'.trc', '.crt'}:
        meta = _parse_trc_metadata(file_name, raw)
    else:
        meta = {'file_name': file_name}
    if not meta.get('fiber_id'):
        meta['fiber_id'] = summary.fiber or 'Fiber'
    if not meta.get('wavelength_nm'):
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        meta['wavelength_nm'] = _snap_nominal_wavelength_nm(int(m.group(1))) if m else None
    if not meta.get('range_km'):
        if trc_trace is not None and isinstance(trc_trace.get('DisplayRange'), (int, float)):
            meta['range_km'] = round(float(trc_trace.get('DisplayRange')) / 1000.0, 4)
        elif sor_meta and isinstance(sor_meta.get('graph_curve_max_km'), (int, float)):
            meta['range_km'] = round(float(sor_meta.get('graph_curve_max_km')), 4)
        elif summary.graph_curve_max_km is not None:
            meta['range_km'] = round(float(summary.graph_curve_max_km), 4)
        elif summary.length_km is not None:
            meta['range_km'] = round(float(summary.length_km), 4)
    meta.setdefault('parser_family', summary.parse_family if summary else '')
    meta.setdefault('parser_family_confidence', summary.parse_family_confidence if summary else '')
    meta.setdefault('parser_family_reason', summary.parse_family_reason if summary else '')
    return meta


def _common_nonempty(values: list[str]) -> str:
    values = [_clean_visible_text(v) for v in values if _clean_visible_text(v)]
    if not values:
        return ''
    counts = Counter(values)
    return counts.most_common(1)[0][0]




def _sor_read_cstr(raw: bytes, pos: int) -> tuple[Optional[str], int]:
    end = raw.find(b'\x00', pos)
    if end < 0:
        return None, pos
    return raw[pos:end].decode('latin1', 'ignore'), end + 1


def _sor_get_uint(raw: bytes, pos: int, nbytes: int) -> int:
    return int.from_bytes(raw[pos: pos + nbytes], 'little', signed=False)


def _sor_get_signed(raw: bytes, pos: int, nbytes: int) -> int:
    return int.from_bytes(raw[pos: pos + nbytes], 'little', signed=True)


def _parse_standard_sor_map(raw: bytes) -> Optional[dict]:
    name, pos = _sor_read_cstr(raw, 0)
    if name != 'Map':
        return None
    if pos + 8 > len(raw):
        return None

    map_rev = _sor_get_uint(raw, pos, 2)
    pos += 2
    map_size = _sor_get_uint(raw, pos, 4)
    pos += 4
    block_count = _sor_get_uint(raw, pos, 2)
    pos += 2

    entries: list[tuple[str, int, int]] = []
    for _ in range(max(block_count - 1, 0)):
        entry_name, pos = _sor_read_cstr(raw, pos)
        if not entry_name or pos + 6 > len(raw):
            break
        entry_rev = _sor_get_uint(raw, pos, 2)
        pos += 2
        entry_size = _sor_get_uint(raw, pos, 4)
        pos += 4
        entries.append((entry_name, entry_rev, entry_size))

    offsets: dict[str, dict] = {}
    offset = map_size
    for entry_name, entry_rev, entry_size in entries:
        offsets[entry_name] = {'offset': offset, 'size': entry_size, 'rev': entry_rev}
        offset += entry_size

    return {
        'format': 2 if map_rev >= 200 else 1,
        'map_rev': map_rev,
        'map_size': map_size,
        'entries': offsets,
    }


def _parse_standard_sor_fxdparams(raw: bytes, map_info: dict) -> Optional[dict]:
    entry = map_info['entries'].get('FxdParams')
    if not entry:
        return None

    fmt = map_info['format']
    off = entry['offset']
    pos = off + len('FxdParams') + 1
    if pos + 64 > len(raw):
        return None

    try:
        unit = raw[pos + 4: pos + 6].decode('latin1', 'ignore')
        raw_wavelength = _sor_get_uint(raw, pos + 6, 2)
        wavelength_nm = _normalize_sor_raw_wavelength_nm(raw_wavelength)
        if fmt == 2:
            sample_spacing_usec = _sor_get_uint(raw, pos + 20, 4) * 1e-8
            num_points = _sor_get_uint(raw, pos + 24, 4)
            index = _sor_get_uint(raw, pos + 28, 4) * 1e-5
            range_km_raw = _sor_get_uint(raw, pos + 40, 4) * 2e-5
        else:
            sample_spacing_usec = _sor_get_uint(raw, pos + 16, 4) * 1e-8
            num_points = _sor_get_uint(raw, pos + 20, 4)
            index = _sor_get_uint(raw, pos + 24, 4) * 1e-5
            range_km_raw = _sor_get_uint(raw, pos + 34, 4) * 2e-5

        # OTDR trace range for reporting is one-way distance, so include the round-trip /2.
        dx_km = sample_spacing_usec * (299792.458 / 1.0e6) / (2.0 * index) if index else 0.0
        range_km = dx_km * num_points if dx_km and num_points else (range_km_raw / 2.0 if range_km_raw else None)
        resolution_m = dx_km * 1000.0
        # Reporting-only metadata. Some vendors store a reflectance/ORL-like
        # threshold in fixed parameters; if an explicit span ORL is absent later we
        # can expose that threshold as a lower-bound display in the report (e.g. <22.75).
        reflection_threshold_db = None
        if pos + 42 <= len(raw):
            thr = _sor_get_uint(raw, pos + 40, 2) * 0.001
            if 0 < thr < 100:
                reflection_threshold_db = round(thr, 3)
        return {
            'unit': unit,
            'wavelength_nm': float(wavelength_nm) if wavelength_nm is not None else None,
            'sample_spacing_usec': sample_spacing_usec,
            'num_points': num_points,
            'index': index,
            'range_km': range_km,
            'resolution_m': resolution_m,
            'reflection_threshold_db': reflection_threshold_db,
        }
    except Exception:
        return None





def _extract_standard_sor_fiber_label(raw: bytes, map_info: dict) -> Optional[str]:
    entry = map_info['entries'].get('WaveMTSParams')
    if not entry:
        return None
    data = raw[entry['offset']: entry['offset'] + entry['size']]
    try:
        decoded = data.decode('latin1', 'ignore')
    except Exception:
        return None
    candidates = re.findall(r'\x00(\d{2,4})\x00', decoded)
    if candidates:
        return Counter(candidates).most_common(1)[0][0]
    m = re.search(r'(?<!\d)(\d{2,4})(?!\d)', decoded)
    if m:
        return m.group(1)
    return None




def _is_standard_sor_terminal_event(row: EventRow) -> bool:
    code = (row.label or '').strip()
    if row.event_type == 'Fiber End':
        return True
    return len(code) >= 2 and code[1] == 'E'


def _is_standard_sor_reflective_event(row: EventRow) -> bool:
    code = (row.label or '').strip()
    if code.startswith(('1', '2')):
        return True
    return row.reflectance_db is not None and (-200.0 < float(row.reflectance_db) < 0.0)


def _choose_standard_sor_end_distance_km(rows: list[EventRow], parsed_end_distance_km: Optional[float]) -> Optional[float]:
    ordered = [r for r in rows if r.distance_km is not None]
    if not ordered:
        return None
    ordered = sorted(ordered, key=lambda r: (r.distance_km or 0.0, r.event_no))
    terminal_rows = [r for r in ordered if _is_standard_sor_terminal_event(r)]
    terminal_distance = terminal_rows[-1].distance_km if terminal_rows else ordered[-1].distance_km
    if parsed_end_distance_km is not None and parsed_end_distance_km > 0:
        if terminal_distance is None:
            return round(parsed_end_distance_km, 6)
        gap = abs(float(parsed_end_distance_km) - float(terminal_distance))
        if gap <= max(1.0, 0.08 * max(float(parsed_end_distance_km), float(terminal_distance))):
            return round(parsed_end_distance_km, 6)
        return round(parsed_end_distance_km, 6)
    if terminal_distance is not None:
        return round(float(terminal_distance), 6)
    return round(float(ordered[-1].distance_km), 6)

def _estimate_total_loss_from_standard_sor_heuristic(rows: list[EventRow]) -> Optional[float]:
    if not rows:
        return None
    ordered = [r for r in rows if r.distance_km is not None]
    if not ordered:
        return None
    ordered = sorted(ordered, key=lambda r: (r.distance_km or 0.0, r.event_no))
    length_km = ordered[-1].distance_km
    if length_km in (None, 0):
        return None

    base_event_loss = 0.0
    slope_values: list[float] = []
    near_end_losses: list[float] = []
    prev_distance = 0.0
    seg_att_capped = 0.0

    for row in ordered:
        slope = row.slope_dbkm
        if slope is not None and 0 < slope < 1:
            slope_values.append(float(slope))
            seg_att_capped += min(float(slope), 0.2) * max((row.distance_km or 0.0) - prev_distance, 0.0)
        prev_distance = row.distance_km or prev_distance

        if row.event_type != 'Fiber End' and row.loss_db is not None and row.loss_db > 0:
            base_event_loss += float(row.loss_db)
            if (length_km - (row.distance_km or 0.0)) <= 2.0:
                near_end_losses.append(float(row.loss_db))

    if not slope_values:
        return round(base_event_loss, 3) if base_event_loss > 0 else None

    q90_slope = _percentile(slope_values, 90)
    max_slope = max(slope_values)
    previous_distance = ordered[-2].distance_km if len(ordered) >= 2 else 0.0
    end_gap_km = max((length_km or 0.0) - (previous_distance or 0.0), 0.0)
    near_end_max = max(near_end_losses) if near_end_losses else 0.0

    if end_gap_km > 10.0:
        fiber_att_loss = q90_slope * length_km
    elif 3.0 <= near_end_max < 10.0 and end_gap_km < 2.0:
        fiber_att_loss = max_slope * length_km
    else:
        fiber_att_loss = seg_att_capped

    total = base_event_loss + fiber_att_loss
    return round(total, 3) if total > 0 else None


def _estimate_route_total_loss_from_standard_sor(rows: list[EventRow], length_km: Optional[float]) -> Optional[float]:
    ordered = [r for r in rows if r.distance_km is not None]
    if not ordered:
        return None
    ordered = sorted(ordered, key=lambda r: (r.distance_km or 0.0, r.event_no))
    if length_km in (None, 0):
        length_km = _choose_standard_sor_end_distance_km(ordered, None)
    if length_km in (None, 0):
        return None

    ordered = [r for r in ordered if (r.distance_km or 0.0) <= float(length_km) + 0.25]
    endpoint_zone_km = max(0.30, min(2.00, 0.02 * float(length_km)))
    cluster_gap_km = max(0.40, min(0.80, 0.01 * float(length_km)))

    fiber_section_loss = 0.0
    discrete_event_loss = 0.0
    prev_distance = 0.0
    valid_slopes: list[float] = []

    distances = [float(r.distance_km or 0.0) for r in ordered]

    for idx, row in enumerate(ordered):
        dist = min(float(row.distance_km or 0.0), float(length_km))
        seg_len = max(dist - prev_distance, 0.0)
        slope = row.slope_dbkm
        if slope is not None:
            slope_f = float(slope)
            if 0 < slope_f < 0.45:
                fiber_section_loss += slope_f * seg_len
                valid_slopes.append(slope_f)
            elif 0.45 <= slope_f < 1.0 and seg_len <= 0.25:
                fiber_section_loss += 0.25 * seg_len
                valid_slopes.append(0.25)
        prev_distance = dist

        loss = row.loss_db
        if loss is None:
            continue
        loss_f = float(loss)
        if loss_f <= 0:
            continue

        et = (row.event_type or '').strip()
        near_start = dist <= endpoint_zone_km
        near_end = (float(length_km) - dist) <= endpoint_zone_km
        terminal = _is_standard_sor_terminal_event(row)
        reflective = _is_standard_sor_reflective_event(row)

        if terminal:
            continue
        if et == 'First Connector':
            continue
        if reflective and (near_start or near_end):
            continue
        if et == 'Connector' and reflective and loss_f > 1.0:
            continue
        if et == 'Connector' and near_start and loss_f > 0.75:
            continue

        # Generic anti-overcounting for SOR files without reliable tail-summary:
        # large non-reflective losses near the edge or packed in a short cluster are
        # often analysis artifacts / duplicated section boundaries rather than route-loss points.
        prev_loss = None
        next_loss = None
        prev_gap = None
        next_gap = None
        if idx > 0:
            prev_row = ordered[idx - 1]
            prev_gap = dist - float(prev_row.distance_km or 0.0)
            prev_loss = float(prev_row.loss_db) if prev_row.loss_db is not None else None
        if idx + 1 < len(ordered):
            next_row = ordered[idx + 1]
            next_gap = float(next_row.distance_km or 0.0) - dist
            next_loss = float(next_row.loss_db) if next_row.loss_db is not None else None

        clustered_high_loss = False
        if loss_f > 1.0 and not reflective:
            if near_end or near_start:
                clustered_high_loss = True
            if prev_gap is not None and prev_gap <= cluster_gap_km and prev_loss is not None and prev_loss > 0.8:
                clustered_high_loss = True
            if next_gap is not None and next_gap <= cluster_gap_km and next_loss is not None and next_loss > 0.8:
                clustered_high_loss = True
            if loss_f >= 3.0:
                clustered_high_loss = True

        if clustered_high_loss:
            continue

        # Keep normal inline events, but cap unusually large non-reflective point loss.
        if not reflective and loss_f > 1.0:
            loss_f = 1.0

        discrete_event_loss += loss_f

    if fiber_section_loss <= 0 and valid_slopes:
        fiber_section_loss = median(valid_slopes) * float(length_km)

    total = fiber_section_loss + discrete_event_loss
    return round(total, 3) if total > 0 else None



def _estimate_stv_like_total_loss_from_standard_sor(rows: list[EventRow], length_km: Optional[float]) -> Optional[float]:
    ordered = [r for r in rows if r.distance_km is not None]
    if not ordered:
        return None
    ordered = sorted(ordered, key=lambda r: (r.distance_km or 0.0, r.event_no))
    if length_km in (None, 0):
        length_km = _choose_standard_sor_end_distance_km(ordered, None)
    if length_km in (None, 0):
        return None

    terminal_label = (ordered[-1].label or '').strip()
    terminal_is_reflective_end = len(terminal_label) >= 2 and terminal_label[1] == 'E' and terminal_label[:1] in {'1', '2'}

    event_sum = 0.0
    fiber_section_loss = 0.0
    prev_distance = 0.0
    positive_connector_losses: list[float] = []
    positive_numeric_events: list[float] = []
    negative_abs_sum = 0.0
    first_numeric_loss: Optional[float] = None
    midline_high_event = False
    q90_slopes: list[float] = []

    for idx, row in enumerate(ordered):
        dist = min(float(row.distance_km or 0.0), float(length_km))
        seg_len = max(dist - prev_distance, 0.0)
        slope = row.slope_dbkm
        if slope is not None:
            slope_f = float(slope)
            if 0 < slope_f < 1.0:
                fiber_section_loss += slope_f * seg_len
                q90_slopes.append(slope_f)
        prev_distance = dist

        loss = row.loss_db
        if loss is None:
            continue
        loss_f = float(loss)
        label = (row.label or '').strip()

        if first_numeric_loss is None:
            first_numeric_loss = loss_f

        # STV-style event table hides terminal end losses encoded as *E.
        if terminal_is_reflective_end and idx == len(ordered) - 1:
            continue
        if len(label) >= 2 and label[1] == 'E':
            continue
        if row.event_type == 'First Connector':
            continue

        if loss_f < 0:
            negative_abs_sum += abs(loss_f)
            continue
        if loss_f == 0:
            continue

        positive_numeric_events.append(loss_f)
        if row.event_type == 'Connector':
            positive_connector_losses.append(loss_f)
        if dist < float(length_km) - 2.0 and loss_f > 2.0:
            midline_high_event = True
        event_sum += loss_f

    # Default STV-like fallback: displayed positive event losses + accumulated section attenuation
    total = event_sum + fiber_section_loss

    # For some long noisy traces with a moderate terminal reflective connector, STV tends
    # to follow a steeper attenuation baseline than the raw center-to-center section sum.
    if q90_slopes:
        q90_section_loss = _percentile(q90_slopes, 90) * float(length_km)
        max_connector = max(positive_connector_losses) if positive_connector_losses else 0.0
        if (
            float(length_km) > 40.0
            and first_numeric_loss is not None and first_numeric_loss < 0
            and 8.0 <= max_connector <= 15.0
            and not midline_high_event
            and q90_section_loss > fiber_section_loss
        ):
            total = event_sum + q90_section_loss

    # For long traces dominated by a very large reflective end connector, the center-sum
    # can overstate fiber attenuation if many gain/negative corrections are present.
    max_connector = max(positive_connector_losses) if positive_connector_losses else 0.0
    if float(length_km) > 40.0 and max_connector >= 20.0 and negative_abs_sum >= 0.5:
        total -= min(negative_abs_sum, 1.5)

    return round(total, 3) if total > 0 else None

def _detect_break_from_series_km(values: list[float], curve_max_km: Optional[float], radius: int = 8) -> Optional[float]:
    if curve_max_km is None or not values or len(values) < 80:
        return None

    smoothed = _moving_average([float(v) for v in values], radius=radius)
    tail = smoothed[int(len(smoothed) * 0.88):]
    pre = smoothed[: max(30, int(len(smoothed) * 0.60))]
    if not tail or not pre:
        return None

    floor = median(tail)
    plateau = _percentile(pre, 90)
    dynamic = plateau - floor
    if dynamic <= 0:
        plateau = max(pre)
        dynamic = plateau - floor
    if dynamic <= 0:
        return None

    threshold = floor + 0.18 * dynamic
    stable_count = max(16, int(len(smoothed) * 0.01))
    start_index = max(10, int(len(smoothed) * 0.05))

    target_index: Optional[int] = None
    for idx in range(start_index, len(smoothed) - stable_count):
        window = smoothed[idx: idx + stable_count]
        if sum(value <= threshold for value in window) >= max(stable_count - 2, 1):
            target_index = idx
            break

    if target_index is None:
        diffs = [smoothed[i] - smoothed[i - 1] for i in range(1, len(smoothed))]
        search_from = max(1, int(len(diffs) * 0.20))
        target_index = min(range(search_from, len(diffs)), key=lambda i: diffs[i])

    break_km = (target_index / max(len(smoothed) - 1, 1)) * curve_max_km
    return round(break_km + 1e-12, 3)


def _parse_standard_sor_datapts(raw: bytes, map_info: dict, fxd: Optional[dict]) -> Optional[dict]:
    if not fxd:
        return None
    entry = map_info['entries'].get('DataPts')
    if not entry:
        return None

    off = entry['offset']
    pos = off + len('DataPts') + 1
    if pos + 12 > len(raw):
        return None

    try:
        num_points = _sor_get_uint(raw, pos, 4)
        pos += 4
        num_traces = _sor_get_signed(raw, pos, 2)
        pos += 2
        _num_points2 = _sor_get_uint(raw, pos, 4)
        pos += 4
        scaling_factor = _sor_get_uint(raw, pos, 2) / 1000.0
        pos += 2
        if num_points <= 0:
            return None

        sample_count = min(num_points, max((len(raw) - pos) // 2, 0))
        if sample_count <= 0:
            return None

        vals = list(struct.unpack('<' + 'H' * sample_count, raw[pos: pos + sample_count * 2]))
        ymax = max(vals)
        fs = 0.001 * scaling_factor
        trace_db = [(ymax - x) * fs for x in vals]

        graph_end_km = _detect_break_from_series_km(trace_db, fxd.get('range_km'))
        range_km = float(fxd.get('range_km') or 0.0)
        # Phase 5: keep calibrated raw trace samples so section loss can be fitted
        # from the actual backscatter curve instead of only from event/slope rows.
        # trace_db is already converted with the DataPts scaling factor above.
        return {
            'graph_end_km': graph_end_km,
            'graph_curve_max_km': round(range_km, 3) if range_km else None,
            'num_points': sample_count,
            'num_traces': num_traces,
            'trace_values_db': trace_db,
            'trace_range_km': range_km if range_km else None,
            'trace_calibrated_db': True,
            'trace_source': 'SOR DataPts',
        }
    except Exception:
        return None


def _decode_standard_sor_event_type(code: str, idx: int, total_events: int, distance_km: Optional[float]) -> str:
    code = (code or '').strip()
    if len(code) >= 2 and code[1] == 'E':
        return 'Fiber End'
    if idx == total_events - 1 and code[:1] in {'0', '1', '2'}:
        return 'Fiber End'
    if idx == 0 and distance_km is not None and distance_km <= 1.5 and code.startswith('1'):
        return 'First Connector'
    if code.startswith(('1', '2')):
        return 'Connector'
    if code.startswith('0'):
        return 'Splice'
    return 'Unknown'







def _find_preterminal_reflective_connector_loss(rows: list[EventRow], length_km: Optional[float]) -> Optional[float]:
    if length_km in (None, 0):
        return None
    ordered = [r for r in rows if r.distance_km is not None]
    if not ordered:
        return None
    ordered = sorted(ordered, key=lambda r: (r.distance_km or 0.0, r.event_no))
    end_zone_km = max(3.0, 0.03 * float(length_km))
    for row in reversed(ordered):
        if _is_standard_sor_terminal_event(row):
            continue
        dist = float(row.distance_km or 0.0)
        if (float(length_km) - dist) > end_zone_km:
            break
        loss = row.loss_db
        if loss is None or loss <= 0:
            continue
        if _is_standard_sor_reflective_event(row):
            return float(loss)
    return None

def _parse_standard_sor_events_with_meta(file_name: str, raw: bytes) -> tuple[list[EventRow], Optional[dict]]:
    map_info = _parse_standard_sor_map(raw)
    if not map_info or 'KeyEvents' not in map_info['entries']:
        return [], None

    fxd = _parse_standard_sor_fxdparams(raw, map_info)
    if not fxd or not fxd.get('index'):
        return [], None

    datapts_meta = _parse_standard_sor_datapts(raw, map_info, fxd)
    fiber_label = _extract_standard_sor_fiber_label(raw, map_info)
    off = map_info['entries']['KeyEvents']['offset']
    pos = off + len('KeyEvents') + 1
    block_end = off + map_info['entries']['KeyEvents']['size']
    if pos + 2 > len(raw):
        return [], None

    try:
        nev = _sor_get_uint(raw, pos, 2)
        pos += 2
        factor = 1e-4 * (299792.458 / 1.0e6) / float(fxd['index'])
        rows: list[EventRow] = []
        for idx in range(nev):
            if pos + 42 > min(len(raw), block_end):
                break

            xid = _sor_get_uint(raw, pos, 2)
            pos += 2
            distance_km = _sor_get_uint(raw, pos, 4) * factor
            pos += 4
            slope_dbkm = _sor_get_signed(raw, pos, 2) * 0.001
            pos += 2
            splice_loss = _sor_get_signed(raw, pos, 2) * 0.001
            pos += 2
            reflectance_db = _sor_get_signed(raw, pos, 4) * 0.001
            pos += 4
            type_code = raw[pos: pos + 8].decode('latin1', 'ignore')
            pos += 8

            _end_prev_km = _sor_get_uint(raw, pos, 4) * factor
            pos += 4
            _start_curr_km = _sor_get_uint(raw, pos, 4) * factor
            pos += 4
            _end_curr_km = _sor_get_uint(raw, pos, 4) * factor
            pos += 4
            _start_next_km = _sor_get_uint(raw, pos, 4) * factor
            pos += 4
            _peak_km = _sor_get_uint(raw, pos, 4) * factor
            pos += 4

            comment, next_pos = _sor_read_cstr(raw, pos)
            if comment is None or next_pos > block_end:
                comment = ''
                next_pos = pos
            pos = next_pos

            event_type = _decode_standard_sor_event_type(type_code, idx, nev, distance_km)
            if reflectance_db <= -1000:
                reflectance_db = None

            loss_db = splice_loss
            if event_type == 'Fiber End' and loss_db >= 20:
                loss_db = None

            rows.append(EventRow(
                file_name=file_name,
                event_no=str(xid),
                event_type=event_type,
                distance_m=distance_km * 1000.0,
                distance_km=round(distance_km, 6),
                wavelength_nm=str(int(round(float(fxd['wavelength_nm'])))),
                loss_db=round(loss_db, 3) if loss_db is not None else None,
                reflectance_db=round(reflectance_db, 3) if reflectance_db is not None else None,
                slope_dbkm=round(slope_dbkm, 3),
                total_loss_db=None,
                note_original=(comment or '').strip(),
                label=type_code.strip(),
            ))

        total_loss_db = None
        parsed_end_distance_km = None
        tail_pos = pos
        if tail_pos + 22 <= block_end:
            parsed_total = _sor_get_signed(raw, tail_pos, 4) * 0.001
            tail_pos += 4
            _loss_start_km = _sor_get_signed(raw, tail_pos, 4) * factor
            tail_pos += 4
            tail_loss_end_km = _sor_get_uint(raw, tail_pos, 4) * factor
            tail_pos += 4
            _orl_db = _sor_get_uint(raw, tail_pos, 2) * 0.001
            tail_pos += 2
            _orl_start_km = _sor_get_signed(raw, tail_pos, 4) * factor
            tail_pos += 4
            _orl_end_km = _sor_get_uint(raw, tail_pos, 4) * factor
            tail_pos += 4
            if 0 < parsed_total < 200:
                total_loss_db = round(parsed_total, 3)
            if tail_loss_end_km > 0:
                parsed_end_distance_km = round(tail_loss_end_km, 6)

        if rows:
            end_distance_km = _choose_standard_sor_end_distance_km(rows, parsed_end_distance_km)

            for row in reversed(rows):
                if _is_standard_sor_terminal_event(row):
                    row.event_type = 'Fiber End'
                    if end_distance_km is not None:
                        row.distance_km = round(end_distance_km, 6)
                        row.distance_m = row.distance_km * 1000.0
                    break

            if total_loss_db is not None and end_distance_km not in (None, 0):
                parsed_att = float(total_loss_db) / float(end_distance_km)
                if not (0.01 <= parsed_att <= 1.00):
                    total_loss_db = None

            if total_loss_db is None and end_distance_km not in (None, 0):
                stv_like_total = _estimate_stv_like_total_loss_from_standard_sor(rows, end_distance_km)
                if stv_like_total is not None:
                    total_loss_db = stv_like_total
                else:
                    route_total = _estimate_route_total_loss_from_standard_sor(rows, end_distance_km)
                    if route_total is not None:
                        total_loss_db = route_total
                    else:
                        heuristic_total = _estimate_total_loss_from_standard_sor_heuristic(rows)
                        if heuristic_total is not None:
                            total_loss_db = heuristic_total

            for row in reversed(rows):
                if row.event_type == 'Fiber End':
                    row.total_loss_db = total_loss_db
                    break
        else:
            end_distance_km = None

        meta = {
            'length_km': round(end_distance_km, 3) if end_distance_km is not None else None,
            'total_loss_db': total_loss_db,
            'orl_db': round(_orl_db, 3) if '_orl_db' in locals() and isinstance(_orl_db, (int, float)) and 0 < float(_orl_db) < 100 else None,
            'graph_end_km': datapts_meta.get('graph_end_km') if datapts_meta else None,
            'graph_curve_max_km': datapts_meta.get('graph_curve_max_km') if datapts_meta else (round(fxd['range_km'], 3) if fxd.get('range_km') else None),
            'wavelength_display': f"{_snap_nominal_wavelength_nm(fxd.get('wavelength_nm'))} nm" if fxd.get('wavelength_nm') else '',
            'parse_mode': 'standard_sor_keyevents',
            'fiber_label': fiber_label,
        }
        return rows, meta
    except Exception:
        return [], None


def parse_standard_sor_events(file_name: str, raw: bytes) -> list[EventRow]:
    rows, _meta = _parse_standard_sor_events_with_meta(file_name, raw)
    return rows




def _decode_keyevent_loss_db(payload: bytes, offset: int) -> Optional[float]:
    """Decode VIAVI/JDSU KeyEvents point loss as signed milli-dB.

    Several MSOR files store negative point-loss corrections in the same 16-bit
    field used for normal positive event losses. Reading that field as unsigned
    turns values like -0.201 dB into 65335, which the old parser then dropped as
    invalid. We keep the signed value here so gain/negative events survive the
    Excel export.
    """
    try:
        raw_loss = struct.unpack_from('<h', payload, offset)[0]
    except Exception:
        return None
    if abs(raw_loss) >= 10000:
        return None
    return round(raw_loss / 1000.0, 3)


def _find_msor_keyevents_record_run(payload: bytes) -> Optional[dict]:
    """Locate the compact VIAVI/JDSU KeyEvents record run.

    The MSOR files from VIAVI/JDSU/Acterna MTS commonly store the first event
    in a small special header area and serialize the remaining events as 44-byte
    records starting with marker 0x0020 + event number.  Older code tried to
    infer the run from `len(payload) - count * 44`, which is unsafe because the
    block can also contain a tail summary.  This scanner looks for the actual
    stable record run instead, so SmartLink XML positions can be enriched with
    the correct event loss values without touching the existing section logic.
    """
    if not payload or len(payload) < 80:
        return None
    try:
        declared_count = struct.unpack_from('<H', payload, 10)[0]
    except Exception:
        declared_count = 0
    record_size = 44
    best_start: Optional[int] = None
    best_len = 0
    limit = min(160, max(0, len(payload) - record_size * 2))
    for start in range(0, limit):
        try:
            marker0, ev0 = struct.unpack_from('<HH', payload, start)
            marker1, ev1 = struct.unpack_from('<HH', payload, start + record_size)
        except Exception:
            continue
        if marker0 != 0x0020 or marker1 != 0x0020:
            continue
        if ev1 != ev0 + 1:
            continue
        cur = 2
        last_ev = ev1
        off = start + record_size * 2
        while off + 4 <= len(payload):
            try:
                marker, ev = struct.unpack_from('<HH', payload, off)
            except Exception:
                break
            if marker != 0x0020 or ev != last_ev + 1:
                break
            cur += 1
            last_ev = ev
            off += record_size
        if cur > best_len:
            best_start = start
            best_len = cur
    if best_start is None or best_len <= 0:
        return None
    return {
        'declared_count': declared_count,
        'record_size': record_size,
        'record_start': best_start,
        'record_count': best_len,
        'first_special_available': len(payload) >= 26,
    }


def _decode_msor_keyevents_records(payload: bytes, expected_count: Optional[int] = None) -> list[dict]:
    """Decode conservative per-event fields from a VIAVI/JDSU KeyEvents block.

    Returned records are ordered to match SmartLink's event order: the special
    first event (event 0/first connector) followed by the compact 44-byte event
    records.  This is intentionally limited to stable fields only: distance raw,
    point loss, slope and vendor label.  It avoids guessing unsupported private
    fields, preserving the old calculation logic where data is not reliable.
    """
    run = _find_msor_keyevents_record_run(payload)
    if not run:
        return []
    declared = int(run.get('declared_count') or 0)
    target_count = expected_count or declared or (int(run['record_count']) + 1)
    if declared and expected_count and abs(declared - expected_count) > 3:
        target_count = min(expected_count, declared)
    target_count = max(0, int(target_count or 0))
    records: list[dict] = []

    # Special first event, usually the first connector at 0 km.  The point loss
    # is stored at byte 20 in the observed VIAVI/JDSU/Acterna MTS block.
    first_loss = _decode_keyevent_loss_db(payload, 20)
    first_label = payload[26:34].decode('latin1', 'ignore').strip('\x00 ').strip() if len(payload) >= 34 else ''
    if target_count != 0:
        records.append({
            'event_no_binary': 1,
            'distance_raw': 0,
            'loss_db': first_loss,
            'slope_dbkm': None,
            'reflectance_db': None,
            'label': first_label,
            'source': 'KeyEvents special first event',
        })

    rec_start = int(run['record_start'])
    record_size = int(run['record_size'])
    rec_count = int(run['record_count'])
    max_extra = max(0, (target_count - 1) if target_count else rec_count)
    for i in range(min(rec_count, max_extra)):
        off = rec_start + i * record_size
        if off + 24 > len(payload):
            break
        try:
            _marker, ev_no = struct.unpack_from('<HH', payload, off)
            dist_raw = struct.unpack_from('<I', payload, off + 4)[0]
            loss_db = _decode_keyevent_loss_db(payload, off + 8)
            slope_raw = struct.unpack_from('<h', payload, off + 10)[0]
            slope_dbkm = round(slope_raw / 1000.0, 3) if abs(slope_raw) < 10000 else None
            refl_raw_i32 = struct.unpack_from('<i', payload, off + 12)[0]
            label = payload[off + 16: off + 24].decode('latin1', 'ignore').strip('\x00 ').strip()
        except Exception:
            continue
        reflectance_db = None
        # In many non-reflective JDSU events this field is zero.  Keep only
        # plausible negative reflectance values to avoid fabricating data.
        if -90000 <= refl_raw_i32 <= -1000:
            reflectance_db = round(refl_raw_i32 / 1000.0, 3)
        records.append({
            'event_no_binary': ev_no,
            'distance_raw': dist_raw,
            'loss_db': loss_db,
            'slope_dbkm': slope_dbkm,
            'reflectance_db': reflectance_db,
            'label': label,
            'source': 'KeyEvents compact record',
        })
    if expected_count and len(records) > expected_count:
        records = records[:expected_count]
    return records


def _parse_keyevents_binary_losses(raw: bytes, expected_count: Optional[int] = None) -> list[Optional[float]]:
    sections = _parse_msor_sections(raw)
    payload = sections.get('KeyEvents')
    if not payload or len(payload) < 40:
        return []
    records = _decode_msor_keyevents_records(payload, expected_count=expected_count)
    if not records:
        return []
    return [rec.get('loss_db') for rec in records]



def _parse_msor_binary_keyevents_events(file_name: str, raw: bytes) -> list[EventRow]:
    """
    Fallback parser for VIAVI/JDSU MSOR files whose events live only in the
    binary KeyEvents block.

    Some SmartOTDR/VIAVI MSOR traces do not embed <smart_link> XML nor a text
    event table. They still store a compact KeyEvents structure where each
    record is 44 bytes and contains:
      - event number
      - distance raw
      - point loss (signed milli-dB)
      - slope (signed milli-dB/km)
      - reflectance (signed milli-dB)
      - vendor label code (8 bytes)

    We keep this parser conservative: if we cannot find a stable record layout,
    we return [] and let the higher-level code fall back as before.
    """
    if Path(file_name).suffix.lower() != '.msor':
        return []
    sections = _parse_msor_sections(raw)
    payload = sections.get('KeyEvents')
    if not payload or len(payload) < 120:
        return []

    try:
        declared_count = struct.unpack_from('<H', payload, 10)[0]
    except Exception:
        declared_count = 0

    record_size = 44
    rec_start = None
    rec_count = 0
    best_len = 0
    for start in range(0, min(128, max(0, len(payload) - record_size * 2))):
        try:
            marker0, ev0 = struct.unpack_from('<HH', payload, start)
            marker1, ev1 = struct.unpack_from('<HH', payload, start + record_size)
        except Exception:
            continue
        if marker0 != 0x0020 or marker1 != 0x0020:
            continue
        if ev1 != ev0 + 1:
            continue
        cur = 2
        last_ev = ev1
        off = start + record_size * 2
        while off + 4 <= len(payload):
            try:
                marker, ev = struct.unpack_from('<HH', payload, off)
            except Exception:
                break
            if marker != 0x0020 or ev != last_ev + 1:
                break
            cur += 1
            last_ev = ev
            off += record_size
        if cur > best_len:
            best_len = cur
            rec_start = start
            rec_count = cur

    if rec_start is None or rec_count <= 0:
        return []

    if declared_count and declared_count > rec_count:
        # Keep the stable run we found. Some files declare one more event in the
        # header than what is serialized in the compact records.
        pass

    meta = _parse_msor_metadata(file_name, raw) or {}
    range_hint = meta.get('range_km')
    try:
        range_hint = float(range_hint) if range_hint not in (None, '') else None
    except Exception:
        range_hint = None

    max_raw = 0
    raws = []
    for i in range(rec_count):
        off = rec_start + i * record_size
        try:
            dist_raw = struct.unpack_from('<I', payload, off + 4)[0]
        except Exception:
            dist_raw = 0
        if dist_raw > 0:
            raws.append(dist_raw)
            max_raw = max(max_raw, dist_raw)
    if not raws:
        return []

    divisors = [1000.0, 10000.0, 100000.0]
    if range_hint and range_hint > 0:
        divisor = min(divisors, key=lambda d: abs((max_raw / d) - range_hint))
    else:
        divisor = 10000.0

    wavelength = _preferred_msor_wavelength(file_name, raw)
    total_loss_db = None
    try:
        summary = _parse_msor_keyevents_summary(raw)
        if summary and isinstance(summary.get('total_loss_db'), (int, float)):
            total_loss_db = round(float(summary['total_loss_db']), 3)
    except Exception:
        total_loss_db = None

    rows: list[EventRow] = []
    for i in range(rec_count):
        off = rec_start + i * record_size
        try:
            _marker, ev_no = struct.unpack_from('<HH', payload, off)
            dist_raw = struct.unpack_from('<I', payload, off + 4)[0]
            loss_raw = struct.unpack_from('<h', payload, off + 8)[0]
            slope_raw = struct.unpack_from('<h', payload, off + 10)[0]
            refl_raw = struct.unpack_from('<i', payload, off + 12)[0]
            label = payload[off + 16: off + 24].decode('latin1', 'ignore').strip('\x00 ').strip()
        except Exception:
            continue

        distance_km = round(dist_raw / divisor, 6) if dist_raw else None
        loss_db = round(loss_raw / 1000.0, 3)
        slope_dbkm = round(slope_raw / 1000.0, 3)
        reflectance_db = round(refl_raw / 1000.0, 3) if refl_raw not in (0, -1000, -1000000) else None

        label_up = label.upper()
        if label_up.startswith('1E') or i == rec_count - 1:
            event_type = 'Fiber End'
        elif reflectance_db is not None and reflectance_db < 0:
            event_type = 'Connector'
        elif label_up.startswith(('1', '2')):
            event_type = 'Connector'
        elif label_up.startswith('0'):
            event_type = 'Splice'
        else:
            event_type = 'Event'

        rows.append(EventRow(
            file_name=file_name,
            event_no=str(ev_no),
            event_type=event_type,
            distance_m=(distance_km * 1000.0) if distance_km is not None else None,
            distance_km=distance_km,
            wavelength_nm=wavelength,
            loss_db=loss_db,
            reflectance_db=reflectance_db,
            slope_dbkm=slope_dbkm,
            total_loss_db=None,
            note_original='KeyEvents binary fallback',
            label=label,
        ))

    if rows and total_loss_db is not None:
        for row in reversed(rows):
            if row.event_type == 'Fiber End':
                row.total_loss_db = total_loss_db
                break

    return rows






def _infer_smart_link_distance_mode(raw_distances: list[float], raw: bytes, unit_hint: str) -> str:
    unit = (unit_hint or '').strip().lower()
    if unit in {'m', 'meter', 'meters'}:
        return 'm'
    if not raw_distances:
        return 'km' if unit == 'km' else 'm'

    max_raw = max(raw_distances)
    med_raw = sorted(raw_distances)[len(raw_distances) // 2]
    curve_hint_km = _extract_curve_max_km(raw)

    # JDSU/Acterna SmartLink quirk: some SOR files declare unit="km"
    # while the stored numeric values are still meters.
    if max_raw > 500:
        return 'm'
    if med_raw > 100:
        return 'm'
    if curve_hint_km is not None:
        diff_if_km = abs(max_raw - curve_hint_km)
        diff_if_m = abs((max_raw / 1000.0) - curve_hint_km)
        if diff_if_m + 0.5 < diff_if_km:
            return 'm'
    return 'km' if unit == 'km' else 'm'


def _estimate_total_loss_from_events(events: list[EventRow]) -> Optional[float]:
    usable: list[float] = []
    for event in events:
        if event.loss_db is None:
            continue
        event_type = event.event_type.lower()
        if 'fiber end' in event_type:
            continue
        if event.loss_db <= 0:
            continue
        usable.append(float(event.loss_db))
    if not usable:
        return None
    return round(sum(usable), 3)

def parse_smart_link_events(file_name: str, raw: bytes) -> list[EventRow]:
    text = raw.decode('latin1', 'ignore')
    block_m = re.search(r'<smart_link>(.*?)</smart_link>', text, re.S | re.I)
    if not block_m:
        return []
    block = block_m.group(1)
    table_m = re.search(r'<event_table\b([^>]*)>(.*?)</event_table>', block, re.S | re.I)
    if not table_m:
        return []
    table_attrs = dict(PAT_ATTR.findall(table_m.group(1)))
    table_body = table_m.group(2)
    unit_hint = (table_attrs.get('unit') or '').strip().lower()
    expected_count = None
    try:
        expected_count = int(table_attrs.get('count', '0') or '0')
    except Exception:
        expected_count = None
    fallback_losses = _parse_keyevents_binary_losses(raw, expected_count=expected_count)
    keyevent_records: list[dict] = []
    try:
        if Path(file_name).suffix.lower() == '.msor':
            _ke_payload = _parse_msor_sections(raw).get('KeyEvents')
            keyevent_records = _decode_msor_keyevents_records(_ke_payload, expected_count=expected_count) if _ke_payload else []
    except Exception:
        keyevent_records = []
    default_wavelength = _preferred_msor_wavelength(file_name, raw, text) if Path(file_name).suffix.lower() == '.msor' else _extract_wavelength_from_text(text)
    sections = _parse_msor_sections(raw)
    pwm_text = sections.get('JDSUpwm_1', b'').decode('latin1', 'ignore')
    m_pwm = re.search(r'Module\s*:\s*\\x00?(1310|1490|1550|1625)\\x00', pwm_text) or re.search(r'Module\s*:\s*.*?\b(1310|1490|1550|1625)\b', pwm_text)
    if m_pwm:
        default_wavelength = m_pwm.group(1)

    rows: list[EventRow] = []
    event_matches = list(PAT_EVENT.finditer(table_body))
    raw_distances: list[float] = []
    for match in event_matches:
        body = match.group(2)
        raw_distance = _to_float(_get_tag(body, 'distance') or _get_tag(body, 'position'))
        if raw_distance is not None:
            raw_distances.append(raw_distance)
    actual_distance_mode = _infer_smart_link_distance_mode(raw_distances, raw, unit_hint)

    for idx, match in enumerate(event_matches):
        attrs = dict(PAT_ATTR.findall(match.group(1)))
        body = match.group(2)
        event_type = attrs.get('eventType') or ''
        if not event_type:
            icon_txt = _get_tag(body, 'icon') or ''
            try:
                icon = int(icon_txt)
            except Exception:
                icon = -1
            if idx == len(event_matches) - 1 or icon == 6:
                event_type = 'Fiber End'
            elif idx == 0 or icon in {4, 69}:
                event_type = 'First Connector'
            elif icon in {3}:
                event_type = 'Splice'
            else:
                event_type = 'Connector'

        wavelength_nm = default_wavelength
        distance_m = _to_float(_get_tag(body, 'distance_m'))
        if distance_m is None:
            raw_distance = _to_float(_get_tag(body, 'distance') or _get_tag(body, 'position'))
            if raw_distance is not None:
                distance_m = raw_distance if actual_distance_mode == 'm' else raw_distance * 1000.0

        loss_db = reflectance_db = slope_dbkm = total_loss_db = None
        laser_match = PAT_LASER.search(body)
        if laser_match:
            laser_attrs = dict(PAT_ATTR.findall(laser_match.group(1)))
            laser_body = laser_match.group(2)
            wavelength_nm = laser_attrs.get('wavelength_nm') or laser_attrs.get('wavelength') or wavelength_nm
            loss_db = _to_float(_get_tag(laser_body, 'loss_dB') or _get_tag(laser_body, 'loss'))
            reflectance_db = _to_float(_get_tag(laser_body, 'reflectance_dB') or _get_tag(laser_body, 'reflectance'))
            slope_dbkm = _to_float(_get_tag(laser_body, 'slope_dBkm') or _get_tag(laser_body, 'slope'))
            total_loss_db = _to_float(_get_tag(laser_body, 'total_loss_dB') or _get_tag(laser_body, 'total_loss'))
        else:
            loss_db = _to_float(_get_tag(body, 'loss_dB') or _get_tag(body, 'loss'))
            reflectance_db = _to_float(_get_tag(body, 'reflectance_dB') or _get_tag(body, 'reflectance'))
            slope_dbkm = _to_float(_get_tag(body, 'slope_dBkm') or _get_tag(body, 'slope'))
            total_loss_db = _to_float(_get_tag(body, 'total_loss_dB') or _get_tag(body, 'total_loss'))

        ke_rec = keyevent_records[idx] if idx < len(keyevent_records) else None
        if loss_db is None and idx < len(fallback_losses):
            loss_db = fallback_losses[idx]
        if ke_rec:
            if slope_dbkm is None and ke_rec.get('slope_dbkm') is not None:
                slope_dbkm = ke_rec.get('slope_dbkm')
            if reflectance_db is None and ke_rec.get('reflectance_db') is not None:
                reflectance_db = ke_rec.get('reflectance_db')

        rows.append(EventRow(
            file_name=file_name,
            event_no=attrs.get('no', str(idx)),
            event_type=_normalize_event_type(event_type),
            distance_m=distance_m,
            distance_km=round(distance_m / 1000.0, 6) if distance_m is not None else None,
            wavelength_nm=wavelength_nm,
            loss_db=loss_db,
            reflectance_db=reflectance_db,
            slope_dbkm=slope_dbkm,
            total_loss_db=total_loss_db,
            note_original='SmartLink + KeyEvents loss enriched' if ke_rec and loss_db is not None else '',
            label=_get_tag(body, 'label') or (str(ke_rec.get('label') or '') if ke_rec else ''),
        ))
    return rows
def _parse_distance_to_km(value: str, unit_hint: Optional[str] = None) -> Optional[float]:
    f = _to_float(value)
    if f is None:
        return None
    unit = (unit_hint or '').lower()
    if 'km' in unit:
        return f
    if unit in {'m', 'meter', 'meters'}:
        return f / 1000.0
    if f > 1000:
        return f / 1000.0
    return f


def parse_text_report_events(file_name: str, raw: bytes) -> list[EventRow]:
    text = raw.decode('latin1', 'ignore')
    lines = [line.strip() for line in text.replace('\r', '\n').split('\n') if line.strip()]
    if not lines:
        return []

    wavelength = _preferred_msor_wavelength(file_name, raw, text) if Path(file_name).suffix.lower() == '.msor' else _extract_wavelength_from_text(text)
    rows: list[EventRow] = []

    # CSV / delimited table support
    sample = '\n'.join(lines[:80])
    if ',' in sample or ';' in sample or '\t' in sample:
        for delim in [',', ';', '\t']:
            try:
                reader = csv.DictReader(io.StringIO('\n'.join(lines)), delimiter=delim)
                fieldnames = [f.lower() for f in (reader.fieldnames or [])]
                if not fieldnames:
                    continue
                has_distance = any('distance' in f or 'loc' in f or 'position' in f for f in fieldnames)
                has_loss = any('loss' in f or 'il' in f for f in fieldnames)
                if not has_distance or not has_loss:
                    continue
                idx = 1
                for row in reader:
                    norm = {k.lower(): (v or '').strip() for k, v in row.items() if k}
                    distance_field = next((norm[k] for k in norm if 'distance' in k or 'loc' in k or 'position' in k), '')
                    unit_hint = next((k for k in norm if 'distance' in k or 'loc' in k or 'position' in k), '')
                    loss_field = next((norm[k] for k in norm if ('loss' in k or 'il' in k) and 'total' not in k), '')
                    event_field = next((norm[k] for k in norm if 'event' in k or 'type' in k), 'Splice')
                    reflect_field = next((norm[k] for k in norm if 'reflect' in k), '')
                    distance_km = _parse_distance_to_km(distance_field, unit_hint)
                    loss_db = _to_float(loss_field)
                    if distance_km is None:
                        continue
                    rows.append(EventRow(
                        file_name=file_name,
                        event_no=str(idx),
                        event_type=_normalize_event_type(event_field or 'Splice'),
                        distance_m=distance_km * 1000.0,
                        distance_km=round(distance_km, 6),
                        wavelength_nm=wavelength,
                        loss_db=loss_db,
                        reflectance_db=_to_float(reflect_field),
                        slope_dbkm=None,
                        total_loss_db=None,
                        note_original='',
                        label='',
                    ))
                    idx += 1
                if rows:
                    return rows
            except Exception:
                continue

    # Free-text report support
    event_regex = re.compile(
        r'^(?:(?P<no>\d{1,3})\s+)?(?P<etype>fiber\s*end|splice|connector|reflective|non-reflective|gainer|end)[^\d\n\r]{0,20}'
        r'(?P<distance>\d+(?:\.\d+)?)\s*(?P<unit>km|m)?[^\n\r]{0,40}?'
        r'(?P<loss>[+-]?\d+(?:\.\d+)?)\s*dB',
        re.I,
    )
    idx = 1
    for line in lines:
        m = event_regex.search(line)
        if not m:
            continue
        distance_km = _parse_distance_to_km(m.group('distance'), m.group('unit'))
        loss_db = _to_float(m.group('loss'))
        if distance_km is None:
            continue
        rows.append(EventRow(
            file_name=file_name,
            event_no=m.group('no') or str(idx),
            event_type=_normalize_event_type(m.group('etype')),
            distance_m=distance_km * 1000.0,
            distance_km=round(distance_km, 6),
            wavelength_nm=wavelength,
            loss_db=loss_db,
            reflectance_db=None,
            slope_dbkm=None,
            total_loss_db=None,
            note_original=line,
            label='',
        ))
        idx += 1
    return rows


def _pick_wavelength(events: list[EventRow]) -> Optional[str]:
    values = [e.wavelength_nm for e in events if e.wavelength_nm]
    if not values:
        return None
    counts = Counter(values)
    preferred_order = ['1550', '1310', '1625', '1490']
    ranked = sorted(
        counts.items(),
        key=lambda kv: (-kv[1], preferred_order.index(kv[0]) if kv[0] in preferred_order else 999, kv[0]),
    )
    return ranked[0][0]


def _scan_summary_value(text: str, patterns: list[str]) -> Optional[float]:
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            val = _to_float(m.group(1))
            unit = m.group(2).lower() if m.lastindex and m.lastindex >= 2 and m.group(2) else ''
            if val is None:
                continue
            if 'km' in unit:
                return val
            if unit == 'm':
                return val / 1000.0
            return val
    return None




def _read_cstr(buf: bytes, offset: int) -> str:
    if offset < 0 or offset >= len(buf):
        return ''
    end = buf.find(b'\x00', offset)
    if end < 0:
        end = len(buf)
    return buf[offset:end].decode('latin1', 'ignore')


def _find_record_offset_by_name(buf: bytes, name: str) -> Optional[int]:
    needle = name.encode('latin1') + b'\x00'
    pos = -1
    while True:
        pos = buf.find(needle, pos + 1)
        if pos < 0:
            return None
        start = pos - 16
        if start < 0:
            continue
        try:
            name_off, _, _, _ = struct.unpack_from('<IIII', buf, start)
        except Exception:
            continue
        if name_off == pos:
            return start


def _parse_trc_record(buf: bytes, offset: int) -> Optional[dict]:
    if offset is None or offset < 0 or offset + 16 > len(buf):
        return None
    try:
        name_off, typ, size, data_off = struct.unpack_from('<IIII', buf, offset)
    except Exception:
        return None
    if name_off < 0 or name_off >= len(buf) or data_off < 0 or data_off > len(buf):
        return None
    name = _read_cstr(buf, name_off)
    if not name:
        return None
    return {
        'off': offset,
        'name_off': name_off,
        'type': typ,
        'size': size,
        'data_off': data_off,
        'name': name,
    }


def _parse_trc_value(buf: bytes, record: dict):
    typ = int(record['type'])
    size = int(record['size'])
    data_off = int(record['data_off'])
    if data_off + size > len(buf):
        size = max(0, len(buf) - data_off)
    if typ == 0:
        count = max(size // 4, 0)
        if data_off + count * 4 > len(buf):
            count = max(0, (len(buf) - data_off) // 4)
        return list(struct.unpack_from('<' + 'I' * count, buf, data_off)) if count else []
    if typ == 1:
        if size == 4 and data_off + 4 <= len(buf):
            return struct.unpack_from('<i', buf, data_off)[0]
        return buf[data_off:data_off + size]
    if typ == 2:
        return buf[data_off:data_off + size]
    if typ == 3:
        if size == 8 and data_off + 8 <= len(buf):
            return struct.unpack_from('<d', buf, data_off)[0]
        return buf[data_off:data_off + size]
    if typ == 4:
        raw = buf[data_off:data_off + size]
        try:
            return raw.decode('utf-16le').rstrip('\x00')
        except Exception:
            return raw.decode('latin1', 'ignore').rstrip('\x00')
    return buf[data_off:data_off + size]


def _parse_trc_object(buf: bytes, offset: int, max_depth: int = 1):
    record = _parse_trc_record(buf, offset)
    if not record:
        return None
    obj = {'__name__': record['name'], '__type__': record['type'], '__off__': offset}
    if record['type'] != 0 or max_depth < 0:
        obj['value'] = _parse_trc_value(buf, record)
        return obj

    child_offsets = _parse_trc_value(buf, record)
    obj['__children__'] = child_offsets
    for child_offset in child_offsets:
        child_record = _parse_trc_record(buf, child_offset)
        if not child_record:
            continue
        if child_record['type'] == 0 and max_depth > 0:
            obj[child_record['name']] = _parse_trc_object(buf, child_offset, max_depth=max_depth - 1)
        else:
            obj[child_record['name']] = _parse_trc_value(buf, child_record)
    return obj




def _parse_trc_record_appregex(buf: bytes, offset: int) -> Optional[dict]:
    if offset is None or offset < 0 or offset >= len(buf):
        return None
    end = buf.find(b'\x00', offset)
    if end < 0:
        return None
    name = buf[offset:end].decode('latin1', 'ignore')
    if not name:
        return None
    pos = end + 1
    if pos + 16 > len(buf):
        return None
    try:
        next_off, typ, size, data_off = struct.unpack_from('<IIII', buf, pos)
    except Exception:
        return None
    if data_off < 0 or data_off > len(buf):
        return None
    if next_off < 0 or next_off > len(buf):
        next_off = 0
    return {
        'off': offset,
        'name': name,
        'name_off': offset,
        'type': typ,
        'size': size,
        'data_off': data_off,
        'next_off': next_off,
    }


def _parse_trc_value_appregex(buf: bytes, record: dict):
    typ = int(record['type'])
    size = int(record['size'])
    data_off = int(record['data_off'])
    if data_off + size > len(buf):
        size = max(0, len(buf) - data_off)
    if typ == 0:
        count = max(size // 4, 0)
        if data_off + count * 4 > len(buf):
            count = max(0, (len(buf) - data_off) // 4)
        return list(struct.unpack_from('<' + 'I' * count, buf, data_off)) if count else []
    if typ == 1:
        if size == 4 and data_off + 4 <= len(buf):
            return struct.unpack_from('<i', buf, data_off)[0]
        return buf[data_off:data_off + size]
    if typ == 2:
        return buf[data_off:data_off + size]
    if typ == 3:
        if size == 8 and data_off + 8 <= len(buf):
            return struct.unpack_from('<d', buf, data_off)[0]
        return buf[data_off:data_off + size]
    if typ == 4:
        raw = buf[data_off:data_off + size]
        try:
            return raw.decode('utf-16le').rstrip('\x00')
        except Exception:
            return raw.decode('latin1', 'ignore').rstrip('\x00')
    return buf[data_off:data_off + size]


def _collect_trc_appregex_records(buf: bytes) -> Optional[list[dict]]:
    start_names = ['OtdrFile', 'RawSamples', 'EventTable']
    starts: list[int] = []
    for name in start_names:
        pos = buf.find(name.encode('latin1') + b'\x00')
        if pos >= 0:
            starts.append(pos)
    if not starts:
        return None

    records: list[dict] = []
    seen: set[int] = set()
    for start in starts:
        off = start
        for _ in range(5000):
            if off <= 0 or off in seen or off >= len(buf):
                break
            seen.add(off)
            rec = _parse_trc_record_appregex(buf, off)
            if not rec:
                break
            records.append(rec)
            nxt = int(rec.get('next_off') or 0)
            if nxt <= 0:
                break
            off = nxt
    records.sort(key=lambda r: int(r['off']))
    return records or None



def _parse_trc_events_appregex(file_name: str, raw: bytes) -> tuple[list[EventRow], Optional[dict]]:
    combined = _decompress_trc_streams(raw)
    if not combined:
        return [], None

    records = _collect_trc_appregex_records(combined)
    if not records:
        return [], None

    by_name: dict[str, list[dict]] = {}
    for rec in records:
        by_name.setdefault(rec['name'], []).append(rec)

    def last_scalar(name: str):
        arr = by_name.get(name) or []
        if not arr:
            return None
        try:
            return _parse_trc_value_appregex(combined, arr[-1])
        except Exception:
            return None

    pulse_val = last_scalar('Pulse')
    wavelength_val = last_scalar('Wavelength')
    wavelength_nm: Optional[str] = None
    if isinstance(pulse_val, float) and 1e-7 < pulse_val < 1e-4:
        wavelength_nm = str(int(round(pulse_val * 1e9)))
    elif isinstance(wavelength_val, (int, float)) and wavelength_val:
        nm = wavelength_val * 1e9 if float(wavelength_val) < 100 else float(wavelength_val)
        if 1000 <= nm <= 1700:
            wavelength_nm = str(int(round(nm)))

    display_range = last_scalar('DisplayRange')
    scale_factor = last_scalar('ScaleFactor')
    spans_length = last_scalar('SpansLength')
    spans_loss = last_scalar('SpansLoss')
    if isinstance(spans_length, (int, float)) and (spans_length <= 0 or spans_length > 1e8):
        spans_length = None
    if isinstance(spans_loss, (int, float)) and (spans_loss < 0 or spans_loss > 1e4):
        spans_loss = None

    raw_samples_blob: Optional[bytes] = None
    raw_rec = (by_name.get('RawSamples') or [None])[-1]
    if raw_rec:
        try:
            raw_val = _parse_trc_value_appregex(combined, raw_rec)
            if isinstance(raw_val, (bytes, bytearray)) and len(raw_val) >= 200:
                raw_samples_blob = bytes(raw_val)
        except Exception:
            pass
    if raw_samples_blob is None:
        big_binary = [r for r in records if int(r.get('type', -1)) == 2 and int(r.get('size', 0)) >= 1000]
        if big_binary:
            best = max(big_binary, key=lambda r: int(r['size']))
            do = int(best['data_off']); sz = int(best['size'])
            if do + sz <= len(combined):
                raw_samples_blob = combined[do:do + sz]

    event_recs = [rec for rec in records if rec['name'].startswith('Event') and rec['name'][5:].isdigit()]
    if not event_recs:
        return [], None

    def event_index(rec: dict) -> int:
        try:
            return int(rec['name'][5:])
        except Exception:
            return 10**9

    event_recs.sort(key=event_index)

    keep_fields = {
        'Type', 'Status', 'SubCursorA', 'CursorA', 'CursorB', 'SubCursorB',
        'Position', 'Length', 'Loss', 'CurveLevel', 'LocalNoise'
    }

    event_dicts: list[dict] = []
    for rec in event_recs:
        start = rec['off']
        data: dict[str, object] = {'__event_name__': rec['name']}
        off = start
        for _ in range(40):
            sub = _parse_trc_record_appregex(combined, off)
            if not sub:
                break
            if off != start and sub['name'].startswith('Event') and sub['name'][5:].isdigit():
                break
            name = sub['name']
            try:
                value = _parse_trc_value_appregex(combined, sub)
            except Exception:
                value = None
            if off == start or name in keep_fields:
                data[name] = value
            if name in {'SubEvents', 'JobInformation'}:
                break
            nxt = int(sub.get('next_off') or 0)
            if nxt <= 0:
                break
            off = nxt
        event_dicts.append(data)

    cumulative_distance_m = 0.0
    cumulative_section_loss = 0.0
    rows: list[EventRow] = []
    rows.append(EventRow(
        file_name=file_name,
        event_no='0',
        event_type='First Connector',
        distance_m=0.0,
        distance_km=0.0,
        wavelength_nm=wavelength_nm,
        loss_db=None,
        reflectance_db=None,
        slope_dbkm=None,
        total_loss_db=None,
        note_original='AppReg Ex',
        label='',
    ))

    final_total_loss: Optional[float] = None
    final_distance_m: Optional[float] = None

    for i in range(1, len(event_dicts)):
        data = event_dicts[i]
        typ = data.get('Type')
        pos = data.get('Position')
        length_val = data.get('Length')
        loss_val = data.get('Loss')
        sub_b = data.get('SubCursorB')

        if typ is None and isinstance(pos, (int, float)) and float(pos) > 0:
            section_distance = float(pos)
            cumulative_distance_m += section_distance
            final_distance_m = cumulative_distance_m
            section_loss = float(length_val) if isinstance(length_val, (int, float)) else None
            if section_loss is not None and section_loss > -0.5:
                cumulative_section_loss += max(section_loss, 0.0)
            rows.append(EventRow(
                file_name=file_name,
                event_no=str(i),
                event_type='Section',
                distance_m=cumulative_distance_m,
                distance_km=round(cumulative_distance_m / 1000.0, 6),
                wavelength_nm=wavelength_nm,
                loss_db=round(section_loss, 3) if section_loss is not None else None,
                reflectance_db=None,
                slope_dbkm=None,
                total_loss_db=None,
                note_original='AppReg Ex section attenuation',
                label='',
            ))
            continue

        if isinstance(typ, int) and typ >= 128:
            end_dist = float(sub_b) if isinstance(sub_b, (int, float)) and float(sub_b) > 0 else cumulative_distance_m
            final_distance_m = end_dist if end_dist > 0 else final_distance_m
            end_loss = float(length_val) if isinstance(length_val, (int, float)) else None
            if end_loss is None:
                final_total_loss = cumulative_section_loss if cumulative_section_loss > 0 else None
                end_note = 'AppReg Ex end (section-sum fallback)'
            else:
                if cumulative_section_loss > 0 and abs(end_loss - cumulative_section_loss) > max(2.0, 0.35 * cumulative_section_loss):
                    final_total_loss = cumulative_section_loss
                    end_note = 'AppReg Ex end (section-sum corrected)'
                else:
                    final_total_loss = end_loss
                    end_note = 'AppReg Ex end'
            rows.append(EventRow(
                file_name=file_name,
                event_no=str(i),
                event_type='Fiber End',
                distance_m=end_dist if end_dist > 0 else None,
                distance_km=round(end_dist / 1000.0, 6) if end_dist > 0 else None,
                wavelength_nm=wavelength_nm,
                loss_db=round(final_total_loss, 3) if final_total_loss is not None else None,
                reflectance_db=None,
                slope_dbkm=None,
                total_loss_db=round(final_total_loss, 3) if final_total_loss is not None else None,
                note_original=end_note,
                label='',
            ))
            continue

        if typ == 0:
            point_dist = float(sub_b) if isinstance(sub_b, (int, float)) and float(sub_b) > 0 else cumulative_distance_m
            point_loss = float(length_val) if isinstance(length_val, (int, float)) else None
            point_reflectance = -abs(float(loss_val)) if isinstance(loss_val, (int, float)) else None
            # Keep high-loss point events as-is. In some AppReg Ex TRC files
            # FastReporter reports a genuinely high point loss at the event
            # location (for example around 20.017 km in s4.trc), and the older
            # heuristic here incorrectly zeroed it only because the value was
            # numerically close to the cumulative section loss.
            point_note = 'AppReg Ex point'
            rows.append(EventRow(
                file_name=file_name,
                event_no=str(i),
                event_type='Reflective Event',
                distance_m=point_dist if point_dist > 0 else None,
                distance_km=round(point_dist / 1000.0, 6) if point_dist > 0 else None,
                wavelength_nm=wavelength_nm,
                loss_db=round(point_loss, 3) if point_loss is not None else None,
                reflectance_db=round(point_reflectance, 3) if point_reflectance is not None else None,
                slope_dbkm=None,
                total_loss_db=None,
                note_original=point_note,
                label='',
            ))

    deduped: list[EventRow] = []
    for row in rows:
        if not deduped:
            deduped.append(row)
            continue
        prev = deduped[-1]
        same_dist = row.distance_m is not None and prev.distance_m is not None and abs(row.distance_m - prev.distance_m) <= 0.5
        if same_dist and row.event_type == prev.event_type and row.loss_db == prev.loss_db and row.reflectance_db == prev.reflectance_db:
            continue
        deduped.append(row)
    rows = deduped

    if final_distance_m is None:
        final_distance_m = max((r.distance_m or 0.0) for r in rows) if rows else None
    if final_total_loss is None:
        final_total_loss = cumulative_section_loss if cumulative_section_loss > 0 else None

    meta = {
        'Wavelength': float(wavelength_nm) if wavelength_nm else wavelength_nm,
        'DisplayRange': float(display_range) if isinstance(display_range, (int, float)) else None,
        'ScaleFactor': float(scale_factor) if isinstance(scale_factor, (int, float)) else None,
        'TraceSpansLength': float(spans_length) if isinstance(spans_length, (int, float)) else None,
        'TraceSpansLoss': float(spans_loss) if isinstance(spans_loss, (int, float)) else None,
        'SpansLength': float(final_distance_m) if isinstance(final_distance_m, (int, float)) else (float(spans_length) if isinstance(spans_length, (int, float)) else None),
        'SpansLoss': float(final_total_loss) if isinstance(final_total_loss, (int, float)) else (float(spans_loss) if isinstance(spans_loss, (int, float)) else None),
        'RawSamples': raw_samples_blob,
        'Range': float(final_distance_m) if isinstance(final_distance_m, (int, float)) else None,
        'parse_mode': 'trc_appregex_sections_points',
    }
    return rows, meta

def _decompress_trc_streams(raw: bytes) -> Optional[bytes]:

    stream_positions = sorted(set(m.start() for m in re.finditer(b'\x78[\x01\x5e\x9c\xda]', raw)))
    if not stream_positions:
        return None

    outputs: list[bytes] = []
    for pos in stream_positions:
        try:
            d = zlib.decompressobj()
            out = d.decompress(raw[pos:])
            if not d.eof or len(out) < 256:
                continue
            # Keep only meaningful streams; skip all-zero padding outputs
            if out.count(0) == len(out):
                continue
            outputs.append(out)
        except Exception:
            continue
    if not outputs:
        return None

    # Heuristic: AppReg/AppReg Ex files often split the object graph across
    # multiple independent zlib streams (metadata, raw samples, event table).
    # Keep all non-empty streams for those files; otherwise keep streams that
    # expose known OTDR object names.
    if raw.startswith(b'AppReg Format Ex') or raw.startswith(b'AppReg Format'):
        return b''.join(outputs)
    meaningful = [out for out in outputs if (b'OtdrData' in out or b'Event0' in out or b'EventTable' in out or b'RawSamples' in out or b'Wavelength' in out)]
    if meaningful:
        outputs = meaningful
    return b''.join(outputs)


def _extract_trc_trace(raw: bytes) -> Optional[dict]:
    combined = _decompress_trc_streams(raw)
    if not combined:
        return None

    trace0_off = _find_record_offset_by_name(combined, 'Trace0')
    if trace0_off is None:
        # Fallback: older/newer files may still expose EventTable directly.
        event_table_off = _find_record_offset_by_name(combined, 'EventTable')
        if event_table_off is None:
            return None
        event_table = _parse_trc_object(combined, event_table_off, max_depth=1)
        return {'combined': combined, 'trace0': {'EventTable': event_table}}

    trace0 = _parse_trc_object(combined, trace0_off, max_depth=1)
    if not trace0:
        return None
    return {'combined': combined, 'trace0': trace0}


def _trc_wavelength_display(trace0: dict) -> str:
    value = trace0.get('Wavelength')
    if isinstance(value, (int, float)):
        nm = value * 1e9 if value < 100 else value
        if 1000 <= nm <= 1700:
            return f'{int(round(nm))} nm'
    return ''


def _extract_trc_graph_end_km(trace0: dict) -> Optional[float]:
    raw_samples = trace0.get('RawSamples')
    range_m = trace0.get('Range')
    if not isinstance(raw_samples, (bytes, bytearray)) or len(raw_samples) < 200:
        return None
    if not isinstance(range_m, (int, float)) or range_m <= 0:
        return None

    sample_count = len(raw_samples) // 2
    if sample_count < 80:
        return None
    values = list(struct.unpack('<' + 'h' * sample_count, raw_samples[:sample_count * 2]))
    smoothed = _moving_average([float(v) for v in values], radius=4)
    tail = smoothed[int(len(smoothed) * 0.88):]
    pre = smoothed[: max(30, int(len(smoothed) * 0.60))]
    if not tail or not pre:
        return None

    floor = median(tail)
    plateau = _percentile(pre, 90)
    dynamic = plateau - floor
    if dynamic <= 0:
        plateau = max(pre)
        dynamic = plateau - floor
    if dynamic <= 0:
        return None

    threshold = floor + 0.18 * dynamic
    stable_count = max(8, int(len(smoothed) * 0.025))
    start_index = max(5, int(len(smoothed) * 0.12))

    target_index: Optional[int] = None
    for idx in range(start_index, len(smoothed) - stable_count):
        window = smoothed[idx: idx + stable_count]
        if sum(value <= threshold for value in window) >= max(stable_count - 1, 1):
            target_index = idx
            break

    if target_index is None:
        diffs = [smoothed[i] - smoothed[i - 1] for i in range(1, len(smoothed))]
        search_from = max(1, int(len(diffs) * 0.20))
        target_index = min(range(search_from, len(diffs)), key=lambda i: diffs[i])

    curve_max_km = float(range_m) / 1000.0
    break_km = (target_index / max(len(smoothed) - 1, 1)) * curve_max_km
    return round(break_km + 1e-12, 3)


def _parse_trc_events(file_name: str, raw: bytes) -> tuple[list[EventRow], Optional[dict]]:
    extracted = _extract_trc_trace(raw)
    if not extracted:
        # AppReg Ex variant used by some EXFO .trc files
        rows, meta = _parse_trc_events_appregex(file_name, raw)
        if rows:
            return rows, meta
        return [], None

    combined = extracted['combined']
    trace0 = extracted['trace0']
    event_table = trace0.get('EventTable')
    event_offsets: list[int] = []

    if isinstance(event_table, dict):
        child_offsets = event_table.get('__children__') or []
        for child_offset in child_offsets:
            child_record = _parse_trc_record(combined, child_offset)
            if child_record and child_record['name'].startswith('Event'):
                event_offsets.append(child_offset)

    if not event_offsets:
        # fallback scan
        for idx in range(0, 256):
            off = _find_record_offset_by_name(combined, f'Event{idx}')
            if off is None:
                break
            event_offsets.append(off)

    rows: list[EventRow] = []
    route_length_m = trace0.get('SpansLength') if isinstance(trace0.get('SpansLength'), (int, float)) else None
    wavelength_display = _trc_wavelength_display(trace0)
    wavelength_nm = wavelength_display.split()[0] if wavelength_display else None

    for idx, event_offset in enumerate(event_offsets):
        event_obj = _parse_trc_object(combined, event_offset, max_depth=1)
        if not isinstance(event_obj, dict):
            continue

        position_m = event_obj.get('Position')
        if not isinstance(position_m, (int, float)):
            continue

        type_code = event_obj.get('Type')
        reflectance = event_obj.get('Reflectance')
        loss_db = event_obj.get('Loss')
        if isinstance(loss_db, float) and loss_db != loss_db:  # NaN
            loss_db = None
        if isinstance(reflectance, float) and reflectance != reflectance:
            reflectance = None

        if isinstance(type_code, int):
            if type_code == 3:
                if idx == 0 or position_m <= 1.0:
                    event_type = 'First Connector'
                elif route_length_m is not None and abs(position_m - route_length_m) <= 200.0:
                    event_type = 'Fiber End'
                else:
                    event_type = 'Connector'
            elif type_code in {1, 2}:
                event_type = 'Splice'
            else:
                event_type = 'Unknown'
        else:
            # Span objects between point events: skip to avoid duplicating segment-loss rows in the Excel format
            continue

        distance_km = round(float(position_m) / 1000.0, 6)
        total_loss_db = None
        if event_type == 'Fiber End' and isinstance(trace0.get('SpansLoss'), (int, float)):
            total_loss_db = float(trace0.get('SpansLoss'))

        rows.append(EventRow(
            file_name=file_name,
            event_no=str(idx),
            event_type=event_type,
            distance_m=float(position_m),
            distance_km=distance_km,
            wavelength_nm=wavelength_nm,
            loss_db=float(loss_db) if isinstance(loss_db, (int, float)) else None,
            reflectance_db=float(reflectance) if isinstance(reflectance, (int, float)) else None,
            slope_dbkm=None,
            total_loss_db=total_loss_db,
            note_original='',
            label='',
        ))

    if not rows:
        rows2, meta2 = _parse_trc_events_appregex(file_name, raw)
        if rows2:
            return rows2, meta2
    return rows, trace0


def parse_trc_events(file_name: str, raw: bytes) -> list[EventRow]:
    rows, _trace0 = _parse_trc_events(file_name, raw)
    return rows

def parse_any_events(file_name: str, raw: bytes) -> list[EventRow]:
    ext = Path(file_name).suffix.lower()
    if ext in {'.trc', '.crt'}:
        rows = parse_trc_events(file_name, raw)
        if rows:
            return rows
    if ext == '.sor':
        rows = parse_standard_sor_events(file_name, raw)
        if rows:
            return rows
    # For VIAVI/JDSU/Acterna MSOR, SmartLink XML often contains clean event
    # positions while the per-event losses live in the binary KeyEvents block.
    # Parse SmartLink first so it can enrich the XML events with KeyEvents data.
    # Generic embedded XML remains the fallback for other vendor XML formats.
    if ext == '.msor':
        rows = parse_smart_link_events(file_name, raw)
        if rows:
            return rows
        rows = parse_embedded_xml_events(file_name, raw)
        if rows:
            return rows
    else:
        rows = parse_embedded_xml_events(file_name, raw)
        if rows:
            return rows
        rows = parse_smart_link_events(file_name, raw)
        if rows:
            return rows
    rows = _parse_msor_binary_keyevents_events(file_name, raw)
    if rows:
        return rows
    return parse_text_report_events(file_name, raw)


def _parse_events_with_context(file_name: str, raw: bytes):
    ext = Path(file_name).suffix.lower()
    text = raw.decode('latin1', 'ignore')
    trc_trace = None
    sor_meta = None
    parse_mode = 'unknown'
    if ext in {'.trc', '.crt'}:
        events, trc_trace = _parse_trc_events(file_name, raw)
        if trc_trace:
            parse_mode = str(trc_trace.get('parse_mode') or 'trc')
    elif ext == '.sor':
        events, sor_meta = _parse_standard_sor_events_with_meta(file_name, raw)
        if sor_meta:
            parse_mode = str(sor_meta.get('parse_mode') or 'standard_sor')
        if not events:
            events = parse_any_events(file_name, raw)
            if events and parse_mode == 'unknown':
                parse_mode = 'sor_fallback'
    else:
        events = parse_any_events(file_name, raw)
        if events:
            if ext == '.msor' and any('keyevents' in (e.note_original or '').lower() for e in events):
                parse_mode = 'msor_smartlink_keyevents_enriched'
            else:
                parse_mode = 'xml_or_text'
    return events, trc_trace, sor_meta, parse_mode, text


def _to_vi_parse_mode(parse_mode: str) -> str:
    mapping = {
        'standard_sor_keyevents': 'SOR chuẩn - KeyEvents',
        'trc_appregex_sections_points': 'TRC - tách điểm theo section',
        'trc_standard': 'TRC chuẩn',
        'sor_fallback': 'SOR - chế độ dự phòng',
        'xml_or_text': 'XML hoặc văn bản',
        'msor_smartlink_keyevents_enriched': 'MSOR VIAVI/JDSU - smart_link + KeyEvents',
        'unknown': 'Chưa xác định',
    }
    return mapping.get(parse_mode, parse_mode)


def _to_vi_confidence(confidence: str) -> str:
    return {'High': 'Cao', 'Medium': 'Trung bình', 'Low': 'Thấp'}.get(confidence, confidence)


def _to_vi_event_type(event_type: str) -> str:
    key = (event_type or '').strip().lower()
    mapping = {
        'fiber end': 'Cuối sợi',
        'section': 'Đoạn tuyến',
        'connector': 'Đầu nối',
        'event': 'Sự kiện',
        'reflective event': 'Sự kiện phản xạ',
        'non-reflective event': 'Sự kiện không phản xạ',
        'non reflective event': 'Sự kiện không phản xạ',
        'splice': 'Mối hàn',
        'splitter': 'Bộ chia',
        'launch': 'Đầu phát',
        'receive': 'Đầu thu',
        'begin': 'Đầu tuyến',
        'end': 'Cuối tuyến',
    }
    return mapping.get(key, event_type)


def extract_raw_event_rows(file_name: str, raw: bytes) -> list[dict]:
    events, trc_trace, sor_meta, parse_mode, _text = _parse_events_with_context(file_name, raw)
    fmt = Path(file_name).suffix.lower().lstrip('.').upper()
    confidence = 'Medium'
    if parse_mode == 'standard_sor_keyevents':
        confidence = 'High'
    elif parse_mode in {'trc_appregex_sections_points', 'trc_standard'}:
        confidence = 'Medium'
    elif parse_mode in {'sor_fallback', 'xml_or_text'}:
        confidence = 'Low'
    rows = []
    for e in events:
        rows.append({
            'file_name': file_name,
            'format': fmt,
            'parse_mode': _to_vi_parse_mode(parse_mode),
            'confidence': _to_vi_confidence(confidence),
            'event_no': e.event_no,
            'event_type': _to_vi_event_type(e.event_type),
            'distance_km': _safe_round3(e.distance_km),
            'loss_db': e.loss_db,
            'reflectance_db': e.reflectance_db,
            'slope_dbkm': e.slope_dbkm,
            'total_loss_db': e.total_loss_db,
            'note_original': e.note_original,
            'label': e.label,
        })
    return rows


def summarize_file(file_name: str, raw: bytes, parsed_context: Optional[tuple] = None) -> FileSummary:
    ext = Path(file_name).suffix.lower()
    if parsed_context is None:
        events, trc_trace, sor_meta, _parse_mode, text = _parse_events_with_context(file_name, raw)
    else:
        events, trc_trace, sor_meta, _parse_mode, text = parsed_context
    if not events:
        raise ValueError(
            f'Chưa bóc được event từ file {file_name}. Với định dạng vendor đặc thù, cần thêm file mẫu để tinh chỉnh parser.'
        )

    selected_wavelength = _pick_wavelength(events)
    filtered = [e for e in events if (selected_wavelength is None or e.wavelength_nm == selected_wavelength)]
    if not filtered:
        filtered = events

    fiber_end_candidates = [e for e in filtered if e.event_type.lower() == 'fiber end']
    if fiber_end_candidates:
        fiber_end = max(fiber_end_candidates, key=lambda e: (e.distance_km is not None, e.distance_km or -1))
    else:
        fiber_end = max(filtered, key=lambda e: (e.distance_km is not None, e.distance_km or -1))

    fiber = sor_meta.get('fiber_label') if sor_meta and sor_meta.get('fiber_label') else Path(file_name).stem
    if not fiber or (ext == '.sor' and str(fiber).strip().isdigit()):
        fiber = Path(file_name).stem

    total_loss_db = fiber_end.total_loss_db if fiber_end.total_loss_db is not None else (sor_meta.get('total_loss_db') if sor_meta else None)
    total_loss_source = 'Tóm tắt SOR / điểm cuối sợi' if total_loss_db is not None else ''
    route_corrected_total_loss_db = None
    if ext == '.msor':
        msor_summary = _parse_msor_keyevents_summary(raw)
        if msor_summary and isinstance(msor_summary.get('total_loss_db'), (int, float)):
            # Prefer the binary KeyEvents route summary for MSOR. In many VIAVI/JDSU
            # files this is the authoritative source for route loss while the smart_link
            # XML only exposes event positions (or zero/empty total_loss_dB fields).
            total_loss_db = float(msor_summary['total_loss_db'])
            total_loss_source = 'Tóm tắt KeyEvents của MSOR'
    if trc_trace is not None and isinstance(trc_trace.get('TraceSpansLoss'), (int, float)):
        if total_loss_db is None:
            total_loss_db = float(trc_trace.get('TraceSpansLoss'))
            total_loss_source = 'Tóm tắt TraceSpansLoss của TRC'
    if total_loss_db is None and trc_trace is not None and isinstance(trc_trace.get('SpansLoss'), (int, float)):
        total_loss_db = float(trc_trace.get('SpansLoss'))
        total_loss_source = 'Tóm tắt SpansLoss của TRC'
    if total_loss_db is None:
        total_loss_db = _scan_summary_value(text, [r'total\s*loss[^\d]*(\d+(?:\.\d+)?)\s*(dB|db)?'])
        if total_loss_db is not None and total_loss_db > 100:
            total_loss_db = None
        elif total_loss_db is not None:
            total_loss_source = 'Quét phần tóm tắt văn bản'
    smart_link_rows = parse_smart_link_events(file_name, raw) if ext == '.sor' else []
    if total_loss_db is None and smart_link_rows:
        total_loss_db = _estimate_total_loss_from_events(filtered)
        if total_loss_db is not None:
            total_loss_source = 'Ước tính từ bảng sự kiện'
    parsed_total_loss_db = round(total_loss_db, 3) if total_loss_db is not None else None
    length_km = _safe_round3(sor_meta.get('length_km')) if sor_meta and sor_meta.get('length_km') is not None else _safe_round3(fiber_end.distance_km)
    if ext == '.msor' and length_km in (None, 0):
        msor_summary = _parse_msor_keyevents_summary(raw)
        # Some MSOR tails expose raw finish distance, but smart_link event positions are
        # usually already reliable. We only keep the finish raw for future tuning and do
        # not attempt a generic raw->km conversion here.
    if trc_trace is not None and isinstance(trc_trace.get('SpansLength'), (int, float)):
        length_km = _safe_round3(float(trc_trace.get('SpansLength')) / 1000.0)
    if length_km is None:
        length_km = _safe_round3(_scan_summary_value(text, [
            r'fiber\s*length[^\d]*(\d+(?:\.\d+)?)\s*(km|m)?',
            r'length[^\d]*(\d+(?:\.\d+)?)\s*(km|m)?',
            r'distance[^\d]*(\d+(?:\.\d+)?)\s*(km|m)?',
        ]))
    if ext == '.sor' and length_km not in (None, 0):
        route_total_loss_db = _estimate_route_total_loss_from_standard_sor(filtered, length_km)
        if route_total_loss_db is not None:
            route_corrected_total_loss_db = round(route_total_loss_db, 3)
            has_strong_keyevents_summary = bool(
                sor_meta
                and sor_meta.get('parse_mode') == 'standard_sor_keyevents'
                and sor_meta.get('total_loss_db') not in (None, 0)
            )
            if total_loss_db in (None, 0):
                total_loss_db = route_total_loss_db
                total_loss_source = 'Hiệu chỉnh theo tuyến'
            else:
                try:
                    parsed_att = float(total_loss_db) / float(length_km)
                except Exception:
                    parsed_att = None
                route_att = route_total_loss_db / float(length_km)
                preterminal_reflective_loss = _find_preterminal_reflective_connector_loss(filtered, length_km)
                suspicious_total = False
                if parsed_att is None or not (0.01 <= parsed_att <= 1.20):
                    suspicious_total = True
                elif not has_strong_keyevents_summary:
                    if parsed_att > max(route_att * 2.40, route_att + 0.35):
                        suspicious_total = True
                    elif parsed_att < min(route_att * 0.25, max(route_att - 0.20, 0.005)):
                        suspicious_total = True

                # Some standard SOR traces carry a tail-summary that includes a very large
                # reflective connector right before the terminal fiber-end marker. That value
                # makes route total loss and average attenuation look far too high in the
                # Excel summary, even though the per-section attenuation is normal.
                if (
                    preterminal_reflective_loss is not None
                    and preterminal_reflective_loss >= 5.0
                    and float(total_loss_db) > float(route_total_loss_db)
                    and (float(total_loss_db) - float(route_total_loss_db)) >= max(4.0, 0.45 * preterminal_reflective_loss)
                    and parsed_att > max(route_att * 1.30, route_att + 0.08)
                ):
                    suspicious_total = True

                if suspicious_total:
                    total_loss_db = route_total_loss_db
                    total_loss_source = 'Hiệu chỉnh theo tuyến'

            if _should_force_route_corrected_total_for_standard_sor(
                summary_total_loss_db=parsed_total_loss_db,
                route_total_loss_db=route_total_loss_db,
                length_km=length_km,
                fiber_end=fiber_end,
                wavelength_display=(sor_meta.get('wavelength_display') if sor_meta else None),
            ):
                total_loss_db = route_total_loss_db
                total_loss_source = 'Hiệu chỉnh theo tuyến'

    attenuation_dbkm = None
    if total_loss_db is not None and length_km not in (None, 0):
        attenuation_dbkm = round(total_loss_db / length_km, 3)

    splice_points: list[tuple[float, float]] = []
    use_all_event_points = bool(sor_meta and sor_meta.get('parse_mode') == 'standard_sor_keyevents')
    trc_point_mode = bool(trc_trace and trc_trace.get('parse_mode') == 'trc_appregex_sections_points')
    for e in filtered:
        et = e.event_type.lower()
        distance = _safe_round3(e.distance_km)
        if distance is None or e.loss_db is None:
            continue
        if trc_point_mode:
            if et in {'connector', 'event', 'reflective event'}:
                splice_points.append((distance, round(e.loss_db, 3)))
        elif use_all_event_points:
            if 'end' not in et:
                splice_points.append((distance, round(e.loss_db, 3)))
        else:
            if ('splice' in et or ('non-reflective' in et)) and e.loss_db is not None:
                splice_points.append((distance, round(e.loss_db, 3)))

    # If no splice tags at all, fall back to non-end, non-section events with loss
    if not splice_points:
        for e in filtered:
            et = e.event_type.lower()
            distance = _safe_round3(e.distance_km)
            if distance is None or e.loss_db is None:
                continue
            if 'end' not in et and et != 'section':
                splice_points.append((distance, round(e.loss_db, 3)))

    wavelength_display = sor_meta.get('wavelength_display') if sor_meta and sor_meta.get('wavelength_display') else (f'{_snap_nominal_wavelength_nm(selected_wavelength)} nm' if selected_wavelength else '')
    if trc_trace is not None:
        wavelength_display = _trc_wavelength_display(trc_trace) or wavelength_display
        graph_end_km = _extract_trc_graph_end_km(trc_trace)
        graph_curve_max_km = _safe_round3(float(trc_trace.get('Range')) / 1000.0) if isinstance(trc_trace.get('Range'), (int, float)) else None
    else:
        if sor_meta is not None:
            graph_end_km = sor_meta.get('graph_end_km')
            graph_curve_max_km = sor_meta.get('graph_curve_max_km')
            if graph_end_km is None and length_km is not None:
                graph_end_km = length_km
            if graph_curve_max_km is None and length_km is not None:
                graph_curve_max_km = length_km
        else:
            graph_end_km, graph_curve_max_km = _estimate_msor_graph_metrics(raw, filtered, length_km)

    family_info = _detect_parser_family(
        file_name, raw, ext, _parse_mode, text=text, sor_meta=sor_meta, summary=None, events=filtered
    )
    return FileSummary(
        file_name=file_name,
        fiber=fiber,
        wavelength_display=wavelength_display,
        total_loss_db=round(total_loss_db, 3) if total_loss_db is not None else None,
        length_km=length_km,
        attenuation_dbkm=attenuation_dbkm,
        splice_points=splice_points,
        end_distance_km=length_km,
        graph_end_km=graph_end_km,
        graph_curve_max_km=graph_curve_max_km,
        source_format=ext.lstrip('.').upper(),
        parsed_total_loss_db=parsed_total_loss_db,
        route_corrected_total_loss_db=route_corrected_total_loss_db,
        loss_source_used=total_loss_source or '',
        parse_family=family_info.get('family', ''),
        parse_family_confidence=family_info.get('confidence', ''),
        parse_family_reason=family_info.get('reason', ''),
    )


def _cluster_distances(distance_values: list[float], deviation_m: float) -> list[DistanceCluster]:
    unique = sorted({round(v, 3) for v in distance_values if v is not None})
    if not unique:
        return []

    deviation_km = max(float(deviation_m), 0.0) / 1000.0
    if deviation_km <= 0:
        return [DistanceCluster(representative_km=v, values_km=[v]) for v in unique]

    clusters: list[list[float]] = []
    current: list[float] = [unique[0]]
    running_mean = unique[0]

    for value in unique[1:]:
        if abs(value - running_mean) <= deviation_km:
            current.append(value)
            running_mean = sum(current) / len(current)
        else:
            clusters.append(current)
            current = [value]
            running_mean = value
    clusters.append(current)

    result: list[DistanceCluster] = []
    for cluster in clusters:
        representative = round(sum(cluster) / len(cluster) + 1e-12, 3)
        result.append(DistanceCluster(representative_km=representative, values_km=cluster))
    return result





def _normalize_graph_check_tolerances(length_tolerance_km: float, graph_reach_tolerance_km: Optional[float], event_shortfall_tolerance_km: Optional[float], overlength_tolerance_km: Optional[float]) -> tuple[float, float, float]:
    legacy_tolerance_km = max(float(length_tolerance_km), 0.0)

    reach_tol = legacy_tolerance_km if graph_reach_tolerance_km is None else max(float(graph_reach_tolerance_km), 0.0)
    shortfall_tol = max(reach_tol, 0.500) if event_shortfall_tolerance_km is None else max(float(event_shortfall_tolerance_km), 0.0)
    # Giữ biến trả về thứ ba để không làm gãy cấu trúc cũ, nhưng từ bản này
    # không còn dùng ngưỡng vượt dài để chặn các core dài hơn tuyến chuẩn.
    over_tol = 0.0
    return reach_tol, shortfall_tol, over_tol


def _assess_graph_length(
    summary: FileSummary,
    expected_route_km: Optional[float],
    jumper_excluded_m: float,
    length_tolerance_km: float,
    graph_reach_tolerance_km: Optional[float] = None,
    event_shortfall_tolerance_km: Optional[float] = None,
    overlength_tolerance_km: Optional[float] = None,
) -> GraphAssessment:
    jumper_km = max(float(jumper_excluded_m), 0.0) / 1000.0
    reach_tol_km, shortfall_tol_km, over_tol_km = _normalize_graph_check_tolerances(
        length_tolerance_km=length_tolerance_km,
        graph_reach_tolerance_km=graph_reach_tolerance_km,
        event_shortfall_tolerance_km=event_shortfall_tolerance_km,
        overlength_tolerance_km=overlength_tolerance_km,
    )

    if summary.graph_end_km is None:
        return GraphAssessment(
            graph_end_km=None,
            jumper_excluded_km=round(jumper_km, 3),
            net_graph_length_km=None,
            expected_route_km=expected_route_km,
            event_length_km=summary.length_km,
            diff_km=None,
            graph_reach_tolerance_km=round(reach_tol_km, 3),
            event_shortfall_tolerance_km=round(shortfall_tol_km, 3),
            overlength_tolerance_km=round(over_tol_km, 3),
            graph_reaches_expected=None,
            verdict='Không đọc được đồ thị',
            reason='File hiện chưa tách được điểm cuối từ đồ thị mini-curve; vẫn có thể dùng event table bình thường.',
        )

    net_length_km = max(summary.graph_end_km - jumper_km, 0.0)
    rounded_net_length_km = round(net_length_km + 1e-12, 3)
    rounded_event_length_km = _safe_round3(summary.length_km)

    if expected_route_km is None or expected_route_km <= 0:
        return GraphAssessment(
            graph_end_km=summary.graph_end_km,
            jumper_excluded_km=round(jumper_km, 3),
            net_graph_length_km=rounded_net_length_km,
            expected_route_km=expected_route_km,
            event_length_km=rounded_event_length_km,
            diff_km=None,
            graph_reach_tolerance_km=round(reach_tol_km, 3),
            event_shortfall_tolerance_km=round(shortfall_tol_km, 3),
            overlength_tolerance_km=round(over_tol_km, 3),
            graph_reaches_expected=None,
            verdict='',
            reason='',
        )

    diff_km = round(net_length_km - expected_route_km, 3)
    graph_reaches_expected = net_length_km >= (expected_route_km - reach_tol_km)

    event_length_km = summary.length_km
    event_missing = event_length_km in (None, 0)

    event_is_enough_direct = False
    event_is_rescued_by_graph = False

    if not event_missing and event_length_km is not None:
        event_is_enough_direct = event_length_km >= (expected_route_km - reach_tol_km)
        event_is_rescued_by_graph = (
            event_length_km < (expected_route_km - reach_tol_km)
            and (expected_route_km - event_length_km) <= shortfall_tol_km
            and graph_reaches_expected
        )

    if event_is_enough_direct:
        verdict = 'Đủ tuyến'
        if graph_reaches_expected:
            reason = 'Chiều dài event table đã đạt hoặc vượt mốc đủ tuyến và đồ thị cũng chạy tới chiều dài chuẩn.'
        else:
            reason = 'Chiều dài event table đã đạt hoặc vượt mốc đủ tuyến; đồ thị mini-curve có thể đọc thiếu nên vẫn tính đủ tuyến theo chiều dài đo.'
    elif event_is_rescued_by_graph:
        verdict = 'Đủ tuyến'
        reason = 'Chiều dài event table ngắn hơn chuẩn nhưng vẫn nằm trong ngưỡng chấp nhận; đồ thị vẫn chạy tới chiều dài chuẩn nên tính đủ tuyến.'
    elif event_missing and graph_reaches_expected:
        verdict = 'Đủ tuyến'
        reason = 'Không có chiều dài event table đáng tin cậy, nhưng đồ thị đã chạy tới chiều dài chuẩn nên vẫn tính đủ tuyến.'
    elif graph_reaches_expected:
        verdict = 'Nghi ngờ đồ thị'
        reason = 'Đồ thị chạy tới chiều dài chuẩn nhưng chiều dài event table ngắn hơn chuẩn quá nhiều; cần kiểm tra lại cách đánh dấu Fiber End hoặc dữ liệu đo.'
    else:
        verdict = 'Đứt'
        reason = 'Chiều dài đo chưa đạt dải đủ tuyến và đồ thị cũng dừng trước chiều dài chuẩn.'

    return GraphAssessment(
        graph_end_km=summary.graph_end_km,
        jumper_excluded_km=round(jumper_km, 3),
        net_graph_length_km=rounded_net_length_km,
        expected_route_km=expected_route_km,
        event_length_km=rounded_event_length_km,
        diff_km=diff_km,
        graph_reach_tolerance_km=round(reach_tol_km, 3),
        event_shortfall_tolerance_km=round(shortfall_tol_km, 3),
        overlength_tolerance_km=round(over_tol_km, 3),
        graph_reaches_expected=graph_reaches_expected,
        verdict=verdict,
        reason=reason,
    )



def _clip_overlap(start_a: float, end_a: float, start_b: float, end_b: float) -> float:
    return max(min(end_a, end_b) - max(start_a, start_b), 0.0)


def _standard_sor_discrete_loss_allowed(row: EventRow, length_km: float, ordered: list[EventRow], idx: int, cluster_gap_km: float, endpoint_zone_km: float) -> bool:
    loss = row.loss_db
    if loss is None:
        return False
    loss_f = float(loss)
    if loss_f <= 0:
        return False

    dist = float(row.distance_km or 0.0)
    et = (row.event_type or '').strip()
    near_start = dist <= endpoint_zone_km
    near_end = (float(length_km) - dist) <= endpoint_zone_km
    terminal = _is_standard_sor_terminal_event(row)
    reflective = _is_standard_sor_reflective_event(row)

    if terminal:
        return False
    if et == 'First Connector':
        return False
    if reflective and (near_start or near_end):
        return False
    if et == 'Connector' and reflective and loss_f > 1.0:
        return False
    if et == 'Connector' and near_start and loss_f > 0.75:
        return False

    prev_loss = None
    next_loss = None
    prev_gap = None
    next_gap = None
    if idx > 0:
        prev_row = ordered[idx - 1]
        prev_gap = dist - float(prev_row.distance_km or 0.0)
        prev_loss = float(prev_row.loss_db) if prev_row.loss_db is not None else None
    if idx + 1 < len(ordered):
        next_row = ordered[idx + 1]
        next_gap = float(next_row.distance_km or 0.0) - dist
        next_loss = float(next_row.loss_db) if next_row.loss_db is not None else None

    clustered_high_loss = False
    if loss_f > 1.0 and not reflective:
        if near_end or near_start:
            clustered_high_loss = True
        if prev_gap is not None and prev_gap <= cluster_gap_km and prev_loss is not None and prev_loss > 0.8:
            clustered_high_loss = True
        if next_gap is not None and next_gap <= cluster_gap_km and next_loss is not None and next_loss > 0.8:
            clustered_high_loss = True
        if loss_f >= 3.0:
            clustered_high_loss = True
    if clustered_high_loss:
        return False
    return True


def _estimate_segment_loss_standard_sor(rows: list[EventRow], start_km: float, end_km: float, route_length_km: Optional[float]) -> tuple[Optional[float], str]:
    ordered = [r for r in rows if r.distance_km is not None]
    if not ordered:
        return None, 'Không có event chuẩn hóa để tính đoạn.'
    ordered = sorted(ordered, key=lambda r: (r.distance_km or 0.0, r.event_no))
    if route_length_km in (None, 0):
        route_length_km = _choose_standard_sor_end_distance_km(ordered, None)
    if route_length_km in (None, 0):
        route_length_km = max(float(r.distance_km or 0.0) for r in ordered)

    route_length_km = float(route_length_km)
    start_km = max(0.0, min(float(start_km), route_length_km))
    end_km = max(start_km, min(float(end_km), route_length_km))
    endpoint_zone_km = max(0.30, min(2.00, 0.02 * route_length_km))
    cluster_gap_km = max(0.40, min(0.80, 0.01 * route_length_km))

    fiber_section_loss = 0.0
    discrete_event_loss = 0.0
    prev_distance = 0.0

    for idx, row in enumerate(ordered):
        dist = min(float(row.distance_km or 0.0), route_length_km)
        seg_len = max(dist - prev_distance, 0.0)
        overlap_km = _clip_overlap(prev_distance, prev_distance + seg_len, start_km, end_km)
        slope = row.slope_dbkm
        if overlap_km > 0 and slope is not None:
            slope_f = float(slope)
            if 0 < slope_f < 0.45:
                fiber_section_loss += slope_f * overlap_km
            elif 0.45 <= slope_f < 1.0 and seg_len <= 0.25:
                fiber_section_loss += 0.25 * overlap_km
        prev_distance = dist

        if not (start_km <= dist <= end_km):
            continue
        if not _standard_sor_discrete_loss_allowed(row, route_length_km, ordered, idx, cluster_gap_km, endpoint_zone_km):
            continue

        loss_f = float(row.loss_db or 0.0)
        reflective = _is_standard_sor_reflective_event(row)
        if not reflective and loss_f > 1.0:
            loss_f = 1.0
        discrete_event_loss += loss_f

    total = fiber_section_loss + discrete_event_loss
    if total <= 0:
        return None, 'Không đủ điểm event hoặc slope hợp lệ để ước tính suy hao đoạn.'
    return round(total, 3), 'Ước tính theo slope từng đoạn và các event hợp lệ trong khoảng đã chọn.'


def _estimate_segment_loss_twopoint(
    rows: list[EventRow],
    start_km: float,
    end_km: float,
    route_length_km: Optional[float],
    total_loss_db_end: Optional[float],
) -> tuple[Optional[float], str]:
    """Two-point estimation using cumulative total-loss interpolation."""
    pts = [(0.0, 0.0)]
    for r in rows:
        if r.distance_km is None or r.total_loss_db is None:
            continue
        try:
            d = float(r.distance_km)
            tl = float(r.total_loss_db)
        except Exception:
            continue
        if d < 0 or tl < 0 or tl > 500:
            continue
        pts.append((d, tl))
    if route_length_km not in (None, 0) and total_loss_db_end not in (None, 0):
        pts.append((float(route_length_km), float(total_loss_db_end)))

    if len(pts) < 3:
        return None, 'Không đủ điểm total_loss để nội suy 2 điểm.'

    pts.sort(key=lambda x: x[0])
    dedup = []
    last_d = None
    cur_max = 0.0
    for d, tl in pts:
        if last_d is not None and abs(d - last_d) < 1e-6:
            if dedup:
                dedup[-1] = (d, max(dedup[-1][1], tl))
            else:
                dedup.append((d, tl))
            continue
        cur_max = max(cur_max, tl)
        dedup.append((d, cur_max))
        last_d = d

    def interp(x: float) -> float:
        x = float(x)
        if x <= dedup[0][0]:
            return dedup[0][1]
        if x >= dedup[-1][0]:
            return dedup[-1][1]
        for i in range(1, len(dedup)):
            d0, v0 = dedup[i - 1]
            d1, v1 = dedup[i]
            if d0 <= x <= d1 and d1 > d0:
                t = (x - d0) / (d1 - d0)
                return v0 + t * (v1 - v0)
        return dedup[-1][1]

    if route_length_km in (None, 0):
        route_length_km = dedup[-1][0]

    start_km = max(0.0, min(float(start_km), float(route_length_km)))
    end_km = max(start_km, min(float(end_km), float(route_length_km)))
    loss = interp(end_km) - interp(start_km)
    if loss <= 0:
        return None, 'Nội suy 2 điểm cho ra loss không dương.'
    return round(loss, 3), 'Nội suy total_loss_db tại 2 điểm rồi lấy hiệu.'


def _estimate_segment_loss_generic(rows: list[EventRow], start_km: float, end_km: float, total_loss_db: Optional[float], length_km: Optional[float]) -> tuple[Optional[float], str]:
    ordered = [r for r in rows if r.distance_km is not None]
    if not ordered:
        return None, 'Không có event để tính đoạn.'
    ordered = sorted(ordered, key=lambda r: (r.distance_km or 0.0, r.event_no))

    fiber_section_loss = 0.0
    discrete_event_loss = 0.0
    prev_distance = 0.0

    for row in ordered:
        dist = float(row.distance_km or 0.0)
        seg_len = max(dist - prev_distance, 0.0)
        overlap_km = _clip_overlap(prev_distance, prev_distance + seg_len, start_km, end_km)
        slope = row.slope_dbkm
        if overlap_km > 0 and slope is not None:
            slope_f = float(slope)
            if 0 < slope_f < 0.60:
                fiber_section_loss += slope_f * overlap_km
        prev_distance = dist

        if not (start_km <= dist <= end_km):
            continue

        et = (row.event_type or '').strip().lower()
        if 'fiber end' in et or 'cuối sợi' in et or et == 'section' or et == 'đoạn tuyến' or 'first connector' in et or 'đầu phát' in et:
            continue
        loss = row.loss_db
        if loss is None:
            continue
        loss_f = float(loss)
        if loss_f <= 0:
            continue
        if loss_f > 5.0:
            continue
        discrete_event_loss += loss_f

    total = fiber_section_loss + discrete_event_loss
    if total > 0:
        return round(total, 3), 'Ước tính theo slope và các event dương nằm trong đoạn đã chọn.'

    if total_loss_db not in (None, 0) and length_km not in (None, 0):
        overall_att = float(total_loss_db) / float(length_km)
        est = overall_att * max(end_km - start_km, 0.0)
        if est > 0:
            return round(est, 3), 'Ước tính theo suy hao trung bình toàn tuyến do file chưa đủ dữ liệu event trong đoạn.'
    return None, 'Không đủ dữ liệu để ước tính suy hao đoạn.'


def _build_segment_assessment(file_name: str, raw: bytes, summary: FileSummary, segment_start_km: Optional[float], segment_end_km: Optional[float]) -> Optional[SegmentAssessment]:
    if segment_start_km is None or segment_end_km is None:
        return None

    start_km = float(segment_start_km)
    end_km = float(segment_end_km)
    if end_km < start_km:
        start_km, end_km = end_km, start_km
    span_km = max(end_km - start_km, 0.0)
    if span_km <= 0:
        return SegmentAssessment(
            start_km=round(start_km, 3),
            end_km=round(end_km, 3),
            span_km=0.0,
            event_count=0,
            segment_total_loss_db=None,
            segment_attenuation_dbkm=None,
            max_positive_event_loss_db=None,
            max_negative_event_loss_db=None,
            note='Khoảng phân tích không hợp lệ.',
            recommendation='Nhập lại mốc đầu và mốc cuối sao cho km cuối lớn hơn km đầu.',
            method='',
        )

    events, _trc_trace, sor_meta, _parse_mode, _text = _parse_events_with_context(file_name, raw)
    if not events:
        return SegmentAssessment(
            start_km=round(start_km, 3),
            end_km=round(end_km, 3),
            span_km=round(span_km, 3),
            event_count=0,
            segment_total_loss_db=None,
            segment_attenuation_dbkm=None,
            max_positive_event_loss_db=None,
            max_negative_event_loss_db=None,
            note='Không bóc được event để phân tích đoạn tuyến.',
            recommendation='Kiểm tra lại file đo hoặc dùng phần mềm gốc để đối chiếu thêm.',
            method='',
        )

    selected_wavelength = _pick_wavelength(events)
    filtered = [e for e in events if (selected_wavelength is None or e.wavelength_nm == selected_wavelength)]
    if not filtered:
        filtered = events

    route_length_km = summary.length_km
    effective_end_km = end_km
    if route_length_km not in (None, 0):
        if start_km >= float(route_length_km):
            return SegmentAssessment(
                start_km=round(start_km, 3),
                end_km=round(end_km, 3),
                span_km=round(span_km, 3),
                event_count=0,
                segment_total_loss_db=None,
                segment_attenuation_dbkm=None,
                max_positive_event_loss_db=None,
                max_negative_event_loss_db=None,
                note='Đoạn đã chọn nằm ngoài chiều dài tuyến hiện có.',
                recommendation='Giảm mốc km hoặc kiểm tra lại chiều dài tuyến chuẩn của file đo.',
                method='',
            )
        effective_end_km = min(end_km, float(route_length_km))

    segment_rows = [e for e in filtered if e.distance_km is not None and start_km <= float(e.distance_km) <= effective_end_km]
    segment_losses = [float(e.loss_db) for e in segment_rows if e.loss_db is not None]
    positive_losses = [loss for loss in segment_losses if loss > 0]
    negative_losses = [loss for loss in segment_losses if loss < 0]
    max_positive_event_loss_db = round(max(positive_losses), 3) if positive_losses else None
    max_negative_event_loss_db = round(min(negative_losses), 3) if negative_losses else None

    actual_span_km = max(effective_end_km - start_km, 0.0)

    if actual_span_km <= 0:
        segment_total_loss_db = None
        segment_attenuation_dbkm = None
        method = ''
        note = 'Đoạn đã chọn không còn chiều dài khả dụng sau khi đối chiếu với file đo.'
        recommendation = 'Kiểm tra lại mốc km đầu/cuối.'
    elif not segment_losses:
        segment_total_loss_db = None
        segment_attenuation_dbkm = None
        method = ''
        note = 'Không có event suy hao trong đoạn đã chọn.'
        recommendation = 'Giữ nguyên mốc đoạn để theo dõi; nếu cần đánh giá chi tiết hơn, đối chiếu lại file gốc.'
    else:
        ext = Path(file_name).suffix.lower()
        parse_mode = sor_meta.get('parse_mode') if sor_meta else ''
        if ext == '.sor' and parse_mode == 'standard_sor_keyevents':
            segment_total_loss_db, method = _estimate_segment_loss_standard_sor(filtered, start_km, effective_end_km, route_length_km)
        else:
            segment_total_loss_db, method = _estimate_segment_loss_generic(filtered, start_km, effective_end_km, summary.total_loss_db, route_length_km)

        segment_attenuation_dbkm = None
        if segment_total_loss_db is not None and actual_span_km > 0:
            segment_attenuation_dbkm = round(segment_total_loss_db / actual_span_km, 3)

        if segment_total_loss_db is None:
            note = 'Chưa đủ dữ liệu event hợp lệ để tính suy hao riêng cho đoạn này.'
            recommendation = 'Đối chiếu lại file đo gốc hoặc giảm ngưỡng lọc event để lấy thêm điểm tham chiếu trong đoạn.'
        else:
            if (segment_attenuation_dbkm is not None and segment_attenuation_dbkm <= 0.25) and (max_positive_event_loss_db or 0) < 0.3:
                note = 'Đoạn tuyến đang ổn định, suy hao điểm và suy hao trung bình đều ở mức tốt.'
                recommendation = 'Tiếp tục vận hành bình thường và theo dõi định kỳ.'
            elif (segment_attenuation_dbkm is not None and segment_attenuation_dbkm <= 0.35) and (max_positive_event_loss_db or 0) < 0.7:
                note = 'Đoạn tuyến có một số điểm suy hao cần theo dõi nhưng chưa thấy dấu hiệu bất thường lớn.'
                recommendation = 'Rà soát các event có suy hao cao nhất trong đoạn và so sánh với lịch sử đo gần nhất.'
            elif (max_positive_event_loss_db or 0) >= 1.0 or (segment_attenuation_dbkm is not None and segment_attenuation_dbkm > 0.45):
                note = 'Đoạn tuyến có suy hao cao hoặc xuất hiện event bất thường cần ưu tiên kiểm tra.'
                recommendation = 'Ưu tiên kiểm tra thực địa các vị trí event lớn trong đoạn, đối chiếu thêm bằng file đo gốc và lịch sử bảo trì.'
            else:
                note = 'Đoạn tuyến có dấu hiệu tăng suy hao so với mức vận hành thông thường.'
                recommendation = 'Theo dõi sát ở lần đo tiếp theo; nếu xu hướng lặp lại, cần kiểm tra mối hàn/đầu nối trong đoạn.'

    return SegmentAssessment(
        start_km=round(start_km, 3),
        end_km=round(effective_end_km, 3),
        span_km=round(actual_span_km, 3),
        event_count=len(segment_rows),
        segment_total_loss_db=segment_total_loss_db,
        segment_attenuation_dbkm=segment_attenuation_dbkm,
        max_positive_event_loss_db=max_positive_event_loss_db,
        max_negative_event_loss_db=max_negative_event_loss_db,
        note=note,
        recommendation=recommendation,
        method=method,
    )

def _extract_segment_event_rows(file_name: str, raw: bytes, segment_start_km: Optional[float], segment_end_km: Optional[float], threshold_db: float) -> list[dict]:
    if segment_start_km is None or segment_end_km is None:
        return []
    start_km = min(float(segment_start_km), float(segment_end_km))
    end_km = max(float(segment_start_km), float(segment_end_km))
    rows = extract_raw_event_rows(file_name, raw)
    return [row for row in rows if row.get('distance_km') is not None and start_km <= float(row['distance_km']) <= end_km and isinstance(row.get('loss_db'), (int, float)) and float(row['loss_db']) > 0]


def _fr_copy_rows_to_sheet(ws, rows: list[list[object]]) -> None:
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)


_TABLES_TEMPLATE_ROWS = [
    ['SheetName', 'TagType', 'Value', 'Address', 'Row', 'Column'],
    ['General Information', 'SingleCell', '<<Identification.FileName>>', '$C$3', 3, 3],
    ['General Information', 'SingleCell', '<<Identification.JobID>>', '$K$3', 3, 11],
    ['General Information', 'SingleCell', '<<Identification.TestDate>>', '$C$4', 4, 3],
    ['General Information', 'SingleCell', '<<Identification.Customer>>', '$K$4', 4, 11],
    ['General Information', 'SingleCell', '<<Identification.TestTime>>', '$C$5', 5, 3],
    ['General Information', 'SingleCell', '<<Identification.Company>>', '$K$5', 5, 11],
    ['General Information', 'SingleCell', '<<Identification.Comments>>', '$C$6', 6, 3],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier1Name>>', '$A$8', 8, 1],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier2Name>>', '$D$8', 8, 4],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier3Name>>', '$G$8', 8, 7],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier4Name>>', '$K$8', 8, 11],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier5Name>>', '$N$8', 8, 14],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier1Value>>', '$A$9', 9, 1],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier2Value>>', '$D$9', 9, 4],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier3Value>>', '$G$9', 9, 7],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier4Value>>', '$K$9', 9, 11],
    ['General Information', 'SingleCell', '<<CustomIdentifiers.Identifier5Value>>', '$N$9', 9, 14],
    ['General Information', 'SingleCell', '<<Identification.LocationA>>', '$C$13', 13, 3],
    ['General Information', 'SingleCell', '<<Identification.LocationB>>', '$K$13', 13, 11],
    ['General Information', 'SingleCell', '<<Identification.OperatorA>>', '$C$14', 14, 3],
    ['General Information', 'SingleCell', '<<Identification.OperatorB>>', '$K$14', 14, 11],
    ['General Information', 'SingleCell', '<<Identification.LocationAUnitModel>>', '$C$15', 15, 3],
    ['General Information', 'SingleCell', '<<Identification.LocationBUnitModel>>', '$K$15', 15, 11],
    ['General Information', 'SingleCell', '<<Identification.LocationAUnitSerialNumber>>', '$C$16', 16, 3],
    ['General Information', 'SingleCell', '<<Identification.LocationBUnitSerialNumber>>', '$K$16', 16, 11],
    ['General Information', 'SingleCell', 'Range (<<GeneralSettings.DistanceUnit>>):', '$D$20', 20, 4],
    ['General Information', 'SingleCell', 'Pulse (<<OTDRSettings.PulseUnit>>):', '$J$20', 20, 10],
    ['General Information', 'SingleCell', 'Position (<<GeneralSettings.DistanceUnit>>)', '$D$23', 23, 4],
    ['General Information', 'MultiRow', '$$OTDRSummary.RealWavelength$$', '$B$20', 20, 2],
    ['General Information', 'MultiRow', '$$OTDRSummary.Range$$', '$E$20', 20, 5],
    ['General Information', 'MultiRow', '$$OTDRSummary.Pulse$$', '$K$20', 20, 11],
    ['General Information', 'MultiRow', '$$OTDRSummary.Duration$$', '$N$20', 20, 14],
    ['General Information', 'MultiRow', '$$OTDRMacrobend.FiberID$$', '$A$24', 24, 1],
    ['General Information', 'MultiRow', '$$OTDRMacrobend.Position$$', '$D$24', 24, 4],
    ['General Information', 'MultiRow', '$$OTDRMacrobend.Wavelength1$$', '$F$24', 24, 6],
    ['General Information', 'MultiRow', '$$OTDRMacrobend.EventLoss1$$', '$I$24', 24, 9],
    ['General Information', 'MultiRow', '$$OTDRMacrobend.Wavelength2$$', '$K$24', 24, 11],
    ['General Information', 'MultiRow', '$$OTDRMacrobend.EventLoss2$$', '$M$24', 24, 13],
    ['Link Results', 'SingleCell', 'Span                                                           Length                                 (<<GeneralSettings.DistanceUnit>>)', '$P$2', 2, 16],
    ['Link Results', 'MultiRow', '$$Identification.FiberID$$', '$A$3', 3, 1],
    ['Link Results', 'MultiRow', '$$OTDRSummary.RealWavelength$$', '$E$3', 3, 5],
    ['Link Results', 'MultiRow', '$$OtdrEventTableCalculatedColumns.MaximumEventLoss$$', '$G$3', 3, 7],
    ['Link Results', 'MultiRow', '$$OtdrEventTableCalculatedColumns.MaximumSectionLoss$$', '$J$3', 3, 10],
    ['Link Results', 'MultiRow', '$$OTDRSummary.AverageSpliceLoss$$', '$L$3', 3, 12],
    ['Link Results', 'MultiRow', '$$OtdrEventTableCalculatedColumns.AverageSectionLoss$$', '$N$3', 3, 14],
    ['Link Results', 'MultiRow', '$$OTDRSummary.SpanLength$$', '$P$3', 3, 16],
    ['Link Results', 'MultiRow', '$$OTDRSummary.SpanLengthPassFailStatus$$', '$R$3', 3, 18],
    ['Link Results', 'MultiRow', '$$OTDRSummary.SpanLoss$$', '$S$3', 3, 19],
    ['Link Results', 'MultiRow', '$$OTDRSummary.SpanLossPassFailStatus$$', '$U$3', 3, 21],
    ['Link Results', 'MultiRow', '$$OTDRSummary.SpanORL$$', '$V$3', 3, 22],
    ['Link Results', 'MultiRow', '$$OTDRSummary.SpanORLPassFailStatus$$', '$X$3', 3, 24],
    ['Link Results', 'MultiRow', '$$OTDRSummary.NumberOfEvents$$', '$Y$3', 3, 25],
    ['Events', 'SingleCell', '<<GeneralSettings.DistanceUnit>>', '$D$3', 3, 4],
    ['Events', 'BlockRow', '##Identification.FiberID##', '$A$5', 5, 1],
    ['Events', 'BlockRow', '##OTDRSummary.RealWavelength##', '$B$5', 5, 2],
    ['Events', 'Block', 'Event [[OTDREventTable.EventNumber]]', '$C$1', 1, 3],
    ['Events', 'Block', '[[OTDREventTable.EventType]]', '$C$2', 2, 3],
    ['Events', 'Block', '[[OTDREventTable.Position]]', '$C$3', 3, 3],
    ['Events', 'Block', '[[##OTDREventTable.EventLoss##]]', '$C$5', 5, 3],
    ['Events', 'Block', '[[##OTDREventTable.Reflectance##]]', '$D$5', 5, 4],
    ['Events', 'Block', '[[##OTDREventTable.EventLossPassFailStatus##]]', '$C$6', 6, 3],
    ['Events', 'Block', '[[##OTDREventTable.ReflectancePassFailStatus##]]', '$D$6', 6, 4],
    ['Events', 'Block', '[[=FRMin(C5:INDIRECT(ADDRESS(ROW()-1,COLUMN())))]]', '$C$7', 7, 3],
    ['Events', 'Block', '[[=FRMin(D5:INDIRECT(ADDRESS(ROW()-1,COLUMN())))]]', '$D$7', 7, 4],
    ['Events', 'Block', '[[=FRMax(C5:INDIRECT(ADDRESS(ROW()-2,COLUMN())))]]', '$C$8', 8, 3],
    ['Events', 'Block', '[[=FRMax(D5:INDIRECT(ADDRESS(ROW()-2,COLUMN())))]]', '$D$8', 8, 4],
    ['Events', 'Block', '[[=FRAverage(C5:INDIRECT(ADDRESS(ROW()-3,COLUMN())))]]', '$C$9', 9, 3],
    ['Events', 'Block', '[[=FRAverage(D5:INDIRECT(ADDRESS(ROW()-3,COLUMN())))]]', '$D$9', 9, 4],
    ['Events', 'Block', '[[=FROccurences(C5:INDIRECT(ADDRESS(ROW()-4,COLUMN())))]]', '$C$10', 10, 3],
    ['Sections', 'SingleCell', '<<GeneralSettings.DistanceUnit>>', '$D$2', 2, 4],
    ['Sections', 'BlockRow', '##Identification.FiberID##', '$A$4', 4, 1],
    ['Sections', 'BlockRow', '##OTDRSummary.RealWavelength##', '$B$4', 4, 2],
    ['Sections', 'Block', 'Section [[OTDREventTable.SectionNumber]]', '$C$1', 1, 3],
    ['Sections', 'Block', '[[OTDREventTable.SectionLength]]', '$C$2', 2, 3],
    ['Sections', 'Block', '[[##OTDREventTable.SectionLoss##]]', '$C$4', 4, 3],
    ['Sections', 'Block', '[[##OTDREventTable.Attenuation##]]', '$D$4', 4, 4],
    ['Sections', 'Block', '[[##OTDREventTable.SectionLossPassFailStatus##]]', '$C$5', 5, 3],
    ['Sections', 'Block', '[[##OTDREventTable.AttenuationPassFailStatus##]]', '$D$5', 5, 4],
    ['Sections', 'Block', '[[=FRMin(C4:INDIRECT(ADDRESS(ROW()-1,COLUMN())))]]', '$C$6', 6, 3],
    ['Sections', 'Block', '[[=FRMin(D4:INDIRECT(ADDRESS(ROW()-1,COLUMN())))]]', '$D$6', 6, 4],
    ['Sections', 'Block', '[[=FRMax(C4:INDIRECT(ADDRESS(ROW()-2,COLUMN())))]]', '$C$7', 7, 3],
    ['Sections', 'Block', '[[=FRMax(D4:INDIRECT(ADDRESS(ROW()-2,COLUMN())))]]', '$D$7', 7, 4],
    ['Sections', 'Block', '[[=FRAverage(C4:INDIRECT(ADDRESS(ROW()-3,COLUMN())))]]', '$C$8', 8, 3],
    ['Sections', 'Block', '[[=FRAverage(D4:INDIRECT(ADDRESS(ROW()-3,COLUMN())))]]', '$D$8', 8, 4],
    ['Sections', 'Block', '[[=FROccurences(C4:INDIRECT(ADDRESS(ROW()-4,COLUMN())))]]', '$C$9', 9, 3],
]

_CONFIG_TEMPLATE_ROWS = [
    ['ApplyThresholds', True],
    ['NumberFormat', False],
    ['CellMerging', False],
    [],
    ['DisplayOnlyMatchedMeasurements', False],
    ['SheetName', 'AddMultipleSheets', 'Tables', 'PrimaryTable', 'PrimaryColumns', 'HasGroups', 'GroupTable', 'GroupColumns'],
    ['Link Results;Events;Sections'],
    [],
    [],
    ['SheetName', 'Range', 'Filter', 'SortColumns'],
    ['Events', '$C$1:$D$6', 'OTDREventTable.IsEvent=true', 'Identification.FiberID;MatchedFilesDetails.Wavelength;MatchedFilesDetails.Direction'],
    ['Sections', '$C$1:$D$5', 'OTDREventTable.IsEvent=false', 'Identification.FiberID;MatchedFilesDetails.Wavelength;MatchedFilesDetails.Direction'],
]

def _fr_event_type_label(row: EventRow) -> str:
    dist = float(row.distance_km or 0.0)
    et = (row.event_type or '').strip().lower()
    if dist <= 0.05 or 'first connector' in et or 'đầu phát' in et:
        return 'Launch Level'
    if 'fiber end' in et or 'cuối sợi' in et:
        return 'Fiber End'
    if 'splice' in et or 'non-reflective' in et or 'không phản xạ' in et or 'mối hàn' in et:
        return 'Non-Reflective'
    if 'connector' in et or 'reflect' in et or 'đầu nối' in et or 'phản xạ' in et:
        return 'Reflective'
    if 'gainer' in et:
        return 'Gainer'
    return row.event_type or 'Unknown'

def _fr_event_status(loss_db: Optional[float], threshold_db: float) -> str:
    if loss_db is None:
        return ''
    try:
        value = float(loss_db)
    except Exception:
        return 'Unknown'
    return 'Pass' if value <= float(threshold_db) else 'Fail'

def _fr_reportable_reflectance_db(reflectance_db: Optional[float]) -> Optional[float]:
    """Reporting-only cleanup for reflectance values.

    Keep parsed reflectance untouched internally, but suppress placeholder values
    like 0 dB in Excel so non-reflective events do not appear as real reflection
    points.
    """
    if reflectance_db is None:
        return None
    try:
        value = float(reflectance_db)
    except Exception:
        return None
    return value if (-200.0 < value < 0.0) else None

def _fr_reflectance_status(reflectance_db: Optional[float]) -> str:
    value = _fr_reportable_reflectance_db(reflectance_db)
    if value is None:
        return ''
    return 'Pass' if value <= -35.0 else 'Fail'

def _fr_section_loss_status(loss_db: Optional[float], threshold_db: float) -> str:
    if loss_db is None:
        return ''
    try:
        value = float(loss_db)
    except Exception:
        return 'Unknown'
    return 'Pass' if value <= max(float(threshold_db), 0.5) else 'Unknown'

def _fr_section_att_status(att_dbkm: Optional[float]) -> str:
    if att_dbkm is None:
        return ''
    try:
        value = float(att_dbkm)
    except Exception:
        return 'Unknown'
    return 'Pass' if value <= 0.35 else 'Unknown'

def _fr_pick_rows_for_file(events: list[EventRow], wavelength_nm: Optional[str]) -> list[EventRow]:
    if wavelength_nm:
        selected = [e for e in events if e.wavelength_nm == wavelength_nm]
        if selected:
            return selected
    picked = _pick_wavelength(events)
    if picked:
        selected = [e for e in events if e.wavelength_nm == picked]
        if selected:
            return selected
    return list(events)



def _extract_orl_db_from_text(text: str) -> Optional[float]:
    """Best-effort parse of explicit ORL / return-loss values from text/XML/JSON-like blobs.

    We only accept candidates that are explicitly labelled as ORL / return loss and we
    avoid common threshold/alarm/limit fields. This is intentionally conservative to
    reduce false positives.
    """
    if not text:
        return None
    patterns = [
        r'(?is)<orl[^>]*>\s*([0-9]{1,2}(?:\.[0-9]+)?)\s*</orl>',
        r'(?is)(?:span[_\s-]*orl|optical[_\s-]*return[_\s-]*loss|return[_\s-]*loss|(?:^|[^A-Za-z])orl(?:$|[^A-Za-z]))\s*[:=]\s*([0-9]{1,2}(?:\.[0-9]+)?)',
        r'(?is)(?:span[_\s-]*orl|optical[_\s-]*return[_\s-]*loss|return[_\s-]*loss|(?:^|[^A-Za-z])orl(?:$|[^A-Za-z]))[^0-9A-Za-z<>{}\[\]\r\n]{0,24}([0-9]{1,2}(?:\.[0-9]+)?)',
    ]
    deny = re.compile(r'(?i)(threshold|alarm|limit|warn|fail|pass|thres|config|target)')
    for pat in patterns:
        for m in re.finditer(pat, text):
            s = max(0, m.start() - 48)
            e = min(len(text), m.end() + 24)
            ctx = text[s:e]
            if deny.search(ctx):
                continue
            try:
                v = float(m.group(1))
            except Exception:
                continue
            if 0 < v < 100:
                return round(v, 3)
    return None


def _extract_orl_db_from_bytes(raw: bytes) -> Optional[float]:
    try:
        text = raw.decode('latin1', 'ignore')
    except Exception:
        return None
    return _extract_orl_db_from_text(text)


def _extract_trc_object_orl_db(raw: bytes) -> Optional[float]:
    """Best-effort ORL extraction from TRC/CRT object graph or decompressed payload."""
    try:
        extracted = _extract_trc_trace(raw)
    except Exception:
        extracted = None
    combined = None
    trace0 = None
    if isinstance(extracted, dict):
        combined = extracted.get('combined')
        trace0 = extracted.get('trace0')

    key_re = re.compile(r'(?i)(^|[_\s-])(span)?orl($|[_\s-])|return[_\s-]*loss|optical[_\s-]*return')
    deny_re = re.compile(r'(?i)(threshold|alarm|limit|warn|config|setting|target)')

    def coerce(v):
        if isinstance(v, (int, float)):
            fv = float(v)
            if 0 < fv < 100:
                return round(fv, 3)
        if isinstance(v, str):
            m = re.search(r'([0-9]{1,2}(?:\.[0-9]+)?)', v)
            if m:
                fv = float(m.group(1))
                if 0 < fv < 100:
                    return round(fv, 3)
        return None

    def walk(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if str(k).startswith('__'):
                    continue
                ks = str(k)
                if key_re.search(ks) and not deny_re.search(ks):
                    cv = coerce(v)
                    if cv is not None:
                        return cv
                out = walk(v)
                if out is not None:
                    return out
        elif isinstance(obj, (list, tuple)):
            for item in obj:
                out = walk(item)
                if out is not None:
                    return out
        return None

    if trace0:
        val = walk(trace0)
        if val is not None:
            return val

    if combined:
        val = _extract_orl_db_from_bytes(combined)
        if val is not None:
            return val
    return _extract_orl_db_from_bytes(raw)


def _extract_sor_vendor_orl_db(raw: bytes, map_info: Optional[dict]) -> Optional[float]:
    """Best-effort explicit ORL extraction from vendor-specific SOR blocks.

    Some SOR files expose ORL in free-form text or vendor extension blocks. We do not
    try to infer ORL from generic float candidates; only explicitly labelled values are
    accepted here.
    """
    if map_info:
        for name in ('NetTestTSI ', 'ARSpecial', 'AREvent', 'WaveMTSParams', 'SupParams', 'GenParams'):
            entry = map_info.get('entries', {}).get(name)
            if not entry:
                continue
            block = raw[entry['offset']: entry['offset'] + entry['size']]
            val = _extract_orl_db_from_bytes(block)
            if val is not None:
                return val
    return _extract_orl_db_from_bytes(raw)


def _extract_msor_xml_orl_db(raw: bytes) -> Optional[float]:
    """Best-effort exact ORL extraction from embedded smart_acq / smart_link XML.

    Many VIAVI/JDSU MSOR files store the measured span ORL in XML as::
        <orl alarm="0">31.48</orl>

    This is a measured span-level value, not the configuration threshold
    (e.g. OTDS:ALAR:ORLM). We only accept sane values in dB.
    """
    try:
        text = raw.decode('latin1', 'ignore')
    except Exception:
        return None
    m = re.search(r'<orl[^>]*>\s*([0-9]+(?:\.[0-9]+)?)\s*</orl>', text, re.I)
    if not m:
        return None
    try:
        v = float(m.group(1))
    except Exception:
        return None
    if 0 < v < 100:
        return round(v, 3)
    return None

def _fr_extract_orl_db(file_name: str, raw: bytes) -> Optional[float]:
    ext = Path(file_name).suffix.lower()
    try:
        if ext == '.msor':
            summary = _parse_msor_keyevents_summary(raw)
            if summary and summary.get('orl_db') not in (None, 0):
                return summary.get('orl_db')
            xml_orl = _extract_msor_xml_orl_db(raw)
            if xml_orl not in (None, 0):
                return xml_orl
            generic = _extract_orl_db_from_bytes(raw)
            if generic not in (None, 0):
                return generic
            return None
        if ext == '.sor':
            # 1) Standard Bellcore KeyEvents tail, when populated
            rows, meta = _parse_standard_sor_events_with_meta(file_name, raw)
            if meta and meta.get('orl_db') not in (None, 0):
                return meta.get('orl_db')
            # 2) Vendor-specific extension blocks / embedded text
            map_info = _parse_standard_sor_map(raw)
            vendor_orl = _extract_sor_vendor_orl_db(raw, map_info)
            if vendor_orl not in (None, 0):
                return vendor_orl
            return None
        if ext in {'.trc', '.crt'}:
            trc_orl = _extract_trc_object_orl_db(raw)
            if trc_orl not in (None, 0):
                return trc_orl
            return None
    except Exception:
        return None
    return None



def _fr_format_span_orl_display(value: Optional[float], lower_bound: bool = False) -> Optional[str]:
    if value in (None, 0):
        return None
    try:
        v = float(value)
    except Exception:
        return None
    if not (0 < v < 100):
        return None
    return f"<{v:.2f}" if lower_bound else f"{v:.2f}"







# ===== Phase 6 ORL advanced analysis =====
def _fr_is_sane_orl_value(value: Optional[float]) -> bool:
    try:
        v = float(value)
    except Exception:
        return False
    return 0.0 < v < 100.0 and math.isfinite(v)


def _fr_round_orl(value: Optional[float]) -> Optional[float]:
    if not _fr_is_sane_orl_value(value):
        return None
    return round(float(value), 3)


def _fr_status_for_measured_orl(value: Optional[float], threshold_db: float) -> str:
    if not _fr_is_sane_orl_value(value):
        return 'Unknown'
    try:
        return 'Pass' if float(value) >= float(threshold_db) else 'Fail'
    except Exception:
        return 'Unknown'


def _fr_extract_measured_orl_candidate(file_name: str, raw: bytes) -> Optional[dict]:
    """Return an exact/measured ORL candidate with provenance.

    Phase 6 keeps this intentionally conservative: only explicitly labelled or
    standard stored span-ORL fields are accepted. Metadata thresholds, alarm
    limits and lower-bound displays are handled separately and never become
    measured ORL.
    """
    ext = Path(file_name).suffix.lower()
    try:
        if ext == '.msor':
            summary = _parse_msor_keyevents_summary(raw)
            if summary and _fr_is_sane_orl_value(summary.get('orl_db')):
                return {
                    'value_db': _fr_round_orl(summary.get('orl_db')),
                    'source_kind': 'measured_orl',
                    'source_detail': 'MSOR KeyEvents summary / span ORL',
                    'source_confidence': 'High',
                    'reason': 'Đọc được ORL span từ vùng KeyEvents/tóm tắt MSOR, không phải threshold.',
                }
            xml_orl = _extract_msor_xml_orl_db(raw)
            if _fr_is_sane_orl_value(xml_orl):
                return {
                    'value_db': _fr_round_orl(xml_orl),
                    'source_kind': 'measured_orl',
                    'source_detail': 'MSOR smart_acq/smart_link XML <orl>',
                    'source_confidence': 'High',
                    'reason': 'Đọc được ORL có nhãn <orl> trong XML nhúng của MSOR.',
                }
            generic = _extract_orl_db_from_bytes(raw)
            if _fr_is_sane_orl_value(generic):
                return {
                    'value_db': _fr_round_orl(generic),
                    'source_kind': 'measured_orl',
                    'source_detail': 'MSOR embedded labelled ORL / return-loss text',
                    'source_confidence': 'Medium',
                    'reason': 'Đọc được giá trị ORL/return loss có nhãn rõ ràng trong payload; đã bỏ qua threshold/alarm/limit.',
                }
            return None
        if ext == '.sor':
            rows, meta = _parse_standard_sor_events_with_meta(file_name, raw)
            if meta and _fr_is_sane_orl_value(meta.get('orl_db')):
                return {
                    'value_db': _fr_round_orl(meta.get('orl_db')),
                    'source_kind': 'measured_orl',
                    'source_detail': 'SOR Bellcore KeyEvents tail ORL',
                    'source_confidence': 'High',
                    'reason': 'Đọc được ORL trong phần tail KeyEvents chuẩn SOR.',
                }
            map_info = _parse_standard_sor_map(raw)
            vendor_orl = _extract_sor_vendor_orl_db(raw, map_info)
            if _fr_is_sane_orl_value(vendor_orl):
                return {
                    'value_db': _fr_round_orl(vendor_orl),
                    'source_kind': 'measured_orl',
                    'source_detail': 'SOR vendor/extension labelled ORL',
                    'source_confidence': 'Medium',
                    'reason': 'Đọc được ORL/return loss có nhãn rõ trong block vendor hoặc text nhúng; đã bỏ qua threshold/alarm/limit.',
                }
            return None
        if ext in {'.trc', '.crt'}:
            trc_orl = _extract_trc_object_orl_db(raw)
            if _fr_is_sane_orl_value(trc_orl):
                return {
                    'value_db': _fr_round_orl(trc_orl),
                    'source_kind': 'measured_orl',
                    'source_detail': 'TRC/CRT object graph labelled ORL',
                    'source_confidence': 'Medium',
                    'reason': 'Đọc được ORL/optical return loss có nhãn trong object graph hoặc payload giải nén.',
                }
            return None
    except Exception:
        return None
    return None


def _fr_physical_orl_trace_diagnostic(ctx: Optional[dict], summary: FileSummary, physical_mode: str) -> dict:
    mode = str(physical_mode or 'disabled').strip().lower()
    rows = (ctx or {}).get('events') or []
    reflectances = []
    for ev in rows:
        val = _fr_reportable_reflectance_db(getattr(ev, 'reflectance_db', None))
        if isinstance(val, (int, float)) and math.isfinite(float(val)):
            reflectances.append(float(val))
    strongest = max(reflectances) if reflectances else None  # closest to 0 dB = strongest reflection

    if mode in {'', 'disabled', 'off', 'none', 'false'}:
        return {
            'mode': 'disabled',
            'attempted': False,
            'value_db': None,
            'status': 'Not attempted',
            'reason': 'Tắt ORL vật lý từ trace. Phase 6 chỉ dùng ORL đo được/lower-bound an toàn.',
            'strongest_reflectance_db': round(strongest, 3) if strongest is not None else None,
        }

    trace = (ctx or {}).get('raw_trace_series') or {}
    if not trace:
        return {
            'mode': mode,
            'attempted': True,
            'value_db': None,
            'status': 'Not available',
            'reason': 'Không có raw trace samples để thử đánh giá ORL vật lý.',
            'strongest_reflectance_db': round(strongest, 3) if strongest is not None else None,
        }

    # Important: do not fake physics. The phase-5 trace may be relative-scaled for
    # section attenuation, but physical ORL needs absolute backscatter/reference
    # calibration. Without those fields, a numeric result would be misleading.
    required_missing = []
    meta = (ctx or {}).get('metadata') or {}
    for key in ('backscatter_coefficient_db', 'launch_power_dbm', 'receiver_calibration_db', 'orl_calibration_db'):
        if key not in meta or meta.get(key) in (None, ''):
            required_missing.append(key)
    return {
        'mode': mode,
        'attempted': True,
        'value_db': None,
        'status': 'Need calibration',
        'reason': 'Chưa tính ORL vật lý từ trace vì thiếu hiệu chuẩn tuyệt đối: ' + ', '.join(required_missing) + '. Raw trace hiện chỉ đủ tin hơn cho section-fit/shape diagnostics, không đủ để biến thành span ORL vật lý.',
        'strongest_reflectance_db': round(strongest, 3) if strongest is not None else None,
    }


def _fr_analyze_orl(
    file_name: str,
    raw: bytes,
    metadata: Optional[dict],
    summary: FileSummary,
    ctx: Optional[dict],
    *,
    orl_pass_threshold_db: float = 28.0,
    orl_source_mode: str = 'auto',
    orl_allow_lower_bound: bool = True,
    orl_lower_bound_status: str = 'Unknown',
    orl_physical_mode: str = 'disabled',
) -> ORLAnalysis:
    source_mode = str(orl_source_mode or 'auto').strip().lower()
    threshold = float(orl_pass_threshold_db or 28.0)
    lower_status = str(orl_lower_bound_status or 'Unknown').strip() or 'Unknown'
    lower_value = None
    if metadata:
        lower_value = _fr_round_orl(metadata.get('reflection_threshold_db'))
    measured = _fr_extract_measured_orl_candidate(file_name, raw)
    has_measured = measured is not None and _fr_is_sane_orl_value(measured.get('value_db'))
    # Conditional phase-6 behavior: physical ORL trace diagnostics are only useful
    # when the file is missing a measured/exact ORL.  If measured ORL exists, do
    # not run/advertise trace-based ORL diagnostics so the report stays decisive.
    requested_physical_mode = str(orl_physical_mode or 'disabled').strip().lower()
    effective_physical_mode = 'disabled' if has_measured and requested_physical_mode not in {'', 'disabled', 'off', 'none', 'false'} else requested_physical_mode
    physical = _fr_physical_orl_trace_diagnostic(ctx, summary, effective_physical_mode)

    def measured_analysis(cand: dict) -> ORLAnalysis:
        value = _fr_round_orl(cand.get('value_db'))
        status = _fr_status_for_measured_orl(value, threshold)
        advanced = 'PASS' if status == 'Pass' else ('FAIL' if status == 'Fail' else 'UNKNOWN')
        rec = 'Có thể dùng ORL này để đánh giá Pass/Fail.' if status in {'Pass', 'Fail'} else 'Không đủ điều kiện đánh giá ORL.'
        return ORLAnalysis(
            file_name=file_name,
            display=_fr_format_span_orl_display(value, lower_bound=False),
            value_db=value,
            status=status,
            advanced_status=advanced,
            source_kind=str(cand.get('source_kind') or 'measured_orl'),
            source_detail=str(cand.get('source_detail') or 'Measured ORL'),
            source_confidence=str(cand.get('source_confidence') or 'Medium'),
            pass_threshold_db=threshold,
            use_for_judgment=status in {'Pass', 'Fail'},
            lower_bound=False,
            reason=str(cand.get('reason') or 'ORL đo được từ file.'),
            recommendation=rec,
            physical_mode=str(physical.get('mode') or orl_physical_mode),
            physical_attempted=bool(physical.get('attempted')),
            physical_value_db=_fr_round_orl(physical.get('value_db')),
            physical_status=str(physical.get('status') or ''),
            physical_reason=str(physical.get('reason') or ''),
            strongest_reflectance_db=physical.get('strongest_reflectance_db'),
        )

    def lower_analysis() -> ORLAnalysis:
        display = _fr_format_span_orl_display(lower_value, lower_bound=True)
        # Status field remains compatible with old Link Results.  The advanced
        # status and Use-for-Judgment are the authoritative phase-6 fields.
        legacy_status = lower_status if lower_status in {'Pass', 'Fail', 'Unknown'} else 'Unknown'
        return ORLAnalysis(
            file_name=file_name,
            display=display,
            value_db=lower_value,
            status=legacy_status,
            advanced_status='LOWER_BOUND_ONLY',
            source_kind='metadata_lower_bound',
            source_detail='FxdParams reflection/ORL-like threshold metadata',
            source_confidence='Low',
            pass_threshold_db=threshold,
            use_for_judgment=False,
            lower_bound=True,
            reason='Chỉ có giá trị tham khảo/threshold dạng lower-bound từ metadata, không phải span ORL đo được thật.',
            recommendation='Không dùng giá trị này để kết luận Pass/Fail ORL; chỉ hiển thị để tham khảo hoặc truy vết file.',
            physical_mode=str(physical.get('mode') or orl_physical_mode),
            physical_attempted=bool(physical.get('attempted')),
            physical_value_db=_fr_round_orl(physical.get('value_db')),
            physical_status=str(physical.get('status') or ''),
            physical_reason=str(physical.get('reason') or ''),
            strongest_reflectance_db=physical.get('strongest_reflectance_db'),
        )

    def none_analysis(reason: str) -> ORLAnalysis:
        return ORLAnalysis(
            file_name=file_name,
            display=None,
            value_db=None,
            status='Unknown',
            advanced_status='NOT_AVAILABLE',
            source_kind='not_available',
            source_detail='',
            source_confidence='None',
            pass_threshold_db=threshold,
            use_for_judgment=False,
            lower_bound=False,
            reason=reason,
            recommendation='Không đánh giá ORL trong báo cáo; cần mở tool gốc/đo lại nếu ORL là tiêu chí bắt buộc.',
            physical_mode=str(physical.get('mode') or orl_physical_mode),
            physical_attempted=bool(physical.get('attempted')),
            physical_value_db=_fr_round_orl(physical.get('value_db')),
            physical_status=str(physical.get('status') or ''),
            physical_reason=str(physical.get('reason') or ''),
            strongest_reflectance_db=physical.get('strongest_reflectance_db'),
        )

    has_lower = bool(orl_allow_lower_bound) and _fr_is_sane_orl_value(lower_value)

    if source_mode == 'exact_only':
        if has_measured:
            return measured_analysis(measured)
        return none_analysis('Đang chọn Chỉ lấy ORL đo được, nhưng file không có exact/measured ORL hợp lệ.')
    if source_mode == 'metadata_only':
        if has_lower:
            return lower_analysis()
        return none_analysis('Đang chọn Chỉ lấy giá trị tham khảo nhưng file không có lower-bound/threshold ORL hợp lệ hoặc đang tắt hiển thị lower-bound.')
    if source_mode == 'prefer_metadata':
        if has_lower:
            return lower_analysis()
        if has_measured:
            return measured_analysis(measured)
        return none_analysis('Không tìm thấy ORL đo được hoặc giá trị tham khảo hợp lệ.')

    # auto/default: exact ORL first; lower-bound only as display fallback.
    if has_measured:
        return measured_analysis(measured)
    if has_lower:
        return lower_analysis()
    return none_analysis('Không tìm thấy ORL đo được; cũng không có lower-bound/threshold hợp lệ để hiển thị tham khảo.')

def _fr_build_context(
    files: Iterable[tuple[str, bytes]],
    threshold_db: float = 0.5,
    expected_route_km: Optional[float] = None,
    jumper_excluded_m: float = 0.0,
    length_tolerance_km: float = 0.300,
    graph_reach_tolerance_km: Optional[float] = None,
    event_shortfall_tolerance_km: Optional[float] = None,
    overlength_tolerance_km: Optional[float] = None,
    segment_start_km: Optional[float] = None,
    segment_end_km: Optional[float] = None,
    orl_pass_threshold_db: float = 28.0,
    orl_source_mode: str = 'auto',
    orl_allow_lower_bound: bool = True,
    orl_lower_bound_status: str = 'Unknown',
    orl_physical_mode: str = 'disabled',
    logs: Optional[list[dict]] = None,
    preview_fast: bool = False,
    max_seconds: Optional[float] = None,
):
    if logs is None:
        logs = []
    summaries: list[FileSummary] = []
    skipped: list[str] = []
    file_payload_map = {}
    contexts: dict[str, dict] = {}
    start_time = time.monotonic()
    for file_name, raw in files:
        if max_seconds is not None and (time.monotonic() - start_time) >= float(max_seconds):
            skipped.append(f'Trace preview timeout: đã dừng sau {int(max_seconds)} giây, các file còn lại chưa dựng đồ thị')
            _fr_log(logs, 'trace-preview', 'WARN', f'Dừng dựng đồ thị sau {int(max_seconds)} giây để tránh chờ quá lâu.')
            break
        file_payload_map[file_name] = raw
        ext = Path(file_name).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            skipped.append(f'{file_name}: định dạng không hỗ trợ')
            _fr_log(logs, 'parse', 'WARN', f'Bỏ qua {file_name}: định dạng không hỗ trợ')
            continue
        try:
            bundle = _fr_get_or_build_parse_bundle(file_name, raw, logs)
            summary = bundle['summary']
            summaries.append(summary)
            metadata = bundle.get('metadata') or {}
            ctx = _fr_init_context(summary, bundle.get('events'), bundle.get('parse_mode'), bundle.get('sor_meta'), bundle.get('trc_trace'), bundle.get('orl_db'), metadata)
            ctx['raw_trace_series'] = _extract_raw_trace_series(file_name, raw, summary, bundle.get('trc_trace'), bundle.get('sor_meta'))
            ctx['section_fit_rows'] = []
            if ctx.get('raw_trace_series'):
                rt = ctx['raw_trace_series']
                _fr_log(logs, 'trace', 'INFO', f"{file_name} | Raw trace source: {rt.get('source')} | points={rt.get('raw_points_total')} | calibrated={rt.get('calibrated_db')}")
            else:
                _fr_log(logs, 'trace', 'WARN', f'{file_name} | Chưa lấy được raw trace để fit section; sẽ dùng fallback event/slope khi cần.')
            orl_analysis = _fr_analyze_orl(
                file_name,
                raw,
                metadata,
                summary,
                ctx,
                orl_pass_threshold_db=orl_pass_threshold_db,
                orl_source_mode=orl_source_mode,
                orl_allow_lower_bound=orl_allow_lower_bound,
                orl_lower_bound_status=orl_lower_bound_status,
                orl_physical_mode=orl_physical_mode,
            )
            ctx['orl_analysis'] = orl_analysis
            ctx['orl_display'] = orl_analysis.display
            ctx['orl_status'] = orl_analysis.status
            ctx['orl_value_db'] = orl_analysis.value_db
            ctx['orl_source_kind'] = orl_analysis.source_kind
            ctx['orl_source_detail'] = orl_analysis.source_detail
            ctx['orl_use_for_judgment'] = orl_analysis.use_for_judgment
            ctx['orl_reason'] = orl_analysis.reason
            _fr_log(logs, 'orl', 'INFO' if orl_analysis.use_for_judgment else 'WARN', f"{file_name} | ORL={orl_analysis.display or '-'} | status={orl_analysis.advanced_status} | source={orl_analysis.source_detail or '-'} | use_for_judgment={orl_analysis.use_for_judgment} | {orl_analysis.reason}")
            ctx['graph_assessment'] = _assess_graph_length(
                summary,
                expected_route_km=expected_route_km,
                jumper_excluded_m=jumper_excluded_m,
                length_tolerance_km=length_tolerance_km,
                graph_reach_tolerance_km=graph_reach_tolerance_km,
                event_shortfall_tolerance_km=event_shortfall_tolerance_km,
                overlength_tolerance_km=overlength_tolerance_km,
            )
            if preview_fast:
                # Trace Viewer chỉ cần trace/events/metadata/ORL tóm tắt.
                # Bỏ qua segment deep-scan để dựng đồ thị nhanh; Excel export vẫn chạy đầy đủ.
                ctx['segment_assessment'] = None
                ctx['segment_event_rows'] = []
            else:
                ctx['segment_assessment'] = _build_segment_assessment(
                    file_name,
                    raw,
                    summary,
                    segment_start_km=segment_start_km,
                    segment_end_km=segment_end_km,
                )
                ctx['segment_event_rows'] = _extract_segment_event_rows(
                    file_name,
                    raw,
                    segment_start_km=segment_start_km,
                    segment_end_km=segment_end_km,
                    threshold_db=threshold_db,
                )
            contexts[file_name] = ctx
        except Exception as exc:
            skipped.append(f'{file_name}: {exc}')
            _fr_log(logs, 'parse', 'ERROR', f'Lỗi file {file_name}: {exc}')
            _fr_log(logs, 'parse', 'ERROR', traceback.format_exc(limit=3))
    summaries.sort(key=lambda s: [int(x) if x.isdigit() else x.lower() for x in re.split(r'(\d+)', s.file_name)])
    return summaries, skipped, file_payload_map, contexts

def _fr_build_common_event_defs(contexts: dict[str, dict], deviation_m: float) -> list[dict]:
    distance_values = [0.0]
    for ctx in contexts.values():
        for row in ctx.get('events', []):
            if row.distance_km is None:
                continue
            distance_values.append(float(row.distance_km))
    clusters = _cluster_distances(distance_values, deviation_m)
    defs: list[dict] = []
    tol = max(float(deviation_m), 0.0) / 1000.0
    for idx, cluster in enumerate(clusters, start=1):
        rep = cluster.representative_km
        matched: list[EventRow] = []
        for ctx in contexts.values():
            for row in ctx.get('events', []):
                if row.distance_km is None:
                    continue
                if abs(float(row.distance_km) - rep) <= max(tol, 0.02):
                    matched.append(row)
        labels = [_fr_event_type_label(r) for r in matched]
        label = Counter(labels).most_common(1)[0][0] if labels else ('Launch Level' if rep <= max(tol, 0.02) else 'Non-Reflective')
        losses = [float(r.loss_db) for r in matched if r.loss_db is not None and float(r.loss_db) > 0]
        max_loss_db = max(losses) if losses else None
        refls = [float(r.reflectance_db) for r in matched if r.reflectance_db is not None]
        has_reflective = any(-200.0 < v < 0.0 for v in refls)
        defs.append({
            'index': idx,
            'distance_km': rep,
            'label': label,
            'cluster': cluster,
            'max_loss_db': max_loss_db,
            'has_reflective': has_reflective,
            'occurrences': len(matched),
        })
    return defs

def _fr_assign_event_to_def(row: EventRow, event_defs: list[dict], deviation_m: float) -> Optional[int]:
    if row.distance_km is None or not event_defs:
        return None
    tol = max(float(deviation_m), 0.0) / 1000.0
    best_idx = None
    best_gap = None
    for item in event_defs:
        gap = abs(float(row.distance_km) - float(item['distance_km']))
        if best_gap is None or gap < best_gap:
            best_gap = gap
            best_idx = int(item['index'])
    if best_gap is None:
        return None
    if best_gap <= max(tol, 0.03):
        return best_idx
    return None

def _fr_build_common_sections(
    event_defs: list[dict],
    summaries: list[FileSummary],
    contexts: dict[str, dict],
    deviation_m: float,
    threshold_db: float,
    section_merge_tolerance_m: Optional[float] = None,
    section_min_length_km: float = 0.0,
    section_event_source: str = 'all',
    section_boundary_priority: str = 'event',
    section_allow_split: bool = False,
) -> list[dict]:
    """Build common section boundaries with genuinely effective controls.

    - section_event_source='all': start from the whole event pool (close to original logic)
      then let priority pruning remove only the weakest boundaries.
    - section_event_source='filtered': keep only boundaries that are strong enough by
      loss / event type / consensus across fibers / slope change.
    - section_boundary_priority='event': bias toward event-driven boundaries.
    - section_boundary_priority='trace': bias toward attenuation-trend breaks.
    - section_min_length_km + section_allow_split: keep or merge very short sections.
    """
    if not summaries:
        return []

    common_end = max(float(s.length_km or 0.0) for s in summaries)
    if common_end <= 0:
        return []
    total_fibers = max(len(summaries), 1)
    merge_tol_km = max(float(section_merge_tolerance_m if section_merge_tolerance_m is not None else deviation_m), 0.0) / 1000.0
    merge_tol_km = max(merge_tol_km, 0.01)

    def _type_score(label: str, has_reflective: bool) -> int:
        ll = (label or '').lower()
        if 'fiber end' in ll or 'cuối sợi' in ll or 'splitter' in ll or 'bộ chia' in ll:
            return 3
        if has_reflective or 'reflect' in ll or 'connector' in ll or 'đầu nối' in ll or 'phản xạ' in ll:
            return 2
        if 'section' in ll or 'đoạn' in ll or 'marker' in ll:
            return 2
        return 0

    def _loss_score(max_loss: Optional[float]) -> int:
        if max_loss is None:
            return 0
        try:
            value = float(max_loss)
        except Exception:
            return 0
        base = max(float(threshold_db), 0.05)
        if value < 0.1:
            return 0
        if value < base:
            return 1
        if value < 2.0 * base:
            return 2
        return 3

    def _consensus_score(occurrences: int) -> int:
        ratio = float(occurrences or 0) / float(total_fibers)
        if ratio < 0.2:
            return 0
        if ratio < 0.5:
            return 1
        if ratio < 0.8:
            return 2
        return 3

    def _slope_change_metrics(dist_km: float) -> tuple[float, int]:
        deltas: list[float] = []
        hits = 0
        for ctx in contexts.values():
            rows = ctx.get('events') or []
            ordered = [r for r in rows if r.distance_km is not None and r.slope_dbkm is not None]
            if len(ordered) < 2:
                continue
            ordered = sorted(ordered, key=lambda r: (float(r.distance_km or 0.0), r.event_no))
            idx = None
            for i, r in enumerate(ordered):
                if float(r.distance_km or 0.0) >= float(dist_km):
                    idx = i
                    break
            if idx is None:
                continue
            before = ordered[idx].slope_dbkm
            after = ordered[idx + 1].slope_dbkm if idx + 1 < len(ordered) else None
            if before is None or after is None:
                continue
            try:
                b = float(before)
                a = float(after)
            except Exception:
                continue
            if not (0.0 < b < 2.0 and 0.0 < a < 2.0):
                continue
            delta = abs(a - b)
            deltas.append(delta)
            if delta >= 0.05:
                hits += 1
        if not deltas:
            return 0.0, 0
        deltas.sort()
        mid = len(deltas) // 2
        med = float(deltas[mid]) if len(deltas) % 2 == 1 else float((deltas[mid - 1] + deltas[mid]) / 2.0)
        return med, hits

    candidates: list[dict] = []
    for item in event_defs:
        d = item.get('distance_km')
        if d is None:
            continue
        d = float(d)
        if d <= max(merge_tol_km, 0.001):
            continue
        if d >= common_end - max(merge_tol_km, 0.03):
            continue
        label = str(item.get('label') or '')
        max_loss = item.get('max_loss_db')
        has_reflective = bool(item.get('has_reflective'))
        occ = int(item.get('occurrences') or 0)
        trace_delta, trace_hits = _slope_change_metrics(d)
        cand = {
            'distance_km': d,
            'label': label,
            'max_loss_db': max_loss,
            'has_reflective': has_reflective,
            'occurrences': occ,
            'loss_score': _loss_score(max_loss),
            'type_score': _type_score(label, has_reflective),
            'consensus_score': _consensus_score(occ),
            'trace_delta': trace_delta,
            'trace_hits': trace_hits,
        }
        if trace_delta < 0.03:
            cand['trace_score'] = 0
        elif trace_delta < 0.07:
            cand['trace_score'] = 1
        elif trace_delta < 0.15:
            cand['trace_score'] = 2
        else:
            cand['trace_score'] = 3
        cand['event_priority_score'] = (0.4 * cand['loss_score']) + (0.3 * cand['type_score']) + (0.2 * cand['consensus_score']) + (0.1 * cand['trace_score'])
        cand['trace_priority_score'] = (0.2 * cand['loss_score']) + (0.15 * cand['type_score']) + (0.2 * cand['consensus_score']) + (0.45 * cand['trace_score'])
        candidates.append(cand)

    if not candidates:
        return [{'index': 1, 'start_km': 0.0, 'end_km': round(common_end, 4), 'length_km': round(common_end, 4)}]

    src = (section_event_source or 'all').strip().lower()
    pr = (section_boundary_priority or 'event').strip().lower()
    kept: list[dict] = []
    for cand in candidates:
        strong_by_signal = cand['loss_score'] >= 2
        strong_by_type = cand['type_score'] >= 2
        strong_by_trace = cand['trace_score'] >= 2 and cand['trace_hits'] >= max(2, total_fibers // 4)
        strong_by_consensus = cand['consensus_score'] >= 2
        if src in {'filtered', 'displayed', 'shown', 'visible'}:
            if strong_by_signal or strong_by_type or strong_by_trace or strong_by_consensus:
                kept.append(cand)
            continue
        # src='all' -> near original behaviour, but priority still has a visible effect.
        if pr in {'trace', 'slope', 'atten', 'attenuation'}:
            if strong_by_trace or cand['trace_priority_score'] >= 1.0 or strong_by_signal or strong_by_type:
                kept.append(cand)
        else:
            if cand['event_priority_score'] >= 0.8 or strong_by_trace:
                kept.append(cand)

    if not kept:
        kept = candidates[:]

    weight_key = 'trace_priority_score' if pr in {'trace', 'slope', 'atten', 'attenuation'} else 'event_priority_score'
    kept = sorted(kept, key=lambda x: float(x['distance_km']))
    groups: list[list[dict]] = []
    for cand in kept:
        if not groups:
            groups.append([cand])
            continue
        if abs(float(cand['distance_km']) - float(groups[-1][-1]['distance_km'])) <= merge_tol_km:
            groups[-1].append(cand)
        else:
            groups.append([cand])

    mids: list[float] = []
    for group in groups:
        weights = [max(float(item.get(weight_key) or 0.0), 0.25) for item in group]
        xs = [float(item['distance_km']) for item in group]
        rep = sum(x * w for x, w in zip(xs, weights)) / sum(weights)
        if rep <= max(merge_tol_km, 0.001):
            continue
        if rep >= common_end - max(merge_tol_km, 0.03):
            continue
        mids.append(round(rep, 4))
    mids = sorted(set(mids))

    boundaries = [0.0]
    for value in mids:
        if value - boundaries[-1] > max(merge_tol_km / 2.0, 0.001):
            boundaries.append(round(value, 4))
    if common_end > boundaries[-1]:
        boundaries.append(round(common_end, 4))

    sections: list[dict] = []
    for i in range(len(boundaries) - 1):
        start_km = float(boundaries[i])
        end_km = float(boundaries[i + 1])
        length_km = round(end_km - start_km, 4)
        if length_km <= 0:
            continue
        if section_min_length_km and length_km < float(section_min_length_km):
            if not section_allow_split:
                if sections:
                    prev = sections[-1]
                    prev['end_km'] = end_km
                    prev['length_km'] = round(prev['end_km'] - prev['start_km'], 4)
                    continue
        sections.append({'index': len(sections) + 1, 'start_km': start_km, 'end_km': end_km, 'length_km': length_km})

    if not sections:
        sections = [{'index': 1, 'start_km': 0.0, 'end_km': round(common_end, 4), 'length_km': round(common_end, 4)}]
    return sections



def _fr_clip_sections_to_range(sections: list[dict], start_km: Optional[float], end_km: Optional[float]) -> list[dict]:
    """Clip prebuilt common sections to a user-selected range.

    Important: when the selected range is fully outside the measured/common section span,
    return an empty list instead of falling back to the full route. The old fallback caused
    confusing exports such as selecting 20-40 km on a 20 km trace but still receiving all
    route sections.
    """
    if not sections:
        return []
    if start_km is None or end_km is None:
        return sections
    try:
        a = float(start_km)
        b = float(end_km)
    except Exception:
        return sections
    if not (math.isfinite(a) and math.isfinite(b)):
        return sections
    if a == b:
        return []
    if a > b:
        a, b = b, a
    clipped: list[dict] = []
    for sec in sections:
        try:
            s = float(sec.get('start_km') or 0.0)
            e = float(sec.get('end_km') or 0.0)
        except Exception:
            continue
        if not (math.isfinite(s) and math.isfinite(e)):
            continue
        if e <= a or s >= b:
            continue
        ns = max(s, a)
        ne = min(e, b)
        if ne <= ns:
            continue
        clipped.append({
            'index': len(clipped) + 1,
            'start_km': round(ns, 4),
            'end_km': round(ne, 4),
            'length_km': round(ne - ns, 4),
            'source_start_km': round(s, 4),
            'source_end_km': round(e, 4),
        })
    return clipped






def _downsample_xy(x: list[float], y: list[float], max_points: int = 2400) -> tuple[list[float], list[float]]:
    if len(x) <= max_points:
        return x, y
    step = max(1, int(math.ceil(len(x) / float(max_points))))
    return x[::step], y[::step]


def _linear_fit_xy(x: list[float], y: list[float]) -> Optional[dict]:
    if len(x) != len(y) or len(x) < 3:
        return None
    n = len(x)
    sx = sum(x)
    sy = sum(y)
    sxx = sum(v * v for v in x)
    sxy = sum(a * b for a, b in zip(x, y))
    denom = n * sxx - sx * sx
    if abs(denom) < 1e-18:
        return None
    slope = (n * sxy - sx * sy) / denom
    intercept = (sy - slope * sx) / n
    residuals = [yy - (intercept + slope * xx) for xx, yy in zip(x, y)]
    sse = sum(r * r for r in residuals)
    y_mean = sy / n
    sst = sum((yy - y_mean) ** 2 for yy in y)
    r2 = 1.0 - (sse / sst) if sst > 1e-18 else 1.0
    rms = math.sqrt(sse / n) if n else None
    max_abs = max((abs(r) for r in residuals), default=None)
    return {
        'slope': slope,
        'intercept': intercept,
        'r2': max(min(r2, 1.0), -999.0),
        'rms': rms,
        'max_abs': max_abs,
        'residuals': residuals,
    }


def _series_from_uniform_samples(values: list[float], range_km: Optional[float]) -> tuple[list[float], list[float]]:
    if not values or range_km in (None, 0):
        return [], []
    try:
        range_f = float(range_km)
    except Exception:
        return [], []
    if range_f <= 0:
        return [], []
    count = len(values)
    if count < 8:
        return [], []
    denom = max(count - 1, 1)
    x = [(i / denom) * range_f for i in range(count)]
    y = [float(v) for v in values]
    return x, y



def _raw_trace_reference_span_km(summary: FileSummary) -> Optional[float]:
    """Best one-way span reference for validating raw-trace distance scale."""
    for value in (
        getattr(summary, 'length_km', None),
        getattr(summary, 'end_distance_km', None),
        getattr(summary, 'graph_end_km', None),
    ):
        try:
            if value not in (None, 0) and math.isfinite(float(value)) and float(value) > 0:
                return float(value)
        except Exception:
            continue
    return None


def _normalize_raw_trace_distance_axis(x: list[float], summary: FileSummary) -> tuple[list[float], dict]:
    """Normalize vendor raw-trace x-axis to one-way span kilometres when detectable.

    Some MSOR/Acterna MiniCurve blocks expose an x-axis that represents the
    two-way OTDR time-of-flight distance.  In those files the curve maximum can
    be roughly 2x the actual span length, e.g. 163 km for an 81.5 km span.  Raw
    section-fit must operate on one-way fibre distance, so we correct only when
    the 2x signature is strong.  Ambiguous scales are kept but marked unusable
    for section loss to avoid silently generating wrong attenuation.
    """
    meta = {
        'reference_span_km': None,
        'raw_last_before_km': None,
        'raw_last_after_km': None,
        'distance_scale_factor': 1.0,
        'distance_scale_status': 'Không có dữ liệu trục km',
        'distance_scale_note': '',
        'distance_scale_ok': False,
    }
    if not x:
        return x, meta
    try:
        x_float = [float(v) for v in x]
        raw_last = max(x_float)
    except Exception:
        meta['distance_scale_status'] = 'Lỗi trục km'
        meta['distance_scale_note'] = 'Không chuyển được trục raw trace sang số km.'
        return x, meta
    meta['raw_last_before_km'] = round(raw_last, 6)
    ref = _raw_trace_reference_span_km(summary)
    if ref in (None, 0):
        meta.update({
            'raw_last_after_km': round(raw_last, 6),
            'distance_scale_status': 'Thiếu span tham chiếu',
            'distance_scale_note': 'Không có span length/end distance để đối chiếu trục raw trace; chỉ dùng thận trọng.',
            'distance_scale_ok': True,
        })
        return x_float, meta
    meta['reference_span_km'] = round(float(ref), 6)
    if raw_last <= 0:
        meta['distance_scale_status'] = 'Trục km không hợp lệ'
        meta['distance_scale_note'] = 'Raw trace last km bằng 0 hoặc âm.'
        return x_float, meta
    ratio = raw_last / float(ref)
    # Normal one-way trace: allow moderate tolerance because some files include
    # launch/receive fibre or a graph range slightly different from the event end.
    if 0.80 <= ratio <= 1.20:
        meta.update({
            'raw_last_after_km': round(raw_last, 6),
            'distance_scale_factor': 1.0,
            'distance_scale_status': 'Khớp span',
            'distance_scale_note': f'Raw last km khớp span tham chiếu (ratio={ratio:.3f}).',
            'distance_scale_ok': True,
        })
        return x_float, meta
    # Strong two-way signature: correct raw axis back to the one-way span.
    if 1.80 <= ratio <= 2.20:
        factor = float(ref) / raw_last
        x_scaled = [v * factor for v in x_float]
        meta.update({
            'raw_last_after_km': round(max(x_scaled), 6),
            'distance_scale_factor': round(factor, 8),
            'distance_scale_status': 'Đã hiệu chỉnh two-way',
            'distance_scale_note': f'Raw last km ≈ {ratio:.3f} lần span; đã scale trục km về one-way span.',
            'distance_scale_ok': True,
        })
        return x_scaled, meta
    meta.update({
        'raw_last_after_km': round(raw_last, 6),
        'distance_scale_factor': 1.0,
        'distance_scale_status': 'Lệch span',
        'distance_scale_note': f'Raw last km/span ratio={ratio:.3f}, không khớp 1x hoặc 2x; không dùng raw-fit để tính loss.',
        'distance_scale_ok': False,
    })
    return x_float, meta

def _calibrate_raw_series_to_db(
    x: list[float],
    raw_y: list[float],
    summary: FileSummary,
    *,
    source: str,
) -> Optional[dict]:
    """Orient and scale vendor raw samples to dB-like loss units when possible.

    SOR DataPts are already calibrated, but TRC RawSamples and MSOR MiniCurve are
    often arbitrary signed amplitudes.  For those families we fit the overall raw
    trend and scale it to the route average attenuation only when the summary loss
    and length are available.  If calibration is not possible, the series is still
    useful for R²-style shape diagnostics, but it must not drive section loss.
    """
    if len(x) != len(raw_y) or len(x) < 8:
        return None
    x_ds, y_ds = _downsample_xy(x, raw_y, max_points=4000)
    overall = _linear_fit_xy(x_ds, y_ds)
    if overall is None:
        return None
    raw_slope = float(overall['slope'])
    orientation = 1.0 if raw_slope >= 0 else -1.0
    calibrated = False
    scale = 1.0
    note = 'raw units; chỉ dùng để đánh giá hình dạng fit, không dùng trực tiếp làm loss section'
    if summary.total_loss_db not in (None, 0) and summary.length_km not in (None, 0) and abs(raw_slope) > 1e-12:
        try:
            overall_att = abs(float(summary.total_loss_db) / float(summary.length_km))
            if 0.001 <= overall_att <= 2.5:
                scale = overall_att / abs(raw_slope)
                calibrated = True
                note = 'đã scale raw amplitude theo attenuation trung bình toàn tuyến'
        except Exception:
            pass
    base = raw_y[0]
    y = [(float(v) - base) * orientation * scale for v in raw_y]
    return {
        'x_km': x,
        'y_db': y,
        'source': source,
        'calibrated_db': calibrated,
        'calibration_note': note,
        'raw_points_total': len(x),
    }


def _extract_raw_trace_series(
    file_name: str,
    raw: bytes,
    summary: FileSummary,
    trc_trace: Optional[dict],
    sor_meta: Optional[dict],
) -> Optional[dict]:
    ext = Path(file_name).suffix.lower()
    cache_key = f"{ext}:{_fr_fast_raw_key(raw)}"
    if cache_key in _TRACE_SERIES_CACHE:
        cached = _TRACE_SERIES_CACHE[cache_key]
        if isinstance(cached, dict):
            return {**cached}
        return cached
    result: Optional[dict] = None
    try:
        if ext == '.sor' and sor_meta and sor_meta.get('trace_values_db'):
            vals = [float(v) for v in sor_meta.get('trace_values_db')]
            range_km = sor_meta.get('trace_range_km') or sor_meta.get('graph_curve_max_km') or summary.length_km
            x, y = _series_from_uniform_samples(vals, range_km)
            if x and y:
                x, dist_meta = _normalize_raw_trace_distance_axis(x, summary)
                return {
                    'x_km': x,
                    'y_db': y,
                    'source': sor_meta.get('trace_source') or 'SOR DataPts',
                    'calibrated_db': bool(sor_meta.get('trace_calibrated_db', True)),
                    'calibration_note': 'DataPts đã có scaling dB từ file SOR',
                    'raw_points_total': len(x),
                    **dist_meta,
                }

        if ext in {'.trc', '.crt'} and trc_trace:
            raw_samples = trc_trace.get('RawSamples')
            if isinstance(raw_samples, (bytes, bytearray)) and len(raw_samples) >= 64:
                sample_count = len(raw_samples) // 2
                vals = list(struct.unpack('<' + 'h' * sample_count, bytes(raw_samples[:sample_count * 2])))
                range_m = trc_trace.get('Range') or trc_trace.get('DisplayRange') or trc_trace.get('SpansLength') or trc_trace.get('TraceSpansLength')
                range_km = float(range_m) / 1000.0 if isinstance(range_m, (int, float)) and float(range_m) > 10 else summary.length_km
                x, raw_y = _series_from_uniform_samples([float(v) for v in vals], range_km)
                x, dist_meta = _normalize_raw_trace_distance_axis(x, summary)
                series = _calibrate_raw_series_to_db(x, raw_y, summary, source='TRC RawSamples')
                if series:
                    series.update(dist_meta)
                    _TRACE_SERIES_CACHE[cache_key] = series
                    return {**series}

        if ext == '.msor':
            vals_i = _extract_mini_curve_values(raw)
            curve_max = _extract_curve_max_km(raw) or summary.graph_curve_max_km or summary.length_km
            if vals_i and curve_max:
                x, raw_y = _series_from_uniform_samples([float(v) for v in vals_i], curve_max)
                x, dist_meta = _normalize_raw_trace_distance_axis(x, summary)
                series = _calibrate_raw_series_to_db(x, raw_y, summary, source='MSOR ActernaMiniCurve')
                if series:
                    series.update(dist_meta)
                    _TRACE_SERIES_CACHE[cache_key] = series
                    return {**series}
    except Exception:
        _TRACE_SERIES_CACHE[cache_key] = None
        return None
    _TRACE_SERIES_CACHE[cache_key] = None
    return None


def _section_fit_confidence(r2: Optional[float], rms: Optional[float], points: int, calibrated: bool, note_prefix: str = '', *, fit_mode: str = 'exact_raw_fit') -> tuple[str, str]:
    """Human-readable confidence for raw/expanded section fitting.

    Exact section fits need fewer points than the old hard 24-point rule because
    MSOR MiniCurve data may be sparse.  Expanded-window fits are explicitly
    labelled as estimates; R²/RMS then describe the expanded fitting window, not
    the tiny section alone.
    """
    mode = (fit_mode or 'exact_raw_fit').strip().lower()
    if mode == 'endpoint_interpolation':
        return 'Ước lượng thấp', f'{note_prefix}Nội suy đầu-cuối từ raw trace; không tính R²/RMS thật.'
    if mode == 'nearest_trace_level_interpolation':
        return 'Ước lượng thấp', f'{note_prefix}Hoàn thiện section kiểu FastReporter: nội suy/điểm trace gần nhất tại hai biên; không tính R²/RMS thật.'
    if mode == 'local_slope_extrapolation':
        return 'Ước lượng thấp', f'{note_prefix}Hoàn thiện section kiểu FastReporter: dùng slope cục bộ từ cửa sổ gần section; không dùng như raw-fit chuẩn.'
    if mode == 'span_attenuation_estimate':
        return 'Ước lượng thấp', f'{note_prefix}Hoàn thiện section kiểu FastReporter: ước lượng theo suy hao trung bình/span attenuation; không có R²/RMS.'
    if points < 3:
        return 'Không đủ điểm', f'{note_prefix}Số điểm fit thấp ({points}).'
    if r2 is None or rms is None:
        return 'Không xác định', f'{note_prefix}Không tính được R²/RMS.'

    # Expanded-window fit is useful when section is too short, but the result is
    # still an estimate for the original short section.
    if mode == 'expanded_window_fit':
        if points < 6:
            return 'Không đủ điểm', f'{note_prefix}Cửa sổ mở rộng vẫn chỉ có {points} điểm.'
        if not calibrated:
            if r2 >= 0.90:
                return 'Shape OK', f'{note_prefix}Cửa sổ mở rộng tuyến tính tốt nhưng chưa scale chắc sang dB.'
            return 'Shape Review', f'{note_prefix}Cửa sổ mở rộng chỉ dùng để tham khảo hình dạng trace.'
        if r2 >= 0.950 and rms <= 0.180:
            return 'Khá - fit mở rộng', f'{note_prefix}Fit mở rộng tốt; dùng để ước lượng section ngắn.'
        if r2 >= 0.800 and rms <= 0.350:
            return 'Trung bình - fit mở rộng', f'{note_prefix}Fit mở rộng dùng được nhưng cần xem lại nếu section quan trọng.'
        return 'Thấp - fit mở rộng', f'{note_prefix}Fit mở rộng nhiễu hoặc không tuyến tính; ưu tiên kiểm tra trace.'

    if points < 8:
        return 'Không đủ điểm', f'{note_prefix}Số điểm fit trong section thấp ({points}).'
    if not calibrated:
        if r2 >= 0.95:
            return 'Shape OK', f'{note_prefix}Raw trace tuyến tính tốt nhưng chưa scale chắc sang dB.'
        if r2 >= 0.85:
            return 'Shape Review', f'{note_prefix}Raw trace có xu hướng nhưng còn nhiễu, chưa scale chắc sang dB.'
        return 'Shape Low', f'{note_prefix}Raw trace không tuyến tính rõ, chỉ dùng tham khảo.'
    if r2 >= 0.985 and rms <= 0.060:
        return 'Cao', f'{note_prefix}Fit tuyến tính rất tốt, residual thấp.'
    if r2 >= 0.950 and rms <= 0.120:
        return 'Khá', f'{note_prefix}Fit tốt, có thể dùng làm loss section.'
    if r2 >= 0.800 and rms <= 0.250:
        return 'Trung bình', f'{note_prefix}Fit dùng được nhưng cần xem lại trace nếu section quan trọng.'
    return 'Thấp', f'{note_prefix}Fit nhiễu hoặc không tuyến tính; ưu tiên fallback event/slope.'


def _weighted_linear_fit_xy(x: list[float], y: list[float], w: Optional[list[float]] = None) -> Optional[dict]:
    """Weighted linear fit used by expanded-window fitting.

    If weights are invalid or uniform, this falls back to the ordinary linear fit
    semantics.  R² and RMS are computed in weighted form so they describe the
    selected fit window while giving more importance to points near the original
    section.
    """
    if len(x) != len(y) or len(x) < 3:
        return None
    if w is None or len(w) != len(x):
        return _linear_fit_xy(x, y)
    triples = []
    for xx, yy, ww in zip(x, y, w):
        try:
            xx = float(xx); yy = float(yy); ww = float(ww)
        except Exception:
            continue
        if math.isfinite(xx) and math.isfinite(yy) and math.isfinite(ww) and ww > 0:
            triples.append((xx, yy, ww))
    if len(triples) < 3:
        return None
    sw = sum(t[2] for t in triples)
    if sw <= 0:
        return None
    x_bar = sum(xx * ww for xx, _yy, ww in triples) / sw
    y_bar = sum(yy * ww for _xx, yy, ww in triples) / sw
    denom = sum(ww * (xx - x_bar) ** 2 for xx, _yy, ww in triples)
    if abs(denom) < 1e-18:
        return None
    slope = sum(ww * (xx - x_bar) * (yy - y_bar) for xx, yy, ww in triples) / denom
    intercept = y_bar - slope * x_bar
    residuals = [yy - (intercept + slope * xx) for xx, yy, _ww in triples]
    sse = sum(ww * (yy - (intercept + slope * xx)) ** 2 for xx, yy, ww in triples)
    sst = sum(ww * (yy - y_bar) ** 2 for _xx, yy, ww in triples)
    r2 = 1.0 - (sse / sst) if sst > 1e-18 else 1.0
    rms = math.sqrt(sse / sw) if sw else None
    max_abs = max((abs(r) for r in residuals), default=None)
    return {
        'slope': slope,
        'intercept': intercept,
        'r2': max(min(r2, 1.0), -999.0),
        'rms': rms,
        'max_abs': max_abs,
        'residuals': residuals,
    }


def _interpolate_y_at(x_all: list[float], y_all: list[float], target_km: float) -> Optional[float]:
    """Linear interpolation of y at target_km using neighbouring real raw points."""
    pts = []
    for xx, yy in zip(x_all, y_all):
        try:
            x = float(xx); y = float(yy)
        except Exception:
            continue
        if math.isfinite(x) and math.isfinite(y):
            pts.append((x, y))
    if len(pts) < 2:
        return None
    pts.sort(key=lambda p: p[0])
    if target_km < pts[0][0] or target_km > pts[-1][0]:
        return None
    for i in range(1, len(pts)):
        x1, y1 = pts[i - 1]
        x2, y2 = pts[i]
        if x1 <= target_km <= x2:
            if abs(x2 - x1) < 1e-12:
                return y1
            return y1 + (target_km - x1) * (y2 - y1) / (x2 - x1)
    return pts[-1][1] if abs(target_km - pts[-1][0]) < 1e-9 else None



def _nearest_y_at(
    x_all: list[float],
    y_all: list[float],
    target_km: float,
    max_gap_km: float,
) -> tuple[Optional[float], Optional[float]]:
    """Return nearest trace level around target when exact interpolation is unavailable.

    This is only used by the FastReporter-style section completion path.  It does
    not change raw trace parsing; it just prevents a whole section cell from
    becoming blank when the boundary is slightly outside the sampled points.
    """
    best: Optional[tuple[float, float]] = None
    for xx, yy in zip(x_all, y_all):
        try:
            x = float(xx); y = float(yy)
        except Exception:
            continue
        if not (math.isfinite(x) and math.isfinite(y)):
            continue
        gap = abs(x - float(target_km))
        if best is None or gap < best[0]:
            best = (gap, y)
    if best is None:
        return None, None
    if best[0] <= max(float(max_gap_km), 0.0):
        return best[1], best[0]
    return None, best[0]


def _section_span_attenuation_estimate(
    summary: FileSummary,
    start_km: float,
    end_km: float,
    section_match_tolerance_m: float = 100.0,
) -> tuple[Optional[float], Optional[float], str, str]:
    """Last-resort FastReporter-like completion using span/average attenuation.

    This function intentionally does not touch event parsing, ORL, raw trace,
    route validation, or other workbook logic.  It only supplies a controlled
    section estimate when the normal fit/event calculation would otherwise leave
    the Sections sheet blank.
    """
    try:
        s = float(start_km); e = float(end_km)
    except Exception:
        return None, None, '', 'Start/End section không hợp lệ.'
    if not (math.isfinite(s) and math.isfinite(e)):
        return None, None, '', 'Start/End section không finite.'
    if e <= s:
        return None, None, '', 'Section có chiều dài bằng 0 hoặc âm.'

    # Respect route/fiber end boundaries.  We can clip the end if the section
    # slightly reaches beyond measured length, but we do not estimate fully after
    # the fibre end.
    route_length = None
    for value in (
        getattr(summary, 'length_km', None),
        getattr(summary, 'end_distance_km', None),
        getattr(summary, 'graph_end_km', None),
    ):
        try:
            if value not in (None, 0) and math.isfinite(float(value)) and float(value) > 0:
                route_length = float(value)
                break
        except Exception:
            continue
    match_tol_km = max(float(section_match_tolerance_m or 0.0), 0.0) / 1000.0
    if route_length is not None:
        if s >= route_length + max(match_tol_km, 0.05):
            return None, None, '', 'Section bắt đầu sau Fiber End/span length nên không ước lượng.'
        e = min(e, route_length)
    if e <= s:
        return None, None, '', 'Section nằm ngoài chiều dài đo sau khi clip theo Fiber End.'
    span = e - s

    candidates: list[tuple[float, str]] = []
    for att_value, label in (
        (getattr(summary, 'attenuation_dbkm', None), 'summary attenuation'),
    ):
        try:
            att = float(att_value)
            if math.isfinite(att) and 0 <= att <= 2.5:
                candidates.append((att, label))
        except Exception:
            pass

    for loss_value, length_value, label in (
        (getattr(summary, 'route_corrected_total_loss_db', None), route_length, 'route-corrected span loss / length'),
        (getattr(summary, 'total_loss_db', None), route_length, 'span loss / length'),
        (getattr(summary, 'parsed_total_loss_db', None), route_length, 'parsed span loss / length'),
    ):
        try:
            loss = float(loss_value)
            length = float(length_value) if length_value not in (None, 0) else None
            if length and math.isfinite(loss) and math.isfinite(length) and length > 0:
                att = loss / length
                if 0 <= att <= 2.5:
                    candidates.append((att, label))
        except Exception:
            pass

    if not candidates:
        return None, None, '', 'Không có attenuation/span loss hợp lệ để ước lượng section.'
    att, method = candidates[0]
    loss = att * span
    return round(loss, 4), round(att, 4), 'span_attenuation_estimate', f'Ước lượng theo {method}; span dùng tính = {span:.4f} km.'

def _median_spacing_km(x_all: list[float]) -> Optional[float]:
    xs = []
    for v in x_all:
        try:
            f = float(v)
        except Exception:
            continue
        if math.isfinite(f):
            xs.append(f)
    xs = sorted(set(xs))
    if len(xs) < 2:
        return None
    diffs = [xs[i] - xs[i - 1] for i in range(1, len(xs)) if xs[i] > xs[i - 1]]
    if not diffs:
        return None
    diffs.sort()
    mid = len(diffs) // 2
    if len(diffs) % 2:
        return diffs[mid]
    return (diffs[mid - 1] + diffs[mid]) / 2.0


def _build_section_fit_result_from_fit(
    *,
    trace_series: dict,
    file_name: str,
    section_index: int,
    start_km: float,
    end_km: float,
    span: float,
    xs: list[float],
    ys: list[float],
    fit: dict,
    calibrated: bool,
    fit_mode: str,
    fit_window_start_km: float,
    fit_window_end_km: float,
    r2_rms_scope: str,
    estimate_level: str,
    note_prefix: str = '',
) -> SectionFitResult:
    slope = float(fit['slope'])
    att = abs(slope) if calibrated else None
    loss = att * span if att is not None else None
    extra = note_prefix or ''
    if att is not None and (att < 0 or att > 2.5):
        att = None
        loss = None
        calibrated = False
        extra += 'Attenuation fit vượt dải hợp lý nên không dùng làm loss section. '
    r2 = round(float(fit['r2']), 5) if fit.get('r2') is not None else None
    rms = round(float(fit['rms']), 5) if fit.get('rms') is not None else None
    max_abs = round(float(fit['max_abs']), 5) if fit.get('max_abs') is not None else None
    conf, note = _section_fit_confidence(r2, rms, len(xs), bool(att is not None), note_prefix=extra, fit_mode=fit_mode)
    return SectionFitResult(
        file_name=file_name,
        section_index=section_index,
        start_km=round(float(start_km), 4),
        end_km=round(float(end_km), 4),
        span_km=round(span, 4),
        source=str(trace_series.get('source') or ''),
        raw_points_total=int(trace_series.get('raw_points_total') or len(trace_series.get('x_km') or [])),
        fit_points_used=len(xs),
        slope_dbkm=round(slope, 6),
        attenuation_dbkm=round(att, 4) if att is not None else None,
        loss_db=round(loss, 4) if loss is not None else None,
        intercept_db=round(float(fit['intercept']), 5) if fit.get('intercept') is not None else None,
        r2=r2,
        rms_residual_db=rms,
        max_abs_residual_db=max_abs,
        confidence=conf,
        used_for_section=False,
        fallback_method='',
        note=f"{note} Nguồn: {trace_series.get('source') or '-'}; {trace_series.get('calibration_note') or ''}; {trace_series.get('distance_scale_note') or ''}".strip(),
        fit_mode=fit_mode,
        fit_window_start_km=round(float(fit_window_start_km), 4),
        fit_window_end_km=round(float(fit_window_end_km), 4),
        r2_rms_scope=r2_rms_scope,
        estimate_level=estimate_level,
    )


def _fit_raw_trace_section(
    trace_series: Optional[dict],
    rows: list[EventRow],
    file_name: str,
    section_index: int,
    start_km: float,
    end_km: float,
    section_match_tolerance_m: float = 100.0,
) -> SectionFitResult:
    span = max(float(end_km) - float(start_km), 0.0)
    base_result = dict(
        file_name=file_name,
        section_index=section_index,
        start_km=round(float(start_km), 4),
        end_km=round(float(end_km), 4),
        span_km=round(span, 4),
        source='',
        raw_points_total=0,
        fit_points_used=0,
        slope_dbkm=None,
        attenuation_dbkm=None,
        loss_db=None,
        intercept_db=None,
        r2=None,
        rms_residual_db=None,
        max_abs_residual_db=None,
        confidence='Không có raw trace',
        used_for_section=False,
        fallback_method='',
        note='File chưa có raw trace samples đủ tin cậy để fit section.',
        fit_mode='event_fallback',
        fit_window_start_km=None,
        fit_window_end_km=None,
        r2_rms_scope='',
        estimate_level='fallback',
    )
    if span <= 0:
        base_result['confidence'] = 'Không hợp lệ'
        base_result['note'] = 'Section có chiều dài bằng 0 hoặc âm.'
        return SectionFitResult(**base_result)
    if not trace_series:
        return SectionFitResult(**base_result)

    x_all = trace_series.get('x_km') or []
    y_all = trace_series.get('y_db') or []
    if len(x_all) != len(y_all) or len(x_all) < 4:
        base_result.update({
            'source': str(trace_series.get('source') or ''),
            'raw_points_total': len(x_all),
            'confidence': 'Không đủ điểm',
            'note': 'Raw trace có quá ít điểm hoặc lỗi độ dài x/y.',
        })
        return SectionFitResult(**base_result)

    if trace_series.get('distance_scale_ok') is False:
        base_result.update({
            'source': str(trace_series.get('source') or ''),
            'raw_points_total': int(trace_series.get('raw_points_total') or len(x_all)),
            'confidence': 'Trục km lệch span',
            'note': (trace_series.get('distance_scale_note') or 'Trục km raw trace không khớp span tham chiếu.') + ' Không dùng raw-fit để tính loss section.',
        })
        return SectionFitResult(**base_result)

    calibrated = bool(trace_series.get('calibrated_db'))
    event_guard_km = max(0.020, min(0.180, float(section_match_tolerance_m) / 1000.0))
    boundary_guard_km = min(max(0.010, span * 0.035), 0.120)
    if span <= 0.20:
        boundary_guard_km = min(max(0.004, span * 0.015), 0.020)
    inner_start = float(start_km) + boundary_guard_km
    inner_end = float(end_km) - boundary_guard_km
    if inner_end <= inner_start:
        inner_start, inner_end = float(start_km), float(end_km)

    event_positions = []
    for r in rows:
        if r.distance_km is None:
            continue
        try:
            d = float(r.distance_km)
        except Exception:
            continue
        if not (float(start_km) < d < float(end_km)):
            continue
        et = (r.event_type or '').strip().lower()
        if et in {'section', 'đoạn tuyến'}:
            continue
        event_positions.append(d)

    finite_points = []
    for xx, yy in zip(x_all, y_all):
        try:
            x = float(xx); y = float(yy)
        except Exception:
            continue
        if math.isfinite(x) and math.isfinite(y):
            finite_points.append((x, y))
    finite_points.sort(key=lambda p: p[0])
    if len(finite_points) < 4:
        base_result.update({
            'source': str(trace_series.get('source') or ''),
            'raw_points_total': int(trace_series.get('raw_points_total') or len(x_all)),
            'confidence': 'Không đủ điểm',
            'note': 'Raw trace sau lọc finite còn quá ít điểm.',
        })
        return SectionFitResult(**base_result)

    def select_window(win_start: float, win_end: float, *, exclude_events: bool, use_inner: bool = False) -> tuple[list[float], list[float]]:
        xs: list[float] = []
        ys: list[float] = []
        a = inner_start if use_inner else win_start
        b = inner_end if use_inner else win_end
        for x, y in finite_points:
            if not (a <= x <= b):
                continue
            if exclude_events and any(abs(x - evd) <= event_guard_km for evd in event_positions):
                continue
            xs.append(x); ys.append(y)
        return xs, ys

    # 1) exact_raw_fit: only points inside this section after guard filtering.
    xs, ys = select_window(float(start_km), float(end_km), exclude_events=True, use_inner=True)
    if len(xs) < 8:
        xs, ys = select_window(float(start_km), float(end_km), exclude_events=False, use_inner=True)
    xs, ys = _downsample_xy(xs, ys, max_points=2400)
    if len(xs) >= 8:
        fit = _linear_fit_xy(xs, ys)
        if fit is not None:
            # Conservative outlier pass for larger exact fits.
            residuals = fit.get('residuals') or []
            rms0 = float(fit.get('rms') or 0.0)
            if len(xs) >= 24 and rms0 > 0:
                clip_limit = max(3.0 * rms0, 0.050 if calibrated else 0.0)
                clipped = [(x, y) for x, y, r in zip(xs, ys, residuals) if abs(float(r)) <= clip_limit]
                if len(clipped) >= max(8, int(0.55 * len(xs))):
                    xs2 = [p[0] for p in clipped]
                    ys2 = [p[1] for p in clipped]
                    fit2 = _linear_fit_xy(xs2, ys2)
                    if fit2 is not None:
                        fit = fit2; xs = xs2; ys = ys2
            exact_candidate = _build_section_fit_result_from_fit(
                trace_series=trace_series, file_name=file_name, section_index=section_index,
                start_km=start_km, end_km=end_km, span=span, xs=xs, ys=ys, fit=fit,
                calibrated=calibrated, fit_mode='exact_raw_fit',
                fit_window_start_km=float(start_km), fit_window_end_km=float(end_km),
                r2_rms_scope='section', estimate_level='exact',
            )
            if _is_fit_result_usable_for_section(exact_candidate):
                return exact_candidate
            # Exact fit exists but did not pass confidence gates; try expanded
            # window next so short/noisy sections still get the best near estimate.

    # 2) expanded_window_fit: borrow points around the section until enough raw
    # samples are available.  This keeps R²/RMS honest because they are computed
    # on real raw points in the expanded window, not on synthetic interpolation.
    x_min, x_max = finite_points[0][0], finite_points[-1][0]
    midpoint = (float(start_km) + float(end_km)) / 2.0
    spacing = _median_spacing_km([p[0] for p in finite_points]) or 0.05
    target_points = 12
    min_points = 6
    max_window_km = 3.0
    # For very sparse traces, allow a slightly larger window so at least 6-12
    # true points can be used, but keep it bounded to avoid representing a whole
    # route as one local section.
    max_window_km = max(max_window_km, min(5.0, spacing * 18.0))
    half_window = max(span / 2.0, spacing * target_points / 2.0, 0.35)
    best = None
    local_slope_candidate: Optional[SectionFitResult] = None
    while half_window <= max_window_km / 2.0 + 1e-9:
        win_start = max(x_min, midpoint - half_window)
        win_end = min(x_max, midpoint + half_window)
        xw, yw = select_window(win_start, win_end, exclude_events=False, use_inner=False)
        if len(xw) >= min_points:
            best = (win_start, win_end, xw, yw)
            if len(xw) >= target_points:
                break
        half_window *= 1.45
    if best is not None:
        win_start, win_end, xw, yw = best
        # Weighted fit: points inside or very near the original section dominate.
        half = max((win_end - win_start) / 2.0, 1e-6)
        weights = []
        for x in xw:
            if float(start_km) <= x <= float(end_km):
                weights.append(1.0)
            else:
                distance_from_section = 0.0
                if x < float(start_km):
                    distance_from_section = float(start_km) - x
                elif x > float(end_km):
                    distance_from_section = x - float(end_km)
                weights.append(max(0.25, 1.0 - (distance_from_section / half)))
        xw, yw = _downsample_xy(xw, yw, max_points=2400)
        if len(weights) != len(xw):
            weights = None
        fit = _weighted_linear_fit_xy(xw, yw, weights)
        if fit is not None:
            expanded_candidate = _build_section_fit_result_from_fit(
                trace_series=trace_series, file_name=file_name, section_index=section_index,
                start_km=start_km, end_km=end_km, span=span, xs=xw, ys=yw, fit=fit,
                calibrated=calibrated, fit_mode='expanded_window_fit',
                fit_window_start_km=win_start, fit_window_end_km=win_end,
                r2_rms_scope='expanded_window', estimate_level='near_estimate',
                note_prefix='Section thiếu điểm nên dùng cửa sổ mở rộng. ',
            )
            if _is_fit_result_usable_for_section(expanded_candidate):
                return expanded_candidate
            # Expanded fit was generated but too noisy/unreliable; keep it as a
            # FastReporter-style local-slope completion candidate.  It is used
            # only after exact/expanded/endpoint interpolation fail, and its
            # provenance is written to Section Fit Quality.
            if expanded_candidate.loss_db is not None and expanded_candidate.attenuation_dbkm is not None and 0 <= expanded_candidate.attenuation_dbkm <= 2.5:
                local_slope_candidate = expanded_candidate

    # 3) endpoint_interpolation: always try to estimate loss/attenuation from the
    # raw curve at the section boundaries.  R²/RMS are intentionally left blank.
    y_start = _interpolate_y_at([p[0] for p in finite_points], [p[1] for p in finite_points], float(start_km))
    y_end = _interpolate_y_at([p[0] for p in finite_points], [p[1] for p in finite_points], float(end_km))
    if calibrated and y_start is not None and y_end is not None and span > 0:
        slope = (float(y_end) - float(y_start)) / span
        att = abs(slope)
        loss = att * span
        if 0 <= att <= 2.5:
            conf, note = _section_fit_confidence(None, None, 2, True, fit_mode='endpoint_interpolation')
            return SectionFitResult(
                file_name=file_name,
                section_index=section_index,
                start_km=round(float(start_km), 4),
                end_km=round(float(end_km), 4),
                span_km=round(span, 4),
                source=str(trace_series.get('source') or ''),
                raw_points_total=int(trace_series.get('raw_points_total') or len(x_all)),
                fit_points_used=2,
                slope_dbkm=round(slope, 6),
                attenuation_dbkm=round(att, 4),
                loss_db=round(loss, 4),
                intercept_db=None,
                r2=None,
                rms_residual_db=None,
                max_abs_residual_db=None,
                confidence=conf,
                used_for_section=False,
                fallback_method='',
                note=f"{note} Nguồn: {trace_series.get('source') or '-'}; {trace_series.get('calibration_note') or ''}; {trace_series.get('distance_scale_note') or ''}".strip(),
                fit_mode='endpoint_interpolation',
                fit_window_start_km=round(float(start_km), 4),
                fit_window_end_km=round(float(end_km), 4),
                r2_rms_scope='not_calculated',
                estimate_level='low_estimate',
            )

    # 4) nearest_trace_level_interpolation: FastReporter-like completion when
    # a boundary falls slightly outside the sampled trace points.  This prevents
    # blank section cells while still recording low confidence / no R².
    if calibrated and span > 0:
        xs_all = [p[0] for p in finite_points]
        ys_all = [p[1] for p in finite_points]
        spacing_for_gap = _median_spacing_km(xs_all) or 0.05
        max_gap_km = max(0.05, min(0.35, max(spacing_for_gap * 3.0, span * 0.60)))
        ys0 = y_start
        ye0 = y_end
        gap_s = 0.0
        gap_e = 0.0
        if ys0 is None:
            ys0, gap_s = _nearest_y_at(xs_all, ys_all, float(start_km), max_gap_km)
        if ye0 is None:
            ye0, gap_e = _nearest_y_at(xs_all, ys_all, float(end_km), max_gap_km)
        if ys0 is not None and ye0 is not None:
            slope = (float(ye0) - float(ys0)) / span
            att = abs(slope)
            loss = att * span
            if 0 <= att <= 2.5:
                conf, note = _section_fit_confidence(None, None, 2, True, fit_mode='nearest_trace_level_interpolation')
                note += f' Khoảng cách điểm gần nhất tới biên: start={float(gap_s or 0):.4f} km, end={float(gap_e or 0):.4f} km.'
                return SectionFitResult(
                    file_name=file_name, section_index=section_index,
                    start_km=round(float(start_km), 4), end_km=round(float(end_km), 4),
                    span_km=round(span, 4), source=str(trace_series.get('source') or ''),
                    raw_points_total=int(trace_series.get('raw_points_total') or len(x_all)),
                    fit_points_used=2, slope_dbkm=round(slope, 6),
                    attenuation_dbkm=round(att, 4), loss_db=round(loss, 4),
                    intercept_db=None, r2=None, rms_residual_db=None, max_abs_residual_db=None,
                    confidence=conf, used_for_section=False, fallback_method='',
                    note=f"{note} Nguồn: {trace_series.get('source') or '-'}; {trace_series.get('calibration_note') or ''}; {trace_series.get('distance_scale_note') or ''}".strip(),
                    fit_mode='nearest_trace_level_interpolation',
                    fit_window_start_km=round(float(start_km), 4),
                    fit_window_end_km=round(float(end_km), 4),
                    r2_rms_scope='not_calculated', estimate_level='low_estimate',
                )

    # 5) local_slope_extrapolation: use the local expanded-window slope when it
    # produced a sane attenuation but failed the stricter R²/RMS gate.
    if local_slope_candidate is not None:
        local_slope_candidate.fit_mode = 'local_slope_extrapolation'
        local_slope_candidate.confidence = 'Ước lượng thấp'
        local_slope_candidate.estimate_level = 'low_estimate'
        local_slope_candidate.r2_rms_scope = local_slope_candidate.r2_rms_scope or 'expanded_window'
        local_slope_candidate.note = (local_slope_candidate.note or '') + ' | FastReporter-style completion: dùng slope cục bộ vì không đủ điều kiện exact/expanded fit chuẩn.'
        return local_slope_candidate

    base_result.update({
        'source': str(trace_series.get('source') or ''),
        'raw_points_total': int(trace_series.get('raw_points_total') or len(x_all)),
        'fit_points_used': len(xs) if 'xs' in locals() else 0,
        'confidence': 'Không đủ điểm',
        'note': 'Không đủ điểm để exact/expanded fit và không nội suy được hai biên section; dùng fallback event/slope nếu có.',
        'fit_mode': 'event_fallback',
        'estimate_level': 'fallback',
    })
    return SectionFitResult(**base_result)


def _is_fit_result_usable_for_section(result: SectionFitResult) -> bool:
    if result.loss_db is None or result.attenuation_dbkm is None:
        return False
    if result.attenuation_dbkm < 0 or result.attenuation_dbkm > 2.5:
        return False
    mode = (getattr(result, 'fit_mode', '') or '').lower()
    if mode == 'exact_raw_fit':
        if result.fit_points_used < 8:
            return False
        if result.r2 is None or result.rms_residual_db is None:
            return False
        return result.r2 >= 0.80 and result.rms_residual_db <= 0.35
    if mode == 'expanded_window_fit':
        if result.fit_points_used < 6:
            return False
        if result.r2 is None or result.rms_residual_db is None:
            return False
        return result.r2 >= 0.75 and result.rms_residual_db <= 0.45
    if mode in {'endpoint_interpolation', 'nearest_trace_level_interpolation', 'local_slope_extrapolation', 'span_attenuation_estimate'}:
        # Completion modes are used only after exact/expanded fit fails.  They
        # are still valid numeric section values, with confidence/provenance kept
        # in Section Fit Quality.
        return True
    return False

def _fr_compute_section_values(summary: FileSummary, rows: list[EventRow], start_km: float, end_km: float, section_match_tolerance_m: float = 100.0, section_measurement_mode: str = 'fit') -> tuple[Optional[float], Optional[float]]:
    route_length_km = summary.length_km
    match_tol_km = max(float(section_match_tolerance_m), 0.0) / 1000.0
    if route_length_km in (None, 0) or start_km >= float(route_length_km) + max(match_tol_km, 0.05):
        return None, None
    effective_end = min(float(end_km), float(route_length_km))
    if effective_end <= start_km:
        return None, None
    mode = (section_measurement_mode or 'fit').strip().lower()
    if mode in {'2point', 'two_point', 'two-point'}:
        loss_db, _method = _estimate_segment_loss_twopoint(rows, start_km, effective_end, route_length_km, summary.total_loss_db)
        if loss_db is None:
            if summary.source_format.upper() == 'SOR':
                loss_db, _method = _estimate_segment_loss_standard_sor(rows, start_km, effective_end, route_length_km)
            else:
                loss_db, _method = _estimate_segment_loss_generic(rows, start_km, effective_end, summary.total_loss_db, route_length_km)
    else:
        if summary.source_format.upper() == 'SOR':
            loss_db, _method = _estimate_segment_loss_standard_sor(rows, start_km, effective_end, route_length_km)
        else:
            loss_db, _method = _estimate_segment_loss_generic(rows, start_km, effective_end, summary.total_loss_db, route_length_km)
    if loss_db is None:
        # FastReporter-style completion: if event/2-point logic cannot produce a
        # value, estimate the section from span/average attenuation instead of
        # leaving the Sections table blank.  The richer provenance is written by
        # _fr_compute_section_values_with_fit; this simple path returns numbers
        # for callers that do not carry SectionFitResult.
        est_loss, est_att, _est_method, _est_note = _section_span_attenuation_estimate(
            summary, start_km, effective_end, section_match_tolerance_m=section_match_tolerance_m
        )
        if est_loss is not None and est_att is not None:
            return est_loss, est_att
        return None, None
    span = max(effective_end - start_km, 0.0)
    if span <= 0:
        return None, None
    att = round(loss_db / span, 3)
    return round(loss_db, 3), att


def _fr_compute_section_values_with_fit(
    summary: FileSummary,
    ctx: dict,
    rows: list[EventRow],
    start_km: float,
    end_km: float,
    *,
    section_index: int,
    section_match_tolerance_m: float = 100.0,
    section_measurement_mode: str = 'fit',
) -> tuple[Optional[float], Optional[float], SectionFitResult]:
    mode = (section_measurement_mode or 'fit').strip().lower()
    trace_series = ctx.get('raw_trace_series') if ctx else None
    fit_result = _fit_raw_trace_section(
        trace_series,
        rows,
        summary.file_name,
        section_index,
        start_km,
        end_km,
        section_match_tolerance_m=section_match_tolerance_m,
    )

    fallback_mode = '2point' if mode in {'2point', 'two_point', 'two-point'} else 'event'
    fallback_loss, fallback_att = _fr_compute_section_values(
        summary,
        rows,
        start_km,
        end_km,
        section_match_tolerance_m=section_match_tolerance_m,
        section_measurement_mode=fallback_mode,
    )
    if mode in {'fit', 'raw_fit', 'raw-trace', 'raw_trace'} and _is_fit_result_usable_for_section(fit_result):
        fit_result.used_for_section = True
        fit_result.fallback_method = ''
        return fit_result.loss_db, fit_result.attenuation_dbkm, fit_result

    fit_result.fallback_method = fallback_mode

    if mode in {'fit', 'raw_fit', 'raw-trace', 'raw_trace'} and fallback_loss is not None:
        fit_result.used_for_section = False
        if fit_result.note:
            fit_result.note += ' | Section value dùng fallback vì fit chưa đạt ngưỡng R²/RMS hoặc chưa calibrated.'
        else:
            fit_result.note = 'Section value dùng fallback vì fit chưa đạt ngưỡng R²/RMS hoặc chưa calibrated.'
        return fallback_loss, fallback_att, fit_result

    # FastReporter-style completion: if both raw-fit and event fallback return
    # empty, fill the section from span/average attenuation when available.  This
    # prevents blank cells while keeping the method/confidence in Fit Quality.
    if mode in {'fit', 'raw_fit', 'raw-trace', 'raw_trace'}:
        est_loss, est_att, est_method, est_note = _section_span_attenuation_estimate(
            summary, start_km, end_km, section_match_tolerance_m=section_match_tolerance_m
        )
        if est_loss is not None and est_att is not None:
            fit_result.loss_db = est_loss
            fit_result.attenuation_dbkm = est_att
            fit_result.fit_mode = est_method or 'span_attenuation_estimate'
            fit_result.fallback_method = est_method or 'span_attenuation_estimate'
            fit_result.confidence = 'Ước lượng thấp'
            fit_result.estimate_level = 'span_estimate'
            fit_result.r2_rms_scope = 'not_calculated'
            fit_result.used_for_section = False
            fit_result.note = ((fit_result.note or '') + ' | FastReporter-style completion: ' + (est_note or 'ước lượng theo suy hao trung bình/span attenuation.')).strip()
            return est_loss, est_att, fit_result

    # Explicit event/2-point mode: still expose R²/RMS diagnostics but do not use them
    # to replace the requested calculation mode.
    fit_result.used_for_section = False
    if mode not in {'fit', 'raw_fit', 'raw-trace', 'raw_trace'}:
        fit_result.note = (fit_result.note or '') + f' | Mode đang chọn = {mode}; fit chỉ dùng làm chẩn đoán chất lượng trace.'
    return fallback_loss, fallback_att, fit_result




def _fr_clear_cells(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _fr_merged_anchor(ws, row: int, col: int) -> tuple[int, int]:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return row, col
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col
    return row, col


def _fr_set_cell_value(ws, row: int, col: int, value) -> None:
    anchor_row, anchor_col = _fr_merged_anchor(ws, row, col)
    ws.cell(anchor_row, anchor_col).value = value


def _fr_insert_rows_preserve_merges(ws, insert_before: int, amount: int) -> None:
    if amount <= 0:
        return
    merged_ranges = [CellRange(str(rng)) for rng in ws.merged_cells.ranges]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.insert_rows(insert_before, amount=amount)
    for rng in merged_ranges:
        if rng.max_row < insert_before:
            shifted = rng
        elif rng.min_row >= insert_before:
            shifted = CellRange(
                min_col=rng.min_col,
                min_row=rng.min_row + amount,
                max_col=rng.max_col,
                max_row=rng.max_row + amount,
            )
        else:
            shifted = CellRange(
                min_col=rng.min_col,
                min_row=rng.min_row,
                max_col=rng.max_col,
                max_row=rng.max_row + amount,
            )
        ws.merge_cells(str(shifted))


def _fr_copy_row_layout(ws, src_row: int, dst_row: int, min_col: int = 1, max_col: Optional[int] = None) -> None:
    if max_col is None:
        max_col = ws.max_column
    for c in range(min_col, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if isinstance(src, MergedCell):
            continue
        dst.value = src.value
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst._style = copy(src._style)
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == src_row and merged.max_row == src_row:
            ws.merge_cells(
                start_row=dst_row,
                start_column=merged.min_col,
                end_row=dst_row,
                end_column=merged.max_col,
            )


def _fr_expand_single_row_table(ws, *, data_start: int, template_row: int, insert_before: int, count: int, reserved_count: int) -> None:
    extra = max(count - reserved_count, 0)
    if extra > 0:
        _fr_insert_rows_preserve_merges(ws, insert_before, extra)
        for offset in range(extra):
            _fr_copy_row_layout(ws, template_row, insert_before + offset)


def _fr_expand_pair_table(ws, *, data_start: int, template_rows: tuple[int, int], insert_before: int, count: int, reserved_pairs: int) -> int:
    extra_pairs = max(count - reserved_pairs, 0)
    if extra_pairs > 0:
        _fr_insert_rows_preserve_merges(ws, insert_before, extra_pairs * 2)
        for pair_index in range(extra_pairs):
            base_row = insert_before + pair_index * 2
            _fr_copy_row_layout(ws, template_rows[0], base_row)
            _fr_copy_row_layout(ws, template_rows[1], base_row + 1)
    return insert_before + extra_pairs * 2


def _fr_expand_horizontal_pair_blocks(ws, required_pairs: int, *, header_merge_rows: tuple[int, ...]) -> None:
    """Expand a sheet made of repeated 2-column blocks starting at column C.

    Some templates keep one or more *blank trailing pairs* at the end of the
    sheet width. If we naively treat ``ws.max_column`` as the number of real
    blocks, newly created event/section pairs inherit those blank columns and the
    header/body formatting breaks around the first expanded pair (this is exactly
    what caused the malformed Event 63+ layout).

    To avoid that, we first detect the *last real pair* by looking for actual
    header/body content or an existing merged header. New pairs are then copied
    from that real template pair instead of the blank trailing columns.
    """
    if required_pairs <= 0:
        return

    max_possible_pairs = max((ws.max_column - 2) // 2, 0)
    current_pairs = _fr_count_real_pair_blocks(ws, header_merge_rows=header_merge_rows)
    if current_pairs <= 0:
        current_pairs = max_possible_pairs
    if required_pairs <= current_pairs:
        return

    template_pair = current_pairs if current_pairs > 0 else 1
    src_col1 = 3 + (template_pair - 1) * 2
    src_col2 = src_col1 + 1
    src_width1 = ws.column_dimensions[get_column_letter(src_col1)].width
    src_width2 = ws.column_dimensions[get_column_letter(src_col2)].width

    for i in range(current_pairs + 1, required_pairs + 1):
        col1 = 3 + (i - 1) * 2
        col2 = col1 + 1
        for r in range(1, ws.max_row + 1):
            src1 = ws.cell(r, src_col1)
            dst1 = ws.cell(r, col1)
            if not isinstance(src1, MergedCell):
                dst1.font = copy(src1.font)
                dst1.fill = copy(src1.fill)
                dst1.border = copy(src1.border)
                dst1.alignment = copy(src1.alignment)
                dst1.number_format = src1.number_format
                dst1.protection = copy(src1.protection)
                dst1._style = copy(src1._style)
                dst1.value = src1.value
            src2 = ws.cell(r, src_col2)
            dst2 = ws.cell(r, col2)
            if not isinstance(src2, MergedCell):
                dst2.font = copy(src2.font)
                dst2.fill = copy(src2.fill)
                dst2.border = copy(src2.border)
                dst2.alignment = copy(src2.alignment)
                dst2.number_format = src2.number_format
                dst2.protection = copy(src2.protection)
                dst2._style = copy(src2._style)
                dst2.value = src2.value

        ws.column_dimensions[get_column_letter(col1)].width = src_width1
        ws.column_dimensions[get_column_letter(col2)].width = src_width2
        for merge_row in header_merge_rows:
            ws.merge_cells(start_row=merge_row, start_column=col1, end_row=merge_row, end_column=col2)


def _fr_count_real_pair_blocks(ws, *, header_merge_rows: tuple[int, ...]) -> int:
    """Return the number of real 2-column blocks starting at column C.

    We intentionally ignore blank trailing pairs that exist only because the
    template workbook has a larger sheet width than the last populated
    Event/Section block.
    """
    max_possible_pairs = max((ws.max_column - 2) // 2, 0)

    def _pair_has_real_content(pair_index: int) -> bool:
        col1 = 3 + (pair_index - 1) * 2
        col2 = col1 + 1
        for merge_row in header_merge_rows:
            for rng in ws.merged_cells.ranges:
                if (
                    rng.min_row == merge_row and rng.max_row == merge_row
                    and rng.min_col == col1 and rng.max_col == col2
                ):
                    if ws.cell(merge_row, col1).value is not None:
                        return True
        probe_rows = (1, 2, 3, 4)
        return any(
            ws.cell(r, c).value is not None
            for r in probe_rows
            for c in (col1, col2)
        )

    actual_pairs = 0
    for i in range(1, max_possible_pairs + 1):
        if _pair_has_real_content(i):
            actual_pairs = i
    return actual_pairs


def _fr_ensure_events_capacity(ws, required_events: int) -> None:
    """Expand the Events sheet horizontally when real event count exceeds the template width."""
    _fr_expand_horizontal_pair_blocks(ws, required_events, header_merge_rows=(1, 2))


def _fr_ensure_sections_capacity(ws, required_sections: int) -> None:
    """Expand the Sections sheet horizontally when real section count exceeds the template width."""
    _fr_expand_horizontal_pair_blocks(ws, required_sections, header_merge_rows=(1,))


def _extract_wavelength_nm_from_display(display: Optional[str]) -> Optional[int]:
    if not display:
        return None
    m = re.search(r"(\d{3,4})", str(display))
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _build_core_metrics(summary: FileSummary, ctx: dict, threshold_db: float = 0.5, section_pairs: Optional[list[tuple[Optional[float], Optional[float]]]] = None, duration_threshold_s: Optional[float] = None) -> dict:
    rows = ctx.get('events', []) or []
    wavelength_nm = _extract_wavelength_nm_from_display(summary.wavelength_display)
    fiber_label = _display_fiber_label(summary, ctx.get('metadata', {}) if isinstance(ctx, dict) else {})
    event_rows = [r for r in rows if r.distance_km is not None]
    event_losses = [float(r.loss_db) for r in rows if r.loss_db is not None and float(r.loss_db) > 0 and _fr_event_type_label(r) not in ('Launch Level', 'Fiber End')]
    red_event_losses = [float(r.loss_db) for r in rows if r.loss_db is not None and abs(float(r.loss_db)) + 1e-12 >= float(threshold_db)]
    splice_losses = [float(r.loss_db) for r in rows if r.loss_db is not None and float(r.loss_db) > 0 and _fr_event_type_label(r) == 'Non-Reflective']
    if section_pairs is None:
        section_pairs = ctx.get('section_pairs') or []
    section_losses = [float(loss) for loss, _att in (section_pairs or []) if isinstance(loss, (int, float))]
    duration_s, duration_status, duration_note = _duration_status_from_meta(ctx, duration_threshold_s)
    return {
        'file_name': summary.file_name,
        'display_file_name': _stv_display_file_name(summary.file_name),
        'source_format': summary.source_format,
        'fiber_label': fiber_label,
        'fiber_raw': summary.fiber,
        'wavelength_nm': wavelength_nm,
        'wavelength_display': summary.wavelength_display,
        'length_km': summary.length_km,
        'total_loss_db': summary.total_loss_db,
        'attenuation_dbkm': summary.attenuation_dbkm,
        'orl_display': ctx.get('orl_display'),
        'orl_status': ctx.get('orl_status', 'Unknown'),
        'orl_advanced_status': getattr(ctx.get('orl_analysis'), 'advanced_status', ''),
        'orl_source_detail': getattr(ctx.get('orl_analysis'), 'source_detail', ctx.get('orl_source_detail', '')),
        'orl_use_for_judgment': getattr(ctx.get('orl_analysis'), 'use_for_judgment', ctx.get('orl_use_for_judgment', False)),
        'loss_source_used': getattr(summary, 'loss_source_used', ''),
        'parse_family': getattr(summary, 'parse_family', ''),
        'parse_family_reason': getattr(summary, 'parse_family_reason', ''),
        'event_count': len(event_rows),
        'max_event_loss_db': round(max(event_losses), 3) if event_losses else None,
        'max_red_event_loss_db': round(max((abs(v) for v in red_event_losses)), 3) if red_event_losses else None,
        'avg_splice_loss_db': round(sum(splice_losses) / len(splice_losses), 3) if splice_losses else None,
        'max_section_loss_db': round(max(section_losses), 3) if section_losses else None,
        'avg_section_loss_db': round(sum(section_losses) / len(section_losses), 3) if section_losses else None,
        'duration_s': round(duration_s, 3) if isinstance(duration_s, (int, float)) else None,
        'duration_threshold_s': duration_threshold_s,
        'duration_status': duration_status,
        'duration_note': duration_note,
    }


def _fr_fill_core_metrics_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict], threshold_db: float = 0.5, section_pairs_by_file: Optional[dict[str, list[tuple[Optional[float], Optional[float]]]]] = None, duration_threshold_s: Optional[float] = None) -> None:
    _fr_write_header_row(ws, [
        'Tệp', 'Định dạng', 'Sợi', 'Bước sóng (nm)', 'Chiều dài (km)', 'Suy hao tổng (dB)',
        'Suy hao TB (dB/km)', 'ORL hiển thị', 'Trạng thái ORL', 'ORL advanced status',
        'Nguồn ORL', 'ORL dùng để kết luận?', 'Nguồn suy hao tổng',
        'Family parser', 'Số event', 'Max event loss', 'Max event đỏ', 'Avg splice loss',
        'Max section loss', 'Avg section loss', 'Duration (s)', 'Duration threshold (s)', 'Duration status', 'Duration note'
    ])
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        section_pairs = (section_pairs_by_file or {}).get(summary.file_name) if section_pairs_by_file else None
        core = _build_core_metrics(summary, ctx, threshold_db=threshold_db, section_pairs=section_pairs, duration_threshold_s=duration_threshold_s)
        values = [
            core['display_file_name'], core['source_format'], core['fiber_label'], core['wavelength_nm'],
            core['length_km'], core['total_loss_db'], core['attenuation_dbkm'], core['orl_display'], core['orl_status'],
            core['orl_advanced_status'], core['orl_source_detail'], 'Yes' if core['orl_use_for_judgment'] else 'No',
            core['loss_source_used'], _to_vi_parser_family(core['parse_family']), core['event_count'], core['max_event_loss_db'],
            core['max_red_event_loss_db'], core['avg_splice_loss_db'], core['max_section_loss_db'], core['avg_section_loss_db'],
            core['duration_s'], core['duration_threshold_s'], core['duration_status'], core['duration_note']
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row, c, value)
        if core.get('duration_status') == 'Fail':
            ws.cell(row, 23).fill = RED_FILL
            ws.cell(row, 24).fill = RED_FILL
        row += 1
    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=70)


def _fr_fill_output_rules_sheet(ws, *, threshold_db: float, deviation_m: float, output_mode: str, section_export_scope: str, section_measurement_mode: str, section_event_source: str, section_boundary_priority: str, expected_route_km: Optional[float], graph_reach_tolerance_km: Optional[float], event_shortfall_tolerance_km: Optional[float]) -> None:
    _fr_write_header_row(ws, ['Nhóm', 'Rule đang khóa trong output'])
    rules = [
        ('Lõi dữ liệu', 'STV và FastReporter cùng dùng một context chuẩn hóa từ parser; không parse lại riêng ở từng đầu ra.'),
        ('Ngưỡng event', f'Ngưỡng event hiện tại = {threshold_db}. FastReporter sự kiện đỏ giữ rule |loss| >= ngưỡng; STV giữ rule theo format STV đã tinh chỉnh.'),
        ('Gom cụm event', f'Sai số gom cụm event dùng chung = {deviation_m} m.'),
        ('Ngưỡng section', f'Ngưỡng Section Loss để tô đỏ = {section_threshold_db if section_threshold_db is not None else "chưa đặt"} dB. Mã ID S+E hiển thị một lần ở tiêu đề mỗi section = Start km + End km; đây chỉ là mã nhận diện layout, không phải chiều dài tuyến và không tạo cột riêng trong dữ liệu core.'),
        ('Duration', f'Ngưỡng duration = {duration_threshold_s if duration_threshold_s is not None else "chưa đặt"} giây. Nếu core có duration thấp hơn ngưỡng thì đánh dấu Fail trong Core Metrics/Link Results.'),
        ('Strict Validation', 'Sheet Strict Validation kiểm tra parser, metadata, route length, attenuation sanity, event coverage, trace scale, fit readiness và ORL readiness. Sheet này chỉ cảnh báo/khóa quyền dùng để kết luận trên giấy, tuyệt đối không sửa số liệu đã tính.'),
        ('Section', f'Section export scope = {section_export_scope}; mode tính section = {section_measurement_mode}; nguồn event dựng section = {section_event_source}; ưu tiên = {section_boundary_priority}. Với mode fit, app ưu tiên raw-trace linear fit, xuất R²/RMS residual và fallback về event/slope nếu fit chưa đủ tin cậy.'),
        ('ORL', 'ORL Analysis phân tách measured ORL / metadata lower-bound / not available. Chỉ measured ORL mới có Use for Judgment = Yes. Lower-bound dạng <xx.xx luôn được ghi rõ là tham khảo, không phải span ORL thật.'),
        ('ORL vật lý từ trace', 'Mặc định tắt. Khi bật diagnostic/experimental, app chỉ kiểm tra điều kiện hiệu chuẩn; nếu thiếu backscatter/launch power/reference calibration thì không tạo số ORL giả.'),
        ('Route check', f'Chiều dài tuyến chuẩn = {expected_route_km}; sai số chạm tuyến = {graph_reach_tolerance_km}; hụt event cho phép = {event_shortfall_tolerance_km}.'),
        ('Output mode', f'Output hiện tại = {output_mode}. Chỉ khác layout, không khác lõi số liệu.')
    ]
    row = 2
    for group, rule in rules:
        ws.cell(row, 1, group)
        ws.cell(row, 2, rule)
        row += 1
    _fr_selective_autofit(ws, max_scan_rows=max(row, 10), max_width=110)



def _duration_status_from_meta(ctx: dict, duration_threshold_s: Optional[float]) -> tuple[Optional[float], str, str]:
    """Return duration, status, note without changing parser logic."""
    meta = (ctx or {}).get('metadata', {}) if isinstance(ctx, dict) else {}
    duration = meta.get('duration_s') if isinstance(meta, dict) else None
    try:
        duration_f = float(duration) if duration is not None else None
    except Exception:
        duration_f = None
    if duration_threshold_s is None:
        return duration_f, 'Not checked', 'Chưa đặt ngưỡng duration.'
    try:
        thr = float(duration_threshold_s)
    except Exception:
        return duration_f, 'Not checked', 'Ngưỡng duration không hợp lệ.'
    if duration_f is None:
        return None, 'Unknown', f'Không đọc được duration từ file; ngưỡng đang đặt {thr:g}s.'
    if duration_f + 1e-12 < thr:
        return duration_f, 'Fail', f'Duration {duration_f:g}s < ngưỡng {thr:g}s.'
    return duration_f, 'Pass', f'Duration {duration_f:g}s >= ngưỡng {thr:g}s.'


def _section_loss_reaches_threshold(loss_db: Optional[float], section_threshold_db: Optional[float]) -> bool:
    if section_threshold_db is None or loss_db is None:
        return False
    try:
        return abs(float(loss_db)) + 1e-12 >= float(section_threshold_db)
    except Exception:
        return False


def _section_start_end_total_km(sec: dict) -> Optional[float]:
    """User-requested section identifier: Start km + End km, e.g. 1+5=6, 5+12=17."""
    try:
        return round(float(sec.get('start_km')) + float(sec.get('end_km')), 6)
    except Exception:
        return None



def _fr_clear_and_hide_unused_section_blocks(ws, first_unused_index: int, last_existing_index: int, *, start_col: int = 3, block: int = 2, max_row: Optional[int] = None) -> None:
    """Hide/clear leftover FastReporter section template blocks beyond real sections.

    This is presentation-only. It prevents old/template section blocks (for example
    Section 43..46) from appearing as blank blue headers when the real calculated
    section list ends earlier. It does not change parser, event, trace, or section
    calculation values.
    """
    try:
        first = int(first_unused_index)
        last = int(last_existing_index)
    except Exception:
        return
    if last < first:
        return
    if max_row is None:
        try:
            max_row = int(ws.max_row or 80)
        except Exception:
            max_row = 80
    max_row = max(max_row, 80)
    # Remove merged headers belonging to unused section columns so Excel will not
    # keep stale merged blue blocks visible after a later export with fewer sections.
    unused_start_col = start_col + (first - 1) * block
    unused_end_col = start_col + last * block - 1
    for rng in list(ws.merged_cells.ranges):
        try:
            if rng.max_col >= unused_start_col and rng.min_col <= unused_end_col:
                ws.unmerge_cells(str(rng))
        except Exception:
            pass
    for i in range(first, last + 1):
        col = start_col + (i - 1) * block
        for c in (col, col + 1):
            try:
                letter = get_column_letter(c)
                ws.column_dimensions[letter].hidden = True
                ws.column_dimensions[letter].width = 0
            except Exception:
                pass
            for r in range(1, max_row + 1):
                try:
                    ws.cell(r, c).value = None
                except Exception:
                    pass

def _fr_fill_link_results(ws, summaries: list[FileSummary], contexts: dict[str, dict], section_pairs_by_file: dict[str, list[tuple[Optional[float], Optional[float]]]], duration_threshold_s: Optional[float] = None) -> None:
    data_start = 3
    reserved_count = 20
    insert_before = 23
    template_row = 22
    _fr_expand_single_row_table(ws, data_start=data_start, template_row=template_row, insert_before=insert_before, count=len(summaries), reserved_count=reserved_count)

    cols = [1, 5, 7, 10, 12, 14, 16, 18, 19, 21, 22, 24, 25, 27, 28]
    data_end = data_start + max(len(summaries), reserved_count) - 1
    for r in range(data_start, data_end + 1):
        for c in cols:
            _fr_set_cell_value(ws, r, c, None)

    _fr_set_cell_value(ws, 2, 27, 'Duration\n(s)')
    _fr_set_cell_value(ws, 2, 28, 'Duration\nStatus')

    row = data_start
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        section_pairs = section_pairs_by_file.get(summary.file_name, [])
        core = _build_core_metrics(summary, ctx, threshold_db=0.5, section_pairs=section_pairs, duration_threshold_s=duration_threshold_s)

        _fr_set_cell_value(ws, row, 1, core['fiber_label'])
        _fr_set_cell_value(ws, row, 5, core['wavelength_nm'])
        _fr_set_cell_value(ws, row, 7, core['max_event_loss_db'])
        _fr_set_cell_value(ws, row, 10, core['max_section_loss_db'])
        _fr_set_cell_value(ws, row, 12, core['avg_splice_loss_db'])
        _fr_set_cell_value(ws, row, 14, core['avg_section_loss_db'])
        _fr_set_cell_value(ws, row, 16, core['length_km'])
        _fr_set_cell_value(ws, row, 18, 'Pass' if core['length_km'] not in (None, 0) else 'Unknown')
        _fr_set_cell_value(ws, row, 19, core['total_loss_db'])
        _fr_set_cell_value(ws, row, 21, 'Pass' if core['total_loss_db'] not in (None, 0) else 'Unknown')
        _fr_set_cell_value(ws, row, 22, core['orl_display'])
        _fr_set_cell_value(ws, row, 24, core['orl_status'])
        _fr_set_cell_value(ws, row, 25, core['event_count'])
        _fr_set_cell_value(ws, row, 27, core['duration_s'])
        _fr_set_cell_value(ws, row, 28, core['duration_status'])
        if core.get('duration_status') == 'Fail':
            ws.cell(row, 27).fill = RED_FILL
            ws.cell(row, 28).fill = RED_FILL
        row += 1


def _fr_fill_events(ws, summaries: list[FileSummary], contexts: dict[str, dict], event_defs: list[dict], deviation_m: float, threshold_db: float) -> None:
    summary_start = _fr_expand_pair_table(
        ws,
        data_start=5,
        template_rows=(43, 44),
        insert_before=45,
        count=len(summaries),
        reserved_pairs=20,
    )
    _fr_ensure_events_capacity(ws, len(event_defs))
    max_events = max(_fr_count_real_pair_blocks(ws, header_merge_rows=(1, 2)), len(event_defs))

    # Header
    for i in range(1, max_events + 1):
        col = 3 + (i - 1) * 2
        ws.cell(1, col).value = f'Event {i}'
        if i <= len(event_defs):
            item = event_defs[i - 1]
            ws.cell(2, col).value = item.get('label')
            ws.cell(3, col).value = round(float(item.get('distance_km') or 0.0), 4)
            ws.cell(3, col + 1).value = 'km'
        else:
            ws.cell(2, col).value = None
            ws.cell(3, col).value = None
            ws.cell(3, col + 1).value = None

    # Clear data and summary area
    _fr_clear_cells(ws, 5, summary_start + 3, 1, ws.max_column)

    rendered_count = len(summaries)
    all_event_cols_loss = [[] for _ in range(max_events)]
    all_event_cols_refl = [[] for _ in range(max_events)]

    for idx, summary in enumerate(summaries):
        data_row = 5 + idx * 2
        status_row = data_row + 1
        ctx = contexts.get(summary.file_name, {})
        rows = ctx.get('events', [])
        cluster_map: dict[int, EventRow] = {}
        for ev in rows:
            di = _fr_assign_event_to_def(ev, event_defs, deviation_m)
            if di is None:
                continue
            cur = cluster_map.get(di)
            if cur is None:
                cluster_map[di] = ev
            else:
                rep = float(event_defs[di - 1]['distance_km'])
                if abs(float(ev.distance_km or 0) - rep) < abs(float(cur.distance_km or 0) - rep):
                    cluster_map[di] = ev

        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        ws.cell(data_row, 1).value = _display_fiber_label(summary, contexts.get(summary.file_name, {}).get('metadata', {}))
        ws.cell(data_row, 2).value = int(m.group(1)) if m else None

        for i in range(1, len(event_defs) + 1):
            col = 3 + (i - 1) * 2
            ev = cluster_map.get(i)
            if not ev:
                continue
            loss = ev.loss_db
            raw_refl = ev.reflectance_db
            refl = _fr_reportable_reflectance_db(raw_refl)
            ws.cell(data_row, col).value = loss
            ws.cell(data_row, col + 1).value = refl
            ws.cell(status_row, col).value = _fr_event_status(loss, threshold_db)
            ws.cell(status_row, col + 1).value = _fr_reflectance_status(raw_refl)
            if isinstance(loss, (int, float)):
                ws.cell(data_row, col).number_format = '0.000'
                if abs(float(loss)) + 1e-12 >= float(threshold_db):
                    ws.cell(data_row, col).fill = RED_FILL
                    ws.cell(status_row, col).fill = RED_FILL
                all_event_cols_loss[i - 1].append(float(loss))
            if isinstance(refl, (int, float)):
                all_event_cols_refl[i - 1].append(float(refl))

    for offs, label in enumerate(['Minimum', 'Maximum', 'Average', 'Occurences']):
        ws.cell(summary_start + offs, 1).value = label

    def _stats(vals, denom):
        if not vals:
            return None, None, None, f'0/{denom}'
        return round(min(vals), 3), round(max(vals), 3), round(sum(vals) / len(vals), 3), f'{len(vals)}/{denom}'

    for i in range(1, max_events + 1):
        col = 3 + (i - 1) * 2
        loss_vals = all_event_cols_loss[i - 1]
        refl_vals = all_event_cols_refl[i - 1]
        loss_min, loss_max, loss_avg, loss_occ = _stats(loss_vals, rendered_count)
        refl_min, refl_max, refl_avg, refl_occ = _stats(refl_vals, rendered_count)
        ws.cell(summary_start, col).value = loss_min
        ws.cell(summary_start, col + 1).value = refl_min
        ws.cell(summary_start + 1, col).value = loss_max
        ws.cell(summary_start + 1, col + 1).value = refl_max
        ws.cell(summary_start + 2, col).value = loss_avg
        ws.cell(summary_start + 2, col + 1).value = refl_avg
        ws.cell(summary_start + 3, col).value = loss_occ
        for rr in [summary_start, summary_start + 1, summary_start + 2]:
            if isinstance(ws.cell(rr, col).value, (int, float)):
                ws.cell(rr, col).number_format = '0.000'
            if isinstance(ws.cell(rr, col + 1).value, (int, float)):
                ws.cell(rr, col + 1).number_format = '0.000'


def _fr_fill_sections(ws, summaries: list[FileSummary], contexts: dict[str, dict], sections: list[dict], threshold_db: float, section_match_tolerance_m: float = 100.0, section_measurement_mode: str = 'fit', section_threshold_db: Optional[float] = None) -> dict[str, list[tuple[Optional[float], Optional[float]]]]:
    """Fill Sections sheet.

    Calculation logic is unchanged.  Layout note: keep the original 2-column
    section block (Loss / Att.) and show the requested Start+End identifier only
    once in the header area, not as a repeated data column for every core.
    """
    real_section_count = len(sections)
    existing_section_blocks = _fr_count_real_pair_blocks(ws, header_merge_rows=(1,))
    _fr_ensure_sections_capacity(ws, real_section_count)
    max_sections = real_section_count
    summary_start = _fr_expand_pair_table(
        ws,
        data_start=4,
        template_rows=(42, 43),
        insert_before=44,
        count=len(summaries),
        reserved_pairs=20,
    )

    # Keep section blocks as 2 columns: Loss / Att.  Older builds temporarily
    # expanded them to 3 columns; this restores the cleaner FastReporter-like
    # layout and avoids a noisy "Tổng" column in every core row.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col >= 3:
            try:
                ws.unmerge_cells(str(rng))
            except Exception:
                pass

    block = 2
    # Make sure real section columns are visible even if an older export hid them.
    for _sec_idx in range(1, max_sections + 1):
        _col = 3 + (_sec_idx - 1) * block
        for _c in (_col, _col + 1):
            try:
                ws.column_dimensions[get_column_letter(_c)].hidden = False
            except Exception:
                pass
    start_col = 3
    for i in range(1, max_sections + 1):
        col = start_col + (i - 1) * block
        ws.cell(1, col).value = f'Section {i}'
        try:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        except Exception:
            pass
        if i <= len(sections):
            sec = sections[i - 1]
            total_km = _section_start_end_total_km(sec)
            ws.cell(2, col).value = sec.get('length_km')
            # Show only one compact identifier in the header, not a full column.
            ws.cell(2, col + 1).value = f'ID S+E {total_km:.3f}' if isinstance(total_km, (int, float)) else 'ID S+E'
            ws.cell(3, col).value = 'Loss\n(dB)'
            ws.cell(3, col + 1).value = 'Att.\n(dB/km)'
        else:
            for c in (col, col + 1):
                ws.cell(2, c).value = None
                ws.cell(3, c).value = None
        ws.column_dimensions[get_column_letter(col)].width = 11
        ws.column_dimensions[get_column_letter(col + 1)].width = 14

    # Hide stale section template blocks beyond the actual calculated section list.
    # This avoids confusing blank columns such as Section 43..46 when real events end earlier.
    _fr_clear_and_hide_unused_section_blocks(
        ws,
        max_sections + 1,
        max(existing_section_blocks, max_sections),
        start_col=start_col,
        block=block,
        max_row=summary_start + max(len(summaries) * 2, 8) + 8,
    )

    _fr_clear_cells(ws, 4, summary_start + 3, 1, start_col + max_sections * block - 1 if max_sections else 2)

    section_pairs_by_file: dict[str, list[tuple[Optional[float], Optional[float]]]] = {}
    rendered_count = len(summaries)
    all_loss_cols = [[] for _ in range(max_sections)]
    all_att_cols = [[] for _ in range(max_sections)]

    for idx, summary in enumerate(summaries):
        data_row = 4 + idx * 2
        status_row = data_row + 1
        ctx = contexts.get(summary.file_name, {})
        ctx['section_fit_rows'] = []
        rows = ctx.get('events', [])
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        ws.cell(data_row, 1).value = _display_fiber_label(summary, contexts.get(summary.file_name, {}).get('metadata', {}))
        ws.cell(data_row, 2).value = int(m.group(1)) if m else None

        pairs: list[tuple[Optional[float], Optional[float]]] = []
        for i in range(1, min(len(sections), max_sections) + 1):
            col = start_col + (i - 1) * block
            sec = sections[i - 1]
            loss_db, att_dbkm, fit_result = _fr_compute_section_values_with_fit(
                summary,
                ctx,
                rows,
                float(sec['start_km']),
                float(sec['end_km']),
                section_index=i,
                section_match_tolerance_m=section_match_tolerance_m,
                section_measurement_mode=section_measurement_mode,
            )
            ctx.setdefault('section_fit_rows', []).append(fit_result)
            pairs.append((loss_db, att_dbkm))
            ws.cell(data_row, col).value = loss_db
            ws.cell(data_row, col + 1).value = att_dbkm
            ws.cell(status_row, col).value = _fr_section_loss_status(loss_db, section_threshold_db if section_threshold_db is not None else threshold_db)
            ws.cell(status_row, col + 1).value = _fr_section_att_status(att_dbkm)
            if _section_loss_reaches_threshold(loss_db, section_threshold_db):
                ws.cell(data_row, col).fill = RED_FILL
                ws.cell(status_row, col).fill = RED_FILL
                ws.cell(status_row, col).value = 'Fail'
            if isinstance(loss_db, (int, float)):
                ws.cell(data_row, col).number_format = '0.000'
                all_loss_cols[i - 1].append(float(loss_db))
            if isinstance(att_dbkm, (int, float)):
                ws.cell(data_row, col + 1).number_format = '0.000'
                all_att_cols[i - 1].append(float(att_dbkm))

        section_pairs_by_file[summary.file_name] = pairs
        ctx['section_pairs'] = pairs

    for offs, label in enumerate(['Minimum', 'Maximum', 'Average', 'Occurences']):
        ws.cell(summary_start + offs, 1).value = label

    def _stats(vals, denom):
        if not vals:
            return None, None, None, f'0/{denom}'
        return round(min(vals), 3), round(max(vals), 3), round(sum(vals) / len(vals), 3), f'{len(vals)}/{denom}'

    for i in range(1, max_sections + 1):
        col = start_col + (i - 1) * block
        loss_min, loss_max, loss_avg, loss_occ = _stats(all_loss_cols[i - 1], rendered_count)
        att_min, att_max, att_avg, _att_occ = _stats(all_att_cols[i - 1], rendered_count)
        ws.cell(summary_start, col).value = loss_min
        ws.cell(summary_start, col + 1).value = att_min
        ws.cell(summary_start + 1, col).value = loss_max
        ws.cell(summary_start + 1, col + 1).value = att_max
        ws.cell(summary_start + 2, col).value = loss_avg
        ws.cell(summary_start + 2, col + 1).value = att_avg
        ws.cell(summary_start + 3, col).value = loss_occ
        for rr in [summary_start, summary_start + 1, summary_start + 2]:
            for c in (col, col + 1):
                if isinstance(ws.cell(rr, c).value, (int, float)):
                    ws.cell(rr, c).number_format = '0.000'

    return section_pairs_by_file



def _fr_fill_general_information_template(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    # Clear editable fields only; keep template layout/styles/labels intact
    scalar_cells = ['C3', 'C4', 'C5', 'K3', 'K4', 'K5', 'C6', 'A9', 'D9', 'G9', 'K9', 'N9', 'C13', 'K13', 'C14', 'K14', 'C15', 'K15', 'C16', 'K16']
    for addr in scalar_cells:
        ws[addr] = None

    metas = [contexts.get(s.file_name, {}).get('metadata', {}) for s in summaries]
    filenames = '; '.join(s.file_name for s in summaries)
    ws['C3'] = filenames

    company = _common_nonempty([m.get('company', '') for m in metas])
    job_id = _common_nonempty([m.get('job_id', '') for m in metas])
    comments = _common_nonempty([m.get('comments', '') for m in metas])
    cable_id = _common_nonempty([m.get('cable_id', '') for m in metas]) or 'Cable'
    fiber_id = _common_nonempty([m.get('fiber_id', '') for m in metas if (m.get('fiber_id') or '').strip().lower() not in {'fiber', ''}]) or '; '.join(_display_fiber_label(s, contexts.get(s.file_name, {}).get('metadata', {})) for s in summaries[:3])
    location_a = _common_nonempty([m.get('location_a', '') for m in metas])
    location_b = _common_nonempty([m.get('location_b', '') for m in metas])
    operator_a = _common_nonempty([m.get('operator_a', '') for m in metas])
    operator_b = _common_nonempty([m.get('operator_b', '') for m in metas])
    unit_model = _common_nonempty([m.get('unit_model', '') for m in metas])
    unit_serial = _common_nonempty([m.get('unit_serial', '') for m in metas])

    ws['K3'] = job_id or None
    ws['K5'] = company or None
    ws['C6'] = comments or None
    ws['A9'] = cable_id
    ws['D9'] = fiber_id
    ws['G9'] = location_a or None
    ws['K9'] = location_b or None
    ws['C13'] = location_a or None
    ws['K13'] = location_b or None
    ws['C14'] = operator_a or None
    ws['K14'] = operator_b or None
    ws['C15'] = unit_model or None
    ws['K15'] = unit_model or None
    ws['C16'] = unit_serial or None
    ws['K16'] = unit_serial or None

    start_row = 20
    last_data_row = 39
    insert_before = 40
    reserved_count = last_data_row - start_row + 1
    _fr_expand_single_row_table(ws, data_start=start_row, template_row=last_data_row, insert_before=insert_before, count=len(summaries), reserved_count=reserved_count)

    parameter_end = start_row + max(len(summaries), reserved_count) - 1
    for r in range(start_row, parameter_end + 1):
        for c in [2, 5, 11, 14]:
            _fr_set_cell_value(ws, r, c, None)

    row = start_row
    for summary in summaries:
        meta = contexts.get(summary.file_name, {}).get('metadata', {})
        wavelength = meta.get('wavelength_nm')
        range_km = meta.get('range_km')
        pulse_us = meta.get('pulse_us')
        duration_s = meta.get('duration_s')

        for col, value in [(2, wavelength), (5, range_km), (11, pulse_us), (14, duration_s)]:
            _fr_set_cell_value(ws, row, col, value)

        if isinstance(range_km, (int, float)):
            ws.cell(*_fr_merged_anchor(ws, row, 5)).number_format = '0.0000'
        if isinstance(pulse_us, (int, float)):
            ws.cell(*_fr_merged_anchor(ws, row, 11)).number_format = '0.0##'
        if isinstance(duration_s, (int, float)):
            ws.cell(*_fr_merged_anchor(ws, row, 14)).number_format = '0.###'
        row += 1

    macro_start = 43 + max(len(summaries) - reserved_count, 0)
    macro_reserved = 1
    _fr_expand_single_row_table(ws, data_start=macro_start, template_row=macro_start, insert_before=macro_start + 1, count=len(summaries), reserved_count=macro_reserved)
    macro_end = macro_start + max(len(summaries), macro_reserved) - 1

    for r in range(macro_start, macro_end + 1):
        for c in [1, 4, 6, 9, 11, 13, 15]:
            _fr_set_cell_value(ws, r, c, None)

    for idx, summary in enumerate(summaries):
        r = macro_start + idx
        meta = contexts.get(summary.file_name, {}).get('metadata', {})
        _fr_set_cell_value(ws, r, 1, _display_fiber_label(summary, meta))
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        if m:
            _fr_set_cell_value(ws, r, 6, int(m.group(1)))
        ws.row_dimensions[r].height = 19




def _fr_safe_sheet_remove(wb: Workbook, title: str) -> None:
    if title in wb.sheetnames:
        ws = wb[title]
        wb.remove(ws)


def _fr_write_header_row(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header)
    ws.freeze_panes = 'A2'


def _fr_fill_app_parameters_sheet(ws, *, threshold_db: float, deviation_m: float, expected_route_km: Optional[float], jumper_excluded_m: float, length_tolerance_km: float, graph_reach_tolerance_km: Optional[float], event_shortfall_tolerance_km: Optional[float], overlength_tolerance_km: Optional[float], segment_start_km: Optional[float], segment_end_km: Optional[float], section_export_scope: str, section_merge_tolerance_m: Optional[float], section_min_length_km: float, section_event_source: str, section_boundary_priority: str, section_allow_split: bool, section_match_tolerance_m: float, section_measurement_mode: str, orl_pass_threshold_db: float, orl_source_mode: str, orl_allow_lower_bound: bool, orl_lower_bound_status: str, orl_physical_mode: str = 'disabled', section_threshold_db: Optional[float] = None, duration_threshold_s: Optional[float] = None) -> None:
    _fr_write_header_row(ws, ['Tham số', 'Giá trị đang dùng', 'Giải thích'])
    rows = [
        ('threshold_db', threshold_db, 'Ngưỡng event dùng cho tô đỏ / lọc event mạnh.'),
        ('section_threshold_db', section_threshold_db, 'Ngưỡng Section Loss dùng riêng để tô đỏ ô section; không thay đổi logic tính section.'),
        ('duration_threshold_s', duration_threshold_s, 'Nếu duration của core nhỏ hơn ngưỡng này thì Core Metrics/Link Results đánh dấu không đạt.'),
        ('deviation_m', deviation_m, 'Sai số gom event chung giữa nhiều fiber.'),
        ('expected_route_km', expected_route_km, 'Chiều dài tuyến chuẩn do người dùng nhập.'),
        ('jumper_excluded_m', jumper_excluded_m, 'Chiều dài jumper loại trừ trước khi so đồ thị với tuyến chuẩn.'),
        ('length_tolerance_km', length_tolerance_km, 'Dung sai legacy giữ tương thích logic cũ.'),
        ('graph_reach_tolerance_km', graph_reach_tolerance_km, 'Dung sai đồ thị chạm tuyến chuẩn.'),
        ('event_shortfall_tolerance_km', event_shortfall_tolerance_km, 'Dung sai event table ngắn hơn tuyến chuẩn nhưng vẫn chấp nhận.'),
        ('overlength_tolerance_km', overlength_tolerance_km, 'Biến tương thích; logic hiện tại không dùng để chặn core dài hơn tuyến chuẩn.'),
        ('segment_start_km', segment_start_km, 'Mốc đầu đoạn phân tích riêng.'),
        ('segment_end_km', segment_end_km, 'Mốc cuối đoạn phân tích riêng.'),
        ('section_merge_tolerance_m', section_merge_tolerance_m, 'Sai số gom boundary section gần nhau.'),
        ('section_min_length_km', section_min_length_km, 'Chiều dài section tối thiểu.'),
        ('section_event_source', section_event_source, 'Nguồn event dựng section: all / filtered.'),
        ('section_boundary_priority', section_boundary_priority, 'Ưu tiên boundary theo event hoặc trace.'),
        ('section_allow_split', section_allow_split, 'Cho phép giữ section ngắn hay gộp lại.'),
        ('section_match_tolerance_m', section_match_tolerance_m, 'Sai số map section chung sang từng fiber.'),
        ('section_measurement_mode', section_measurement_mode, 'Cách tính loss section: fit = raw-trace fit + fallback event/slope; event = logic event/slope cũ; 2point = nội suy đầu-cuối.'),
        ('orl_pass_threshold_db', orl_pass_threshold_db, 'Ngưỡng Pass/Fail cho exact ORL.'),
        ('orl_source_mode', orl_source_mode, 'Nguồn lấy ORL: exact / metadata / auto.'),
        ('orl_allow_lower_bound', orl_allow_lower_bound, 'Có cho phép hiện ORL kiểu <xx.xx hay không.'),
        ('orl_lower_bound_status', orl_lower_bound_status, 'Status legacy gán cho ORL fallback kiểu <xx.xx; Phase 6 vẫn đánh Use For Judgment = No với lower-bound.'),
        ('orl_physical_mode', orl_physical_mode, 'ORL vật lý từ trace: disabled/diagnostic/experimental. Mặc định tắt; nếu bật mà thiếu hiệu chuẩn tuyệt đối thì app chỉ ghi lý do, không tự tạo số ORL.'),
    ]
    for r_idx, (name, value, desc) in enumerate(rows, start=2):
        ws.cell(r_idx, 1, name)
        ws.cell(r_idx, 2, value if value is not None else '')
        ws.cell(r_idx, 3, desc)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 80


def _fr_fill_route_analysis_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    headers = [
        'Fiber ID', 'Wavelength (nm)', 'Expected Route (km)', 'Event Length (km)', 'Graph End (km)',
        'Jumper Excluded (km)', 'Net Graph Length (km)', 'Diff vs Expected (km)',
        'Graph Reach Tol (km)', 'Event Shortfall Tol (km)', 'Graph Reaches Expected',
        'Verdict', 'Reason'
    ]
    _fr_write_header_row(ws, headers)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        ga = ctx.get('graph_assessment')
        meta = ctx.get('metadata', {})
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        wavelength = int(m.group(1)) if m else None
        ws.cell(row, 1, _display_fiber_label(summary, meta))
        ws.cell(row, 2, wavelength)
        if ga is not None:
            ws.cell(row, 3, ga.expected_route_km)
            ws.cell(row, 4, ga.event_length_km)
            ws.cell(row, 5, ga.graph_end_km)
            ws.cell(row, 6, ga.jumper_excluded_km)
            ws.cell(row, 7, ga.net_graph_length_km)
            ws.cell(row, 8, ga.diff_km)
            ws.cell(row, 9, ga.graph_reach_tolerance_km)
            ws.cell(row, 10, ga.event_shortfall_tolerance_km)
            ws.cell(row, 11, 'Yes' if ga.graph_reaches_expected else ('No' if ga.graph_reaches_expected is False else ''))
            ws.cell(row, 12, ga.verdict)
            ws.cell(row, 13, ga.reason)
        row += 1
    for col in 'ABCDEFGHIJ':
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 80


def _fr_fill_segment_analysis_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    headers = [
        'Fiber ID', 'Wavelength (nm)', 'Start (km)', 'End (km)', 'Span (km)', 'Event Count',
        'Segment Loss (dB)', 'Segment Att (dB/km)', 'Max Positive Event Loss (dB)',
        'Max Negative Event Loss (dB)', 'Method', 'Note', 'Recommendation'
    ]
    _fr_write_header_row(ws, headers)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        sa = ctx.get('segment_assessment')
        meta = ctx.get('metadata', {})
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        wavelength = int(m.group(1)) if m else None
        ws.cell(row, 1, _display_fiber_label(summary, meta))
        ws.cell(row, 2, wavelength)
        if sa is not None:
            ws.cell(row, 3, sa.start_km)
            ws.cell(row, 4, sa.end_km)
            ws.cell(row, 5, sa.span_km)
            ws.cell(row, 6, sa.event_count)
            ws.cell(row, 7, sa.segment_total_loss_db)
            ws.cell(row, 8, sa.segment_attenuation_dbkm)
            ws.cell(row, 9, sa.max_positive_event_loss_db)
            ws.cell(row, 10, sa.max_negative_event_loss_db)
            ws.cell(row, 11, sa.method)
            ws.cell(row, 12, sa.note)
            ws.cell(row, 13, sa.recommendation)
        row += 1
    for col in 'ABCDEFGHIJ':
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['K'].width = 45
    ws.column_dimensions['L'].width = 55
    ws.column_dimensions['M'].width = 70


def _fr_fill_segment_events_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    headers = ['Fiber ID', 'Wavelength (nm)', 'Event No', 'Type', 'Distance (km)', 'Loss (dB)', 'Reflectance (dB)', 'Slope (dB/km)', 'Status']
    _fr_write_header_row(ws, headers)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        seg_rows = ctx.get('segment_event_rows') or []
        meta = ctx.get('metadata', {})
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        wavelength = int(m.group(1)) if m else None
        if not seg_rows:
            ws.cell(row, 1, _display_fiber_label(summary, meta))
            ws.cell(row, 2, wavelength)
            ws.cell(row, 9, 'Không có event dương trong đoạn đã chọn')
            row += 1
            continue
        for item in seg_rows:
            ws.cell(row, 1, _display_fiber_label(summary, meta))
            ws.cell(row, 2, wavelength)
            ws.cell(row, 3, item.get('event_no', ''))
            ws.cell(row, 4, item.get('type', ''))
            ws.cell(row, 5, item.get('distance_km', ''))
            ws.cell(row, 6, item.get('loss_db', ''))
            ws.cell(row, 7, item.get('reflectance_db', ''))
            ws.cell(row, 8, item.get('slope_dbkm', ''))
            ws.cell(row, 9, item.get('status', ''))
            row += 1
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 28

def _fr_style_sheet_headers(ws, header_rows: tuple[int, ...] = (1,)) -> None:
    bold = Font(bold=True)
    for row_idx in header_rows:
        for cell in ws[row_idx]:
            if cell.value not in (None, ''):
                cell.fill = HEADER_FILL if row_idx == 1 else SUBHEADER_FILL
                cell.font = bold


def _fr_selective_autofit(ws, max_scan_rows: int = 120, max_width: int = 60) -> None:
    upper = min(ws.max_row, max_scan_rows)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        current = ws.column_dimensions[letter].width or 0
        if current >= max_width:
            continue
        best = int(current or 0)
        for row_idx in range(1, upper + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            text_val = str(value).replace('\n', ' ')
            best = max(best, min(len(text_val) + 2, max_width))
        if best > current:
            ws.column_dimensions[letter].width = best




def _stv_display_file_name(file_name: str) -> str:
    return Path(file_name).name


def _stv_display_loss_for_row(row: EventRow, is_terminal_row: bool = False) -> Optional[float]:
    if row.loss_db is None:
        return None
    try:
        loss = float(row.loss_db)
    except Exception:
        return None
    if loss == 0:
        return None
    label = (row.label or '').strip()
    if row.event_type == 'First Connector':
        return None
    if is_terminal_row:
        return None
    if len(label) >= 2 and label[1] == 'E':
        return None
    return round(loss, 3)


def _stv_display_loss_for_row(row: EventRow, threshold_db: float = 0.5, is_terminal_row: bool = False) -> Optional[float]:
    """Giá trị hiển thị trên bảng STV: chỉ giữ suy hao điểm dương vượt ngưỡng."""
    if row.loss_db is None:
        return None
    try:
        loss = float(row.loss_db)
    except Exception:
        return None
    label = (row.label or '').strip()
    if row.event_type == 'First Connector':
        return None
    if is_terminal_row:
        return None
    if len(label) >= 2 and label[1] == 'E':
        return None
    if loss <= 0:
        return None
    if loss + 1e-12 < float(threshold_db):
        return None
    return round(loss, 3)




def _stv_assessment_fill(status: str):
    if status == 'Đạt':
        return STV_DAT_FILL
    if status == 'Đứt':
        return STV_DUT_FILL
    if status == 'Suy hao':
        return STV_SUY_HAO_FILL
    return None


def _stv_auto_assessment_status(ga, attenuation_dbkm, point_losses: list[float]) -> str:
    """Tự động đánh giá STV theo đúng thứ tự người dùng yêu cầu.

    Thứ tự:
    1) Giữ nguyên logic đánh giá chiều dài tuyến hiện tại: nếu có chiều dài tuyến chuẩn
       và logic hiện tại không kết luận 'Đủ tuyến' thì đánh giá 'Đứt'.
    2) Nếu chiều dài đủ điều kiện, xét suy hao trung bình: <= 0.27 dB/km mới đạt.
    3) Nếu suy hao trung bình đạt, xét suy hao điểm: nếu có trên 5 điểm suy hao và
       tồn tại điểm suy hao lớn >= 1 dB thì đánh giá 'Suy hao'.
    Chỉ trả về 3 trạng thái: Đạt / Đứt / Suy hao.
    """
    try:
        expected_route = float(getattr(ga, 'expected_route_km', None)) if ga is not None and getattr(ga, 'expected_route_km', None) not in (None, '') else None
    except Exception:
        expected_route = None
    if expected_route is not None and expected_route > 0:
        verdict = str(getattr(ga, 'verdict', '') or '').strip()
        if verdict != 'Đủ tuyến':
            return 'Đứt'

    try:
        att = float(attenuation_dbkm)
    except Exception:
        att = None
    if att is None or att > 0.27 + 1e-12:
        return 'Suy hao'

    clean_losses: list[float] = []
    for value in point_losses or []:
        try:
            v = float(value)
        except Exception:
            continue
        if v > 0:
            clean_losses.append(v)
    if len(clean_losses) > 5 and any(v >= 1.0 - 1e-12 for v in clean_losses):
        return 'Suy hao'
    return 'Đạt'





def _stv_fill_main_sheet(
    ws,
    summaries: list[FileSummary],
    contexts: dict[str, dict],
    event_defs: list[dict],
    deviation_m: float,
    threshold_db: float,
    expected_route_km: Optional[float],
    jumper_excluded_m: float,
    graph_reach_tolerance_km: Optional[float],
    event_shortfall_tolerance_km: Optional[float],
    stv_total_core: Optional[int] = None,
    stv_used_core: Optional[int] = None,
) -> None:
    distance_list = [float(item.get('distance_km') or 0.0) for item in (event_defs or [])]
    event_start_col = 10

    ws.title = 'Bảng sự kiện'
    ws.cell(2, 1, 'Định dạng:')
    ws.cell(2, 2, '.MSOR / .SOR / .TRC')

    ws.cell(3, 1, 'Ngày:')
    ws.cell(3, 2, datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    ws.cell(3, 4, 'Chiều dài tuyến chuẩn, km:')
    ws.cell(3, 5, expected_route_km)
    ws.cell(3, 6, 'Sai số chạm tuyến của đồ thị, km:')
    ws.cell(3, 7, graph_reach_tolerance_km)

    ws.cell(4, 1, 'Sai số gom cụm, m:')
    ws.cell(4, 2, deviation_m)
    ws.cell(4, 4, 'Độ dài đấu nhảy loại trừ, m:')
    ws.cell(4, 5, jumper_excluded_m)
    ws.cell(4, 6, 'Mức hụt event cho phép, km:')
    ws.cell(4, 7, event_shortfall_tolerance_km)
    ws.cell(4, event_start_col, 'Khoảng cách, km')
    if distance_list:
        ws.merge_cells(start_row=4, start_column=event_start_col, end_row=4, end_column=event_start_col + len(distance_list) - 1)

    headers = [
        'Trạng thái đồ thị', 'Tệp', 'Định dạng', 'Sợi', 'Bước sóng',
        'Suy hao tổng, dB', 'Chiều dài, km', 'Suy hao TB, dB/km', 'Đánh giá'
    ]
    for idx, label in enumerate(headers, start=1):
        ws.cell(5, idx, label)
    for idx, distance in enumerate(distance_list, start=event_start_col):
        cell = ws.cell(5, idx, distance)
        cell.number_format = '0.000############'

    fixed_widths = {1: 18, 2: 16, 3: 10, 4: 14, 5: 12, 6: 14, 7: 11, 8: 16, 9: 12}
    for c, w in fixed_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for col in range(event_start_col, event_start_col + len(distance_list)):
        ws.column_dimensions[get_column_letter(col)].width = 13

    for row_idx, summary in enumerate(summaries, start=6):
        ctx = contexts.get(summary.file_name) or {}
        core = _build_core_metrics(summary, ctx, threshold_db=threshold_db)
        ga = ctx.get('graph_assessment')
        total_loss = summary.route_corrected_total_loss_db if summary.route_corrected_total_loss_db is not None else core['total_loss_db']
        length_km = core['length_km']
        if total_loss is not None and length_km not in (None, 0):
            attenuation = round(total_loss / length_km, 3)
        else:
            attenuation = core['attenuation_dbkm']

        route_verdict = ''
        if ga is not None and ga.expected_route_km not in (None, 0):
            route_verdict = ga.verdict or ''

        ws.cell(row_idx, 1, route_verdict)
        ws.cell(row_idx, 2, core['display_file_name'])
        ws.cell(row_idx, 3, core['source_format'])
        ws.cell(row_idx, 4, core['fiber_label'])
        ws.cell(row_idx, 5, core['wavelength_display'])
        ws.cell(row_idx, 6, total_loss)
        ws.cell(row_idx, 7, length_km)
        ws.cell(row_idx, 8, attenuation)
        ws.cell(row_idx, 6).number_format = '0.000'
        ws.cell(row_idx, 7).number_format = '0.000############'
        ws.cell(row_idx, 8).number_format = '0.000'

        wavelength_nm = str(core['wavelength_nm']) if core['wavelength_nm'] else None
        rows = _fr_pick_rows_for_file(ctx.get('events', []), wavelength_nm)
        ordered = sorted([r for r in rows if r.distance_km is not None], key=lambda r: (float(r.distance_km or 0.0), str(r.event_no)))

        if distance_list:
            ws.cell(row_idx, event_start_col, 'Đầu tuyến')

        bucketed: dict[int, float] = {}
        for idx_row, row in enumerate(ordered):
            is_terminal = idx_row == len(ordered) - 1 and row.event_type == 'Fiber End'
            value = _stv_display_loss_for_row(row, threshold_db=threshold_db, is_terminal_row=is_terminal)
            if value is None:
                continue
            def_idx = _fr_assign_event_to_def(row, event_defs, deviation_m)
            if def_idx is None:
                continue
            prev = bucketed.get(def_idx)
            if prev is None or value > prev:
                bucketed[def_idx] = value

        assessment = _stv_auto_assessment_status(ga, attenuation, list(bucketed.values()))
        assessment_cell = ws.cell(row_idx, 9, assessment)
        assessment_fill = _stv_assessment_fill(assessment)
        if assessment_fill is not None:
            assessment_cell.fill = assessment_fill

        for def_idx, value in bucketed.items():
            cell = ws.cell(row_idx, event_start_col + (def_idx - 1), value)
            cell.number_format = '0.000'
            cell.fill = RED_FILL

        if summary.end_distance_km not in (None, 0) and distance_list:
            end_idx = None
            best_gap = None
            for item in event_defs:
                gap = abs(float(summary.end_distance_km) - float(item.get('distance_km') or 0.0))
                if best_gap is None or gap < best_gap:
                    best_gap = gap
                    end_idx = int(item.get('index') or 0)
            if end_idx:
                end_col = event_start_col + (end_idx - 1)
                if ws.cell(row_idx, end_col).value in (None, ''):
                    ws.cell(row_idx, end_col, 'Cuối tuyến')

    # First and second KPI tables restoration
    from openpyxl.styles import Font, PatternFill, Alignment

    font_big = Font(name='Times New Roman', size=11, bold=True)
    font_val = Font(name='Times New Roman', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    c_total = 2
    c_used = 3
    c_dat = 4
    c_dut = 5
    c_suyhao = 6
    cols = [c_total, c_used, c_dat, c_dut, c_suyhao]

    start_row = row_idx + 2
    total_row = start_row + 1
    second_header_row = total_row + 2
    second_value_row = second_header_row + 1

    total_core = stv_total_core if stv_total_core is not None else len(summaries)
    assessment_range = f"I6:I{row_idx}"

    labels = ['Tổng core', 'Core\nsử dụng', 'Core đạt', 'Core\nđứt', 'Core suy\nhao cao']
    fills = [
        PatternFill(fill_type='solid', fgColor='FFFFFF'),
        STV_DAT_FILL,
        STV_SUY_HAO_FILL,
        STV_DUT_FILL,
        PatternFill(fill_type='solid', fgColor='FFC000')
    ]

    for col, label, fill in zip(cols, labels, fills):
        cell = ws.cell(start_row, col, label)
        cell.fill = fill
        cell.font = font_big
        cell.alignment = center
        ws.cell(total_row, col).fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
        ws.cell(total_row, col).font = font_val
        ws.cell(total_row, col).alignment = center

    ws.cell(total_row, c_total, total_core)
    ws.cell(total_row, c_used, f'={get_column_letter(c_total)}{total_row}-SUM({get_column_letter(c_dat)}{total_row}:{get_column_letter(c_suyhao)}{total_row})')
    ws.cell(total_row, c_dat, f'=COUNTIF({assessment_range},"Đạt")')
    ws.cell(total_row, c_dut, f'=COUNTIF({assessment_range},"Đứt")')
    ws.cell(total_row, c_suyhao, f'=COUNTIF({assessment_range},"Suy hao")')

    for col in cols:
        ws.cell(total_row, col).number_format = '0'

    _stv_set_block_border(ws, start_row, c_total, total_row, c_suyhao)
    ws.row_dimensions[start_row].height = 54
    ws.row_dimensions[total_row].height = 26

    # Second KPI table: Tỉ lệ khai thác / DKD Core / DKD yêu cầu / Đánh giá.
    second_labels = ['Tỉ lệ khai thác', 'DKD Core', 'DKD yêu cầu', 'Đánh giá', '']
    second_fill = HEADER_FILL
    for i, col in enumerate(cols):
        cell = ws.cell(second_header_row, col, second_labels[i])
        cell.fill = second_fill
        cell.font = font_big
        cell.alignment = center
        value_cell = ws.cell(second_value_row, col)
        value_cell.font = font_val
        value_cell.alignment = center
        value_cell.fill = PatternFill(fill_type='solid', fgColor='FFFFFF')

    ws.cell(second_value_row, c_total, f'=IF({get_column_letter(c_total)}{total_row}=0,0,{get_column_letter(c_used)}{total_row}/{get_column_letter(c_total)}{total_row}*100)')
    ws.cell(second_value_row, c_used, f'=IF({get_column_letter(c_total)}{total_row}=0,0,SUM({get_column_letter(c_used)}{total_row}:{get_column_letter(c_dat)}{total_row})/{get_column_letter(c_total)}{total_row}*100)')
    ws.cell(second_value_row, c_dat, f'=IF(OR({get_column_letter(c_total)}{second_value_row}<30,{get_column_letter(c_total)}{second_value_row}=30),70,IF(OR({get_column_letter(c_total)}{second_value_row}<50,{get_column_letter(c_total)}{second_value_row}=50),80,85))')
    ws.cell(second_value_row, c_dut, f'=IF(OR({get_column_letter(c_used)}{second_value_row}>{get_column_letter(c_dat)}{second_value_row},{get_column_letter(c_used)}{second_value_row}={get_column_letter(c_dat)}{second_value_row}),"Đạt","Không đạt")')
    ws.cell(second_value_row, c_suyhao, '')

    for col in (c_total, c_used, c_dat):
        ws.cell(second_value_row, col).number_format = '0'

    # Pre-color final status from the generated values; formula remains editable.
    cnt_dat = 0
    cnt_dut = 0
    cnt_suyhao = 0
    for summary in summaries:
        ctx = contexts.get(summary.file_name) or {}
        ga = ctx.get('graph_assessment')
        core = _build_core_metrics(summary, ctx, threshold_db=threshold_db)
        
        # Re-run same bucket logic as main loop to get values for assessment
        wavelength_nm = str(core['wavelength_nm']) if core['wavelength_nm'] else None
        rows = _fr_pick_rows_for_file(ctx.get('events', []), wavelength_nm)
        ordered = sorted([r for r in rows if r.distance_km is not None], key=lambda r: (float(r.distance_km or 0.0), str(r.event_no)))
        bucketed: dict[int, float] = {}
        for idx_row, row in enumerate(ordered):
            is_terminal = idx_row == len(ordered) - 1 and row.event_type == 'Fiber End'
            value = _stv_display_loss_for_row(row, threshold_db=threshold_db, is_terminal_row=is_terminal)
            if value is None:
                continue
            def_idx = _fr_assign_event_to_def(row, event_defs, deviation_m)
            if def_idx is None:
                continue
            prev = bucketed.get(def_idx)
            if prev is None or value > prev:
                bucketed[def_idx] = value

        total_loss = summary.route_corrected_total_loss_db if summary.route_corrected_total_loss_db is not None else core['total_loss_db']
        length_km = core['length_km']
        if total_loss is not None and length_km not in (None, 0):
            attenuation = round(total_loss / length_km, 3)
        else:
            attenuation = core['attenuation_dbkm']

        assessment = _stv_auto_assessment_status(ga, attenuation, list(bucketed.values()))
        if assessment == 'Đạt':
            cnt_dat += 1
        elif assessment == 'Đứt':
            cnt_dut += 1
        elif assessment == 'Suy hao':
            cnt_suyhao += 1

    total_core_val = stv_total_core if stv_total_core is not None else len(summaries)
    used_core_val = stv_used_core if stv_used_core is not None else (total_core_val - (cnt_dat + cnt_dut + cnt_suyhao))
    ti_le_khai_thac = (used_core_val / total_core_val * 100) if total_core_val > 0 else 0
    dkd_core = ((used_core_val + cnt_dat) / total_core_val * 100) if total_core_val > 0 else 0
    dkd_yeu_cau = 70 if ti_le_khai_thac <= 30 else (80 if ti_le_khai_thac <= 50 else 85)
    overall_text = 'Đạt' if dkd_core >= dkd_yeu_cau else 'Không đạt'

    result_fill = STV_DAT_FILL if overall_text == 'Đạt' else PatternFill(fill_type='solid', fgColor='ED7D31')
    ws.cell(second_value_row, c_dut).fill = result_fill
    ws.cell(second_value_row, c_suyhao).fill = result_fill
    ws.cell(second_value_row, c_dut).font = Font(bold=True, size=16, name='Times New Roman')
    ws.cell(second_value_row, c_suyhao).font = Font(bold=True, size=16, name='Times New Roman')

    _stv_set_block_border(ws, second_header_row, c_total, second_value_row, c_suyhao)
    ws.row_dimensions[second_header_row].height = 48
    ws.row_dimensions[second_value_row].height = 26

    for col in cols:
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 15)

def _stv_set_block_border(ws, r_start, c_start, r_end, c_end):
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in range(r_start, r_end + 1):
        for c in range(c_start, c_end + 1):
            ws.cell(r, c).border = border

def _stv_fill_graph_check_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    ws.title = 'Kiểm tra đồ thị'
    headers = [
        'Tệp', 'Định dạng', 'Sợi', 'Bước sóng', 'Điểm cuối đồ thị, km', 'Đấu nhảy loại trừ, km',
        'Chiều dài đồ thị thực, km', 'Chiều dài event, km', 'Chiều dài tuyến chuẩn, km',
        'Chênh lệch, km', 'Sai số chạm tuyến, km', 'Mức hụt event cho phép, km',
        'Đồ thị chạm tuyến', 'Trạng thái', 'Lý do', 'Suy hao đã đọc, dB', 'Suy hao hiệu chỉnh tuyến, dB',
        'Nguồn suy hao đang dùng'
    ]
    for idx, label in enumerate(headers, start=1):
        ws.cell(1, idx, label)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name) or {}
        ga = ctx.get('graph_assessment')
        ws.cell(row, 1, _stv_display_file_name(summary.file_name))
        ws.cell(row, 2, summary.source_format)
        ws.cell(row, 3, summary.fiber)
        ws.cell(row, 4, summary.wavelength_display)
        if ga is not None:
            ws.cell(row, 5, ga.graph_end_km)
            ws.cell(row, 6, ga.jumper_excluded_km)
            ws.cell(row, 7, ga.net_graph_length_km)
            ws.cell(row, 8, ga.event_length_km)
            ws.cell(row, 9, ga.expected_route_km)
            ws.cell(row, 10, ga.diff_km)
            ws.cell(row, 11, ga.graph_reach_tolerance_km)
            ws.cell(row, 12, ga.event_shortfall_tolerance_km)
            if ga.graph_reaches_expected is True:
                ws.cell(row, 13, 'Có')
            elif ga.graph_reaches_expected is False:
                ws.cell(row, 13, 'Không')
            ws.cell(row, 14, ga.verdict)
            ws.cell(row, 15, ga.reason)
        ws.cell(row, 16, summary.parsed_total_loss_db)
        ws.cell(row, 17, summary.route_corrected_total_loss_db)
        ws.cell(row, 18, summary.loss_source_used)
        row += 1
    for c in range(1, 19):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions['O'].width = 60
    ws.column_dimensions['R'].width = 28


def _stv_fill_skipped_sheet(ws, skipped: list[str]) -> None:
    ws.title = 'Tệp bỏ qua'
    ws.cell(1, 1, 'Tệp')
    ws.cell(1, 2, 'Lý do')
    row = 2
    for item in skipped:
        if ': ' in item:
            file_name, reason = item.split(': ', 1)
        else:
            file_name, reason = item, ''
        ws.cell(row, 1, _stv_display_file_name(file_name))
        ws.cell(row, 2, reason)
        row += 1
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 100


def _stv_fill_raw_events_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    ws.title = 'Sự kiện thô'
    headers = ['Tệp', 'Định dạng', 'Chế độ đọc', 'Độ tin cậy', 'STT sự kiện', 'Loại sự kiện', 'Khoảng cách, km', 'Suy hao, dB', 'Phản xạ, dB', 'Độ dốc, dB/km', 'Suy hao tổng, dB', 'Ghi chú', 'Nhãn']
    for idx, label in enumerate(headers, start=1):
        ws.cell(1, idx, label)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name) or {}
        parse_mode = str(ctx.get('parse_mode') or 'unknown')
        if parse_mode == 'standard_sor_keyevents':
            confidence = 'High'
        elif parse_mode in {'trc_appregex_sections_points', 'trc_standard'}:
            confidence = 'Medium'
        else:
            confidence = 'Low'
        rows = ctx.get('events') or []
        for e in rows:
            ws.cell(row, 1, _stv_display_file_name(summary.file_name))
            ws.cell(row, 2, summary.source_format)
            ws.cell(row, 3, _to_vi_parse_mode(parse_mode))
            ws.cell(row, 4, _to_vi_confidence(confidence))
            ws.cell(row, 5, e.event_no)
            ws.cell(row, 6, _to_vi_event_type(e.event_type))
            ws.cell(row, 7, _safe_round3(e.distance_km))
            ws.cell(row, 8, e.loss_db)
            ws.cell(row, 9, e.reflectance_db)
            ws.cell(row, 10, e.slope_dbkm)
            ws.cell(row, 11, e.total_loss_db)
            ws.cell(row, 12, e.note_original)
            ws.cell(row, 13, e.label)
            row += 1
    for c in range(1, 14):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions['L'].width = 40


def _stv_build_workbook(
    summaries: list[FileSummary],
    contexts: dict[str, dict],
    event_defs: list[dict],
    deviation_m: float,
    threshold_db: float,
    expected_route_km: Optional[float],
    jumper_excluded_m: float,
    graph_reach_tolerance_km: Optional[float],
    event_shortfall_tolerance_km: Optional[float],
    skipped: list[str],
    logs: Optional[list[dict]] = None,
) -> BytesIO:
    wb = Workbook()
    ws_main = wb.active
    _stv_fill_main_sheet(
        ws_main,
        summaries,
        contexts,
        event_defs,
        deviation_m=deviation_m,
        threshold_db=threshold_db,
        expected_route_km=expected_route_km,
        jumper_excluded_m=jumper_excluded_m,
        graph_reach_tolerance_km=graph_reach_tolerance_km,
        event_shortfall_tolerance_km=event_shortfall_tolerance_km,
    )
    ws_graph = wb.create_sheet('Kiểm tra đồ thị')
    _stv_fill_graph_check_sheet(ws_graph, summaries, contexts)
    ws_skipped = wb.create_sheet('Tệp bỏ qua')
    _stv_fill_skipped_sheet(ws_skipped, skipped)
    ws_raw = wb.create_sheet('Sự kiện thô')
    _stv_fill_raw_events_sheet(ws_raw, summaries, contexts)
    if sections:
        _fr_precompute_section_fit_quality(summaries, contexts, sections, section_match_tolerance_m=section_match_tolerance_m, section_measurement_mode=section_measurement_mode)
        ws_fit = wb.create_sheet('Section Fit Quality')
        _fr_fill_section_fit_quality_sheet(ws_fit, summaries, contexts)
    ws_raw_diag = wb.create_sheet('Raw Trace Diagnostics')
    _fr_fill_raw_trace_diagnostics_sheet(ws_raw_diag, summaries, contexts)
    ws_orl = wb.create_sheet('ORL Analysis')
    _fr_fill_orl_analysis_sheet(ws_orl, summaries, contexts)
    ws_parser = wb.create_sheet('Parser Diagnostics')
    _fr_fill_parser_diagnostics_sheet(ws_parser, summaries, contexts, skipped)
    ws_vendor = wb.create_sheet('Vendor Compatibility')
    _fr_fill_vendor_compatibility_matrix_sheet(ws_vendor, summaries, contexts, skipped)
    ws_strict = wb.create_sheet('Strict Validation')
    _fr_fill_strict_validation_sheet(ws_strict, summaries, contexts, expected_route_km=expected_route_km, length_tolerance_km=length_tolerance_km if 'length_tolerance_km' in locals() else 0.300, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km, duration_threshold_s=duration_threshold_s if 'duration_threshold_s' in locals() else None, skipped=skipped)
    ws_core = wb.create_sheet('Core Metrics')
    _fr_fill_core_metrics_sheet(ws_core, summaries, contexts, threshold_db=threshold_db, section_pairs_by_file=None, duration_threshold_s=duration_threshold_s)
    ws_rules = wb.create_sheet('Output Rules')
    _fr_fill_output_rules_sheet(ws_rules, threshold_db=threshold_db, deviation_m=deviation_m, output_mode='stv', section_export_scope='all', section_measurement_mode='fit', section_event_source='all', section_boundary_priority='event', expected_route_km=expected_route_km, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km)
    _fr_apply_workbook_polish(wb)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _fr_apply_workbook_polish(wb: Workbook) -> None:
    freeze_map = {
        'General Information': 'A20',
        'Link Results': 'A3',
        'Events': 'C5',
        'Sections': 'C4',
        'App Parameters': 'A2',
        'Route Analysis': 'A2',
        'Segment Analysis': 'A2',
        'Segment Events': 'A2',
        'Run Log': 'A2',
        'Core Metrics': 'A2',
        'Output Rules': 'A2',
        'Section Fit Quality': 'A2',
        'ORL Analysis': 'A2',
        'Raw Trace Diagnostics': 'A2',
        'Parser Diagnostics': 'A2',
        'Vendor Compatibility': 'A2',
        'Strict Validation': 'A2',
    }
    header_map = {
        'General Information': (18, 20, 23, 24),
        'Link Results': (2,),
        'Events': (1, 2, 3, 4),
        'Sections': (1, 2, 3),
        'App Parameters': (1,),
        'Route Analysis': (1,),
        'Segment Analysis': (1,),
        'Segment Events': (1,),
        'Run Log': (1,),
        'Core Metrics': (1,),
        'Output Rules': (1,),
        'Section Fit Quality': (1,),
        'ORL Analysis': (1,),
        'Raw Trace Diagnostics': (1,),
        'Parser Diagnostics': (1,),
        'Vendor Compatibility': (1,),
        'Strict Validation': (1,),
    }
    selective_fit = {'App Parameters', 'Route Analysis', 'Segment Analysis', 'Segment Events', 'Run Log', 'Core Metrics', 'Output Rules', 'Section Fit Quality', 'ORL Analysis', 'Raw Trace Diagnostics', 'Parser Diagnostics', 'Vendor Compatibility', 'Strict Validation', 'Kiểm tra đồ thị', 'Tệp bỏ qua', 'Sự kiện thô'}
    for ws in wb.worksheets:
        if ws.title in freeze_map:
            ws.freeze_panes = freeze_map[ws.title]
        if ws.title in header_map:
            _fr_style_sheet_headers(ws, header_map[ws.title])
        ws.auto_filter.ref = ws.dimensions
        if ws.title in selective_fit:
            _fr_selective_autofit(ws)

        # Làm tròn tất cả các số thập phân và chỉnh định dạng số Excel về 2 chữ số thập phân
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is not None:
                    if isinstance(val, float):
                        cell.value = round(val, 2)
                        if cell.number_format is None or cell.number_format == 'General':
                            cell.number_format = '0.00'
                    
                    fmt = cell.number_format
                    if fmt:
                        if '%' in fmt:
                            base_fmt = fmt.replace('%', '')
                            if '.' in base_fmt:
                                parts = base_fmt.split('.')
                                if len(parts) == 2 and any(char in parts[1] for char in ('0', '#')):
                                    cell.number_format = parts[0] + '.00%'
                        else:
                            if '.' in fmt:
                                parts = fmt.split('.')
                                if len(parts) == 2 and any(char in parts[1] for char in ('0', '#')):
                                    cell.number_format = parts[0] + '.00'




def _fr_precompute_section_fit_quality(
    summaries: list[FileSummary],
    contexts: dict[str, dict],
    sections: list[dict],
    *,
    section_match_tolerance_m: float = 100.0,
    section_measurement_mode: str = 'fit',
) -> None:
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        ctx['section_fit_rows'] = []
        rows = ctx.get('events', [])
        for i, sec in enumerate(sections, start=1):
            _loss, _att, fit_result = _fr_compute_section_values_with_fit(
                summary,
                ctx,
                rows,
                float(sec.get('start_km') or 0.0),
                float(sec.get('end_km') or 0.0),
                section_index=i,
                section_match_tolerance_m=section_match_tolerance_m,
                section_measurement_mode=section_measurement_mode,
            )
            ctx.setdefault('section_fit_rows', []).append(fit_result)


def _fr_fill_section_fit_quality_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    headers = [
        'Tệp', 'Fiber', 'Wavelength', 'Section', 'Start km', 'End km', 'Span km',
        'Nguồn raw trace', 'Tổng điểm raw', 'Điểm dùng fit', 'Fit mode',
        'Fit window start km', 'Fit window end km', 'R²/RMS scope', 'Estimate level',
        'Slope fit (dB/km)', 'Attenuation dùng fit (dB/km)', 'Loss dùng fit (dB)',
        'R²', 'RMS residual (dB)', 'Max |residual| (dB)', 'Confidence nền',
        'Đã dùng cho Sections?', 'Fallback', 'Ghi chú'
    ]
    _fr_write_header_row(ws, headers)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        fit_rows = ctx.get('section_fit_rows') or []
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        for fit in fit_rows:
            ws.cell(row, 1, summary.file_name)
            ws.cell(row, 2, _display_fiber_label(summary, ctx.get('metadata', {})))
            ws.cell(row, 3, int(m.group(1)) if m else None)
            ws.cell(row, 4, fit.section_index)
            ws.cell(row, 5, fit.start_km)
            ws.cell(row, 6, fit.end_km)
            ws.cell(row, 7, fit.span_km)
            ws.cell(row, 8, fit.source)
            ws.cell(row, 9, fit.raw_points_total)
            ws.cell(row, 10, fit.fit_points_used)
            ws.cell(row, 11, getattr(fit, 'fit_mode', '') or '')
            ws.cell(row, 12, getattr(fit, 'fit_window_start_km', None))
            ws.cell(row, 13, getattr(fit, 'fit_window_end_km', None))
            ws.cell(row, 14, getattr(fit, 'r2_rms_scope', '') or '')
            ws.cell(row, 15, getattr(fit, 'estimate_level', '') or '')
            ws.cell(row, 16, fit.slope_dbkm)
            ws.cell(row, 17, fit.attenuation_dbkm)
            ws.cell(row, 18, fit.loss_db)
            ws.cell(row, 19, fit.r2)
            ws.cell(row, 20, fit.rms_residual_db)
            ws.cell(row, 21, fit.max_abs_residual_db)
            ws.cell(row, 22, fit.confidence)
            ws.cell(row, 23, 'Có' if fit.used_for_section else 'Không')
            ws.cell(row, 24, fit.fallback_method)
            ws.cell(row, 25, fit.note)
            if fit.used_for_section:
                for c in range(1, 26):
                    ws.cell(row, c).fill = GREEN_FILL
            elif fit.confidence in {'Thấp', 'Không đủ điểm', 'Không có raw trace', 'Không xác định', 'Trục km lệch span'}:
                for c in range(1, 26):
                    ws.cell(row, c).fill = LOG_WARN_FILL
            row += 1
    for c in range(5, 22):
        for rr in range(2, row):
            if isinstance(ws.cell(rr, c).value, (int, float)):
                ws.cell(rr, c).number_format = '0.0000' if c in {5, 6, 7, 12, 13, 16, 17, 18, 20, 21} else '0.00000'
    widths = {
        'A': 34, 'B': 18, 'C': 12, 'D': 10, 'E': 12, 'F': 12, 'G': 12,
        'H': 24, 'I': 14, 'J': 14, 'K': 20, 'L': 18, 'M': 18, 'N': 18, 'O': 18,
        'P': 16, 'Q': 18, 'R': 16, 'S': 12, 'T': 16, 'U': 18, 'V': 22,
        'W': 20, 'X': 12, 'Y': 100,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=100)



def _fr_fill_orl_analysis_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    headers = [
        'Tệp', 'Fiber', 'Wavelength', 'ORL hiển thị', 'ORL value (dB)',
        'Legacy status', 'Advanced status', 'Nguồn ORL', 'Chi tiết nguồn', 'Độ tin cậy nguồn',
        'Ngưỡng đạt (dB)', 'Use for judgment', 'Lower-bound?', 'Lý do', 'Khuyến nghị',
        'Physical mode', 'Physical attempted', 'Physical ORL (dB)', 'Physical status',
        'Strongest reflectance event (dB)', 'Physical reason'
    ]
    _fr_write_header_row(ws, headers)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        analysis = ctx.get('orl_analysis')
        if not isinstance(analysis, ORLAnalysis):
            analysis = ORLAnalysis(
                file_name=summary.file_name,
                display=ctx.get('orl_display'),
                value_db=ctx.get('orl_value_db'),
                status=ctx.get('orl_status', 'Unknown'),
                advanced_status='UNKNOWN',
                source_kind=ctx.get('orl_source_kind', ''),
                source_detail=ctx.get('orl_source_detail', ''),
                source_confidence='',
                pass_threshold_db=28.0,
                use_for_judgment=bool(ctx.get('orl_use_for_judgment')),
                lower_bound=False,
                reason=ctx.get('orl_reason', ''),
                recommendation='',
                physical_mode='disabled',
                physical_attempted=False,
                physical_value_db=None,
                physical_status='',
                physical_reason='',
                strongest_reflectance_db=None,
            )
        meta = ctx.get('metadata', {})
        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        values = [
            summary.file_name,
            _display_fiber_label(summary, meta),
            int(m.group(1)) if m else None,
            analysis.display,
            analysis.value_db,
            analysis.status,
            analysis.advanced_status,
            analysis.source_kind,
            analysis.source_detail,
            analysis.source_confidence,
            analysis.pass_threshold_db,
            'Yes' if analysis.use_for_judgment else 'No',
            'Yes' if analysis.lower_bound else 'No',
            analysis.reason,
            analysis.recommendation,
            analysis.physical_mode,
            'Yes' if analysis.physical_attempted else 'No',
            analysis.physical_value_db,
            analysis.physical_status,
            analysis.strongest_reflectance_db,
            analysis.physical_reason,
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row, c, value)
        if analysis.advanced_status == 'PASS':
            fill = GREEN_FILL
        elif analysis.advanced_status == 'FAIL':
            fill = RED_FILL
        elif analysis.advanced_status in {'LOWER_BOUND_ONLY', 'NOT_AVAILABLE', 'UNKNOWN'}:
            fill = LOG_WARN_FILL
        else:
            fill = LOG_INFO_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row, c).fill = fill
        for c in (5, 11, 18, 20):
            if isinstance(ws.cell(row, c).value, (int, float)):
                ws.cell(row, c).number_format = '0.000'
        row += 1
    widths = {
        'A': 34, 'B': 18, 'C': 12, 'D': 14, 'E': 14, 'F': 14, 'G': 18,
        'H': 22, 'I': 42, 'J': 16, 'K': 14, 'L': 18, 'M': 14,
        'N': 72, 'O': 72, 'P': 18, 'Q': 18, 'R': 16, 'S': 18, 'T': 24, 'U': 90,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=90)


def _fr_yes_no(value) -> str:
    return 'Có' if bool(value) else 'Không'


def _fr_pct(num: int, den: int) -> Optional[float]:
    try:
        if int(den) <= 0:
            return None
        return round(float(num) / float(den), 4)
    except Exception:
        return None


def _fr_is_yes_text(value) -> bool:
    text = str(value or '').strip().lower()
    return text.startswith('có') or text in {'yes', 'true', '1', 'ok'}


def _fr_vendor_family_from_context(summary: FileSummary, ctx: dict) -> str:
    meta = (ctx or {}).get('metadata') or {}
    company = str(meta.get('company') or '').lower()
    model = str(meta.get('unit_model') or '').lower()
    parser_family = str(getattr(summary, 'parse_family', '') or (ctx or {}).get('parser_family') or '').lower()
    source_format = str(getattr(summary, 'source_format', '') or '').lower()
    haystack = ' '.join([company, model, parser_family, source_format])
    if any(k in haystack for k in ['viavi', 'jdsu', 'acterna', 'mts ', 'mts-', 'mts2000', 'mts 2000']):
        return 'VIAVI/JDSU/Acterna'
    if any(k in haystack for k in ['exfo', 'ftb-', 'ftb ', 'fastreporter', 'iolm']):
        return 'EXFO'
    if 'yokogawa' in haystack or 'aq7' in haystack or 'aq1' in haystack:
        return 'Yokogawa'
    if 'anritsu' in haystack or 'mt1000' in haystack or 'mw907' in haystack:
        return 'Anritsu'
    if str(getattr(summary, 'source_format', '') or '').upper() == 'SOR':
        return 'Generic Telcordia SOR'
    if str(getattr(summary, 'source_format', '') or '').upper() == 'TRC':
        return 'TRC/Native trace'
    return 'Unknown'


def _fr_event_health_metrics(summary: FileSummary, ctx: dict) -> dict:
    events = list((ctx or {}).get('events') or [])
    length = None
    try:
        length = float(summary.length_km) if summary.length_km is not None else None
    except Exception:
        length = None
    event_count = len(events)
    loss_count = sum(1 for e in events if e.loss_db is not None)
    refl_count = sum(1 for e in events if e.reflectance_db is not None)
    slope_count = sum(1 for e in events if e.slope_dbkm is not None)
    position_count = sum(1 for e in events if e.distance_km is not None)
    fiber_end_count = sum(1 for e in events if 'end' in str(e.event_type or '').lower() or 'fiber end' in str(e.label or '').lower())
    gainer_count = 0
    after_end_count = 0
    invalid_position_count = 0
    max_event_loss = None
    min_event_loss = None
    for e in events:
        if e.loss_db is not None:
            try:
                val = float(e.loss_db)
                if val < 0:
                    gainer_count += 1
                max_event_loss = val if max_event_loss is None else max(max_event_loss, val)
                min_event_loss = val if min_event_loss is None else min(min_event_loss, val)
            except Exception:
                pass
        if e.distance_km is not None:
            try:
                d = float(e.distance_km)
                if d < -0.001:
                    invalid_position_count += 1
                if length is not None and d > length + 0.5:
                    after_end_count += 1
            except Exception:
                invalid_position_count += 1
    return {
        'event_count': event_count,
        'position_count': position_count,
        'loss_count': loss_count,
        'reflectance_count': refl_count,
        'slope_count': slope_count,
        'fiber_end_count': fiber_end_count,
        'loss_coverage_pct': _fr_pct(loss_count, event_count),
        'reflectance_coverage_pct': _fr_pct(refl_count, event_count),
        'slope_coverage_pct': _fr_pct(slope_count, event_count),
        'gainer_count': gainer_count,
        'after_end_count': after_end_count,
        'invalid_position_count': invalid_position_count,
        'max_event_loss': max_event_loss,
        'min_event_loss': min_event_loss,
    }


def _fr_parser_final_status(summary: FileSummary, ctx: dict, diag: dict, profile: dict, event_metrics: dict) -> tuple[str, str, str, str]:
    """Return (status, level, use_for_judgment, recommendation) for operational control.

    This function only classifies existing parser outputs; it does not change
    parser/calculation logic.
    """
    event_count = int(event_metrics.get('event_count') or 0)
    loss_count = int(event_metrics.get('loss_count') or 0)
    loss_cov = event_metrics.get('loss_coverage_pct')
    can_draw = bool(profile.get('can_draw_graph'))
    can_fit = _fr_is_yes_text(profile.get('can_use_for_fit'))
    raw_scale_ok = bool((diag or {}).get('distance_scale_ok'))
    family_conf = str(getattr(summary, 'parse_family_confidence', '') or (ctx or {}).get('parser_family_confidence') or '').lower()
    status = 'PASS_READ_PARTIAL'
    level = 'WARN'
    use_for_judgment = 'Có điều kiện'
    recommendation = 'Dùng được cho báo cáo cơ bản; kiểm tra thêm các sheet Diagnostics trước khi kết luận sâu.'

    if event_count <= 0 and not can_draw:
        return ('FAIL_INSUFFICIENT_DATA', 'FAIL', 'Không', 'Không đủ event/trace để kết luận. Cần mở bằng phần mềm hãng hoặc kiểm tra lại file đo.')
    if event_count <= 0 and can_draw:
        return ('PASS_GRAPH_ONLY', 'WARN', 'Chỉ xem đồ thị', 'Chỉ có dữ liệu đồ thị; không dùng để kết luận event/section loss.')
    if event_count > 0 and loss_count <= 0:
        if can_draw:
            return ('PASS_EVENT_POSITION_ONLY', 'WARN', 'Chỉ tham khảo', 'Có vị trí event nhưng chưa có event loss; chỉ dùng kiểm tra vị trí, không kết luận suy hao event.')
        return ('PASS_EVENT_ONLY', 'WARN', 'Chỉ tham khảo', 'Có event nhưng thiếu loss/trace đủ tin cậy; cần kiểm tra bằng phần mềm hãng.')
    if event_count > 0 and isinstance(loss_cov, float) and loss_cov < 0.70:
        return ('PASS_READ_PARTIAL', 'WARN', 'Có điều kiện', 'Event loss chưa đủ coverage; chỉ dùng phần có dữ liệu, cần kiểm tra event thiếu loss.')
    if can_fit and raw_scale_ok and (not isinstance(loss_cov, float) or loss_cov >= 0.90) and family_conf in {'high', 'medium'}:
        return ('PASS_READ_FULL', 'OK', 'Có', 'Đọc tốt: metadata/event/trace đủ điều kiện để dùng cho kết luận chính theo các rule hiện tại.')
    if can_draw and (not isinstance(loss_cov, float) or loss_cov >= 0.80):
        return ('PASS_READ_PARTIAL', 'WARN', 'Có điều kiện', 'Event đọc tốt nhưng trace/scale chưa đủ để fit sâu; dùng Excel đầy đủ và diagnostics khi kết luận section.')
    if can_draw:
        return ('PASS_GRAPH_ONLY', 'WARN', 'Chỉ xem đồ thị', 'Dùng tốt để xem trace/event nhanh; không thay thế phân tích Excel đầy đủ.')
    return (status, level, use_for_judgment, recommendation)


def _fr_parser_warning_text(summary: FileSummary, ctx: dict, diag: dict, profile: dict, event_metrics: dict) -> str:
    warnings: list[str] = []
    conf = str(getattr(summary, 'parse_family_confidence', '') or (ctx or {}).get('parser_family_confidence') or '')
    if conf.lower() in {'low', ''}:
        warnings.append('Độ tin cậy family parser thấp/chưa rõ.')
    if event_metrics.get('event_count') and event_metrics.get('loss_count') != event_metrics.get('event_count'):
        warnings.append(f"Event loss coverage {event_metrics.get('loss_count')}/{event_metrics.get('event_count')}.")
    if event_metrics.get('after_end_count'):
        warnings.append(f"Có {event_metrics.get('after_end_count')} event nằm sau span length + 0.5 km.")
    if event_metrics.get('invalid_position_count'):
        warnings.append(f"Có {event_metrics.get('invalid_position_count')} event có vị trí bất thường.")
    if event_metrics.get('gainer_count'):
        warnings.append(f"Có {event_metrics.get('gainer_count')} event loss âm/gainer.")
    if diag and diag.get('candidate_found') and not diag.get('distance_scale_ok'):
        warnings.append(str(diag.get('distance_scale_note') or diag.get('reject_reason') or 'Trace scale chưa tin cậy.'))
    if profile and not _fr_is_yes_text(profile.get('can_use_for_fit')):
        warnings.append('Trace hiện không đủ điều kiện dùng fit sâu.')
    try:
        att = float(summary.attenuation_dbkm) if summary.attenuation_dbkm is not None else None
        if att is not None and (att < 0 or att > 2.0):
            warnings.append(f'Attenuation bất thường: {att:.3f} dB/km.')
        elif att is not None and (att < 0.05 or att > 0.80):
            warnings.append(f'Attenuation ngoài dải kiểm tra thông thường: {att:.3f} dB/km.')
    except Exception:
        pass
    orl = (ctx or {}).get('orl_analysis')
    if orl is not None and not getattr(orl, 'use_for_judgment', False):
        warnings.append('ORL không phải measured ORL dùng để kết luận.')
    if not warnings:
        return 'OK'
    return ' | '.join(dict.fromkeys(warnings))


def _fr_fill_parser_diagnostics_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict], skipped: Optional[list[str]] = None) -> None:
    headers = [
        'Tệp', 'Fiber/Core', 'Định dạng', 'Vendor family', 'Model máy đo', 'Serial máy',
        'Parser family', 'Độ tin cậy parser', 'Parse mode', 'Final parser status', 'Mức cảnh báo',
        'Dùng để kết luận?', 'Khuyến nghị vận hành',
        'Bước sóng', 'Range đo (km)', 'Pulse (µs)', 'Duration (s)',
        'Span length (km)', 'Graph end (km)', 'Fiber end (km)',
        'Parsed span loss (dB)', 'Route-corrected loss (dB)', 'Loss đang dùng (dB)', 'Nguồn loss',
        'Số event', 'Event có position', 'Event có loss', 'Event loss coverage', 'Event có reflectance',
        'Reflectance coverage', 'Event có slope', 'Slope coverage', 'Fiber End count',
        'Event sau tuyến', 'Event loss âm/gainer', 'Max event loss (dB)', 'Min event loss (dB)',
        'Trace availability', 'Trace source type', 'Trace source rank', 'Graph mode',
        'Trace source', 'Trace points', 'First km', 'Last km', 'Reference span km',
        'Raw last trước scale', 'Raw last sau scale', 'Distance scale status', 'Distance scale OK?',
        'Can draw graph?', 'Can use for fit?', 'Can calculate R²/RMS?', 'Trace quality',
        'ORL display', 'ORL source', 'ORL dùng kết luận?', 'ORL status', 'ORL reason',
        'Route status', 'Route reason', 'Parser reason', 'Warnings'
    ]
    _fr_write_header_row(ws, headers)
    row = 2

    status_fills = {
        'OK': GREEN_FILL,
        'WARN': LOG_WARN_FILL,
        'FAIL': LOG_ERROR_FILL,
    }

    for summary in summaries:
        ctx = contexts.get(summary.file_name) or {}
        meta = ctx.get('metadata') or {}
        diag = _raw_trace_candidate_diagnostics(summary, ctx)
        profile = _trace_profile_for_diagnostics(summary, ctx, diag)
        event_metrics = _fr_event_health_metrics(summary, ctx)
        final_status, level, use_for_judgment, recommendation = _fr_parser_final_status(summary, ctx, diag, profile, event_metrics)
        warnings = _fr_parser_warning_text(summary, ctx, diag, profile, event_metrics)
        ga = ctx.get('graph_assessment')
        orl = ctx.get('orl_analysis')
        values = [
            summary.file_name,
            _display_fiber_label(summary, meta),
            summary.source_format,
            _fr_vendor_family_from_context(summary, ctx),
            meta.get('unit_model') or '',
            meta.get('unit_serial') or '',
            _to_vi_parser_family(getattr(summary, 'parse_family', '') or ctx.get('parser_family') or ''),
            getattr(summary, 'parse_family_confidence', '') or ctx.get('parser_family_confidence') or '',
            _to_vi_parse_mode(ctx.get('parse_mode') or ''),
            final_status,
            level,
            use_for_judgment,
            recommendation,
            summary.wavelength_display,
            meta.get('range_km'),
            meta.get('pulse_us'),
            meta.get('duration_s'),
            summary.length_km,
            summary.graph_end_km,
            summary.end_distance_km,
            summary.parsed_total_loss_db,
            summary.route_corrected_total_loss_db,
            summary.total_loss_db,
            summary.loss_source_used,
            event_metrics.get('event_count'),
            event_metrics.get('position_count'),
            event_metrics.get('loss_count'),
            event_metrics.get('loss_coverage_pct'),
            event_metrics.get('reflectance_count'),
            event_metrics.get('reflectance_coverage_pct'),
            event_metrics.get('slope_count'),
            event_metrics.get('slope_coverage_pct'),
            event_metrics.get('fiber_end_count'),
            event_metrics.get('after_end_count'),
            event_metrics.get('gainer_count'),
            event_metrics.get('max_event_loss'),
            event_metrics.get('min_event_loss'),
            profile.get('trace_availability'),
            profile.get('trace_source_type'),
            profile.get('trace_source_rank'),
            profile.get('graph_mode'),
            diag.get('candidate_source') or diag.get('source') or '',
            diag.get('x_count') or diag.get('candidate_points'),
            diag.get('first_km'),
            diag.get('last_km'),
            diag.get('reference_span_km'),
            diag.get('raw_last_before_km'),
            diag.get('raw_last_after_km'),
            diag.get('distance_scale_status') or diag.get('raw_status'),
            _fr_yes_no(diag.get('distance_scale_ok')),
            _fr_yes_no(profile.get('can_draw_graph')),
            profile.get('can_use_for_fit'),
            profile.get('can_calculate_r2_rms'),
            profile.get('trace_quality_level'),
            getattr(orl, 'display', ctx.get('orl_display') or ''),
            getattr(orl, 'source_detail', ctx.get('orl_source_detail') or ''),
            _fr_yes_no(getattr(orl, 'use_for_judgment', ctx.get('orl_use_for_judgment') or False)),
            getattr(orl, 'advanced_status', ctx.get('orl_status') or ''),
            getattr(orl, 'reason', ctx.get('orl_reason') or ''),
            getattr(ga, 'verdict', ''),
            getattr(ga, 'reason', ''),
            getattr(summary, 'parse_family_reason', '') or ctx.get('parser_family_reason') or '',
            warnings,
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row, c, value)
        fill = status_fills.get(level, LOG_INFO_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row, c).fill = fill
        # Numeric formats
        for c in [14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 28, 30, 32, 36, 37, 43, 44, 45, 46, 47, 48]:
            cell = ws.cell(row, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.000'
        for c in [28, 30, 32]:
            cell = ws.cell(row, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'
        row += 1

    for item in skipped or []:
        file_name, reason = (item.split(': ', 1) + [''])[:2] if ': ' in item else (item, '')
        values = [file_name, '', '', 'Unknown', '', '', '', '', '', 'FAIL_SKIPPED', 'FAIL', 'Không', 'File bị bỏ qua, không dùng để kết luận.'] + [''] * (len(headers) - 14) + [reason]
        for c, value in enumerate(values[:len(headers)], start=1):
            ws.cell(row, c, value)
            ws.cell(row, c).fill = LOG_ERROR_FILL
        row += 1

    widths = {
        'A': 32, 'B': 18, 'C': 12, 'D': 22, 'E': 26, 'F': 16, 'G': 34, 'H': 18,
        'I': 28, 'J': 28, 'K': 14, 'L': 18, 'M': 62, 'N': 12, 'O': 14,
        'P': 12, 'Q': 13, 'R': 15, 'S': 14, 'T': 14, 'U': 17, 'V': 18,
        'W': 16, 'X': 24, 'Y': 12, 'Z': 16,
        'AA': 14, 'AB': 16, 'AC': 16, 'AD': 17, 'AE': 14, 'AF': 14, 'AG': 14,
        'AH': 14, 'AI': 16, 'AJ': 16, 'AK': 16, 'AL': 24, 'AM': 18, 'AN': 14,
        'AO': 18, 'AP': 24, 'AQ': 14, 'AR': 14, 'AS': 14, 'AT': 16, 'AU': 18,
        'AV': 18, 'AW': 20, 'AX': 18, 'AY': 16, 'AZ': 18, 'BA': 18, 'BB': 18,
        'BC': 18, 'BD': 20, 'BE': 18, 'BF': 18, 'BG': 60, 'BH': 60, 'BI': 60, 'BJ': 86,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=90)





def _fr_vendor_rulebook_rows() -> list[dict]:
    """Phase 7.4: static compatibility guide.

    This matrix is intentionally diagnostic-only. It does not change parsing,
    event calculations, section fitting, ORL rules, or workbook formulas.
    """
    return [
        {
            'row_type': 'RULEBOOK',
            'vendor_family': 'EXFO',
            'source_format': 'SOR/MSOR/iOLM export',
            'support_level': 'SUPPORTED_IF_STANDARD_BLOCKS',
            'metadata_support': 'OK',
            'event_position_support': 'OK',
            'event_loss_support': 'OK nếu KeyEvents/Events chuẩn có loss',
            'reflectance_support': 'OK nếu file có reflectance',
            'trace_graph_support': 'OK nếu có DataPts/raw trace',
            'raw_fit_support': 'OK nếu trace scale khớp span',
            'orl_support': 'Measured only',
            'production_usage': 'Dùng kết luận khi Parser Diagnostics = PASS_READ_FULL hoặc PASS_READ_PARTIAL có điều kiện rõ.',
            'known_limits': 'iOLM/proprietary package có thể cần export SOR chuẩn để đọc đủ.',
            'required_action': 'Nếu FAIL/partial: export lại dạng SOR chuẩn hoặc gửi file mẫu để bổ sung parser.',
        },
        {
            'row_type': 'RULEBOOK',
            'vendor_family': 'VIAVI/JDSU/Acterna',
            'source_format': 'MSOR/SOR',
            'support_level': 'SUPPORTED_CONDITIONAL',
            'metadata_support': 'OK',
            'event_position_support': 'OK với smart_link/XML/KeyEvents',
            'event_loss_support': 'OK nếu ghép được KeyEvents nhị phân',
            'reflectance_support': 'LIMITED theo block có sẵn',
            'trace_graph_support': 'OK với ActernaMiniCurve/display trace',
            'raw_fit_support': 'CONDITIONAL: chỉ dùng khi distance scale OK và đủ điểm',
            'orl_support': 'Measured only; lower-bound chỉ tham khảo',
            'production_usage': 'Dùng báo cáo cơ bản khi event loss coverage tốt; fit sâu phụ thuộc trace diagnostics.',
            'known_limits': 'MiniCurve có thể lệch span hoặc là display trace, không mặc định coi là raw trace thật.',
            'required_action': 'Kiểm tra Parser Diagnostics + Raw Trace Diagnostics trước khi kết luận section fit sâu.',
        },
        {
            'row_type': 'RULEBOOK',
            'vendor_family': 'Yokogawa',
            'source_format': 'SOR/proprietary',
            'support_level': 'LIMITED_UNTIL_VALIDATED',
            'metadata_support': 'OK nếu Telcordia block chuẩn',
            'event_position_support': 'OK nếu KeyEvents chuẩn',
            'event_loss_support': 'CONDITIONAL',
            'reflectance_support': 'CONDITIONAL',
            'trace_graph_support': 'CONDITIONAL',
            'raw_fit_support': 'CONDITIONAL sau khi có file mẫu validate',
            'orl_support': 'Measured only',
            'production_usage': 'Không dùng kết luận chính nếu chưa có PASS_READ_FULL/PARTIAL có kiểm chứng.',
            'known_limits': 'Chưa đủ bộ golden test Yokogawa trong app.',
            'required_action': 'Bổ sung file mẫu tốt/xấu để khóa parser và sanity rule.',
        },
        {
            'row_type': 'RULEBOOK',
            'vendor_family': 'Anritsu',
            'source_format': 'SOR/proprietary',
            'support_level': 'LIMITED_UNTIL_VALIDATED',
            'metadata_support': 'OK nếu Telcordia block chuẩn',
            'event_position_support': 'OK nếu KeyEvents chuẩn',
            'event_loss_support': 'CONDITIONAL',
            'reflectance_support': 'CONDITIONAL',
            'trace_graph_support': 'CONDITIONAL',
            'raw_fit_support': 'CONDITIONAL sau khi có file mẫu validate',
            'orl_support': 'Measured only',
            'production_usage': 'Không dùng kết luận chính nếu chưa có PASS_READ_FULL/PARTIAL có kiểm chứng.',
            'known_limits': 'Chưa đủ bộ golden test Anritsu trong app.',
            'required_action': 'Bổ sung file mẫu tốt/xấu để khóa parser và sanity rule.',
        },
        {
            'row_type': 'RULEBOOK',
            'vendor_family': 'Generic Telcordia SOR',
            'source_format': 'SOR',
            'support_level': 'SUPPORTED_BY_STANDARD_BLOCKS',
            'metadata_support': 'OK nếu block chuẩn đầy đủ',
            'event_position_support': 'OK',
            'event_loss_support': 'OK nếu KeyEvents có loss',
            'reflectance_support': 'OK nếu file có reflectance',
            'trace_graph_support': 'OK nếu có DataPts',
            'raw_fit_support': 'OK nếu trace scale OK và đủ điểm',
            'orl_support': 'Measured only',
            'production_usage': 'Dùng kết luận khi Parser Diagnostics cho phép.',
            'known_limits': 'Vendor có thể nhúng private block không chuẩn.',
            'required_action': 'Nếu thiếu trường quan trọng, kiểm tra lại bằng phần mềm hãng.',
        },
        {
            'row_type': 'RULEBOOK',
            'vendor_family': 'TRC/Native trace',
            'source_format': 'TRC',
            'support_level': 'TRACE_ONLY_CONDITIONAL',
            'metadata_support': 'LIMITED',
            'event_position_support': 'CONDITIONAL',
            'event_loss_support': 'CONDITIONAL',
            'reflectance_support': 'CONDITIONAL',
            'trace_graph_support': 'OK nếu đọc được điểm trace',
            'raw_fit_support': 'CONDITIONAL',
            'orl_support': 'Unknown nếu không có measured ORL',
            'production_usage': 'Ưu tiên dùng để xem trace; kết luận phụ thuộc Parser Diagnostics.',
            'known_limits': 'TRC có nhiều biến thể native; không mặc định đủ event table.',
            'required_action': 'Cần cross-check với file SOR/MSOR cùng core nếu có.',
        },
        {
            'row_type': 'RULEBOOK',
            'vendor_family': 'Unknown',
            'source_format': 'Any',
            'support_level': 'UNSUPPORTED_UNTIL_CLASSIFIED',
            'metadata_support': 'UNKNOWN',
            'event_position_support': 'UNKNOWN',
            'event_loss_support': 'UNKNOWN',
            'reflectance_support': 'UNKNOWN',
            'trace_graph_support': 'UNKNOWN',
            'raw_fit_support': 'NO',
            'orl_support': 'NO',
            'production_usage': 'Không dùng kết luận chính.',
            'known_limits': 'Không đủ dấu hiệu vendor/parser.',
            'required_action': 'Đưa vào Skipped/Diagnostics và bổ sung parser sau khi có file mẫu.',
        },
    ]


def _fr_vendor_support_row(summary: FileSummary, ctx: dict) -> dict:
    meta = ctx.get('metadata') or {}
    diag = _raw_trace_candidate_diagnostics(summary, ctx)
    profile = _trace_profile_for_diagnostics(summary, ctx, diag)
    event_metrics = _fr_event_health_metrics(summary, ctx)
    final_status, level, use_for_judgment, recommendation = _fr_parser_final_status(summary, ctx, diag, profile, event_metrics)
    vendor = _fr_vendor_family_from_context(summary, ctx)
    event_count = int(event_metrics.get('event_count') or 0)
    position_count = int(event_metrics.get('position_count') or 0)
    loss_count = int(event_metrics.get('loss_count') or 0)
    refl_count = int(event_metrics.get('reflectance_count') or 0)
    can_draw = bool(profile.get('can_draw_graph'))
    can_fit = _fr_is_yes_text(profile.get('can_use_for_fit'))
    can_r2 = _fr_is_yes_text(profile.get('can_calculate_r2_rms'))
    trace_scale_ok = bool(diag.get('distance_scale_ok'))
    orl = ctx.get('orl_analysis')
    orl_use = bool(getattr(orl, 'use_for_judgment', ctx.get('orl_use_for_judgment') or False))

    if final_status == 'PASS_READ_FULL':
        support_level = 'SUPPORTED_FULL'
    elif final_status in {'PASS_READ_PARTIAL', 'PASS_EVENT_ONLY'}:
        support_level = 'SUPPORTED_CONDITIONAL'
    elif final_status in {'PASS_GRAPH_ONLY', 'PASS_EVENT_POSITION_ONLY'}:
        support_level = 'VIEW_OR_REFERENCE_ONLY'
    else:
        support_level = 'NOT_SAFE_FOR_PRODUCTION_JUDGMENT'

    def count_status(count: int, total: int, label: str) -> str:
        if total <= 0:
            return 'NO_DATA'
        if count >= total:
            return f'OK {count}/{total}'
        if count > 0:
            return f'PARTIAL {count}/{total}'
        return f'MISSING 0/{total}'

    evidence = []
    evidence.append(f"status={final_status}")
    evidence.append(f"event_loss={loss_count}/{event_count}")
    evidence.append(f"trace={profile.get('trace_source_type') or profile.get('trace_availability')}")
    evidence.append(f"scale_ok={_fr_yes_no(trace_scale_ok)}")
    if getattr(summary, 'parse_family_reason', '') or ctx.get('parser_family_reason'):
        evidence.append(f"reason={getattr(summary, 'parse_family_reason', '') or ctx.get('parser_family_reason')}")

    limitations = []
    if loss_count < event_count:
        limitations.append('Event loss chưa đủ coverage.')
    if can_draw and not can_fit:
        limitations.append('Trace chỉ dùng xem/fit có điều kiện; chưa đủ fit sâu.')
    if diag.get('candidate_found') and not trace_scale_ok:
        limitations.append('Distance scale chưa đạt rule tin cậy.')
    if not orl_use:
        limitations.append('ORL không phải measured ORL dùng kết luận.')
    if vendor in {'Yokogawa', 'Anritsu'}:
        limitations.append('Vendor chưa có đủ golden test trong app.')
    if not limitations:
        limitations.append('Không phát hiện hạn chế chính trong lớp compatibility.')

    return {
        'row_type': 'FILE_CHECK',
        'vendor_family': vendor,
        'source_format': summary.source_format,
        'file_core': f"{summary.file_name} | {_display_fiber_label(summary, meta)}",
        'model': meta.get('unit_model') or '',
        'parser_family': _to_vi_parser_family(getattr(summary, 'parse_family', '') or ctx.get('parser_family') or ''),
        'parser_confidence': getattr(summary, 'parse_family_confidence', '') or ctx.get('parser_family_confidence') or '',
        'support_level': support_level,
        'final_status': final_status,
        'warning_level': level,
        'use_for_judgment': use_for_judgment,
        'metadata_support': 'OK' if meta else 'LIMITED',
        'event_position_support': count_status(position_count, event_count, 'position'),
        'event_loss_support': count_status(loss_count, event_count, 'loss'),
        'reflectance_support': count_status(refl_count, event_count, 'reflectance'),
        'trace_graph_support': 'OK' if can_draw else 'NO',
        'raw_fit_support': 'OK' if can_fit else ('CONDITIONAL/NO' if can_draw else 'NO'),
        'r2_rms_support': 'OK' if can_r2 else 'NO',
        'distance_scale_status': diag.get('distance_scale_status') or diag.get('raw_status') or '',
        'orl_support': 'MEASURED_OK' if orl_use else 'REFERENCE/UNKNOWN',
        'production_usage': recommendation,
        'known_limits': ' | '.join(dict.fromkeys(limitations)),
        'required_action': 'Cho phép dùng theo cột Dùng để kết luận; nếu WARN/FAIL phải xem Parser Diagnostics trước khi chốt báo cáo.',
        'evidence': ' | '.join(evidence),
    }


def _fr_fill_vendor_compatibility_matrix_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict], skipped: Optional[list[str]] = None) -> None:
    headers = [
        'Loại dòng', 'Vendor family', 'Định dạng', 'File/Core', 'Model máy đo',
        'Parser family', 'Độ tin cậy parser', 'Mức hỗ trợ', 'Final parser status', 'Mức cảnh báo',
        'Dùng để kết luận?', 'Metadata', 'Event position', 'Event loss', 'Reflectance',
        'Trace graph', 'Raw fit', 'R²/RMS', 'Distance scale status', 'ORL',
        'Khuyến nghị vận hành', 'Giới hạn đã biết', 'Việc cần làm', 'Bằng chứng nhận diện'
    ]
    _fr_write_header_row(ws, headers)
    row = 2
    fills = {
        'SUPPORTED_FULL': GREEN_FILL,
        'SUPPORTED_CONDITIONAL': LOG_WARN_FILL,
        'VIEW_OR_REFERENCE_ONLY': LOG_WARN_FILL,
        'NOT_SAFE_FOR_PRODUCTION_JUDGMENT': LOG_ERROR_FILL,
        'SUPPORTED_IF_STANDARD_BLOCKS': LOG_INFO_FILL,
        'SUPPORTED_BY_STANDARD_BLOCKS': LOG_INFO_FILL,
        'LIMITED_UNTIL_VALIDATED': LOG_WARN_FILL,
        'TRACE_ONLY_CONDITIONAL': LOG_WARN_FILL,
        'UNSUPPORTED_UNTIL_CLASSIFIED': LOG_ERROR_FILL,
    }

    for rule in _fr_vendor_rulebook_rows():
        values = [
            rule.get('row_type'), rule.get('vendor_family'), rule.get('source_format'), '', '',
            '', '', rule.get('support_level'), '', '', '',
            rule.get('metadata_support'), rule.get('event_position_support'), rule.get('event_loss_support'),
            rule.get('reflectance_support'), rule.get('trace_graph_support'), rule.get('raw_fit_support'),
            '', '', rule.get('orl_support'), rule.get('production_usage'), rule.get('known_limits'),
            rule.get('required_action'), 'Quy tắc compatibility tĩnh của app'
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row, c, value)
            ws.cell(row, c).fill = fills.get(rule.get('support_level'), LOG_INFO_FILL)
        row += 1

    row += 1
    for summary in summaries:
        ctx = contexts.get(summary.file_name) or {}
        r = _fr_vendor_support_row(summary, ctx)
        values = [
            r.get('row_type'), r.get('vendor_family'), r.get('source_format'), r.get('file_core'), r.get('model'),
            r.get('parser_family'), r.get('parser_confidence'), r.get('support_level'), r.get('final_status'), r.get('warning_level'),
            r.get('use_for_judgment'), r.get('metadata_support'), r.get('event_position_support'), r.get('event_loss_support'),
            r.get('reflectance_support'), r.get('trace_graph_support'), r.get('raw_fit_support'), r.get('r2_rms_support'),
            r.get('distance_scale_status'), r.get('orl_support'), r.get('production_usage'), r.get('known_limits'),
            r.get('required_action'), r.get('evidence')
        ]
        fill = fills.get(r.get('support_level'), LOG_INFO_FILL)
        if r.get('warning_level') == 'FAIL':
            fill = LOG_ERROR_FILL
        for c, value in enumerate(values, start=1):
            ws.cell(row, c, value)
            ws.cell(row, c).fill = fill
        row += 1

    for item in skipped or []:
        file_name, reason = (item.split(': ', 1) + [''])[:2] if ': ' in item else (item, '')
        values = ['FILE_CHECK', 'Unknown', '', file_name, '', '', '', 'NOT_SAFE_FOR_PRODUCTION_JUDGMENT', 'FAIL_SKIPPED', 'FAIL', 'Không'] + [''] * 9 + ['File bị bỏ qua.', reason, 'Cần kiểm tra định dạng/file lỗi.', reason]
        for c, value in enumerate(values[:len(headers)], start=1):
            ws.cell(row, c, value)
            ws.cell(row, c).fill = LOG_ERROR_FILL
        row += 1

    widths = {
        'A': 16, 'B': 24, 'C': 18, 'D': 42, 'E': 24, 'F': 34, 'G': 18, 'H': 28,
        'I': 28, 'J': 14, 'K': 18, 'L': 16, 'M': 22, 'N': 22, 'O': 22, 'P': 18,
        'Q': 18, 'R': 14, 'S': 26, 'T': 18, 'U': 72, 'V': 72, 'W': 64, 'X': 86,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=90)


def _raw_trace_candidate_diagnostics(summary: FileSummary, ctx: dict) -> dict:
    """Create a human-readable diagnostic summary for raw trace availability.

    This does not change parsing or fitting logic. It only explains why raw-fit
    could or could not run for a file.
    """
    file_name = summary.file_name
    ext = Path(file_name).suffix.lower()
    sor_meta = ctx.get('sor_meta') or {}
    trc_trace = ctx.get('trc_trace') or {}
    trace_series = ctx.get('raw_trace_series') or None
    fit_rows = ctx.get('section_fit_rows') or []

    candidate_found = False
    candidate_source = ''
    candidate_points = None
    candidate_note = ''

    try:
        if ext == '.sor':
            if sor_meta.get('trace_values_db'):
                candidate_found = True
                candidate_source = sor_meta.get('trace_source') or 'SOR DataPts'
                candidate_points = len(sor_meta.get('trace_values_db') or [])
                candidate_note = 'SOR parser đọc được trace_values_db từ DataPts.'
            elif sor_meta:
                candidate_note = 'SOR metadata đọc được nhưng chưa có trace_values_db/DataPts dùng được.'
            else:
                candidate_note = 'Không có SOR metadata/DataPts trong context.'
        elif ext in {'.trc', '.crt'}:
            raw_samples = trc_trace.get('RawSamples') if isinstance(trc_trace, dict) else None
            if isinstance(raw_samples, (bytes, bytearray)) and len(raw_samples) >= 64:
                candidate_found = True
                candidate_source = 'TRC RawSamples'
                candidate_points = len(raw_samples) // 2
                candidate_note = 'TRC parser tìm thấy RawSamples nhị phân.'
            elif trc_trace:
                candidate_note = 'TRC metadata có nhưng không thấy RawSamples đủ dài.'
            else:
                candidate_note = 'Không có TRC trace metadata trong context.'
        elif ext == '.msor':
            if trace_series and trace_series.get('source'):
                candidate_found = True
                candidate_source = str(trace_series.get('source') or 'MSOR raw trace')
                candidate_points = int(trace_series.get('raw_points_total') or 0)
                candidate_note = 'MSOR raw/mini-curve đã được chuẩn hóa thành raw_trace_series.'
            else:
                candidate_note = 'Chưa trích được MiniCurve/ActernaMiniCurve hoặc raw trace tương đương từ MSOR.'
        else:
            candidate_note = 'Định dạng không có rule raw trace riêng.'
    except Exception as exc:
        candidate_note = f'Lỗi khi đọc ứng viên raw trace: {exc}'

    x = (trace_series or {}).get('x_km') or []
    y = (trace_series or {}).get('y_db') or []
    x_count = len(x)
    y_count = len(y)
    first_km = last_km = min_level = max_level = None
    try:
        if x:
            first_km = round(float(x[0]), 6)
            last_km = round(float(x[-1]), 6)
        if y:
            finite_y = [float(v) for v in y if isinstance(v, (int, float)) and math.isfinite(float(v))]
            if finite_y:
                min_level = round(min(finite_y), 6)
                max_level = round(max(finite_y), 6)
    except Exception:
        pass

    used_count = sum(1 for f in fit_rows if getattr(f, 'used_for_section', False))
    fallback_count = sum(1 for f in fit_rows if getattr(f, 'fallback_method', '') or not getattr(f, 'used_for_section', False))
    total_fit_rows = len(fit_rows)

    raw_status = 'Chưa đọc được trace trong app'
    reject_reason = ''
    recommendation = ''
    if trace_series:
        if x_count != y_count or x_count < 16:
            raw_status = 'Có trace nhưng không đủ điểm'
            reject_reason = 'x/y lệch độ dài hoặc số điểm dưới ngưỡng tối thiểu.'
            recommendation = 'Kiểm tra block raw samples/DataPts và cách scale khoảng cách.'
        elif trace_series.get('distance_scale_ok') is False:
            raw_status = 'Có trace nhưng scale km chưa tin cậy'
            reject_reason = trace_series.get('distance_scale_note') or 'Trục km raw trace không khớp span tham chiếu.'
            recommendation = 'Không dùng raw-fit cho loss section cho tới khi xác minh scale khoảng cách hoặc bổ sung parser đúng block.'
        else:
            raw_status = 'Có trace'
            if bool(trace_series.get('calibrated_db')):
                raw_status = 'Có trace đã scale dB'
                recommendation = 'Có thể dùng cho raw-trace fit nếu từng section đủ điểm, scale km OK và residual đạt.'
            else:
                raw_status = 'Có trace dạng shape/tham khảo'
                recommendation = 'Dùng để kiểm tra hình dạng trace; cần hiệu chuẩn dB chắc hơn nếu muốn dùng làm loss chính.'
            if total_fit_rows and used_count == 0:
                reject_reason = 'Raw trace có/hoặc có ứng viên nhưng chưa có section nào dùng raw-fit; xem Section Fit Quality để biết lý do từng section.'
    else:
        reject_reason = candidate_note or 'Không tạo được raw_trace_series.'
        recommendation = 'Nếu phần mềm gốc vẫn vẽ được trace, cần mở rộng parser để đọc đúng block raw trace của hãng/file này.'

    return {
        'candidate_found': candidate_found,
        'candidate_source': candidate_source,
        'candidate_points': candidate_points,
        'candidate_note': candidate_note,
        'raw_status': raw_status,
        'reject_reason': reject_reason,
        'recommendation': recommendation,
        'x_count': x_count,
        'y_count': y_count,
        'first_km': first_km,
        'last_km': last_km,
        'min_level': min_level,
        'max_level': max_level,
        'calibrated': bool((trace_series or {}).get('calibrated_db')),
        'calibration_note': (trace_series or {}).get('calibration_note') or '',
        'reference_span_km': (trace_series or {}).get('reference_span_km'),
        'raw_last_before_km': (trace_series or {}).get('raw_last_before_km'),
        'raw_last_after_km': (trace_series or {}).get('raw_last_after_km'),
        'distance_scale_factor': (trace_series or {}).get('distance_scale_factor'),
        'distance_scale_status': (trace_series or {}).get('distance_scale_status') or '',
        'distance_scale_note': (trace_series or {}).get('distance_scale_note') or '',
        'distance_scale_ok': (trace_series or {}).get('distance_scale_ok'),
        'source': (trace_series or {}).get('source') or candidate_source,
        'total_fit_rows': total_fit_rows,
        'used_count': used_count,
        'fallback_count': fallback_count,
    }



def _trace_source_type_and_rank(source: str, ext: str = '') -> tuple[str, int, str]:
    """Classify trace source for human-facing diagnostics.

    Important wording rule for Phase 7.1: when the app cannot expose a trace,
    say that the app has not decoded a usable trace yet; do not imply the
    measurement file itself definitely lacks trace data.
    """
    src = (source or '').strip().lower()
    ext = (ext or '').strip().lower()
    if 'sor' in src and ('datapts' in src or 'data' in src or 'trace' in src):
        return 'full_raw_points', 1, 'Raw trace đầy đủ'
    if ext == '.sor' and src:
        return 'full_raw_points', 1, 'Raw trace đầy đủ'
    if 'trc' in src or ext in {'.trc', '.crt'}:
        return 'native_trace', 2, 'Native trace'
    if 'minicurve' in src or 'mini' in src or 'acterna' in src or 'display' in src or 'curve' in src:
        return 'mini_curve', 3, 'Mini/display trace'
    if src:
        return 'vendor_trace', 3, 'Trace vendor-specific'
    return 'not_decoded', 9, 'Chưa giải mã được trace'


def _trace_profile_for_diagnostics(summary: FileSummary, ctx: dict, diag: dict) -> dict:
    """Return explicit Phase 7.1 trace availability/quality labels.

    The profile separates four questions that used to be conflated:
    1) Did the app find a trace-like block?
    2) Can the app draw something useful?
    3) Can the app use it for fit/section loss?
    4) Can the app compute real R²/RMS from raw points?
    """
    trace_series = (ctx or {}).get('raw_trace_series') or None
    ext = Path(getattr(summary, 'file_name', '') or '').suffix.lower()
    source = (trace_series or {}).get('source') or diag.get('candidate_source') or ''
    source_type, rank, source_label = _trace_source_type_and_rank(str(source), ext)
    x_count = int(diag.get('x_count') or 0)
    y_count = int(diag.get('y_count') or 0)
    calibrated = bool(diag.get('calibrated'))
    distance_ok = diag.get('distance_scale_ok')
    candidate_found = bool(diag.get('candidate_found'))
    candidate_note = str(diag.get('candidate_note') or '')
    raw_status = str(diag.get('raw_status') or '')

    availability = 'Chưa giải mã được trace trong app'
    quality = 'Không có dữ liệu trace dùng được'
    graph_mode = 'Event/Section schematic'
    can_draw = False
    can_fit = 'Không'
    can_r2 = 'Không'
    interpretation = ''

    if trace_series and x_count == y_count and x_count >= 2:
        can_draw = True
        if source_type == 'mini_curve':
            availability = 'Có mini/display trace'
            graph_mode = 'Mini trace'
            quality = 'Trung bình - trace rút gọn'
            interpretation = 'Vẽ được đồ thị gần đúng từ mini/display curve; không phải raw sample chi tiết nhất.'
        elif source_type == 'native_trace':
            availability = 'Có native trace'
            graph_mode = 'Native trace'
            quality = 'Khá'
            interpretation = 'Trace lấy từ định dạng native; độ tin cậy phụ thuộc scale km/dB.'
        elif source_type == 'full_raw_points':
            availability = 'Có raw trace/data points'
            graph_mode = 'Raw trace'
            quality = 'Cao' if calibrated else 'Trung bình - chưa chắc scale dB'
            interpretation = 'Trace có thể dùng tốt hơn cho fit nếu scale km/dB hợp lệ.'
        else:
            availability = 'Có trace vendor-specific'
            graph_mode = 'Vendor trace'
            quality = 'Trung bình - cần kiểm chứng parser'
            interpretation = 'Trace đọc từ block riêng; cần đối chiếu với phần mềm gốc nếu dùng làm số chính.'

        if distance_ok is False:
            quality = 'Thấp - scale km chưa tin cậy'
            can_fit = 'Không'
            can_r2 = 'Không'
            interpretation += ' Trục km chưa khớp span nên chỉ nên vẽ/kiểm tra, không dùng fit loss.'
        else:
            if calibrated:
                can_fit = 'Có, nếu section đủ điểm/residual đạt'
            else:
                can_fit = 'Thận trọng - chỉ kiểm tra hình dạng nếu chưa scale dB chắc'
            if x_count >= 8:
                can_r2 = 'Có, nếu fit dùng điểm raw thật'
            elif x_count >= 3:
                can_r2 = 'Hạn chế - chỉ fit cửa sổ lớn, không đủ cho section nhỏ'
            else:
                can_r2 = 'Không'
    else:
        if candidate_found:
            availability = 'Có ứng viên trace nhưng chưa chuẩn hóa được'
            quality = 'Chưa dùng được'
            graph_mode = 'Event/Section schematic'
            interpretation = candidate_note or 'App tìm thấy dấu hiệu trace nhưng chưa tạo được x/y series đáng tin cậy.'
        else:
            availability = 'Chưa tìm thấy block trace đọc được'
            quality = 'Không có dữ liệu trace dùng được'
            graph_mode = 'Event/Section schematic nếu có event/section'
            interpretation = 'Điều này không khẳng định file không có trace; chỉ nói parser hiện tại chưa đọc được trace đủ tin cậy.'

    # If there is no trace but the file still has events/sections, the app can
    # still draw a schematic route later in Trace Viewer.
    events = (ctx or {}).get('events') or []
    if not can_draw and events:
        can_draw = True
        graph_mode = 'Event/Section schematic'
        if not interpretation:
            interpretation = 'Có thể vẽ sơ đồ event/section, nhưng không phải trace OTDR thật.'

    return {
        'trace_availability': availability,
        'trace_source_type': source_type,
        'trace_source_rank': rank,
        'trace_quality_level': quality,
        'graph_mode': graph_mode,
        'can_draw_graph': can_draw,
        'can_use_for_fit': can_fit,
        'can_calculate_r2_rms': can_r2,
        'trace_interpretation_note': interpretation.strip(),
        'source_label': source_label,
    }

def _fr_fill_raw_trace_diagnostics_sheet(ws, summaries: list[FileSummary], contexts: dict[str, dict]) -> None:
    headers = [
        'Tệp', 'Định dạng', 'Parse mode', 'Fiber', 'Wavelength',
        'Trace availability', 'Trace source type', 'Trace source rank', 'Trace quality level',
        'Graph mode', 'Can draw graph?', 'Can use for fit?', 'Can calculate R²/RMS?',
        'Raw block found?', 'Ứng viên nguồn raw', 'Số điểm ứng viên', 'Trạng thái raw trace',
        'Nguồn raw_trace_series', 'x points', 'y points', 'First km', 'Last km',
        'Min level', 'Max level', 'Đã scale dB?', 'Ghi chú scale/hiệu chuẩn',
        'Span tham chiếu km', 'Raw last trước scale', 'Raw last sau scale', 'Hệ số scale km',
        'Trạng thái scale km', 'Ghi chú scale km', 'Scale km OK?',
        'Section fit rows', 'Sections dùng raw-fit/ước lượng', 'Sections fallback/không dùng raw-fit',
        'Lý do không dùng/ghi chú', 'Khuyến nghị', 'Trace interpretation note'
    ]
    _fr_write_header_row(ws, headers)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name) or {}
        diag = _raw_trace_candidate_diagnostics(summary, ctx)
        profile = _trace_profile_for_diagnostics(summary, ctx, diag)
        metadata = ctx.get('metadata') or {}
        values = [
            summary.file_name,
            summary.source_format,
            _to_vi_parse_mode(ctx.get('parse_mode') or ''),
            _display_fiber_label(summary, metadata),
            summary.wavelength_display,
            profile['trace_availability'],
            profile['trace_source_type'],
            profile['trace_source_rank'],
            profile['trace_quality_level'],
            profile['graph_mode'],
            'Có' if profile['can_draw_graph'] else 'Không',
            profile['can_use_for_fit'],
            profile['can_calculate_r2_rms'],
            'Có' if diag['candidate_found'] else 'Không',
            diag['candidate_source'],
            diag['candidate_points'],
            diag['raw_status'],
            diag['source'],
            diag['x_count'],
            diag['y_count'],
            diag['first_km'],
            diag['last_km'],
            diag['min_level'],
            diag['max_level'],
            'Có' if diag['calibrated'] else 'Không',
            diag['calibration_note'],
            diag['reference_span_km'],
            diag['raw_last_before_km'],
            diag['raw_last_after_km'],
            diag['distance_scale_factor'],
            diag['distance_scale_status'],
            diag['distance_scale_note'],
            'Có' if diag['distance_scale_ok'] is True else ('Không' if diag['distance_scale_ok'] is False else ''),
            diag['total_fit_rows'],
            diag['used_count'],
            diag['fallback_count'],
            diag['reject_reason'],
            diag['recommendation'],
            profile['trace_interpretation_note'],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, value)
        if diag['raw_status'].startswith('Không') or diag['used_count'] == 0:
            fill = LOG_WARN_FILL
        elif diag['raw_status'].startswith('Có trace đã scale') and diag['used_count'] > 0:
            fill = GREEN_FILL
        else:
            fill = LOG_INFO_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).fill = fill
        row += 1
    # Phase 7.1: numeric columns moved after new trace-profile fields.
    for col in list(range(19, 25)) + list(range(27, 31)) + list(range(34, 37)):
        for rr in range(2, row):
            if isinstance(ws.cell(rr, col).value, (int, float)):
                ws.cell(rr, col).number_format = '0.0000'
    widths = {
        'A': 34, 'B': 14, 'C': 22, 'D': 18, 'E': 14,
        'F': 30, 'G': 22, 'H': 12, 'I': 28, 'J': 26, 'K': 16, 'L': 30, 'M': 30,
        'N': 16, 'O': 24, 'P': 16, 'Q': 28, 'R': 24, 'S': 12, 'T': 12,
        'U': 12, 'V': 12, 'W': 14, 'X': 14, 'Y': 12, 'Z': 46,
        'AA': 16, 'AB': 18, 'AC': 18, 'AD': 16, 'AE': 24, 'AF': 58, 'AG': 14,
        'AH': 16, 'AI': 22, 'AJ': 26, 'AK': 55, 'AL': 70, 'AM': 78,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=90)


def _fr_fill_run_log_sheet(ws, logs: list[dict], skipped: list[str]) -> None:
    _fr_write_header_row(ws, ['Time', 'Stage', 'Level', 'Message'])
    row = 2
    for item in logs:
        ws.cell(row, 1, item.get('time', ''))
        ws.cell(row, 2, item.get('stage', ''))
        ws.cell(row, 3, item.get('level', ''))
        ws.cell(row, 4, item.get('message', ''))
        fill = LOG_INFO_FILL
        if item.get('level') == 'WARN':
            fill = LOG_WARN_FILL
        elif item.get('level') == 'ERROR':
            fill = LOG_ERROR_FILL
        for c in range(1, 5):
            ws.cell(row, c).fill = fill
        row += 1
    if skipped:
        for msg in skipped:
            ws.cell(row, 1, datetime.now().strftime('%H:%M:%S'))
            ws.cell(row, 2, 'input')
            ws.cell(row, 3, 'WARN')
            ws.cell(row, 4, msg)
            for c in range(1, 5):
                ws.cell(row, c).fill = LOG_WARN_FILL
            row += 1
    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=90)


def build_workbook_from_uploads(
    files: Iterable[tuple[str, bytes]],
    threshold_db: float = 0.5,
    section_threshold_db: Optional[float] = None,
    duration_threshold_s: Optional[float] = None,
    deviation_m: float = 100.0,
    expected_route_km: Optional[float] = None,
    jumper_excluded_m: float = 0.0,
    length_tolerance_km: float = 0.300,
    graph_reach_tolerance_km: Optional[float] = None,
    event_shortfall_tolerance_km: Optional[float] = None,
    overlength_tolerance_km: Optional[float] = None,
    segment_start_km: Optional[float] = None,
    segment_end_km: Optional[float] = None,
    section_export_scope: str = 'all',
    section_merge_tolerance_m: Optional[float] = None,
    section_min_length_km: float = 0.0,
    section_event_source: str = 'all',
    section_boundary_priority: str = 'event',
    section_allow_split: bool = False,
    section_match_tolerance_m: float = 100.0,
    section_measurement_mode: str = 'fit',
    orl_pass_threshold_db: float = 28.0,
    orl_source_mode: str = 'auto',
    orl_allow_lower_bound: bool = True,
    orl_lower_bound_status: str = 'Unknown',
    orl_physical_mode: str = 'disabled',
    output_mode: str = 'fastreporter',
) -> BytesIO:
    t0 = time.perf_counter()
    logs: list[dict] = []
    _fr_log(logs, 'run', 'INFO', 'Bắt đầu dựng workbook.')
    files = list(files)
    summaries, skipped, _file_payload_map, contexts = _fr_build_context(
        files,
        threshold_db=threshold_db,
        expected_route_km=expected_route_km,
        jumper_excluded_m=jumper_excluded_m,
        length_tolerance_km=length_tolerance_km,
        graph_reach_tolerance_km=graph_reach_tolerance_km,
        event_shortfall_tolerance_km=event_shortfall_tolerance_km,
        overlength_tolerance_km=overlength_tolerance_km,
        segment_start_km=segment_start_km,
        segment_end_km=segment_end_km,
        orl_pass_threshold_db=orl_pass_threshold_db,
        orl_source_mode=orl_source_mode,
        orl_allow_lower_bound=orl_allow_lower_bound,
        orl_lower_bound_status=orl_lower_bound_status,
        orl_physical_mode=orl_physical_mode,
        logs=logs,
    )
    if not summaries:
        message = ' | '.join(skipped) if skipped else 'Không có file hợp lệ nào được tải lên.'
        raise ValueError(message)

    event_defs = _fr_build_common_event_defs(contexts, deviation_m=deviation_m)
    _fr_log(logs, 'run', 'INFO', 'Phase 2: dùng chung event defs và lõi dữ liệu chuẩn hóa cho cả STV và FastReporter.')
    sections = _fr_build_common_sections(event_defs, summaries, contexts, deviation_m=deviation_m, threshold_db=threshold_db, section_merge_tolerance_m=section_merge_tolerance_m, section_min_length_km=section_min_length_km, section_event_source=section_event_source, section_boundary_priority=section_boundary_priority, section_allow_split=section_allow_split)
    if str(section_export_scope).lower() == 'selected_range' and segment_start_km is not None and segment_end_km is not None:
        sections = _fr_clip_sections_to_range(sections, segment_start_km, segment_end_km)
        try:
            _fr_log(logs, 'run', 'INFO', f'Giới hạn xuất Sections theo đoạn người dùng chọn: {min(float(segment_start_km), float(segment_end_km)):.3f} - {max(float(segment_start_km), float(segment_end_km)):.3f} km | Số section sau khi cắt: {len(sections)}')
        except Exception:
            pass
    elif str(section_export_scope).lower() == 'selected_range':
        _fr_log(logs, 'run', 'WARN', 'Đã chọn chỉ xuất section theo đoạn đã chọn nhưng chưa nhập đủ Đoạn bắt đầu/Đoạn kết thúc. App giữ nguyên toàn bộ section của tuyến.')

    if str(output_mode).lower() == 'stv':
        _fr_log(logs, 'run', 'INFO', 'Xuất theo định dạng STV tinh chỉnh dùng chung lõi tính toán hiện tại.')
        return _stv_build_workbook(
            summaries,
            contexts,
            event_defs,
            deviation_m=deviation_m,
            threshold_db=threshold_db,
            expected_route_km=expected_route_km,
            jumper_excluded_m=jumper_excluded_m,
            graph_reach_tolerance_km=graph_reach_tolerance_km if graph_reach_tolerance_km is not None else length_tolerance_km,
            event_shortfall_tolerance_km=event_shortfall_tolerance_km if event_shortfall_tolerance_km is not None else length_tolerance_km,
            skipped=skipped,
            logs=logs,
        )

    template_path = Path(__file__).with_name('2.xlsx')
    if not template_path.exists():
        raise FileNotFoundError('Không tìm thấy template Excel (2.xlsx) trong thư mục chương trình.')

    wb = load_workbook(template_path)
    for warn in _fr_validate_template_or_raise(template_path, wb):
        _fr_log(logs, 'template', 'WARN', warn)
    _fr_log(logs, 'template', 'INFO', f'Template OK: {template_path.name}')

    # General Information (fill directly from per-file metadata, giữ nguyên layout template)
    ws_info = wb['General Information']
    _fr_fill_general_information_template(ws_info, summaries, contexts)

    # Sections first to get per-file section pairs
    section_pairs_by_file = _fr_fill_sections(wb['Sections'], summaries, contexts, sections, threshold_db, section_match_tolerance_m=section_match_tolerance_m, section_measurement_mode=section_measurement_mode, section_threshold_db=section_threshold_db)

    # Link results
    _fr_fill_link_results(wb['Link Results'], summaries, contexts, section_pairs_by_file, duration_threshold_s=duration_threshold_s)

    # Events
    _fr_fill_events(wb['Events'], summaries, contexts, event_defs, deviation_m, threshold_db)

    # Export all control parameters and the calculations they drive without changing core algorithms.
    for extra_name in ['App Parameters', 'Route Analysis', 'Segment Analysis', 'Segment Events', 'Section Fit Quality', 'Raw Trace Diagnostics', 'ORL Analysis', 'Parser Diagnostics', 'Vendor Compatibility', 'Strict Validation', 'Run Log', 'Core Metrics', 'Output Rules']:
        _fr_safe_sheet_remove(wb, extra_name)
    ws_params = wb.create_sheet('App Parameters')
    _fr_fill_app_parameters_sheet(
        ws_params,
        threshold_db=threshold_db,
        section_threshold_db=section_threshold_db,
        duration_threshold_s=duration_threshold_s,
        deviation_m=deviation_m,
        expected_route_km=expected_route_km,
        jumper_excluded_m=jumper_excluded_m,
        length_tolerance_km=length_tolerance_km,
        graph_reach_tolerance_km=graph_reach_tolerance_km,
        event_shortfall_tolerance_km=event_shortfall_tolerance_km,
        overlength_tolerance_km=overlength_tolerance_km,
        segment_start_km=segment_start_km,
        segment_end_km=segment_end_km,
        section_export_scope=section_export_scope,
        section_merge_tolerance_m=section_merge_tolerance_m,
        section_min_length_km=section_min_length_km,
        section_event_source=section_event_source,
        section_boundary_priority=section_boundary_priority,
        section_allow_split=section_allow_split,
        section_match_tolerance_m=section_match_tolerance_m,
        section_measurement_mode=section_measurement_mode,
        orl_pass_threshold_db=orl_pass_threshold_db,
        orl_source_mode=orl_source_mode,
        orl_allow_lower_bound=orl_allow_lower_bound,
        orl_lower_bound_status=orl_lower_bound_status,
        orl_physical_mode=orl_physical_mode,
    )
    ws_route = wb.create_sheet('Route Analysis')
    _fr_fill_route_analysis_sheet(ws_route, summaries, contexts)
    ws_segment = wb.create_sheet('Segment Analysis')
    _fr_fill_segment_analysis_sheet(ws_segment, summaries, contexts)
    ws_segment_events = wb.create_sheet('Segment Events')
    _fr_fill_segment_events_sheet(ws_segment_events, summaries, contexts)
    ws_fit = wb.create_sheet('Section Fit Quality')
    _fr_fill_section_fit_quality_sheet(ws_fit, summaries, contexts)
    ws_raw_diag = wb.create_sheet('Raw Trace Diagnostics')
    _fr_fill_raw_trace_diagnostics_sheet(ws_raw_diag, summaries, contexts)
    ws_orl = wb.create_sheet('ORL Analysis')
    _fr_fill_orl_analysis_sheet(ws_orl, summaries, contexts)
    ws_parser = wb.create_sheet('Parser Diagnostics')
    _fr_fill_parser_diagnostics_sheet(ws_parser, summaries, contexts, skipped)
    ws_vendor = wb.create_sheet('Vendor Compatibility')
    _fr_fill_vendor_compatibility_matrix_sheet(ws_vendor, summaries, contexts, skipped)
    ws_strict = wb.create_sheet('Strict Validation')
    _fr_fill_strict_validation_sheet(ws_strict, summaries, contexts, expected_route_km=expected_route_km, length_tolerance_km=length_tolerance_km if 'length_tolerance_km' in locals() else 0.300, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km, duration_threshold_s=duration_threshold_s if 'duration_threshold_s' in locals() else None, skipped=skipped)
    ws_core = wb.create_sheet('Core Metrics')
    _fr_fill_core_metrics_sheet(ws_core, summaries, contexts, threshold_db=threshold_db, section_pairs_by_file=section_pairs_by_file, duration_threshold_s=duration_threshold_s)
    ws_rules = wb.create_sheet('Output Rules')
    _fr_fill_output_rules_sheet(ws_rules, threshold_db=threshold_db, deviation_m=deviation_m, output_mode='fastreporter', section_export_scope=section_export_scope, section_measurement_mode=section_measurement_mode, section_event_source=section_event_source, section_boundary_priority=section_boundary_priority, expected_route_km=expected_route_km, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km)
    ws_log = wb.create_sheet('Run Log')
    _fr_log(logs, 'run', 'INFO', f'Tổng file hợp lệ: {len(summaries)} | Event defs: {len(event_defs)} | Sections: {len(sections)}')
    _fr_log(logs, 'run', 'INFO', f'Thời gian dựng workbook: {time.perf_counter() - t0:.2f}s')
    _fr_fill_run_log_sheet(ws_log, logs, skipped)
    _fr_apply_workbook_polish(wb)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===== Phase 3 safe overrides =====
def _fr_fill_output_rules_sheet(ws, *, threshold_db: float, deviation_m: float, output_mode: str, section_export_scope: str, section_measurement_mode: str, section_event_source: str, section_boundary_priority: str, expected_route_km: Optional[float], graph_reach_tolerance_km: Optional[float], event_shortfall_tolerance_km: Optional[float], segment_start_km: Optional[float] = None, segment_end_km: Optional[float] = None, section_threshold_db: Optional[float] = None, duration_threshold_s: Optional[float] = None) -> None:
    rules = [
        ('Giai đoạn', 'Phase 7.5 - Strict Validation Layer: thêm lớp kiểm tra sanity/production-readiness; chỉ ghi cảnh báo và quyền dùng dữ liệu, không thay đổi logic tính toán.'),
        ('Lõi dữ liệu', 'STV và FastReporter cùng dùng một context chuẩn hóa từ parser; không parse lại riêng ở từng đầu ra.'),
        ('Ngưỡng event', f'Ngưỡng event hiện tại = {threshold_db} dB. FastReporter Events/STV chỉ hiện loss đạt ngưỡng |loss| >= ngưỡng; riêng hàng Minimum/Maximum/Average/Occurrences của Events tính trên toàn bộ event đã parse trong cùng cụm, kể cả event bị ẩn dưới ngưỡng.'),
        ('Gom cụm event', f'Deviation = {deviation_m} m. STV và FastReporter cùng bám vào event definitions sau khi gom cụm.'),
        ('Ngưỡng section', f'Ngưỡng Section Loss để tô đỏ = {section_threshold_db if section_threshold_db is not None else "chưa đặt"} dB. Mã ID S+E hiển thị một lần ở tiêu đề mỗi section = Start km + End km; đây chỉ là mã nhận diện layout, không phải chiều dài tuyến và không tạo cột riêng trong dữ liệu core.'),
        ('Duration', f'Ngưỡng duration = {duration_threshold_s if duration_threshold_s is not None else "chưa đặt"} giây. Nếu core có duration thấp hơn ngưỡng thì đánh dấu Fail trong Core Metrics/Link Results.'),
        ('Strict Validation', 'Sheet Strict Validation kiểm tra parser, metadata, route length, attenuation sanity, event coverage, trace scale, fit readiness và ORL readiness. Sheet này chỉ cảnh báo/khóa quyền dùng để kết luận trên giấy, tuyệt đối không sửa số liệu đã tính.'),
        ('Section', f'Section export scope = {section_export_scope}; mode tính section = {section_measurement_mode}; nguồn event dựng section = {section_event_source}; ưu tiên = {section_boundary_priority}. Với mode fit, app ưu tiên raw-trace linear fit, xuất R²/RMS residual và fallback về event/slope nếu fit chưa đủ tin cậy.'),
        ('ORL', 'ORL Analysis phân tách measured ORL / metadata lower-bound / not available. Chỉ measured ORL mới có Use for Judgment = Yes. Lower-bound dạng <xx.xx luôn được ghi rõ là tham khảo, không phải span ORL thật.'),
        ('ORL vật lý từ trace', 'Mặc định tắt. Khi bật diagnostic/experimental, app chỉ kiểm tra điều kiện hiệu chuẩn; nếu thiếu backscatter/launch power/reference calibration thì không tạo số ORL giả.'),
    ]
    if str(section_export_scope).lower() == 'selected_range' and segment_start_km is not None and segment_end_km is not None:
        a = min(float(segment_start_km), float(segment_end_km))
        b = max(float(segment_start_km), float(segment_end_km))
        rules.append(('Phạm vi section', f'Chỉ xuất section giao với đoạn người dùng chọn: {a:.3f} - {b:.3f} km; section bị cắt đúng biên start/end khi xuất.'))
    elif str(section_export_scope).lower() == 'selected_range':
        rules.append(('Phạm vi section', 'Đang chọn chỉ xuất section theo đoạn đã chọn nhưng chưa nhập đủ Đoạn bắt đầu / Đoạn kết thúc. App giữ nguyên section toàn tuyến.'))
    else:
        rules.append(('Phạm vi section', 'Đang xuất section toàn tuyến.'))
    rules.extend([
        ('Route check', f'Chiều dài tuyến chuẩn = {expected_route_km}; sai số reach = {graph_reach_tolerance_km}; hụt event cho phép = {event_shortfall_tolerance_km}.'),
        ('STV', 'Bảng sự kiện giữ format vận hành: các cột summary cố định, giữ Đầu tuyến/Cuối tuyến, chỉ hiện loss đạt ngưỡng.'),
        ('FastReporter', 'Events sheet tập trung vào event đỏ; Sections có thể xuất toàn tuyến hoặc theo đoạn đã chọn; Link Results và Core Metrics lấy từ lõi chuẩn hóa.'),
    ])
    ws.cell(1, 1, 'Mục')
    ws.cell(1, 2, 'Rule đang áp dụng')
    row = 2
    for label, value in rules:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        row += 1
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 120


def _fr_fill_events(ws, summaries: list[FileSummary], contexts: dict[str, dict], event_defs: list[dict], deviation_m: float, threshold_db: float) -> None:
    summary_start = _fr_expand_pair_table(
        ws,
        data_start=5,
        template_rows=(43, 44),
        insert_before=45,
        count=len(summaries),
        reserved_pairs=20,
    )

    # Build representative event map per fiber first, then keep only event columns
    # that contain at least one "red" loss (|loss| >= threshold). This keeps the
    # Events sheet focused on the events the user actually cares about, without
    # changing the parsing / clustering logic upstream.
    per_summary_cluster: list[dict[int, EventRow]] = []
    visible_event_indices: list[int] = []
    visible_set: set[int] = set()
    thr = float(threshold_db)

    for summary in summaries:
        ctx = contexts.get(summary.file_name, {})
        rows = ctx.get('events', [])
        cluster_map: dict[int, EventRow] = {}
        for ev in rows:
            di = _fr_assign_event_to_def(ev, event_defs, deviation_m)
            if di is None:
                continue
            cur = cluster_map.get(di)
            if cur is None:
                cluster_map[di] = ev
            else:
                rep = float(event_defs[di - 1]['distance_km'])
                if abs(float(ev.distance_km or 0) - rep) < abs(float(cur.distance_km or 0) - rep):
                    cluster_map[di] = ev
        per_summary_cluster.append(cluster_map)
        for di, ev in cluster_map.items():
            loss = ev.loss_db
            if isinstance(loss, (int, float)) and abs(float(loss)) + 1e-12 >= thr:
                if di not in visible_set:
                    visible_set.add(di)
                    visible_event_indices.append(di)

    visible_event_indices.sort()
    _fr_ensure_events_capacity(ws, len(visible_event_indices))
    max_events = max(_fr_count_real_pair_blocks(ws, header_merge_rows=(1, 2)), len(visible_event_indices))

    # Header
    for out_i in range(1, max_events + 1):
        col = 3 + (out_i - 1) * 2
        ws.cell(1, col).value = f'Event {out_i}'
        if out_i <= len(visible_event_indices):
            src_i = visible_event_indices[out_i - 1]
            item = event_defs[src_i - 1]
            ws.cell(2, col).value = item.get('label')
            ws.cell(3, col).value = round(float(item.get('distance_km') or 0.0), 4)
            ws.cell(3, col + 1).value = 'km'
        else:
            ws.cell(2, col).value = None
            ws.cell(3, col).value = None
            ws.cell(3, col + 1).value = None

    # Clear data and summary area
    _fr_clear_cells(ws, 5, summary_start + 3, 1, ws.max_column)

    rendered_count = len(summaries)

    # Hai lớp dữ liệu được tách riêng:
    #   - displayed_*: chỉ để ghi bảng chính, vẫn giữ rule chỉ hiện event vượt ngưỡng.
    #   - stats_*: dùng cho Minimum / Maximum / Average / Occurrences, lấy toàn bộ
    #     event đại diện trong cùng cụm, kể cả event bị ẩn vì loss dưới ngưỡng.
    # Như vậy người vận hành nhìn bảng vẫn gọn, nhưng hàng thống kê không bị lệch
    # do chỉ tính các ô đang hiện màu đỏ.
    displayed_event_cols_loss = [[] for _ in range(max_events)]
    displayed_event_cols_refl = [[] for _ in range(max_events)]
    stats_event_cols_loss = [[] for _ in range(max_events)]
    stats_event_cols_refl = [[] for _ in range(max_events)]

    for idx, summary in enumerate(summaries):
        data_row = 5 + idx * 2
        status_row = data_row + 1
        cluster_map = per_summary_cluster[idx]

        m = re.search(r'(\d{3,4})', summary.wavelength_display or '')
        ws.cell(data_row, 1).value = _display_fiber_label(summary, contexts.get(summary.file_name, {}).get('metadata', {}))
        ws.cell(data_row, 2).value = int(m.group(1)) if m else None

        for out_i, src_i in enumerate(visible_event_indices, start=1):
            col = 3 + (out_i - 1) * 2
            ev = cluster_map.get(src_i)
            if not ev:
                continue

            loss = ev.loss_db
            raw_refl = ev.reflectance_db
            refl = _fr_reportable_reflectance_db(raw_refl)

            # Thống kê lấy toàn bộ event thuộc cột/cụm này, không phụ thuộc việc
            # ô đó có được hiển thị trong bảng chính hay bị ẩn dưới ngưỡng.
            if isinstance(loss, (int, float)):
                stats_event_cols_loss[out_i - 1].append(float(loss))
            if isinstance(refl, (int, float)):
                stats_event_cols_refl[out_i - 1].append(float(refl))

            # Bảng chính vẫn chỉ hiện event vượt ngưỡng để không làm rối báo cáo.
            if not isinstance(loss, (int, float)) or abs(float(loss)) + 1e-12 < thr:
                continue
            ws.cell(data_row, col).value = loss
            ws.cell(data_row, col + 1).value = refl
            ws.cell(status_row, col).value = _fr_event_status(loss, threshold_db)
            ws.cell(status_row, col + 1).value = _fr_reflectance_status(raw_refl)
            ws.cell(data_row, col).number_format = '0.000'
            ws.cell(data_row, col).fill = RED_FILL
            ws.cell(status_row, col).fill = RED_FILL
            displayed_event_cols_loss[out_i - 1].append(float(loss))
            if isinstance(refl, (int, float)):
                displayed_event_cols_refl[out_i - 1].append(float(refl))

    for offs, label in enumerate(['Minimum', 'Maximum', 'Average', 'Occurences']):
        ws.cell(summary_start + offs, 1).value = label

    def _stats(vals, denom):
        if not vals:
            return None, None, None, f'0/{denom}'
        return round(min(vals), 3), round(max(vals), 3), round(sum(vals) / len(vals), 3), f'{len(vals)}/{denom}'

    for out_i in range(1, max_events + 1):
        col = 3 + (out_i - 1) * 2
        loss_vals = stats_event_cols_loss[out_i - 1]
        refl_vals = stats_event_cols_refl[out_i - 1]
        loss_min, loss_max, loss_avg, loss_occ = _stats(loss_vals, rendered_count)
        refl_min, refl_max, refl_avg, refl_occ = _stats(refl_vals, rendered_count)
        ws.cell(summary_start, col).value = loss_min
        ws.cell(summary_start, col + 1).value = refl_min
        ws.cell(summary_start + 1, col).value = loss_max
        ws.cell(summary_start + 1, col + 1).value = refl_max
        ws.cell(summary_start + 2, col).value = loss_avg
        ws.cell(summary_start + 2, col + 1).value = refl_avg
        ws.cell(summary_start + 3, col).value = loss_occ
        for rr in [summary_start, summary_start + 1, summary_start + 2]:
            if isinstance(ws.cell(rr, col).value, (int, float)):
                ws.cell(rr, col).number_format = '0.000'
            if isinstance(ws.cell(rr, col + 1).value, (int, float)):
                ws.cell(rr, col + 1).number_format = '0.000'


def _stv_display_loss_for_row(row: EventRow, threshold_db: float = 0.5, is_terminal_row: bool = False) -> Optional[float]:
    """Phase 3 override: STV only shows point losses that hit the threshold; negatives are kept if magnitude is large enough."""
    if row.loss_db is None:
        return None
    try:
        loss = float(row.loss_db)
    except Exception:
        return None
    label = (row.label or '').strip()
    if row.event_type == 'First Connector':
        return None
    if is_terminal_row:
        return None
    if len(label) >= 2 and label[1] == 'E':
        return None
    if abs(loss) + 1e-12 < float(threshold_db):
        return None
    return round(loss, 3)



# ---------------------------------------------------------------------------
# Phase 7.5 - Strict Validation Layer (diagnostic-only)
# ---------------------------------------------------------------------------
def _fr_float_or_none(value):
    try:
        if value is None:
            return None
        f = float(value)
        if not math.isfinite(f):
            return None
        return f
    except Exception:
        return None


def _fr_validation_level_rank(level: str) -> int:
    level = str(level or '').upper()
    if level == 'FAIL':
        return 3
    if level == 'WARN':
        return 2
    if level in {'INFO', 'CHECK'}:
        return 1
    return 0


def _fr_validation_fill(level: str):
    level = str(level or '').upper()
    if level == 'OK':
        return GREEN_FILL
    if level == 'FAIL':
        return LOG_ERROR_FILL
    if level == 'WARN':
        return LOG_WARN_FILL
    return LOG_INFO_FILL


def _fr_make_validation_row(summary: FileSummary, ctx: dict, category: str, rule_code: str, rule_name: str, level: str, measured_value='', expected_rule='', impact='', action='', evidence='') -> dict:
    meta = (ctx or {}).get('metadata') or {}
    diag = _raw_trace_candidate_diagnostics(summary, ctx)
    profile = _trace_profile_for_diagnostics(summary, ctx, diag)
    event_metrics = _fr_event_health_metrics(summary, ctx)
    final_status, parser_level, use_for_judgment, recommendation = _fr_parser_final_status(summary, ctx, diag, profile, event_metrics)
    return {
        'file': getattr(summary, 'file_name', '') or '',
        'fiber': _display_fiber_label(summary, meta),
        'wavelength': getattr(summary, 'wavelength_display', '') or '',
        'vendor': _fr_vendor_family_from_context(summary, ctx),
        'parser_family': _to_vi_parser_family(getattr(summary, 'parse_family', '') or (ctx or {}).get('parser_family') or ''),
        'final_parser_status': final_status,
        'parser_level': parser_level,
        'use_for_judgment': use_for_judgment,
        'category': category,
        'rule_code': rule_code,
        'rule_name': rule_name,
        'level': level,
        'measured_value': measured_value,
        'expected_rule': expected_rule,
        'impact': impact,
        'action': action,
        'evidence': evidence,
        'recommendation': recommendation,
    }


def _fr_strict_validation_rows_for_file(
    summary: FileSummary,
    ctx: dict,
    *,
    expected_route_km: Optional[float] = None,
    length_tolerance_km: float = 0.300,
    graph_reach_tolerance_km: Optional[float] = None,
    event_shortfall_tolerance_km: Optional[float] = None,
    duration_threshold_s: Optional[float] = None,
) -> list[dict]:
    """Build strict validation rows without changing any calculation.

    Phase 7.5 is a guardrail/reporting layer only.  It classifies whether the
    already-parsed data is safe for production judgement; it never edits parser
    output, section values, event loss, ORL, trace scale, or Excel formulas.
    """
    rows: list[dict] = []
    ctx = ctx or {}
    meta = ctx.get('metadata') or {}
    diag = _raw_trace_candidate_diagnostics(summary, ctx)
    profile = _trace_profile_for_diagnostics(summary, ctx, diag)
    event_metrics = _fr_event_health_metrics(summary, ctx)
    final_status, parser_level, use_for_judgment, recommendation = _fr_parser_final_status(summary, ctx, diag, profile, event_metrics)
    ga = ctx.get('graph_assessment')
    orl = ctx.get('orl_analysis')

    def add(category, code, name, level, measured='', expected='', impact='', action='', evidence=''):
        rows.append(_fr_make_validation_row(summary, ctx, category, code, name, level, measured, expected, impact, action, evidence))

    # Parser / vendor readiness
    if parser_level == 'FAIL' or str(final_status).startswith('FAIL'):
        add('Parser', 'SV-PARSER-001', 'Parser final status', 'FAIL', final_status, 'Không được FAIL', 'Không đủ dữ liệu để kết luận vận hành.', 'Mở bằng phần mềm hãng hoặc bổ sung parser trước khi dùng.', recommendation)
    elif final_status == 'PASS_READ_FULL':
        add('Parser', 'SV-PARSER-001', 'Parser final status', 'OK', final_status, 'PASS_READ_FULL hoặc PASS_READ_PARTIAL có điều kiện', 'Đủ điều kiện parser để dùng theo rule hiện tại.', 'Có thể dùng, vẫn cần kiểm tra rule WARN nếu có.', recommendation)
    else:
        add('Parser', 'SV-PARSER-001', 'Parser final status', 'WARN', final_status, 'PASS_READ_FULL là tốt nhất; PARTIAL/GRAPH_ONLY phải có ghi chú', 'Chỉ dùng có điều kiện hoặc tham khảo tùy cột Dùng để kết luận.', 'Không chốt kết luận sâu nếu còn rule FAIL/WARN trọng yếu.', recommendation)

    conf = str(getattr(summary, 'parse_family_confidence', '') or ctx.get('parser_family_confidence') or '').lower()
    if conf in {'high', 'medium'}:
        add('Parser', 'SV-PARSER-002', 'Parser family confidence', 'OK', conf, 'high/medium', 'Nguồn parser đủ nhận diện.', 'Không cần xử lý thêm.', getattr(summary, 'parse_family_reason', '') or ctx.get('parser_family_reason') or '')
    else:
        add('Parser', 'SV-PARSER-002', 'Parser family confidence', 'WARN', conf or 'unknown', 'high/medium', 'Vendor/parser chưa khóa chắc, nguy cơ thiếu block riêng.', 'Đối chiếu bằng phần mềm hãng hoặc bổ sung golden test.', getattr(summary, 'parse_family_reason', '') or ctx.get('parser_family_reason') or '')

    # Metadata checks
    length = _fr_float_or_none(getattr(summary, 'length_km', None))
    if length is not None and length > 0:
        add('Metadata', 'SV-META-001', 'Span length hợp lệ', 'OK', round(length, 6), '> 0 km', 'Có chiều dài tuyến để kiểm tra section/event.', 'Không cần xử lý thêm.')
    else:
        add('Metadata', 'SV-META-001', 'Span length hợp lệ', 'FAIL', getattr(summary, 'length_km', None), '> 0 km', 'Không có span length đáng tin cậy.', 'Không dùng để kết luận route/section; kiểm tra file gốc.')

    wl_text = str(getattr(summary, 'wavelength_display', '') or '')
    wl_match = re.search(r'(\d{3,4})', wl_text)
    wl_num = int(wl_match.group(1)) if wl_match else None
    known_wavelengths = {850, 1300, 1310, 1490, 1550, 1625, 1650}
    if wl_num in known_wavelengths:
        add('Metadata', 'SV-META-002', 'Wavelength hợp lệ', 'OK', wl_num, '850/1300/1310/1490/1550/1625/1650 nm', 'Bước sóng nhận diện được.', 'Không cần xử lý thêm.')
    elif wl_num is not None:
        add('Metadata', 'SV-META-002', 'Wavelength hợp lệ', 'WARN', wl_num, 'Wavelength OTDR phổ biến', 'Wavelength lạ, có thể do vendor/private block hoặc parser.', 'Kiểm tra lại metadata bằng phần mềm hãng.')
    else:
        add('Metadata', 'SV-META-002', 'Wavelength hợp lệ', 'WARN', wl_text, 'Có wavelength', 'Thiếu wavelength, khó so sánh ngưỡng attenuation theo chuẩn.', 'Kiểm tra tên file/metadata.')

    duration = _fr_float_or_none(meta.get('duration_s'))
    if duration_threshold_s is not None:
        th = _fr_float_or_none(duration_threshold_s)
        if duration is None:
            add('Metadata', 'SV-META-003', 'Duration theo ngưỡng', 'WARN', 'Unknown', f'>= {th:g} s', 'Không có duration nên không thể áp rule duration chắc chắn.', 'Kiểm tra file gốc hoặc phần mềm hãng.')
        elif th is not None and duration < th:
            add('Metadata', 'SV-META-003', 'Duration theo ngưỡng', 'FAIL', round(duration, 3), f'>= {th:g} s', 'Core không đạt yêu cầu thời gian đo theo ngưỡng vận hành.', 'Đo lại hoặc không dùng core này để nghiệm thu.')
        else:
            add('Metadata', 'SV-META-003', 'Duration theo ngưỡng', 'OK', round(duration, 3), f'>= {th:g} s', 'Duration đạt rule người dùng nhập.', 'Không cần xử lý thêm.')
    else:
        if duration is None:
            add('Metadata', 'SV-META-003', 'Duration có đọc được?', 'INFO', 'Unknown', 'Không bắt buộc nếu chưa đặt ngưỡng', 'Không ảnh hưởng nếu chưa dùng duration để đánh giá.', 'Có thể đặt ngưỡng duration khi vận hành.')
        else:
            add('Metadata', 'SV-META-003', 'Duration có đọc được?', 'OK', round(duration, 3), '> 0 s', 'Duration đọc được để tham chiếu.', 'Không cần xử lý thêm.')

    # Route/length sanity
    expected = _fr_float_or_none(expected_route_km)
    tol = _fr_float_or_none(length_tolerance_km)
    if expected is not None and expected > 0 and length is not None:
        diff = abs(length - expected)
        if tol is not None and diff <= tol:
            add('Route', 'SV-ROUTE-001', 'Span length so với tuyến chuẩn', 'OK', f'{length:.3f} km; lệch {diff:.3f} km', f'≤ {tol:.3f} km', 'Chiều dài đo khớp tuyến chuẩn.', 'Không cần xử lý thêm.')
        else:
            add('Route', 'SV-ROUTE-001', 'Span length so với tuyến chuẩn', 'WARN', f'{length:.3f} km; lệch {diff:.3f} km', f'≤ {tol:.3f} km', 'Chiều dài đo lệch tuyến chuẩn; có thể thiếu/thừa tuyến hoặc scale sai.', 'Kiểm tra route/fiber end/đấu nhảy/đo lại nếu cần.')
    else:
        add('Route', 'SV-ROUTE-001', 'Span length so với tuyến chuẩn', 'INFO', 'Không đủ dữ liệu', 'Cần chiều dài tuyến chuẩn để so sánh', 'Không kết luận đủ/thiếu tuyến bằng rule này.', 'Nhập Chiều dài tuyến chuẩn nếu cần kiểm soát route.')

    att = _fr_float_or_none(getattr(summary, 'attenuation_dbkm', None))
    if att is None:
        add('Route', 'SV-ROUTE-002', 'Attenuation sanity', 'WARN', 'Unknown', 'Có attenuation', 'Thiếu suy hao trung bình để kiểm tra vật lý.', 'Kiểm tra span loss/length trong file.')
    elif att < 0 or att > 2.0:
        add('Route', 'SV-ROUTE-002', 'Attenuation sanity', 'FAIL', f'{att:.3f} dB/km', '0–2.0 dB/km hard sanity', 'Số liệu attenuation bất thường, nguy cơ sai đơn vị/scale/parser.', 'Không dùng để kết luận cho tới khi đối chiếu phần mềm hãng.')
    elif att < 0.05 or att > 0.80:
        add('Route', 'SV-ROUTE-002', 'Attenuation sanity', 'WARN', f'{att:.3f} dB/km', '0.05–0.80 dB/km operational check', 'Attenuation ngoài dải thường gặp, cần kiểm tra tuyến/tham số.', 'Đối chiếu với file gốc và điều kiện tuyến.')
    else:
        add('Route', 'SV-ROUTE-002', 'Attenuation sanity', 'OK', f'{att:.3f} dB/km', '0.05–0.80 dB/km operational check', 'Attenuation trong dải kiểm tra thông thường.', 'Không cần xử lý thêm.')

    if ga is not None and getattr(ga, 'expected_route_km', None) not in (None, 0):
        verdict = getattr(ga, 'verdict', '') or ''
        level = 'OK' if str(verdict).upper() in {'PASS', 'ĐẠT'} else 'WARN'
        add('Route', 'SV-ROUTE-003', 'Graph/Event reach tuyến chuẩn', level, f'{verdict}; {getattr(ga, "reason", "")}', f'sai số graph={graph_reach_tolerance_km}; hụt event={event_shortfall_tolerance_km}', 'Kiểm tra tuyến đủ/hụt theo graph/event.', 'Nếu WARN, xem sheet Route Analysis/Kiểm tra đồ thị.')

    # Event health
    event_count = int(event_metrics.get('event_count') or 0)
    position_count = int(event_metrics.get('position_count') or 0)
    loss_count = int(event_metrics.get('loss_count') or 0)
    if event_count > 0:
        add('Event', 'SV-EVENT-001', 'Có event table', 'OK', event_count, '> 0 event', 'Có cơ sở phân tích event/section.', 'Không cần xử lý thêm.')
    else:
        add('Event', 'SV-EVENT-001', 'Có event table', 'FAIL', event_count, '> 0 event', 'Không có event table để kết luận event/section.', 'Mở phần mềm hãng hoặc bổ sung parser.')

    if event_count > 0 and position_count == event_count:
        add('Event', 'SV-EVENT-002', 'Event position coverage', 'OK', f'{position_count}/{event_count}', '100%', 'Tất cả event có vị trí.', 'Không cần xử lý thêm.')
    elif event_count > 0 and position_count > 0:
        add('Event', 'SV-EVENT-002', 'Event position coverage', 'WARN', f'{position_count}/{event_count}', '100%', 'Một số event thiếu vị trí, có thể ảnh hưởng gom cụm/section.', 'Kiểm tra event thiếu position.')
    else:
        add('Event', 'SV-EVENT-002', 'Event position coverage', 'FAIL', f'{position_count}/{event_count}', '100%', 'Không đủ vị trí event.', 'Không dùng event/section để kết luận.')

    if event_count > 0:
        cov = loss_count / event_count if event_count else 0
        if cov >= 0.90:
            level = 'OK'
            impact = 'Event loss coverage tốt.'
            action = 'Không cần xử lý thêm.'
        elif cov >= 0.70:
            level = 'WARN'
            impact = 'Event loss coverage trung bình; vẫn cần kiểm tra event thiếu loss.'
            action = 'Không chốt event thiếu loss; xem Parser Diagnostics.'
        elif loss_count > 0:
            level = 'WARN'
            impact = 'Event loss coverage thấp; nguy cơ thiếu block vendor/private.'
            action = 'Đối chiếu bằng phần mềm hãng trước khi kết luận.'
        else:
            level = 'FAIL'
            impact = 'Không có event loss để kết luận suy hao event.'
            action = 'Bổ sung parser event loss hoặc kiểm tra file gốc.'
        add('Event', 'SV-EVENT-003', 'Event loss coverage', level, f'{loss_count}/{event_count} ({cov:.1%})', '>=90% OK; 70–90% WARN; <70% WARN/FAIL', impact, action)

    after_end = int(event_metrics.get('after_end_count') or 0)
    invalid_pos = int(event_metrics.get('invalid_position_count') or 0)
    if after_end == 0 and invalid_pos == 0:
        add('Event', 'SV-EVENT-004', 'Event position sanity', 'OK', f'after_end={after_end}; invalid={invalid_pos}', 'Không có event âm/sau tuyến bất thường', 'Event position không phát hiện lỗi hard sanity.', 'Không cần xử lý thêm.')
    else:
        add('Event', 'SV-EVENT-004', 'Event position sanity', 'WARN', f'after_end={after_end}; invalid={invalid_pos}', '0', 'Có event vị trí bất thường, có thể là ghost/event sau cuối tuyến.', 'Không dùng event bất thường làm section chính nếu chưa kiểm tra.')

    gainer = int(event_metrics.get('gainer_count') or 0)
    if gainer > 0:
        add('Event', 'SV-EVENT-005', 'Gainer/loss âm', 'INFO', gainer, 'Cho phép nhưng phải ghi nhận', 'Gainer có thể hợp lý trong OTDR nhưng cần biết khi đánh giá event đỏ.', 'Không tô đỏ theo abs nếu rule không yêu cầu; kiểm tra event liên quan.')

    # Trace / fit readiness
    can_draw = bool(profile.get('can_draw_graph'))
    can_fit = _fr_is_yes_text(profile.get('can_use_for_fit'))
    can_r2 = _fr_is_yes_text(profile.get('can_calculate_r2_rms'))
    if can_draw:
        add('Trace', 'SV-TRACE-001', 'Trace/graph drawable', 'OK', profile.get('trace_availability'), 'Có thể vẽ graph/schematic', 'Có thể xem đồ thị hoặc sơ đồ event.', 'Không cần xử lý thêm.')
    else:
        add('Trace', 'SV-TRACE-001', 'Trace/graph drawable', 'WARN', profile.get('trace_availability'), 'Có thể vẽ graph/schematic', 'Không vẽ được trace/schematic đáng tin cậy.', 'Kiểm tra Raw Trace Diagnostics.')

    if diag.get('candidate_found'):
        if diag.get('distance_scale_ok') is True:
            add('Trace', 'SV-TRACE-002', 'Distance scale sanity', 'OK', diag.get('distance_scale_status') or diag.get('raw_status'), 'Scale km khớp span/rule correction', 'Trục km trace đủ tin cậy theo rule hiện tại.', 'Không cần xử lý thêm.', diag.get('distance_scale_note') or '')
        elif diag.get('distance_scale_ok') is False:
            add('Trace', 'SV-TRACE-002', 'Distance scale sanity', 'WARN', diag.get('distance_scale_status') or diag.get('raw_status'), 'Scale km khớp span/rule correction', 'Có trace nhưng trục km chưa đủ tin cậy để fit sâu.', 'Chỉ dùng xem nhanh; không dùng fit/R² nếu chưa kiểm tra.', diag.get('distance_scale_note') or diag.get('reject_reason') or '')
        else:
            add('Trace', 'SV-TRACE-002', 'Distance scale sanity', 'INFO', diag.get('distance_scale_status') or diag.get('raw_status'), 'Có thông tin scale', 'Chưa đủ dữ liệu kết luận scale.', 'Xem Raw Trace Diagnostics.')
    else:
        add('Trace', 'SV-TRACE-002', 'Distance scale sanity', 'INFO', 'Không có trace candidate', 'Không bắt buộc nếu chỉ đánh giá event', 'Không có trace để kiểm tra scale.', 'Dùng event/metadata nếu đủ, hoặc mở phần mềm hãng.')

    if can_fit and can_r2:
        add('Trace', 'SV-TRACE-003', 'Fit/R² readiness', 'OK', f'fit={profile.get("can_use_for_fit")}; r2={profile.get("can_calculate_r2_rms")}', 'Fit và R²/RMS khả dụng nếu section đủ điểm', 'Có thể xem Section Fit Quality để kết luận sâu.', 'Không cần xử lý thêm.')
    elif can_draw:
        add('Trace', 'SV-TRACE-003', 'Fit/R² readiness', 'WARN', f'fit={profile.get("can_use_for_fit")}; r2={profile.get("can_calculate_r2_rms")}', 'Fit/R² đầy đủ là tốt nhất', 'Graph có thể xem nhưng fit sâu hạn chế.', 'Kết luận section dựa trên Section Fit Quality và fallback note.')
    else:
        add('Trace', 'SV-TRACE-003', 'Fit/R² readiness', 'INFO', 'Không có trace fit', 'Fit/R² không bắt buộc nếu chỉ xuất event', 'Không có fit sâu.', 'Không dùng R²/RMS để kết luận.')

    # ORL judgement readiness
    orl_use = bool(getattr(orl, 'use_for_judgment', ctx.get('orl_use_for_judgment') or False)) if orl is not None else False
    orl_display = getattr(orl, 'display', ctx.get('orl_display') or '') if orl is not None else (ctx.get('orl_display') or '')
    orl_status = getattr(orl, 'advanced_status', ctx.get('orl_status') or '') if orl is not None else (ctx.get('orl_status') or '')
    if orl_use:
        add('ORL', 'SV-ORL-001', 'ORL measured/use-for-judgment', 'OK', f'{orl_display}; {orl_status}', 'Measured ORL và use_for_judgment=True', 'ORL có thể dùng kết luận theo ngưỡng.', 'Không cần xử lý thêm.')
    elif orl_display:
        add('ORL', 'SV-ORL-001', 'ORL measured/use-for-judgment', 'WARN', f'{orl_display}; {orl_status}', 'Measured ORL nếu muốn kết luận ORL', 'ORL hiện chỉ tham khảo/lower-bound/unknown, không phải kết luận measured chắc chắn.', 'Không dùng ORL để fail/pass chính nếu source không measured.')
    else:
        add('ORL', 'SV-ORL-001', 'ORL measured/use-for-judgment', 'INFO', 'Unknown', 'Measured ORL nếu cần đánh giá ORL', 'Không có ORL để kết luận.', 'Bỏ qua ORL hoặc kiểm tra bằng phần mềm hãng.')

    # Section fit health if it has already been precomputed.
    fit_rows = ctx.get('section_fit_rows') or []
    if fit_rows:
        total_fit = len(fit_rows)
        used_fit = 0
        fallback = 0
        low_conf = 0
        for fit in fit_rows:
            if getattr(fit, 'used_for_section', False):
                used_fit += 1
            if getattr(fit, 'fallback_method', ''):
                fallback += 1
            if str(getattr(fit, 'confidence', '') or '') in {'Thấp', 'Không đủ điểm', 'Không có raw trace', 'Không xác định', 'Trục km lệch span'}:
                low_conf += 1
        if used_fit == total_fit and low_conf == 0:
            add('Section', 'SV-SECTION-001', 'Section fit/fallback coverage', 'OK', f'used={used_fit}/{total_fit}; fallback={fallback}; low_conf={low_conf}', 'Tất cả section dùng fit/estimate đạt confidence', 'Section Fit Quality không phát hiện điểm nghẽn chính.', 'Không cần xử lý thêm.')
        elif used_fit > 0:
            add('Section', 'SV-SECTION-001', 'Section fit/fallback coverage', 'WARN', f'used={used_fit}/{total_fit}; fallback={fallback}; low_conf={low_conf}', 'Ít fallback/low confidence nhất có thể', 'Một số section đang dùng fallback/low confidence.', 'Khi kết luận section, xem cột Fit mode/Estimate level/R²/RMS.')
        else:
            add('Section', 'SV-SECTION-001', 'Section fit/fallback coverage', 'WARN', f'used={used_fit}/{total_fit}; fallback={fallback}; low_conf={low_conf}', 'Có section đủ fit/estimate', 'Không có section đủ điều kiện dùng fit/estimate tin cậy.', 'Không kết luận section sâu bằng trace; kiểm tra event/fallback.')
    else:
        add('Section', 'SV-SECTION-001', 'Section fit/fallback coverage', 'INFO', 'Chưa có section_fit_rows', 'Có khi xuất FastReporter Sections/fit', 'Không có dữ liệu fit section để strict-check.', 'Nếu cần kết luận section sâu, dùng FastReporter mode và xem Section Fit Quality.')

    return rows


def _fr_strict_validation_overall(rows: list[dict]) -> tuple[str, str, str, str]:
    max_rank = 0
    for r in rows:
        max_rank = max(max_rank, _fr_validation_level_rank(r.get('level')))
    if max_rank >= 3:
        return ('FAIL', 'Không', 'Không đủ điều kiện vận hành chính thức nếu chưa xử lý rule FAIL.', 'Mở sheet Strict Validation, xử lý các dòng FAIL trước khi dùng kết luận.')
    if max_rank >= 2:
        return ('WARN', 'Có điều kiện', 'Được dùng có điều kiện; phải đọc các cảnh báo trước khi chốt báo cáo.', 'Xử lý hoặc ghi chú các rule WARN quan trọng trong báo cáo.')
    return ('OK', 'Có', 'Đạt lớp kiểm soát strict validation hiện tại.', 'Có thể dùng theo logic tính toán hiện tại.')


def _fr_fill_strict_validation_sheet(
    ws,
    summaries: list[FileSummary],
    contexts: dict[str, dict],
    *,
    expected_route_km: Optional[float] = None,
    length_tolerance_km: float = 0.300,
    graph_reach_tolerance_km: Optional[float] = None,
    event_shortfall_tolerance_km: Optional[float] = None,
    duration_threshold_s: Optional[float] = None,
    skipped: Optional[list[str]] = None,
) -> None:
    headers = [
        'Tệp', 'Fiber/Core', 'Wavelength', 'Vendor family', 'Parser family',
        'Final parser status', 'Parser level', 'Parser cho phép kết luận?',
        'Overall strict status', 'Overall dùng kết luận?', 'Overall note',
        'Nhóm kiểm tra', 'Mã rule', 'Tên rule', 'Mức', 'Giá trị đọc được',
        'Rule kỳ vọng', 'Ảnh hưởng nếu sai', 'Hành động khuyến nghị', 'Bằng chứng/Ghi chú'
    ]
    _fr_write_header_row(ws, headers)
    row = 2
    for summary in summaries:
        ctx = contexts.get(summary.file_name) or {}
        validation_rows = _fr_strict_validation_rows_for_file(
            summary,
            ctx,
            expected_route_km=expected_route_km,
            length_tolerance_km=length_tolerance_km,
            graph_reach_tolerance_km=graph_reach_tolerance_km,
            event_shortfall_tolerance_km=event_shortfall_tolerance_km,
            duration_threshold_s=duration_threshold_s,
        )
        overall_status, overall_use, overall_note, overall_action = _fr_strict_validation_overall(validation_rows)
        for item in validation_rows:
            values = [
                item.get('file'), item.get('fiber'), item.get('wavelength'), item.get('vendor'), item.get('parser_family'),
                item.get('final_parser_status'), item.get('parser_level'), item.get('use_for_judgment'),
                overall_status, overall_use, overall_note,
                item.get('category'), item.get('rule_code'), item.get('rule_name'), item.get('level'), item.get('measured_value'),
                item.get('expected_rule'), item.get('impact'), item.get('action') or overall_action, item.get('evidence'),
            ]
            for c, value in enumerate(values, start=1):
                ws.cell(row, c, value)
            fill = _fr_validation_fill(item.get('level'))
            for c in range(1, len(headers) + 1):
                ws.cell(row, c).fill = fill
            row += 1
    for item in skipped or []:
        file_name, reason = (item.split(': ', 1) + [''])[:2] if ': ' in item else (item, '')
        values = [file_name, '', '', 'Unknown', '', 'FAIL_SKIPPED', 'FAIL', 'Không', 'FAIL', 'Không', 'File bị bỏ qua.', 'Input', 'SV-INPUT-001', 'Skipped/unsupported file', 'FAIL', reason, 'File phải đọc được bởi parser', 'Không có dữ liệu để tính.', 'Kiểm tra định dạng hoặc bổ sung parser.', reason]
        for c, value in enumerate(values, start=1):
            ws.cell(row, c, value)
            ws.cell(row, c).fill = LOG_ERROR_FILL
        row += 1
    widths = {
        'A': 34, 'B': 18, 'C': 12, 'D': 22, 'E': 34, 'F': 26, 'G': 14, 'H': 20,
        'I': 18, 'J': 20, 'K': 64, 'L': 16, 'M': 18, 'N': 34, 'O': 12,
        'P': 28, 'Q': 44, 'R': 70, 'S': 72, 'T': 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    _fr_selective_autofit(ws, max_scan_rows=max(row, 20), max_width=90)

def _stv_build_workbook(
    summaries: list[FileSummary],
    contexts: dict[str, dict],
    event_defs: list[dict],
    deviation_m: float,
    threshold_db: float,
    expected_route_km: Optional[float],
    jumper_excluded_m: float,
    graph_reach_tolerance_km: Optional[float],
    event_shortfall_tolerance_km: Optional[float],
    skipped: list[str],
    logs: Optional[list[dict]] = None,
    section_export_scope: str = 'all',
    section_measurement_mode: str = 'fit',
    section_event_source: str = 'all',
    section_boundary_priority: str = 'event',
    segment_start_km: Optional[float] = None,
    segment_end_km: Optional[float] = None,
    sections: Optional[list[dict]] = None,
    section_match_tolerance_m: float = 100.0,
    section_threshold_db: Optional[float] = None,
    duration_threshold_s: Optional[float] = None,
    stv_total_core: Optional[int] = None,
    stv_used_core: Optional[int] = None,
) -> BytesIO:
    wb = Workbook()
    ws_main = wb.active
    _stv_fill_main_sheet(
        ws_main,
        summaries,
        contexts,
        event_defs,
        deviation_m=deviation_m,
        threshold_db=threshold_db,
        expected_route_km=expected_route_km,
        jumper_excluded_m=jumper_excluded_m,
        graph_reach_tolerance_km=graph_reach_tolerance_km,
        event_shortfall_tolerance_km=event_shortfall_tolerance_km,
        stv_total_core=stv_total_core,
        stv_used_core=stv_used_core,
    )
    ws_graph = wb.create_sheet('Kiểm tra đồ thị')
    _stv_fill_graph_check_sheet(ws_graph, summaries, contexts)
    ws_skipped = wb.create_sheet('Tệp bỏ qua')
    _stv_fill_skipped_sheet(ws_skipped, skipped)
    ws_raw = wb.create_sheet('Sự kiện thô')
    _stv_fill_raw_events_sheet(ws_raw, summaries, contexts)
    if sections:
        _fr_precompute_section_fit_quality(summaries, contexts, sections, section_match_tolerance_m=section_match_tolerance_m, section_measurement_mode=section_measurement_mode)
        ws_fit = wb.create_sheet('Section Fit Quality')
        _fr_fill_section_fit_quality_sheet(ws_fit, summaries, contexts)
    ws_raw_diag = wb.create_sheet('Raw Trace Diagnostics')
    _fr_fill_raw_trace_diagnostics_sheet(ws_raw_diag, summaries, contexts)
    ws_orl = wb.create_sheet('ORL Analysis')
    _fr_fill_orl_analysis_sheet(ws_orl, summaries, contexts)
    ws_parser = wb.create_sheet('Parser Diagnostics')
    _fr_fill_parser_diagnostics_sheet(ws_parser, summaries, contexts, skipped)
    ws_vendor = wb.create_sheet('Vendor Compatibility')
    _fr_fill_vendor_compatibility_matrix_sheet(ws_vendor, summaries, contexts, skipped)
    ws_strict = wb.create_sheet('Strict Validation')
    _fr_fill_strict_validation_sheet(ws_strict, summaries, contexts, expected_route_km=expected_route_km, length_tolerance_km=length_tolerance_km if 'length_tolerance_km' in locals() else 0.300, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km, duration_threshold_s=duration_threshold_s if 'duration_threshold_s' in locals() else None, skipped=skipped)
    ws_core = wb.create_sheet('Core Metrics')
    _fr_fill_core_metrics_sheet(ws_core, summaries, contexts, threshold_db=threshold_db, section_pairs_by_file=None, duration_threshold_s=duration_threshold_s)
    ws_rules = wb.create_sheet('Output Rules')
    _fr_fill_output_rules_sheet(ws_rules, threshold_db=threshold_db, deviation_m=deviation_m, output_mode='stv', section_export_scope=section_export_scope, section_measurement_mode=section_measurement_mode, section_event_source=section_event_source, section_boundary_priority=section_boundary_priority, expected_route_km=expected_route_km, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km, segment_start_km=segment_start_km, segment_end_km=segment_end_km, section_threshold_db=section_threshold_db, duration_threshold_s=duration_threshold_s)
    _fr_apply_workbook_polish(wb)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_workbook_from_uploads(
    files: Iterable[tuple[str, bytes]],
    threshold_db: float = 0.5,
    section_threshold_db: Optional[float] = None,
    duration_threshold_s: Optional[float] = None,
    deviation_m: float = 100.0,
    expected_route_km: Optional[float] = None,
    jumper_excluded_m: float = 0.0,
    length_tolerance_km: float = 0.300,
    graph_reach_tolerance_km: Optional[float] = None,
    event_shortfall_tolerance_km: Optional[float] = None,
    overlength_tolerance_km: Optional[float] = None,
    segment_start_km: Optional[float] = None,
    segment_end_km: Optional[float] = None,
    section_export_scope: str = 'all',
    section_merge_tolerance_m: Optional[float] = None,
    section_min_length_km: float = 0.0,
    section_event_source: str = 'all',
    section_boundary_priority: str = 'event',
    section_allow_split: bool = False,
    section_match_tolerance_m: float = 100.0,
    section_measurement_mode: str = 'fit',
    orl_pass_threshold_db: float = 28.0,
    orl_source_mode: str = 'auto',
    orl_allow_lower_bound: bool = True,
    orl_lower_bound_status: str = 'Unknown',
    orl_physical_mode: str = 'disabled',
    output_mode: str = 'fastreporter',
    stv_total_core: Optional[int] = None,
    stv_used_core: Optional[int] = None,
) -> BytesIO:
    t0 = time.perf_counter()
    logs: list[dict] = []
    _fr_log(logs, 'run', 'INFO', 'Bắt đầu dựng workbook.')
    files = list(files)
    summaries, skipped, _file_payload_map, contexts = _fr_build_context(
        files,
        threshold_db=threshold_db,
        expected_route_km=expected_route_km,
        jumper_excluded_m=jumper_excluded_m,
        length_tolerance_km=length_tolerance_km,
        graph_reach_tolerance_km=graph_reach_tolerance_km,
        event_shortfall_tolerance_km=event_shortfall_tolerance_km,
        overlength_tolerance_km=overlength_tolerance_km,
        segment_start_km=segment_start_km,
        segment_end_km=segment_end_km,
        orl_pass_threshold_db=orl_pass_threshold_db,
        orl_source_mode=orl_source_mode,
        orl_allow_lower_bound=orl_allow_lower_bound,
        orl_lower_bound_status=orl_lower_bound_status,
        orl_physical_mode=orl_physical_mode,
        logs=logs,
    )
    if not summaries:
        message = ' | '.join(skipped) if skipped else 'Không có file hợp lệ nào được tải lên.'
        raise ValueError(message)

    event_defs = _fr_build_common_event_defs(contexts, deviation_m=deviation_m)
    _fr_log(logs, 'run', 'INFO', 'Phase 6: ORL nâng cao; phân biệt measured ORL / metadata lower-bound / not available, thêm ORL Analysis và không tự tính ORL vật lý nếu thiếu hiệu chuẩn.')
    sections = _fr_build_common_sections(event_defs, summaries, contexts, deviation_m=deviation_m, threshold_db=threshold_db, section_merge_tolerance_m=section_merge_tolerance_m, section_min_length_km=section_min_length_km, section_event_source=section_event_source, section_boundary_priority=section_boundary_priority, section_allow_split=section_allow_split)
    if str(section_export_scope).lower() == 'selected_range' and segment_start_km is not None and segment_end_km is not None:
        sections = _fr_clip_sections_to_range(sections, segment_start_km, segment_end_km)
        try:
            _fr_log(logs, 'run', 'INFO', f'Giới hạn xuất Sections theo đoạn người dùng chọn: {min(float(segment_start_km), float(segment_end_km)):.3f} - {max(float(segment_start_km), float(segment_end_km)):.3f} km | Số section sau khi cắt: {len(sections)}')
        except Exception:
            pass
    elif str(section_export_scope).lower() == 'selected_range':
        _fr_log(logs, 'run', 'WARN', 'Đã chọn chỉ xuất section theo đoạn đã chọn nhưng chưa nhập đủ Đoạn bắt đầu/Đoạn kết thúc. App giữ nguyên toàn bộ section của tuyến.')

    if str(output_mode).lower() == 'stv':
        _fr_log(logs, 'run', 'INFO', 'Xuất theo định dạng STV tinh chỉnh dùng chung lõi tính toán hiện tại.')
        return _stv_build_workbook(
            summaries,
            contexts,
            event_defs,
            deviation_m=deviation_m,
            threshold_db=threshold_db,
            expected_route_km=expected_route_km,
            jumper_excluded_m=jumper_excluded_m,
            graph_reach_tolerance_km=graph_reach_tolerance_km if graph_reach_tolerance_km is not None else length_tolerance_km,
            event_shortfall_tolerance_km=event_shortfall_tolerance_km if event_shortfall_tolerance_km is not None else length_tolerance_km,
            skipped=skipped,
            logs=logs,
            section_export_scope=section_export_scope,
            section_measurement_mode=section_measurement_mode,
            section_event_source=section_event_source,
            section_boundary_priority=section_boundary_priority,
            segment_start_km=segment_start_km,
            segment_end_km=segment_end_km,
            sections=sections,
            section_match_tolerance_m=section_match_tolerance_m,
            section_threshold_db=section_threshold_db,
            duration_threshold_s=duration_threshold_s,
            stv_total_core=stv_total_core,
            stv_used_core=stv_used_core,
        )

    template_path = Path(__file__).with_name('2.xlsx')
    if not template_path.exists():
        raise FileNotFoundError('Không tìm thấy template Excel (2.xlsx) trong thư mục chương trình.')

    wb = load_workbook(template_path)
    for warn in _fr_validate_template_or_raise(template_path, wb):
        _fr_log(logs, 'template', 'WARN', warn)
    _fr_log(logs, 'template', 'INFO', f'Template OK: {template_path.name}')

    # General Information (fill directly from per-file metadata, giữ nguyên layout template)
    ws_info = wb['General Information']
    _fr_fill_general_information_template(ws_info, summaries, contexts)

    # Sections first to get per-file section pairs
    section_pairs_by_file = _fr_fill_sections(wb['Sections'], summaries, contexts, sections, threshold_db, section_match_tolerance_m=section_match_tolerance_m, section_measurement_mode=section_measurement_mode, section_threshold_db=section_threshold_db)

    # Link results
    _fr_fill_link_results(wb['Link Results'], summaries, contexts, section_pairs_by_file, duration_threshold_s=duration_threshold_s)

    # Events
    _fr_fill_events(wb['Events'], summaries, contexts, event_defs, deviation_m, threshold_db)

    # Export all control parameters and the calculations they drive without changing core algorithms.
    for extra_name in ['App Parameters', 'Route Analysis', 'Segment Analysis', 'Segment Events', 'Section Fit Quality', 'Raw Trace Diagnostics', 'ORL Analysis', 'Parser Diagnostics', 'Vendor Compatibility', 'Strict Validation', 'Run Log', 'Core Metrics', 'Output Rules']:
        _fr_safe_sheet_remove(wb, extra_name)
    ws_params = wb.create_sheet('App Parameters')
    _fr_fill_app_parameters_sheet(
        ws_params,
        threshold_db=threshold_db,
        section_threshold_db=section_threshold_db,
        duration_threshold_s=duration_threshold_s,
        deviation_m=deviation_m,
        expected_route_km=expected_route_km,
        jumper_excluded_m=jumper_excluded_m,
        length_tolerance_km=length_tolerance_km,
        graph_reach_tolerance_km=graph_reach_tolerance_km,
        event_shortfall_tolerance_km=event_shortfall_tolerance_km,
        overlength_tolerance_km=overlength_tolerance_km,
        segment_start_km=segment_start_km,
        segment_end_km=segment_end_km,
        section_export_scope=section_export_scope,
        section_merge_tolerance_m=section_merge_tolerance_m,
        section_min_length_km=section_min_length_km,
        section_event_source=section_event_source,
        section_boundary_priority=section_boundary_priority,
        section_allow_split=section_allow_split,
        section_match_tolerance_m=section_match_tolerance_m,
        section_measurement_mode=section_measurement_mode,
        orl_pass_threshold_db=orl_pass_threshold_db,
        orl_source_mode=orl_source_mode,
        orl_allow_lower_bound=orl_allow_lower_bound,
        orl_lower_bound_status=orl_lower_bound_status,
        orl_physical_mode=orl_physical_mode,
    )
    ws_route = wb.create_sheet('Route Analysis')
    _fr_fill_route_analysis_sheet(ws_route, summaries, contexts)
    ws_segment = wb.create_sheet('Segment Analysis')
    _fr_fill_segment_analysis_sheet(ws_segment, summaries, contexts)
    ws_segment_events = wb.create_sheet('Segment Events')
    _fr_fill_segment_events_sheet(ws_segment_events, summaries, contexts)
    ws_fit = wb.create_sheet('Section Fit Quality')
    _fr_fill_section_fit_quality_sheet(ws_fit, summaries, contexts)
    ws_raw_diag = wb.create_sheet('Raw Trace Diagnostics')
    _fr_fill_raw_trace_diagnostics_sheet(ws_raw_diag, summaries, contexts)
    ws_orl = wb.create_sheet('ORL Analysis')
    _fr_fill_orl_analysis_sheet(ws_orl, summaries, contexts)
    ws_parser = wb.create_sheet('Parser Diagnostics')
    _fr_fill_parser_diagnostics_sheet(ws_parser, summaries, contexts, skipped)
    ws_vendor = wb.create_sheet('Vendor Compatibility')
    _fr_fill_vendor_compatibility_matrix_sheet(ws_vendor, summaries, contexts, skipped)
    ws_strict = wb.create_sheet('Strict Validation')
    _fr_fill_strict_validation_sheet(ws_strict, summaries, contexts, expected_route_km=expected_route_km, length_tolerance_km=length_tolerance_km if 'length_tolerance_km' in locals() else 0.300, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km, duration_threshold_s=duration_threshold_s if 'duration_threshold_s' in locals() else None, skipped=skipped)
    ws_core = wb.create_sheet('Core Metrics')
    _fr_fill_core_metrics_sheet(ws_core, summaries, contexts, threshold_db=threshold_db, section_pairs_by_file=section_pairs_by_file, duration_threshold_s=duration_threshold_s)
    ws_rules = wb.create_sheet('Output Rules')
    _fr_fill_output_rules_sheet(ws_rules, threshold_db=threshold_db, deviation_m=deviation_m, output_mode='fastreporter', section_export_scope=section_export_scope, section_measurement_mode=section_measurement_mode, section_event_source=section_event_source, section_boundary_priority=section_boundary_priority, expected_route_km=expected_route_km, graph_reach_tolerance_km=graph_reach_tolerance_km, event_shortfall_tolerance_km=event_shortfall_tolerance_km, segment_start_km=segment_start_km, segment_end_km=segment_end_km, section_threshold_db=section_threshold_db, duration_threshold_s=duration_threshold_s)
    ws_log = wb.create_sheet('Run Log')
    _fr_log(logs, 'run', 'INFO', f'Tổng file hợp lệ: {len(summaries)} | Event defs: {len(event_defs)} | Sections: {len(sections)}')
    _fr_log(logs, 'run', 'INFO', f'Thời gian dựng workbook: {time.perf_counter() - t0:.2f}s')
    _fr_fill_run_log_sheet(ws_log, logs, skipped)
    _fr_apply_workbook_polish(wb)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Phase 7.2 - Trace Viewer payload
# ---------------------------------------------------------------------------
def _phase72_safe_float(value, digits: int = 6):
    try:
        if value is None:
            return None
        f = float(value)
        if not math.isfinite(f):
            return None
        return round(f, digits)
    except Exception:
        return None


def _phase72_downsample_for_graph(x_values: list[float], y_values: list[float], max_points: int = 1600) -> tuple[list[float], list[float]]:
    """Return a lightweight graph series for the browser.

    This keeps Phase 7.2 responsive on phone/Tailscale while preserving the
    overall OTDR shape.  The original trace remains unchanged in the parser.
    """
    if not x_values or not y_values or len(x_values) != len(y_values):
        return [], []
    pairs = []
    for xx, yy in zip(x_values, y_values):
        try:
            x = float(xx); y = float(yy)
        except Exception:
            continue
        if math.isfinite(x) and math.isfinite(y):
            pairs.append((x, y))
    if not pairs:
        return [], []
    pairs.sort(key=lambda p: p[0])
    if len(pairs) <= max_points:
        return [round(p[0], 6) for p in pairs], [round(p[1], 6) for p in pairs]
    step = max(1, math.ceil(len(pairs) / max_points))
    sampled = pairs[::step]
    if sampled[-1] != pairs[-1]:
        sampled.append(pairs[-1])
    return [round(p[0], 6) for p in sampled], [round(p[1], 6) for p in sampled]


def _phase72_estimate_display_slope(x_values: list[float], y_values: list[float]) -> Optional[float]:
    """Estimate the display trend of a trace series.

    Some vendor mini/display curves store cumulative loss/amplitude in the
    opposite sign compared with the visual OTDR trace.  This helper is used only
    for browser display orientation; it does not change raw data used by the
    parser or Excel calculations.
    """
    pairs: list[tuple[float, float]] = []
    for xx, yy in zip(x_values or [], y_values or []):
        try:
            x = float(xx); y = float(yy)
        except Exception:
            continue
        if math.isfinite(x) and math.isfinite(y):
            pairs.append((x, y))
    if len(pairs) < 2:
        return None
    pairs.sort(key=lambda p: p[0])
    x0, x1 = pairs[0][0], pairs[-1][0]
    if abs(x1 - x0) < 1e-9:
        return None
    if len(pairs) < 4:
        return (pairs[-1][1] - pairs[0][1]) / (x1 - x0)
    x_mean = sum(p[0] for p in pairs) / len(pairs)
    y_mean = sum(p[1] for p in pairs) / len(pairs)
    den = sum((p[0] - x_mean) ** 2 for p in pairs)
    if den <= 1e-12:
        return None
    num = sum((p[0] - x_mean) * (p[1] - y_mean) for p in pairs)
    return num / den


def _phase72_orient_graph_series_downward(x_values: list[float], y_values: list[float]) -> tuple[list[float], str, str, str]:
    """Return a display series whose OTDR trend visually falls with distance.

    For a real OTDR trace, the displayed backscatter level should generally go
    downward as distance increases.  Some decoded mini-curves are stored as
    cumulative loss/normalized amplitude and therefore numerically rise with
    distance.  In that case we mirror the *display copy* of y around the first
    valid point.  This is intentionally display-only: section fit, Events, ORL
    and Excel calculations still use the original parsed values.
    """
    if not x_values or not y_values or len(x_values) != len(y_values):
        return y_values, 'unknown', 'Không đủ điểm để xác định chiều đồ thị.', 'Trace level'
    slope = _phase72_estimate_display_slope(x_values, y_values)
    finite_y = [float(y) for y in y_values if isinstance(y, (int, float)) and math.isfinite(float(y))]
    if slope is None or not finite_y:
        return y_values, 'unknown', 'Không đủ dữ liệu để xác định chiều đồ thị.', 'Trace level'
    y_range = max(finite_y) - min(finite_y)
    if y_range <= 1e-9:
        return y_values, 'flat', 'Đồ thị gần như phẳng, giữ nguyên chiều hiển thị.', 'Trace level'
    # Positive slope means the decoded display series rises with distance; OTDR
    # visual convention should fall with distance.  Use a small threshold to
    # avoid flipping nearly-flat noisy curves.
    if slope > max(1e-6, y_range * 1e-5 / max(max(x_values) - min(x_values), 1e-9)):
        anchor = float(y_values[0])
        oriented = [round(2 * anchor - float(y), 6) for y in y_values]
        return (
            oriented,
            'flipped_for_otdr_display',
            'Dữ liệu trace/mini-curve có xu hướng tăng theo km nên app đã đảo chiều hiển thị để đồ thị OTDR đi xuống. Việc này chỉ áp dụng cho đồ thị, không đổi dữ liệu tính toán.',
            'Trace level hiển thị',
        )
    return (
        [round(float(y), 6) for y in y_values],
        'native_downward',
        'Chiều dữ liệu phù hợp quy ước OTDR, giữ nguyên hiển thị.',
        'Trace level',
    )



def _phase72_build_preview_sections_for_file(
    summary: FileSummary,
    events: list[EventRow],
    *,
    threshold_db: float = 0.5,
    section_merge_tolerance_m: Optional[float] = None,
    section_min_length_km: float = 0.0,
    section_event_source: str = 'all',
    segment_start_km: Optional[float] = None,
    segment_end_km: Optional[float] = None,
) -> list[dict]:
    """Cheap per-file section boundaries for Trace Viewer preview.

    Full Excel export still uses the common-section pipeline.  The browser only
    needs visual boundaries, so this creates sections directly from the selected
    file's event positions and runs in near-linear time.
    """
    length = _phase72_safe_float(summary.length_km or summary.end_distance_km or summary.graph_end_km, 6)
    if length is None or length <= 0:
        dists = [float(e.distance_km) for e in events if e.distance_km is not None]
        length = max(dists) if dists else 1.0
    merge_tol_km = max(float(section_merge_tolerance_m or 80.0) / 1000.0, 0.001)
    min_len = max(float(section_min_length_km or 0.0), 0.0)
    use_filtered = str(section_event_source or '').lower() in {'filtered', 'important', 'significant'}
    boundaries: list[float] = [0.0]
    for e in events or []:
        if e.distance_km is None:
            continue
        try:
            d = float(e.distance_km)
        except Exception:
            continue
        if not math.isfinite(d) or d <= 0 or d >= float(length):
            continue
        important = (
            (e.loss_db is not None and abs(float(e.loss_db)) >= float(threshold_db))
            or (e.reflectance_db is not None)
            or ('end' in (e.event_type or '').lower())
            or ('reflect' in (e.event_type or '').lower())
        )
        if use_filtered and not important:
            continue
        boundaries.append(round(d, 6))
    boundaries.append(round(float(length), 6))
    boundaries = sorted(set(boundaries))
    merged: list[float] = []
    for b in boundaries:
        if not merged or abs(b - merged[-1]) > merge_tol_km:
            merged.append(b)
        else:
            # Keep the later boundary if it is the end of line, otherwise retain
            # the earlier one to avoid creating tiny visual sections.
            if abs(b - float(length)) <= 1e-6:
                merged[-1] = b
    sections: list[dict] = []
    for i in range(len(merged) - 1):
        st = float(merged[i]); en = float(merged[i + 1])
        if en <= st:
            continue
        if min_len and (en - st) < min_len and sections:
            sections[-1]['end_km'] = round(en, 4)
            sections[-1]['length_km'] = round(sections[-1]['end_km'] - sections[-1]['start_km'], 4)
            continue
        sections.append({'index': len(sections) + 1, 'start_km': round(st, 4), 'end_km': round(en, 4), 'length_km': round(en - st, 4)})
    if not sections:
        sections = [{'index': 1, 'start_km': 0.0, 'end_km': round(float(length), 4), 'length_km': round(float(length), 4)}]
    if segment_start_km is not None and segment_end_km is not None:
        return _fr_clip_sections_to_range(sections, segment_start_km, segment_end_km)
    return sections


def _phase72_build_event_section_schematic(summary: FileSummary, events: list[EventRow]) -> tuple[list[float], list[float]]:
    """Build a non-raw schematic when no usable x/y trace exists.

    This is intentionally labelled as schematic in the API.  It is only for
    visual navigation of event/section positions, not for raw-fit/R²/RMS.
    """
    length = _phase72_safe_float(summary.length_km or summary.end_distance_km or summary.graph_end_km, 6)
    if length is None or length <= 0:
        distances = [float(e.distance_km) for e in events if e.distance_km is not None]
        length = max(distances) if distances else 1.0
    base_att = None
    try:
        if summary.total_loss_db is not None and length and length > 0:
            base_att = float(summary.total_loss_db) / float(length)
    except Exception:
        base_att = None
    if base_att is None or not math.isfinite(base_att) or base_att <= 0 or base_att > 2.5:
        base_att = float(summary.attenuation_dbkm or 0.22 or 0.22)
        if base_att <= 0 or base_att > 2.5:
            base_att = 0.22
    important = sorted({0.0, float(length)} | {float(e.distance_km) for e in events if e.distance_km is not None and 0 <= float(e.distance_km) <= float(length)})
    x_out: list[float] = []
    y_out: list[float] = []
    cumulative_event_loss = 0.0
    event_loss_by_pos: dict[float, float] = {}
    for e in events:
        if e.distance_km is None or e.loss_db is None:
            continue
        try:
            d = round(float(e.distance_km), 6)
            loss = max(float(e.loss_db), 0.0)
        except Exception:
            continue
        if 0 <= d <= float(length):
            event_loss_by_pos[d] = event_loss_by_pos.get(d, 0.0) + loss
    for d in important:
        d = float(d)
        y_before = -(base_att * d + cumulative_event_loss)
        x_out.append(round(d, 6)); y_out.append(round(y_before, 6))
        loss_at = event_loss_by_pos.get(round(d, 6), 0.0)
        if loss_at > 0:
            cumulative_event_loss += loss_at
            x_out.append(round(d, 6)); y_out.append(round(-(base_att * d + cumulative_event_loss), 6))
    return x_out, y_out



